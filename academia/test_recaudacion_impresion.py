from datetime import date
from decimal import Decimal
from types import SimpleNamespace
from unittest.mock import Mock, patch
from io import BytesIO
from openpyxl import load_workbook
from django.test import SimpleTestCase
from academia.models import Abono
from academia.views_pagos import _construir_hoja_recaudacion, _build_recaudacion_excel_response


class ResumenImpresionTests(SimpleTestCase):
    @staticmethod
    def hoja_excel(cantidad):
        return {
            'curso': SimpleNamespace(nombre='Servicio técnico y microsoldadura'),
            'fecha': date(2026, 9, 5), 'dia_semana': 'SÁBADO',
            'ciudad': 'Guayaquil', 'jornada_label': 'Presencial – Sábados intensivos – 05/09/2026',
            'responsable': 'Responsable de prueba',
            'total_efectivo': Decimal('10'), 'total_transferencia_impresion': Decimal('15'),
            'total_payphone': Decimal('20'), 'total_recaudado': Decimal('45'),
            'items': [dict(estudiante=SimpleNamespace(nombre_completo=f'Estudiante Apellido {i}'),
                           modulo=1, cuota_sugerida=Decimal('25'), recaudado=Decimal('0'),
                           forma_pago='Efectivo', banco='N/A') for i in range(cantidad)],
        }

    def test_excel_paginas_completas_sin_perder_estudiantes_ni_duplicar_totales(self):
        hojas = [self.hoja_excel(11), self.hoja_excel(3)]
        response = _build_recaudacion_excel_response('prueba.xlsx', 'Recaudación', hojas)
        ws = load_workbook(BytesIO(response.content)).active
        self.assertEqual([b.id for b in ws.row_breaks.brk], [13, 35])
        self.assertEqual(ws['A17'].value, 11)
        nombres = [ws.cell(row, 2).value for row in range(1, ws.max_row + 1)]
        self.assertEqual(sum(str(v).startswith('Estudiante Apellido') for v in nombres), 14)
        self.assertEqual(nombres.count('TOTAL RECAUDADO'), 2)
        self.assertIsNone(ws['J30'].value)
        self.assertIsNone(ws['J31'].value)
        self.assertEqual(ws['B32'].value, 'OBSERVACIONES ADICIONALES')
        self.assertIn('B32:J35', {str(r) for r in ws.merged_cells.ranges})
        self.assertIn('$A$1:$J$13', str(ws.print_area))
        self.assertIn('$A$14:$J$35', str(ws.print_area))
        self.assertIn('$A$36:$J$57', str(ws.print_area))
        self.assertEqual(ws.page_setup.fitToHeight, 1)

    def test_excel_celeste_solo_deuda_vencida_y_blanco_para_pagados(self):
        hoja = self.hoja_excel(2)
        hoja['items'][0].update(pago_pendiente_fecha=True, recaudado=Decimal('5'))
        hoja['items'][1].update(pago_pendiente_fecha=False, recuperacion='Ya pagada', observaciones='retira')
        response = _build_recaudacion_excel_response('prueba.xlsx', 'Recaudación', [hoja])
        ws = load_workbook(BytesIO(response.content)).active
        for col in range(1, 11):
            self.assertEqual(ws.cell(4, col).fill.fgColor.rgb, 'FFBDEBFA')
            self.assertEqual(ws.cell(5, col).fill.fgColor.rgb, 'FFFFFFFF')

    def test_separa_payphone_en_pagos_mixtos_sin_alterar_total_anterior(self):
        abonos = [
            Abono(monto=Decimal('30'), metodo='efectivo', monto_2=Decimal('20'),
                  metodo_2='tarjeta', banco_2='payphone'),
            Abono(monto=Decimal('15'), metodo='transferencia', banco='guayaquil'),
        ]
        matricula = SimpleNamespace(
            abonos=Mock(), estudiante=SimpleNamespace(nombre_completo='Estudiante de prueba'),
            pk=1, observaciones='', talla_camiseta='', jornada=None, registrado_por_id=None,
        )
        matricula.abonos.filter.return_value = abonos
        plan = {'modulo': 1, 'cuota_sugerida': Decimal('25'), 'saldo_modulo': Decimal('50')}
        with patch('academia.views_pagos._plan_recaudacion_matricula', return_value=plan), \
             patch('academia.views_pagos._modulo_recaudacion_label', return_value='Módulo 1'), \
             patch('academia.views_pagos._recuperacion_recaudacion_label', return_value=''):
            hoja = _construir_hoja_recaudacion(SimpleNamespace(nombre='Curso'), [matricula], date(2026, 9, 5))
        self.assertEqual(hoja['total_efectivo'], Decimal('10'))
        self.assertEqual(hoja['total_transferencia_impresion'], Decimal('15'))
        self.assertEqual(hoja['total_payphone'], Decimal('20'))
        self.assertEqual(hoja['total_recaudado'], Decimal('45'))
        self.assertEqual(hoja['total_transferencia'], Decimal('35'))
