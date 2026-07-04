"""
Registro Administrativo: dashboard financiero.

Solo accesible por administradores. Muestra:
- KPIs del mes (ingresos, egresos, balance)
- Gráfico de barras: ingresos vs egresos por mes (últimos 6 meses)
- Top egresos del mes por categoría
- Movimientos recientes (egresos + ingresos mezclados)
- CRUD de egresos
"""
import calendar
import json
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.sessions.models import Session
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.text import slugify
from django.utils import timezone
from django.views.decorators.http import require_POST

from .forms import EgresoForm
from .models import (
    Abono, AbonoArchivado, Adicional, AdicionalArchivado, CategoriaEgreso,
    CierreAdministrativo, CierreCurso, Comprobante, Egreso, Matricula, Curso,
)
from .permisos import admin_requerido


MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]


def _rango_mes(anio, mes):
    """Devuelve (primer_dia, ultimo_dia) del mes dado."""
    primer = date(anio, mes, 1)
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    ultimo = date(anio, mes, ultimo_dia)
    return primer, ultimo


# Minutos sin actividad tras los cuales dejamos de considerar a alguien "en línea".
EN_LINEA_MINUTOS = 5


def _usuarios_actividad():
    """Lista de usuarios con su última conexión y si están en línea.

    - "Última conexión": la más reciente entre el `last_login` de Django y la
      última actividad registrada en su sesión (middleware UltimaActividad).
    - "En línea": tuvo actividad en los últimos EN_LINEA_MINUTOS minutos y su
      sesión sigue vigente.
    """
    User = get_user_model()
    ahora = timezone.now()

    # Mapear cada usuario con su última actividad vista en sesiones vigentes.
    actividad_por_uid = {}
    for ses in Session.objects.filter(expire_date__gte=ahora):
        try:
            data = ses.get_decoded()
        except Exception:
            continue
        uid = data.get('_auth_user_id')
        if not uid:
            continue
        try:
            uid = int(uid)
        except (TypeError, ValueError):
            continue
        marca = data.get('ultima_actividad')
        dt = None
        if marca:
            try:
                dt = datetime.fromisoformat(marca)
            except (ValueError, TypeError):
                dt = None
        actual = actividad_por_uid.get(uid)
        if actual is None or (dt and (actual is None or dt > actual)):
            actividad_por_uid[uid] = dt or actual

    usuarios = []
    for u in User.objects.filter(is_active=True):
        act = actividad_por_uid.get(u.id)
        # Última conexión = max(last_login, última actividad de sesión)
        candidatos = [d for d in (u.last_login, act) if d is not None]
        ultima_conexion = max(candidatos) if candidatos else None
        en_linea = bool(act and (ahora - act).total_seconds() <= EN_LINEA_MINUTOS * 60)
        nombre = (f'{u.first_name} {u.last_name}'.strip() or u.username)
        usuarios.append({
            'id': u.id,
            'nombre': nombre,
            'username': u.username,
            'es_admin': u.is_superuser or u.is_staff,
            'ultima_conexion': ultima_conexion,
            'en_linea': en_linea,
        })

    # Orden: en línea primero, luego por última conexión más reciente.
    usuarios.sort(
        key=lambda x: (
            not x['en_linea'],
            -(x['ultima_conexion'].timestamp() if x['ultima_conexion'] else 0),
        )
    )
    return usuarios


def _ingresos_periodo(desde, hasta):
    """
    Suma todos los ingresos en el rango de fechas.
    Combina:
      - Abonos (de matrículas registradas en el sistema): por fecha de abono.
      - Abonos ARCHIVADOS (de cursos cerrados): por fecha de abono. Esto es
        clave: cuando se cierra un curso, los Abonos vivos se borran, pero el
        dinero que entró NO debe desaparecer del registro administrativo. Por
        eso sumamos también los AbonoArchivado del periodo.
      - Comprobantes "manuales" (cargados desde el módulo Comprobantes sin
        crear matrícula): por fecha de inscripción, solo el pago_abono.
        IMPORTANTE: se EXCLUYEN los comprobantes vinculados a una matrícula,
        porque su pago_abono refleja el valor_pagado de esa matrícula y ese
        dinero ya está contado en `abonos`. Si no se excluyera, cada Abono
        contaría doble (una vez en `abonos` y otra en `ventas`).
      - Adicionales (certificados, examen supletorio, camisas extra): por fecha.
    """
    abonos_qs = Abono.objects.filter(fecha__gte=desde, fecha__lte=hasta).select_related('matricula__estudiante', 'matricula__curso')
    abonos = abonos_qs.aggregate(s=Sum('monto'))['s'] or Decimal('0.00')

    # Abonos archivados (de cierres de curso) — conservan su fecha original
    abonos_archivados_qs = AbonoArchivado.objects.filter(fecha__gte=desde, fecha__lte=hasta)
    abonos_archivados = abonos_archivados_qs.aggregate(s=Sum('monto'))['s'] or Decimal('0.00')

    # Solo comprobantes manuales (sin matrícula vinculada).
    # Los comprobantes-espejo de matrículas ya están contados via Abonos.
    ventas_qs = Comprobante.objects.filter(
        fecha_inscripcion__gte=desde, fecha_inscripcion__lte=hasta,
        matricula__isnull=True,
    )
    ventas = ventas_qs.aggregate(s=Sum('pago_abono'))['s'] or Decimal('0.00')

    adicionales_vivos_qs = Adicional.objects.filter(fecha__gte=desde, fecha__lte=hasta).select_related('estudiante')
    adicionales_vivos = adicionales_vivos_qs.aggregate(s=Sum('valor'))['s'] or Decimal('0.00')
    
    adicionales_archivados_qs = AdicionalArchivado.objects.filter(fecha__gte=desde, fecha__lte=hasta)
    adicionales_archivados = adicionales_archivados_qs.aggregate(s=Sum('valor'))['s'] or Decimal('0.00')
    
    adicionales = adicionales_vivos + adicionales_archivados

    return {
        'abonos': abonos,
        'abonos_archivados': abonos_archivados,
        'ventas': ventas,
        'adicionales': adicionales,
        'adicionales_vivos': adicionales_vivos,
        'adicionales_archivados': adicionales_archivados,
        'total': abonos + abonos_archivados + ventas + adicionales,
        'qs_abonos': abonos_qs.order_by('-fecha'),
        'qs_abonos_archivados': abonos_archivados_qs.order_by('-fecha'),
        'qs_ventas': ventas_qs.order_by('-fecha_inscripcion'),
        'qs_adicionales_vivos': adicionales_vivos_qs.order_by('-fecha'),
        'qs_adicionales_archivados': adicionales_archivados_qs.order_by('-fecha'),
    }


def _cierres_curso_periodo(desde, hasta):
    """
    Devuelve los cierres de curso ejecutados dentro del periodo, con sus totales.
    Sirve para reflejar en el dashboard administrativo qué cursos se cerraron
    y cuánto dinero representaron.
    """
    qs = CierreCurso.objects.filter(
        fecha_cierre__date__gte=desde, fecha_cierre__date__lte=hasta
    ).order_by('-fecha_cierre')

    cierres = list(qs)
    total_cobrado = sum((c.total_cobrado for c in cierres), Decimal('0.00'))
    total_facturado = sum((c.total_facturado for c in cierres), Decimal('0.00'))
    total_matriculas = sum((c.total_matriculas for c in cierres), 0)

    return {
        'cierres': cierres,
        'count': len(cierres),
        'total_cobrado': total_cobrado,
        'total_facturado': total_facturado,
        'total_matriculas': total_matriculas,
    }


def _adicionales_periodo(desde, hasta):
    """
    Estadísticas de Adicionales registrados en el rango.
    Incluye activos y archivados para que el dashboard no pierda ingresos
    después de ejecutar un cierre adicional.
    """
    qs = Adicional.objects.filter(fecha__gte=desde, fecha__lte=hasta)
    arch_qs = AdicionalArchivado.objects.filter(fecha__gte=desde, fecha__lte=hasta)

    total_vivos = qs.aggregate(s=Sum('valor'))['s'] or Decimal('0.00')
    total_archivados = arch_qs.aggregate(s=Sum('valor'))['s'] or Decimal('0.00')
    total = total_vivos + total_archivados
    count = qs.count() + arch_qs.count()
    interno = (
        (qs.filter(estudiante__isnull=False).aggregate(s=Sum('valor'))['s'] or Decimal('0.00'))
        + (arch_qs.filter(origen_label='Estudiante interno').aggregate(s=Sum('valor'))['s'] or Decimal('0.00'))
    )
    externo = (
        (qs.filter(persona_externa__isnull=False).aggregate(s=Sum('valor'))['s'] or Decimal('0.00'))
        + (arch_qs.filter(origen_label='Persona externa').aggregate(s=Sum('valor'))['s'] or Decimal('0.00'))
    )

    # Desglose por tipo
    por_tipo_qs = (qs.values('tipo_adicional')
                     .annotate(total=Sum('valor'), count=Count('id'))
                     .order_by('-total'))
    por_tipo_arch_qs = (
        arch_qs.values('tipo_adicional')
        .annotate(total=Sum('valor'), count=Count('id'))
        .order_by('-total')
    )
    tipos_dict = {t[0]: t[1] for t in Adicional.TIPOS_ADICIONAL}
    por_tipo_map = {}
    for r in por_tipo_qs:
        codigo = r['tipo_adicional']
        por_tipo_map[codigo] = {
            'codigo': codigo,
            'label': tipos_dict.get(codigo, codigo),
            'total': r['total'] or Decimal('0.00'),
            'count': r['count'] or 0,
        }
    for r in por_tipo_arch_qs:
        codigo = r['tipo_adicional']
        item = por_tipo_map.setdefault(codigo, {
            'codigo': codigo,
            'label': tipos_dict.get(codigo, codigo),
            'total': Decimal('0.00'),
            'count': 0,
        })
        item['total'] += r['total'] or Decimal('0.00')
        item['count'] += r['count'] or 0
    por_tipo = sorted(por_tipo_map.values(), key=lambda item: item['total'], reverse=True)

    return {
        'total': total,
        'activos': total_vivos,
        'archivados': total_archivados,
        'count': count,
        'interno': interno,
        'externo': externo,
        'por_tipo': por_tipo,
    }


def _adicionales_archivados_periodo(desde, hasta):
    """Adicionales movidos al archivo durante el mes seleccionado."""
    qs = AdicionalArchivado.objects.filter(
        archivado_en__date__gte=desde,
        archivado_en__date__lte=hasta,
    )
    total = qs.aggregate(s=Sum('valor'))['s'] or Decimal('0.00')
    interno = qs.filter(origen_label='Estudiante interno').aggregate(s=Sum('valor'))['s'] or Decimal('0.00')
    externo = qs.filter(origen_label='Persona externa').aggregate(s=Sum('valor'))['s'] or Decimal('0.00')
    return {
        'total': total,
        'count': qs.count(),
        'interno': interno,
        'externo': externo,
    }


# ─────────────────────────────────────────────────────────
# Análisis por TIPO DE PAGO (Abono / Pago Completo / Por
# Módulo / Clase de Recuperación)  ─ usado en gráficos
# ─────────────────────────────────────────────────────────

# Etiquetas oficiales y colores para cada tipo de pago.
# Usamos los mismos códigos de Abono.TIPOS_PAGO.
TIPOS_PAGO_INFO = [
    ('abono',          'Abono',                 '#1a237e'),  # azul
    ('pago_completo',  'Pago Completo',         '#2e7d32'),  # verde
    ('por_modulo',     'Por Módulo',            '#f0ad4e'),  # naranja
    ('solo_modulo',    'Solo Módulo',           '#7e57c2'),  # morado
    ('recuperacion',   'Clase de Recuperación', '#c62828'),  # rojo
]


def _tipos_pago_periodo(desde, hasta):
    """
    Devuelve un dict {codigo_tipo: {'label', 'total', 'count', 'color'}}
    con la suma y conteo de abonos por tipo de pago en el rango.

    Solo se cuentan abonos cuyo cuenta_para_saldo=True para coherencia con
    los ingresos del mes (las recuperaciones cobradas aparte se reportan
    por separado en `recuperaciones_aparte`).
    """
    qs = (Abono.objects
          .filter(fecha__gte=desde, fecha__lte=hasta)
          .values('tipo_pago')
          .annotate(total=Sum('monto'), count=Count('id')))

    base = {
        codigo: {
            'codigo': codigo,
            'label': label,
            'color': color,
            'total': Decimal('0.00'),
            'count': 0,
        }
        for codigo, label, color in TIPOS_PAGO_INFO
    }
    for r in qs:
        c = r['tipo_pago']
        if c in base:
            base[c]['total'] = r['total'] or Decimal('0.00')
            base[c]['count'] = r['count'] or 0
    return base


def _recuperaciones_periodo(desde, hasta):
    """
    Devuelve estadísticas de clases de recuperación cobradas en el rango.
    Suma TODOS los abonos tipo recuperación (cuenten o no para saldo) para
    dar visibilidad real a cuánto se está facturando por recuperaciones.
    """
    qs = Abono.objects.filter(
        tipo_pago='recuperacion',
        fecha__gte=desde, fecha__lte=hasta,
    )
    total = qs.aggregate(s=Sum('monto'))['s'] or Decimal('0.00')
    cuentan = qs.filter(cuenta_para_saldo=True).aggregate(s=Sum('monto'))['s'] or Decimal('0.00')
    aparte = qs.filter(cuenta_para_saldo=False).aggregate(s=Sum('monto'))['s'] or Decimal('0.00')
    return {
        'total': total,
        'cuentan_para_saldo': cuentan,
        'aparte': aparte,
        'count': qs.count(),
    }


def _egresos_periodo(desde, hasta):
    """Suma egresos en el rango."""
    return Egreso.objects.filter(
        fecha__gte=desde, fecha__lte=hasta
    ).aggregate(s=Sum('monto'))['s'] or Decimal('0.00')


def _retiros_periodo(desde, hasta):
    """Suma la deuda perdonada de las matrículas en retiro voluntario en el rango."""
    retiros_qs = Matricula.objects.filter(
        estado='retiro_voluntario',
        fecha_matricula__gte=desde, 
        fecha_matricula__lte=hasta
    )
    total = Decimal('0.00')
    for r in retiros_qs:
        total += (r.valor_curso or Decimal('0.00')) - (r.valor_pagado or Decimal('0.00'))
    return total


def _egresos_por_categoria_periodo(desde, hasta):
    """Devuelve lista [{categoria, total, color, icono}, …]."""
    qs = (Egreso.objects
          .filter(fecha__gte=desde, fecha__lte=hasta)
          .values('categoria__id', 'categoria__nombre',
                  'categoria__color', 'categoria__icono')
          .annotate(total=Sum('monto'))
          .order_by('-total'))
    return [
        {
            'id': r['categoria__id'],
            'nombre': r['categoria__nombre'],
            'color': r['categoria__color'],
            'icono': r['categoria__icono'],
            'total': r['total'] or Decimal('0.00'),
        }
        for r in qs
    ]




@admin_requerido
def admin_dashboard(request):
    """Panel principal del Registro Administrativo."""
    hoy = timezone.localdate()

    # Permite dos modos de periodo:
    #  1) Rango personalizado (estilo extracto bancario):
    #     ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD
    #  2) Mes/Año (compatibilidad anterior): ?anio=2026&mes=4
    desde_str = (request.GET.get('desde') or '').strip()
    hasta_str = (request.GET.get('hasta') or '').strip()
    modo_rango = False
    desde = hasta = None

    if desde_str and hasta_str:
        try:
            desde = date.fromisoformat(desde_str)
            hasta = date.fromisoformat(hasta_str)
            if hasta < desde:
                desde, hasta = hasta, desde
            modo_rango = True
        except ValueError:
            desde = hasta = None
            modo_rango = False

    if modo_rango:
        # El mes/año de referencia (para series y cierres) se toma del fin de rango.
        anio, mes = hasta.year, hasta.month
    else:
        try:
            anio = int(request.GET.get('anio', hoy.year))
            mes = int(request.GET.get('mes', hoy.month))
            if not (1 <= mes <= 12):
                mes = hoy.month
        except (TypeError, ValueError):
            anio, mes = hoy.year, hoy.month
        desde, hasta = _rango_mes(anio, mes)

    # ── Datos del mes seleccionado ──
    ingresos = _ingresos_periodo(desde, hasta)
    egresos_total = _egresos_periodo(desde, hasta)
    balance = ingresos['total'] - egresos_total
    top_categorias = _egresos_por_categoria_periodo(desde, hasta)

    # ── NUEVO: Tipos de pago del mes (Abono / Pago Completo / Por Módulo / Recuperación) ──
    tipos_pago_mes_dict = _tipos_pago_periodo(desde, hasta)
    # Lista ordenada (mismo orden que TIPOS_PAGO_INFO) para iterar en el template
    tipos_pago_mes = [tipos_pago_mes_dict[c] for c, _l, _col in TIPOS_PAGO_INFO]
    total_tipos_pago_mes = sum(
        (x['total'] for x in tipos_pago_mes), Decimal('0.00')
    )
    # Datos para el gráfico circular (pie chart) por mes
    pie_tipos_pago = [
        {
            'codigo': x['codigo'],
            'label': x['label'],
            'color': x['color'],
            'total': float(x['total']),
            'count': x['count'],
        }
        for x in tipos_pago_mes
    ]

    # ── NUEVO: Estadísticas de Clases de Recuperación del mes ──
    recuperaciones_mes = _recuperaciones_periodo(desde, hasta)

    # ── NUEVO: Estadísticas de Adicionales del mes ──
    adicionales_mes = _adicionales_periodo(desde, hasta)
    adicionales_archivados_mes = _adicionales_archivados_periodo(desde, hasta)

    # ── Comparación con periodo anterior ──
    if modo_rango:
        # Periodo anterior de igual duración, justo antes del rango actual.
        dur = (hasta - desde).days + 1
        hasta_prev = desde - timedelta(days=1)
        desde_prev = hasta_prev - timedelta(days=dur - 1)
    elif mes == 1:
        desde_prev, hasta_prev = _rango_mes(anio - 1, 12)
    else:
        desde_prev, hasta_prev = _rango_mes(anio, mes - 1)
    ing_prev = _ingresos_periodo(desde_prev, hasta_prev)['total']
    egr_prev = _egresos_periodo(desde_prev, hasta_prev)
    bal_prev = ing_prev - egr_prev
    recup_prev = _recuperaciones_periodo(desde_prev, hasta_prev)['total']
    adic_prev = _adicionales_periodo(desde_prev, hasta_prev)['total']

    def variacion(actual, anterior):
        if anterior == 0:
            return None
        return float(((actual - anterior) / abs(anterior)) * 100)

    # ── Datos para gráfico (últimos 6 meses) ──
    # Ahora incluimos también el monto facturado por clases de recuperación
    # y el desglose por tipo de pago para el gráfico circular por mes.
    serie_meses = []
    for i in range(5, -1, -1):
        # contar i meses hacia atrás desde el mes seleccionado
        m = mes - i
        a = anio
        while m <= 0:
            m += 12
            a -= 1
        d, h = _rango_mes(a, m)
        ing = _ingresos_periodo(d, h)['total']
        egr = _egresos_periodo(d, h)
        ret = _retiros_periodo(d, h)
        rec = _recuperaciones_periodo(d, h)['total']
        adic = _adicionales_periodo(d, h)['total']
        tp_dict = _tipos_pago_periodo(d, h)
        tipos_pago_mes_serie = [
            {
                'codigo': c,
                'label': tp_dict[c]['label'],
                'color': tp_dict[c]['color'],
                'total': float(tp_dict[c]['total']),
                'count': tp_dict[c]['count'],
            }
            for c, _l, _col in TIPOS_PAGO_INFO
        ]
        total_tp = sum((x['total'] for x in tipos_pago_mes_serie), 0.0)
        serie_meses.append({
            'label': f'{MESES_ES[m][:3]} {a}',
            'mes_nombre': f'{MESES_ES[m]} {a}',
            'ingresos': float(ing),
            'egresos': float(egr),
            'retiros': float(ret),
            'recuperaciones': float(rec),
            'adicionales': float(adic),
            'balance': float(ing - egr),
            'tipos_pago': tipos_pago_mes_serie,
            'total_tipos_pago': total_tp,
        })

    # ── Total acumulado histórico (todo el sistema) ──
    total_abonos_hist = Abono.objects.aggregate(s=Sum('monto'))['s'] or Decimal('0.00')
    # Sumar también los abonos archivados (de cursos cerrados) para que el
    # histórico financiero no baje al ejecutar cierres de curso.
    total_abonos_arch_hist = AbonoArchivado.objects.aggregate(s=Sum('monto'))['s'] or Decimal('0.00')
    # Mismo criterio que `_ingresos_periodo`: solo comprobantes manuales
    # (sin matrícula vinculada) para no duplicar lo que ya está en Abonos.
    total_ventas_hist = Comprobante.objects.filter(
        matricula__isnull=True
    ).aggregate(s=Sum('pago_abono'))['s'] or Decimal('0.00')
    total_adicionales_hist = (
        (Adicional.objects.aggregate(s=Sum('valor'))['s'] or Decimal('0.00'))
        + (AdicionalArchivado.objects.aggregate(s=Sum('valor'))['s'] or Decimal('0.00'))
    )
    total_egresos_hist = Egreso.objects.aggregate(s=Sum('monto'))['s'] or Decimal('0.00')
    total_ingresos_hist = (
        total_abonos_hist + total_abonos_arch_hist
        + total_ventas_hist + total_adicionales_hist
    )
    balance_hist = total_ingresos_hist - total_egresos_hist

    # Por cobrar: saldos pendientes (informativo, no se cuenta como ingreso).
    # Se compone de dos fuentes que no se solapan:
    #   - Saldos pendientes de matrículas activas (la fuente de verdad para
    #     matrículas registradas en el sistema).
    #   - Diferencias de comprobantes manuales (sin matrícula vinculada).
    por_cobrar_matriculas = Decimal('0.00')
    for m_ in Matricula.objects.exclude(estado='retiro_voluntario'):
        s = m_.saldo
        if s > 0:
            por_cobrar_matriculas += s
    por_cobrar_comprobantes_manuales = Comprobante.objects.filter(
        matricula__isnull=True
    ).aggregate(s=Sum('diferencia'))['s'] or Decimal('0.00')
    por_cobrar_comprobantes = por_cobrar_matriculas + por_cobrar_comprobantes_manuales

    # Retiros Voluntarios (acumulado histórico de saldo perdonado)
    # Calculamos la diferencia entre el valor del curso y lo que pagaron de las matrículas en retiro
    retiros_qs = Matricula.objects.filter(estado='retiro_voluntario')
    total_retiros = Decimal('0.00')
    for r in retiros_qs:
        vc = r.valor_curso or Decimal('0.00')
        vp = r.valor_pagado or Decimal('0.00')
        total_retiros += (vc - vp)

    # ── Movimientos recientes del mes (últimos 10 egresos) ──
    egresos_recientes = (Egreso.objects
                         .filter(fecha__gte=desde, fecha__lte=hasta)
                         .select_related('categoria', 'registrado_por')
                         .order_by('-fecha', '-creado')[:10])

    # Lista de meses para el selector (últimos 24 meses)
    meses_selector = []
    for i in range(0, 24):
        m = hoy.month - i
        a = hoy.year
        while m <= 0:
            m += 12
            a -= 1
        meses_selector.append({
            'anio': a, 'mes': m,
            'label': f'{MESES_ES[m]} {a}',
            'seleccionado': (a == anio and m == mes),
        })

    # ── Cierres de CURSO del periodo (reflejar su dinero en el panel admin) ──
    cierres_curso_mes = _cierres_curso_periodo(desde, hasta)

    # ── Cierres ADMINISTRATIVOS previos (cortes de caja ya hechos) ──
    cierres_admin = CierreAdministrativo.objects.all()[:12]
    # ¿Ya existe un cierre administrativo para este mes?
    cierre_admin_existente = CierreAdministrativo.objects.filter(
        anio=anio, mes=mes
    ).order_by('-fecha_cierre').first()
    tiene_valores_corte = ingresos['total'] != Decimal('0.00') or egresos_total != Decimal('0.00')

    # ── Listas detalladas para modales ──
    qs_egresos_mes = Egreso.objects.filter(fecha__gte=desde, fecha__lte=hasta).select_related('categoria', 'registrado_por').order_by('-fecha', '-creado')
    
    qs_retiros_mes = Matricula.objects.filter(estado='retiro_voluntario', fecha_matricula__gte=desde, fecha_matricula__lte=hasta).select_related('estudiante', 'curso')
    retiros_detalle = []
    for r in qs_retiros_mes:
        saldo_perdonado = (r.valor_curso or Decimal('0.00')) - (r.valor_pagado or Decimal('0.00'))
        retiros_detalle.append({'matricula': r, 'saldo_perdonado': saldo_perdonado})

    qs_recuperaciones_mes = Abono.objects.filter(tipo_pago='recuperacion', fecha__gte=desde, fecha__lte=hasta).select_related('matricula__estudiante', 'matricula__curso').order_by('-fecha')

    # Por cobrar detalle (matriculas activas con saldo y comprobantes manuales con diferencia)
    por_cobrar_matriculas_detalle = [
        m_ for m_ in Matricula.objects.exclude(estado='retiro_voluntario').select_related('estudiante', 'curso')
        if m_.saldo > 0
    ]
    por_cobrar_comprobantes_detalle = Comprobante.objects.filter(matricula__isnull=True).exclude(diferencia=0)

    # ── NUEVO: Lista combinada de todos los movimientos (Ingresos y Egresos) ──
    movimientos_rango = []
    
    for e in qs_egresos_mes:
        movimientos_rango.append({
            'fecha': e.fecha,
            'tipo': 'egreso',
            'categoria': f"Egreso: {e.categoria.nombre}" if e.categoria else 'Egreso',
            'concepto': e.concepto,
            'involucrado': e.registrado_por.get_full_name() if e.registrado_por else 'N/A',
            'monto': e.monto
        })
        
    for a in ingresos['qs_abonos']:
        curso = a.matricula.curso.nombre if a.matricula and a.matricula.curso else 'Sin curso'
        est = a.matricula.estudiante if a.matricula else None
        est_nombre = f"{est.nombres}" if est else 'N/A'
        movimientos_rango.append({
            'fecha': a.fecha.date() if isinstance(a.fecha, datetime) else a.fecha,
            'tipo': 'ingreso',
            'categoria': f"Abono ({a.get_tipo_pago_display()})",
            'concepto': f"{curso} - {a.observaciones}" if a.observaciones else curso,
            'involucrado': est_nombre,
            'monto': a.monto
        })
        
    for a in ingresos['qs_abonos_archivados']:
        movimientos_rango.append({
            'fecha': a.fecha.date() if isinstance(a.fecha, datetime) else a.fecha,
            'tipo': 'ingreso',
            'categoria': 'Abono (Archivado)',
            'concepto': a.matricula_archivada.curso_nombre if hasattr(a, 'matricula_archivada') and a.matricula_archivada else 'Curso Archivado',
            'involucrado': f"{a.matricula_archivada.nombres}" if hasattr(a, 'matricula_archivada') and a.matricula_archivada else "N/A",
            'monto': a.monto
        })
        
    for v in ingresos['qs_ventas']:
        movimientos_rango.append({
            'fecha': v.fecha_inscripcion.date() if isinstance(v.fecha_inscripcion, datetime) else v.fecha_inscripcion,
            'tipo': 'ingreso',
            'categoria': 'Venta (Comprobante)',
            'concepto': v.concepto_pago or 'Comprobante manual',
            'involucrado': v.nombre_estudiante or 'N/A',
            'monto': v.pago_abono
        })
        
    for ad in ingresos['qs_adicionales_vivos']:
        if ad.estudiante:
            inv = f"{ad.estudiante.nombres}"
        elif ad.persona_externa:
            inv = f"{ad.persona_externa.nombres}"
        else:
            inv = "Desconocido"
        movimientos_rango.append({
            'fecha': ad.fecha,
            'tipo': 'ingreso',
            'categoria': 'Adicional',
            'concepto': ad.get_tipo_adicional_display() + (f" ({ad.observaciones})" if ad.observaciones else ""),
            'involucrado': inv,
            'monto': ad.valor
        })
        
    for ad in ingresos['qs_adicionales_archivados']:
        movimientos_rango.append({
            'fecha': ad.fecha,
            'tipo': 'ingreso',
            'categoria': 'Adicional (Archivado)',
            'concepto': ad.tipo_adicional_label,
            'involucrado': ad.persona_nombre,
            'monto': ad.valor
        })
        
    movimientos_rango.sort(key=lambda x: x['fecha'] or date.today(), reverse=True)

    # ── NUEVO: Recaudación por curso (Activos) ──
    # Se obtienen todos los cursos que tienen matrículas activas
    cursos_activos = Curso.objects.filter(matriculas__estado='activa').distinct()
    
    # Abonos del mes actual, agrupados por curso
    abonos_mes_por_curso = Abono.objects.filter(
        fecha__gte=desde, fecha__lte=hasta, matricula__estado='activa'
    ).values('matricula__curso_id').annotate(total_mes=Sum('monto'))
    abonos_mes_dict = {
        item['matricula__curso_id']: item['total_mes'] 
        for item in abonos_mes_por_curso if item['matricula__curso_id']
    }
    
    recaudacion_cursos_activos = []
    for curso in cursos_activos:
        mats = curso.matriculas.filter(estado='activa')
        recaudado_mes = abonos_mes_dict.get(curso.id, Decimal('0.00'))
        
        recaudado_total = Decimal('0.00')
        saldo_pendiente = Decimal('0.00')
        recaudacion_esperada = Decimal('0.00')
        num_matriculados = 0
        
        for m in mats:
            num_matriculados += 1
            recaudacion_esperada += m.valor_neto
            recaudado_total += (m.valor_pagado or Decimal('0.00'))
            saldo_pendiente += m.saldo
            
        recaudacion_cursos_activos.append({
            'curso': curso,
            'recaudado_mes': recaudado_mes,
            'recaudado_total': recaudado_total,
            'saldo_pendiente': saldo_pendiente,
            'recaudacion_esperada': recaudacion_esperada,
            'num_matriculados': num_matriculados,
        })
    recaudacion_cursos_activos.sort(key=lambda x: x['recaudado_mes'], reverse=True)

    # ── NUEVO: Recaudación por curso (Cerrados / Archivados) ──
    recaudacion_cursos_cerrados = []
    cursos_cerrados = CierreCurso.objects.all()
    
    abonos_arch_mes_por_curso = AbonoArchivado.objects.filter(
        fecha__gte=desde, fecha__lte=hasta
    ).values('cierre_id').annotate(total_mes=Sum('monto'))
    
    abonos_arch_mes_dict = {
        item['cierre_id']: item['total_mes'] 
        for item in abonos_arch_mes_por_curso if item['cierre_id']
    }
    
    for curso_c in cursos_cerrados:
        mats_c = curso_c.matriculas_archivadas.all()
        recaudado_mes = abonos_arch_mes_dict.get(curso_c.id, Decimal('0.00'))
        
        recaudado_total = Decimal('0.00')
        saldo_pendiente = Decimal('0.00')
        recaudacion_esperada = Decimal('0.00')
        num_matriculados = 0
        
        for m in mats_c:
            num_matriculados += 1
            recaudacion_esperada += m.valor_neto
            recaudado_total += m.valor_pagado
            saldo_pendiente += m.saldo
            
        # Para mantener compatibilidad con dashboard.html que espera item.curso.nombre
        class DummyCurso:
            def __init__(self, nombre):
                self.nombre = nombre

        recaudacion_cursos_cerrados.append({
            'curso': DummyCurso(curso_c.curso_nombre),
            'recaudado_mes': recaudado_mes,
            'recaudado_total': recaudado_total,
            'saldo_pendiente': saldo_pendiente,
            'recaudacion_esperada': recaudacion_esperada,
            'num_matriculados': num_matriculados,
            'fecha_cierre': curso_c.fecha_cierre,
        })
    recaudacion_cursos_cerrados.sort(key=lambda x: x['fecha_cierre'], reverse=True)

    # ── Usuarios: última conexión y quién está en línea ──
    usuarios_actividad = _usuarios_actividad()
    usuarios_en_linea = sum(1 for u in usuarios_actividad if u['en_linea'])

    return render(request, 'admin_panel/dashboard.html', {
        'anio': anio,
        'mes': mes,
        'mes_nombre': MESES_ES[mes],
        'desde': desde,
        'hasta': hasta,
        'modo_rango': modo_rango,
        'desde_iso': desde.isoformat(),
        'hasta_iso': hasta.isoformat(),
        'usuarios_actividad': usuarios_actividad,
        'usuarios_en_linea': usuarios_en_linea,
        'ingresos': ingresos,
        'egresos_total': egresos_total,
        'balance': balance,
        'top_categorias': top_categorias,
        'egresos_recientes': egresos_recientes,
        'meses_selector': meses_selector,
        # Comparativa
        'ing_prev': ing_prev,
        'egr_prev': egr_prev,
        'bal_prev': bal_prev,
        'var_ingresos': variacion(ingresos['total'], ing_prev),
        'var_egresos': variacion(egresos_total, egr_prev),
        'var_balance': variacion(balance, bal_prev),
        'var_recuperaciones': variacion(recuperaciones_mes['total'], recup_prev),
        # Histórico
        'total_ingresos_hist': total_ingresos_hist,
        'total_egresos_hist': total_egresos_hist,
        'balance_hist': balance_hist,
        'total_abonos_hist': total_abonos_hist,
        'total_ventas_hist': total_ventas_hist,
        'total_adicionales_hist': total_adicionales_hist,
        'por_cobrar': por_cobrar_comprobantes,
        'total_retiros': total_retiros,
        # Detalles para modales
        'qs_egresos_mes': qs_egresos_mes,
        'retiros_detalle': retiros_detalle,
        'qs_recuperaciones_mes': qs_recuperaciones_mes,
        'por_cobrar_matriculas_detalle': por_cobrar_matriculas_detalle,
        'por_cobrar_comprobantes_detalle': por_cobrar_comprobantes_detalle,
        # NUEVO: Tipos de pago del mes + Recuperaciones
        'tipos_pago_mes': tipos_pago_mes,
        'total_tipos_pago_mes': total_tipos_pago_mes,
        'pie_tipos_pago_json': json.dumps(pie_tipos_pago),
        'recuperaciones_mes': recuperaciones_mes,
        'movimientos_rango': movimientos_rango,
        # ★ NUEVO: Recaudación por Curso
        'recaudacion_cursos_activos': recaudacion_cursos_activos,
        'recaudacion_cursos_cerrados': recaudacion_cursos_cerrados,
        # ★ NUEVO: Adicionales del mes (KPI con +)
        'adicionales_mes': adicionales_mes,
        'adicionales_archivados_mes': adicionales_archivados_mes,
        'var_adicionales': variacion(adicionales_mes['total'], adic_prev),
        # ★ NUEVO: Cierres de curso y administrativos
        'cierres_curso_mes': cierres_curso_mes,
        'cierres_admin': cierres_admin,
        'cierre_admin_existente': cierre_admin_existente,
        'tiene_valores_corte': tiene_valores_corte,
        # Gráfico (JSON serializable)
        'serie_meses_json': json.dumps(serie_meses),
    })


@admin_requerido
def egresos_lista(request):
    """Lista de todos los egresos con filtros."""
    qs = Egreso.objects.select_related('categoria', 'registrado_por')

    # Filtros
    categoria_id = request.GET.get('categoria', '').strip()
    desde = request.GET.get('desde', '').strip()
    hasta = request.GET.get('hasta', '').strip()
    q = request.GET.get('q', '').strip()

    if categoria_id:
        qs = qs.filter(categoria_id=categoria_id)
    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)
    if q:
        qs = qs.filter(concepto__icontains=q)

    total_filtrado = qs.aggregate(s=Sum('monto'))['s'] or Decimal('0.00')

    return render(request, 'admin_panel/egresos_lista.html', {
        'egresos': qs[:200],  # limitar a 200 para no quemar render
        'total_filtrado': total_filtrado,
        'categorias': CategoriaEgreso.objects.filter(activo=True),
        'filtros': {
            'categoria': categoria_id,
            'desde': desde,
            'hasta': hasta,
            'q': q,
        },
    })


@admin_requerido
def egreso_crear(request):
    if request.method == 'POST':
        form = EgresoForm(request.POST)
        if form.is_valid():
            egreso = form.save(commit=False)
            egreso.registrado_por = request.user
            egreso.save()
            messages.success(
                request,
                f'Egreso registrado: {egreso.concepto} (${egreso.monto}).'
            )
            return redirect('academia:admin_egresos_lista')
    else:
        form = EgresoForm(initial={'fecha': timezone.localdate()})
    return render(request, 'admin_panel/egreso_form.html', {
        'form': form,
        'modo': 'crear',
        'titulo': 'Registrar nuevo egreso',
    })


@admin_requerido
def egreso_editar(request, pk):
    egreso = get_object_or_404(Egreso, pk=pk)
    if request.method == 'POST':
        form = EgresoForm(request.POST, instance=egreso)
        if form.is_valid():
            form.save()
            messages.success(request, 'Egreso actualizado.')
            return redirect('academia:admin_egresos_lista')
    else:
        form = EgresoForm(instance=egreso)
    return render(request, 'admin_panel/egreso_form.html', {
        'form': form,
        'egreso': egreso,
        'modo': 'editar',
        'titulo': f'Editar egreso #{egreso.pk}',
    })


@admin_requerido
@require_POST
def egreso_eliminar(request, pk):
    egreso = get_object_or_404(Egreso, pk=pk)
    concepto = egreso.concepto
    monto = egreso.monto
    egreso.delete()
    messages.success(
        request,
        f'Egreso eliminado: {concepto} (${monto}).'
    )
    return redirect('academia:admin_egresos_lista')


# ─────────────────────────────────────────────────────────
# Exportación a CSV
# ─────────────────────────────────────────────────────────

import csv


def _csv_response(filename):
    """
    Crea una respuesta HTTP CSV. Usa BOM UTF-8 para que Excel
    abra bien las tildes y la ñ en Windows.
    """
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')  # BOM para Excel
    return response


@admin_requerido
def export_libro_mayor(request):
    """
    Exporta el Libro Mayor (todos los movimientos del rango) a CSV.
    """
    hoy = timezone.localdate()
    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')
    
    try:
        desde = date.fromisoformat(desde_str) if desde_str else hoy.replace(day=1)
        hasta = date.fromisoformat(hasta_str) if hasta_str else hoy
    except ValueError:
        desde, hasta = hoy.replace(day=1), hoy

    ingresos = _ingresos_periodo(desde, hasta)
    qs_egresos = Egreso.objects.filter(fecha__gte=desde, fecha__lte=hasta).select_related('categoria', 'registrado_por')

    movimientos_rango = []
    
    for e in qs_egresos:
        movimientos_rango.append({
            'fecha': e.fecha,
            'tipo': 'Egreso',
            'categoria': f"Egreso: {e.categoria.nombre}" if e.categoria else 'Egreso',
            'concepto': e.concepto,
            'involucrado': e.registrado_por.get_full_name() if e.registrado_por else 'N/A',
            'monto': -e.monto
        })
        
    for a in ingresos['qs_abonos']:
        curso = a.matricula.curso.nombre if a.matricula and a.matricula.curso else 'Sin curso'
        est = a.matricula.estudiante if a.matricula else None
        est_nombre = f"{est.nombres}" if est else 'N/A'
        movimientos_rango.append({
            'fecha': a.fecha.date() if hasattr(a.fecha, 'date') else a.fecha,
            'tipo': 'Ingreso',
            'categoria': f"Abono ({a.get_tipo_pago_display()})",
            'concepto': f"{curso} - {a.concepto}" if a.concepto else curso,
            'involucrado': est_nombre,
            'monto': a.monto
        })
        
    for a in ingresos['qs_abonos_archivados']:
        movimientos_rango.append({
            'fecha': a.fecha.date() if hasattr(a.fecha, 'date') else a.fecha,
            'tipo': 'Ingreso',
            'categoria': 'Abono (Archivado)',
            'concepto': a.matricula_archivada.curso_nombre if hasattr(a, 'matricula_archivada') and a.matricula_archivada else 'Curso Archivado',
            'involucrado': f"{a.matricula_archivada.nombres}" if hasattr(a, 'matricula_archivada') and a.matricula_archivada else "N/A",
            'monto': a.monto
        })
        
    for v in ingresos['qs_ventas']:
        movimientos_rango.append({
            'fecha': v.fecha_inscripcion.date() if hasattr(v.fecha_inscripcion, 'date') else v.fecha_inscripcion,
            'tipo': 'Ingreso',
            'categoria': 'Venta (Comprobante)',
            'concepto': v.concepto_pago or 'Comprobante manual',
            'involucrado': v.nombre_estudiante or 'N/A',
            'monto': v.pago_abono
        })
        
    for ad in ingresos['qs_adicionales_vivos']:
        if ad.estudiante:
            inv = f"{ad.estudiante.nombres}"
        elif ad.persona_externa:
            inv = f"{ad.persona_externa.nombres}"
        else:
            inv = "Desconocido"
        movimientos_rango.append({
            'fecha': ad.fecha.date() if hasattr(ad.fecha, 'date') else ad.fecha,
            'tipo': 'Ingreso',
            'categoria': 'Adicional',
            'concepto': ad.get_tipo_adicional_display() + (f" ({ad.descripcion})" if ad.descripcion else ""),
            'involucrado': inv,
            'monto': ad.valor
        })
        
    for ad in ingresos['qs_adicionales_archivados']:
        movimientos_rango.append({
            'fecha': ad.fecha.date() if hasattr(ad.fecha, 'date') else ad.fecha,
            'tipo': 'Ingreso',
            'categoria': 'Adicional (Archivado)',
            'concepto': ad.tipo_adicional_label,
            'involucrado': ad.persona_nombre,
            'monto': ad.valor
        })
        
    movimientos_rango.sort(key=lambda x: x['fecha'] or date.today(), reverse=True)

    response = _csv_response(f'libro_mayor_{desde.strftime("%Y%m%d")}_{hasta.strftime("%Y%m%d")}.csv')
    w = csv.writer(response)

    w.writerow([f'LIBRO MAYOR - DETALLE DE MOVIMIENTOS'])
    w.writerow([f'Período: {desde.strftime("%d/%m/%Y")} a {hasta.strftime("%d/%m/%Y")}'])
    w.writerow([f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
    w.writerow([])
    w.writerow(['Fecha', 'Tipo', 'Categoría', 'Concepto', 'Involucrado', 'Monto'])

    for m in movimientos_rango:
        fecha_str = m['fecha'].strftime("%d/%m/%Y") if hasattr(m['fecha'], 'strftime') else str(m['fecha'])
        w.writerow([
            fecha_str,
            m['tipo'],
            m['categoria'],
            m['concepto'],
            m['involucrado'],
            f"{m['monto']:.2f}"
        ])

    return response


@admin_requerido
def export_reporte_mes(request):
    """
    Exporta el reporte financiero completo del mes:
    - Resumen (ingresos, egresos, balance)
    - Detalle de egresos
    - Detalle de ingresos por abonos
    - Detalle de ingresos por ventas/comprobantes
    Todo en un solo CSV.
    """
    hoy = timezone.localdate()
    try:
        anio = int(request.GET.get('anio', hoy.year))
        mes = int(request.GET.get('mes', hoy.month))
        if not (1 <= mes <= 12):
            mes = hoy.month
    except (TypeError, ValueError):
        anio, mes = hoy.year, hoy.month

    desde, hasta = _rango_mes(anio, mes)
    nombre_mes = MESES_ES[mes]

    ingresos = _ingresos_periodo(desde, hasta)
    egresos_total = _egresos_periodo(desde, hasta)
    balance = ingresos['total'] - egresos_total

    response = _csv_response(
        f'reporte_{anio}_{mes:02d}_{nombre_mes.lower()}.csv'
    )
    w = csv.writer(response)

    # ── Encabezado ─────────────────────────────────────
    w.writerow([f'REPORTE FINANCIERO — {nombre_mes.upper()} {anio}'])
    w.writerow([f'Período: {desde.strftime("%d/%m/%Y")} a {hasta.strftime("%d/%m/%Y")}'])
    w.writerow([f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
    w.writerow([])

    # ── Resumen ────────────────────────────────────────
    w.writerow(['RESUMEN'])
    w.writerow(['Concepto', 'Monto (USD)'])
    w.writerow(['Ingresos por abonos (matrículas)', f'{ingresos["abonos"]:.2f}'])
    w.writerow(['Ingresos por ventas (comprobantes)', f'{ingresos["ventas"]:.2f}'])
    w.writerow(['TOTAL INGRESOS', f'{ingresos["total"]:.2f}'])
    w.writerow(['TOTAL EGRESOS', f'{egresos_total:.2f}'])
    w.writerow([
        'BALANCE NETO',
        f'{balance:.2f}',
        '(GANANCIA)' if balance > 0 else ('(PÉRDIDA)' if balance < 0 else '(EQUILIBRIO)')
    ])
    w.writerow([])

    # ── Egresos detallados ─────────────────────────────
    w.writerow(['EGRESOS DETALLADOS'])
    w.writerow(['Fecha', 'Categoría', 'Concepto', 'Monto (USD)', 'Notas', 'Registrado por'])
    egresos = (Egreso.objects
               .filter(fecha__gte=desde, fecha__lte=hasta)
               .select_related('categoria', 'registrado_por')
               .order_by('fecha'))
    for e in egresos:
        registrador = ''
        if e.registrado_por:
            registrador = e.registrado_por.get_full_name() or e.registrado_por.username
        w.writerow([
            e.fecha.strftime('%d/%m/%Y'),
            e.categoria.nombre,
            e.concepto,
            f'{e.monto:.2f}',
            e.notas or '',
            registrador,
        ])
    if not egresos:
        w.writerow(['(Sin egresos en este período)'])
    w.writerow([])

    # ── Ingresos por abonos ────────────────────────────
    w.writerow(['INGRESOS — ABONOS DE MATRÍCULAS'])
    w.writerow(['Fecha', 'Recibo', 'Estudiante', 'Cédula', 'Curso', 'Método', 'Banco', 'Monto (USD)'])
    abonos = (Abono.objects
              .filter(fecha__gte=desde, fecha__lte=hasta)
              .select_related('matricula__estudiante', 'matricula__curso')
              .order_by('fecha'))
    for a in abonos:
        est = a.matricula.estudiante
        w.writerow([
            a.fecha.strftime('%d/%m/%Y'),
            a.numero_recibo or '',
            est.nombre_completo,
            est.cedula,
            a.matricula.curso.nombre,
            a.get_metodo_display(),
            a.get_banco_display() if a.banco else '',
            f'{a.monto:.2f}',
        ])
    if not abonos:
        w.writerow(['(Sin abonos en este período)'])
    w.writerow([])

    # ── Ingresos por comprobantes ──────────────────────
    # Listar solo comprobantes manuales (sin matrícula vinculada). Los
    # comprobantes-espejo de matrículas ya están reflejados en la sección
    # de ABONOS arriba; listarlos aquí duplicaría el ingreso al ojo del
    # lector del reporte.
    w.writerow(['INGRESOS — VENTAS POR COMPROBANTE (solo ventas manuales)'])
    w.writerow(['Fecha insc.', 'Cliente', 'Celular', 'Curso', 'Modalidad',
                'Pago/Abono (USD)', 'Diferencia (USD)', 'Vendedora'])
    comprobantes = (Comprobante.objects
                    .filter(fecha_inscripcion__gte=desde, fecha_inscripcion__lte=hasta,
                            matricula__isnull=True)
                    .select_related('curso', 'vendedora')
                    .order_by('fecha_inscripcion'))
    for c in comprobantes:
        vendedora = c.vendedora_nombre or (c.vendedora.username if c.vendedora else '')
        w.writerow([
            c.fecha_inscripcion.strftime('%d/%m/%Y'),
            c.nombre_persona,
            c.celular,
            c.curso.nombre,
            c.get_modalidad_display(),
            f'{c.pago_abono:.2f}',
            f'{c.diferencia:.2f}',
            vendedora,
        ])
    if not comprobantes:
        w.writerow(['(Sin comprobantes en este período)'])

    return response


@admin_requerido
def export_egresos(request):
    """
    Exporta los egresos respetando los filtros activos
    (mismos parámetros GET que la lista).
    """
    qs = Egreso.objects.select_related('categoria', 'registrado_por')

    categoria_id = request.GET.get('categoria', '').strip()
    desde = request.GET.get('desde', '').strip()
    hasta = request.GET.get('hasta', '').strip()
    q = request.GET.get('q', '').strip()

    if categoria_id:
        qs = qs.filter(categoria_id=categoria_id)
    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)
    if q:
        qs = qs.filter(concepto__icontains=q)

    qs = qs.order_by('-fecha', '-creado')

    response = _csv_response(
        f'egresos_{timezone.now().strftime("%Y%m%d_%H%M")}.csv'
    )
    w = csv.writer(response)
    w.writerow(['EGRESOS — Exportación filtrada'])
    if desde or hasta or categoria_id or q:
        filtros = []
        if desde: filtros.append(f'desde {desde}')
        if hasta: filtros.append(f'hasta {hasta}')
        if categoria_id:
            cat = CategoriaEgreso.objects.filter(pk=categoria_id).first()
            if cat: filtros.append(f'categoría: {cat.nombre}')
        if q: filtros.append(f'búsqueda: "{q}"')
        w.writerow(['Filtros: ' + ', '.join(filtros)])
    w.writerow([f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'])
    w.writerow([])

    w.writerow(['Fecha', 'Categoría', 'Concepto', 'Monto (USD)', 'Notas', 'Registrado por'])

    total = Decimal('0.00')
    for e in qs:
        registrador = ''
        if e.registrado_por:
            registrador = e.registrado_por.get_full_name() or e.registrado_por.username
        w.writerow([
            e.fecha.strftime('%d/%m/%Y'),
            e.categoria.nombre,
            e.concepto,
            f'{e.monto:.2f}',
            e.notas or '',
            registrador,
        ])
        total += e.monto

    w.writerow([])
    w.writerow(['', '', 'TOTAL', f'{total:.2f}'])

    return response


# ═════════════════════════════════════════════════════════════════
# CIERRE ADMINISTRATIVO (corte de caja del periodo)
# ═════════════════════════════════════════════════════════════════

@admin_requerido
def cierre_admin_preview(request):
    """
    Vista previa del cierre administrativo de un mes: muestra el detalle de
    ingresos (incluidos los archivados de cursos cerrados), egresos y balance,
    antes de congelarlo.
    """
    hoy = timezone.localdate()
    try:
        anio = int(request.GET.get('anio', hoy.year))
        mes = int(request.GET.get('mes', hoy.month))
        if not (1 <= mes <= 12):
            mes = hoy.month
    except (TypeError, ValueError):
        anio, mes = hoy.year, hoy.month

    desde, hasta = _rango_mes(anio, mes)

    ingresos = _ingresos_periodo(desde, hasta)
    egresos_total = _egresos_periodo(desde, hasta)
    egresos_categorias = _egresos_por_categoria_periodo(desde, hasta)
    cierres_curso = _cierres_curso_periodo(desde, hasta)
    balance = ingresos['total'] - egresos_total

    existente = CierreAdministrativo.objects.filter(
        anio=anio, mes=mes
    ).order_by('-fecha_cierre').first()
    cortes_existentes_count = CierreAdministrativo.objects.filter(anio=anio, mes=mes).count()
    tiene_valores_corte = ingresos['total'] != Decimal('0.00') or egresos_total != Decimal('0.00')

    meses_selector = []
    for i in range(0, 24):
        m = hoy.month - i
        a = hoy.year
        while m <= 0:
            m += 12
            a -= 1
        meses_selector.append({
            'anio': a, 'mes': m,
            'label': f'{MESES_ES[m]} {a}',
            'seleccionado': (a == anio and m == mes),
        })

    return render(request, 'admin_panel/cierre_admin_preview.html', {
        'anio': anio,
        'mes': mes,
        'mes_nombre': MESES_ES[mes],
        'desde': desde,
        'hasta': hasta,
        'ingresos': ingresos,
        'egresos_total': egresos_total,
        'egresos_categorias': egresos_categorias,
        'cierres_curso': cierres_curso,
        'balance': balance,
        'existente': existente,
        'cortes_existentes_count': cortes_existentes_count,
        'tiene_valores_corte': tiene_valores_corte,
        'meses_selector': meses_selector,
    })


@admin_requerido
@require_POST
def cierre_admin_ejecutar(request):
    """Congela el corte de caja del mes en un CierreAdministrativo."""
    try:
        anio = int(request.POST.get('anio'))
        mes = int(request.POST.get('mes'))
        if not (1 <= mes <= 12):
            raise ValueError
    except (TypeError, ValueError):
        messages.error(request, 'Periodo no válido.')
        return redirect('academia:admin_dashboard')

    etiqueta = (request.POST.get('etiqueta', '').strip())[:80]
    observaciones = request.POST.get('observaciones', '').strip()
    admin_password = request.POST.get('admin_password', '')

    if not request.user.check_password(admin_password):
        messages.error(request, 'Contraseña de administrador incorrecta. Corte de caja abortado por seguridad.')
        return redirect(f'/admin-panel/cierre/?anio={anio}&mes={mes}')

    desde, hasta = _rango_mes(anio, mes)
    ingresos = _ingresos_periodo(desde, hasta)
    egresos_total = _egresos_periodo(desde, hasta)
    egresos_categorias = _egresos_por_categoria_periodo(desde, hasta)
    cierres_curso = _cierres_curso_periodo(desde, hasta)
    balance = ingresos['total'] - egresos_total

    if ingresos['total'] == Decimal('0.00') and egresos_total == Decimal('0.00'):
        messages.warning(
            request,
            f'No se creó el corte de {MESES_ES[mes]} {anio} porque ese mes no tiene ingresos ni egresos.'
        )
        return redirect(f'/admin-panel/cierre/?anio={anio}&mes={mes}')

    # Serializar el desglose de egresos por categoría
    egresos_detalle = [
        {
            'categoria': e['nombre'],
            'total': float(e['total']),
            'color': e.get('color', ''),
            'icono': e.get('icono', ''),
        }
        for e in egresos_categorias
    ]

    cierre = CierreAdministrativo.objects.create(
        anio=anio, mes=mes,
        etiqueta=etiqueta,
        fecha_desde=desde,
        fecha_hasta=hasta,
        ingreso_abonos=ingresos['abonos'],
        ingreso_abonos_archivados=ingresos['abonos_archivados'],
        ingreso_ventas=ingresos['ventas'],
        ingreso_adicionales=ingresos['adicionales'],
        ingreso_total=ingresos['total'],
        egreso_total=egresos_total,
        egresos_detalle_json=json.dumps(egresos_detalle),
        balance_neto=balance,
        cierres_curso_incluidos=cierres_curso['count'],
        monto_cierres_curso=cierres_curso['total_cobrado'],
        observaciones=observaciones,
        cerrado_por=request.user if request.user.is_authenticated else None,
    )

    messages.success(
        request,
        f'✅ Cierre administrativo de {MESES_ES[mes]} {anio} guardado. '
        f'Ingresos ${ingresos["total"]:.2f} − Egresos ${egresos_total:.2f} = '
        f'Balance ${balance:.2f}.'
    )
    return redirect('academia:cierre_admin_detalle', pk=cierre.pk)


@admin_requerido
def cierre_admin_historial(request):
    """Lista todos los cierres administrativos (cortes de caja)."""
    qs = CierreAdministrativo.objects.all()

    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()
    if anio.isdigit():
        qs = qs.filter(anio=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        # Filtramos por el campo `mes` del propio cierre cuando existe;
        # si el cierre es anual (mes=NULL), se descarta del filtrado por mes.
        qs = qs.filter(mes=int(mes))

    cierres = list(qs)
    total_ingresos = sum((c.ingreso_total for c in cierres), Decimal('0.00'))
    total_egresos = sum((c.egreso_total for c in cierres), Decimal('0.00'))
    total_balance = sum((c.balance_neto for c in cierres), Decimal('0.00'))

    anios = sorted(
        set(CierreAdministrativo.objects.values_list('anio', flat=True)),
        reverse=True
    )

    return render(request, 'admin_panel/cierre_admin_historial.html', {
        'cierres': cierres,
        'total_ingresos': total_ingresos,
        'total_egresos': total_egresos,
        'total_balance': total_balance,
        'anios': anios,
        'filtro_anio': anio,
        'filtro_mes': mes,
    })


@admin_requerido
def cierre_admin_detalle(request, pk):
    """Detalle de un cierre administrativo congelado."""
    cierre = get_object_or_404(CierreAdministrativo, pk=pk)

    egresos_detalle = []
    if cierre.egresos_detalle_json:
        try:
            egresos_detalle = json.loads(cierre.egresos_detalle_json)
        except (ValueError, TypeError):
            egresos_detalle = []

    # Cierres de curso que cayeron en el mismo periodo (informativo)
    cierres_curso = _cierres_curso_periodo(cierre.fecha_desde, cierre.fecha_hasta)

    return render(request, 'admin_panel/cierre_admin_detalle.html', {
        'cierre': cierre,
        'egresos_detalle': egresos_detalle,
        'cierres_curso': cierres_curso,
    })


def _cierre_admin_egresos_detalle(cierre):
    if not cierre.egresos_detalle_json:
        return []
    try:
        data = json.loads(cierre.egresos_detalle_json)
    except (ValueError, TypeError):
        return []
    return data if isinstance(data, list) else []


def _cierre_admin_filename(cierre, extension):
    base = slugify(cierre.encabezado) or f'corte_{cierre.anio}_{cierre.mes or 0:02d}'
    return f'{base}_{cierre.fecha_cierre.strftime("%Y%m%d")}.{extension}'


@admin_requerido
def cierre_admin_export_excel(request, pk):
    """Exporta el corte administrativo congelado como Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    cierre = get_object_or_404(CierreAdministrativo, pk=pk)
    egresos_detalle = _cierre_admin_egresos_detalle(cierre)
    cierres_curso = _cierres_curso_periodo(cierre.fecha_desde, cierre.fecha_hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen'

    header_fill = PatternFill('solid', fgColor='1A237E')
    header_font = Font(bold=True, color='FFFFFF')
    title_font = Font(bold=True, size=15, color='1A237E')
    money_font = Font(bold=True, color='2E7D32')
    danger_font = Font(bold=True, color='C62828')
    total_fill = PatternFill('solid', fgColor='E8EAF6')
    thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    ws.merge_cells('A1:D1')
    ws['A1'] = cierre.encabezado
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(['Periodo', f'{cierre.fecha_desde:%d/%m/%Y} - {cierre.fecha_hasta:%d/%m/%Y}', 'Guardado', cierre.fecha_cierre.strftime('%d/%m/%Y %H:%M')])
    ws.append(['Responsable', cierre.cerrado_por.get_full_name() or cierre.cerrado_por.username if cierre.cerrado_por else '', 'Observaciones', cierre.observaciones or ''])
    ws.append([])

    def write_section(title, rows, start_row):
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, color='1A237E')
        header_row = start_row + 1
        for col, header in enumerate(('Concepto', 'Monto'), start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin
            cell.alignment = Alignment(horizontal='center')
        row_idx = header_row + 1
        for label, amount, kind in rows:
            ws.cell(row=row_idx, column=1, value=label).border = thin
            amount_cell = ws.cell(row=row_idx, column=2, value=float(amount or Decimal('0.00')))
            amount_cell.border = thin
            amount_cell.number_format = '"$"#,##0.00'
            if kind == 'danger':
                amount_cell.font = danger_font
            else:
                amount_cell.font = money_font
            if kind == 'total':
                for c in (1, 2):
                    ws.cell(row=row_idx, column=c).fill = total_fill
                    ws.cell(row=row_idx, column=c).font = Font(bold=True, color='1A237E')
            row_idx += 1
        return row_idx + 1

    next_row = 6
    next_row = write_section('Ingresos congelados', [
        ('Abonos (operación viva)', cierre.ingreso_abonos, 'money'),
        ('Abonos de cursos cerrados', cierre.ingreso_abonos_archivados, 'money'),
        ('Ventas / comprobantes', cierre.ingreso_ventas, 'money'),
        ('Adicionales', cierre.ingreso_adicionales, 'money'),
        ('TOTAL INGRESOS', cierre.ingreso_total, 'total'),
    ], next_row)
    next_row = write_section('Resumen final', [
        ('TOTAL INGRESOS', cierre.ingreso_total, 'money'),
        ('TOTAL EGRESOS', cierre.egreso_total, 'danger'),
        ('BALANCE NETO', cierre.balance_neto, 'total'),
    ], next_row)

    for col, width in enumerate((34, 16, 18, 36), start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws2 = wb.create_sheet('Egresos')
    ws2.append(['Categoría', 'Total'])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
    for e in egresos_detalle:
        ws2.append([f'{e.get("icono", "")} {e.get("categoria", "")}'.strip(), float(e.get('total') or 0)])
        ws2.cell(row=ws2.max_row, column=2).number_format = '"$"#,##0.00'
    ws2.append(['TOTAL EGRESOS', float(cierre.egreso_total)])
    for cell in ws2[ws2.max_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True, color='C62828')
    ws2.column_dimensions['A'].width = 34
    ws2.column_dimensions['B'].width = 16

    ws3 = wb.create_sheet('Cierres de curso')
    headers = ['Curso', 'Alcance', 'Jornada', 'Cerrado el', 'Matrículas', 'Facturado', 'Cobrado', 'Pendiente']
    ws3.append(headers)
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')
    for c in cierres_curso['cierres']:
        ws3.append([
            c.curso_nombre,
            c.get_alcance_display(),
            c.jornada_descripcion or '',
            c.fecha_cierre.strftime('%d/%m/%Y %H:%M'),
            c.total_matriculas,
            float(c.total_facturado),
            float(c.total_cobrado),
            float(c.total_pendiente),
        ])
        for col in (6, 7, 8):
            ws3.cell(row=ws3.max_row, column=col).number_format = '"$"#,##0.00'
    widths = [30, 14, 30, 18, 12, 14, 14, 14]
    for col, width in enumerate(widths, start=1):
        ws3.column_dimensions[get_column_letter(col)].width = width

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    response = HttpResponse(
        out.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{_cierre_admin_filename(cierre, "xlsx")}"'
    return response


@admin_requerido
def cierre_admin_export_pdf(request, pk):
    """Exporta el corte administrativo congelado como PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    cierre = get_object_or_404(CierreAdministrativo, pk=pk)
    egresos_detalle = _cierre_admin_egresos_detalle(cierre)
    cierres_curso = _cierres_curso_periodo(cierre.fecha_desde, cierre.fecha_hasta)

    out = BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4, leftMargin=28, rightMargin=28, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(cierre.encabezado, styles['Title']),
        Paragraph(
            f'Periodo: {cierre.fecha_desde:%d/%m/%Y} - {cierre.fecha_hasta:%d/%m/%Y} · '
            f'Guardado: {cierre.fecha_cierre:%d/%m/%Y %H:%M}',
            styles['Normal'],
        ),
        Spacer(1, 10),
    ]
    if cierre.observaciones:
        story.extend([Paragraph(f'Observaciones: {cierre.observaciones}', styles['Normal']), Spacer(1, 8)])

    def money(value):
        return f'${Decimal(str(value or 0)):,.2f}'

    resumen = [
        ['Concepto', 'Monto'],
        ['Ingresos', money(cierre.ingreso_total)],
        ['Egresos', money(cierre.egreso_total)],
        ['Balance neto', money(cierre.balance_neto)],
    ]
    ingresos = [
        ['Ingreso congelado', 'Monto'],
        ['Abonos (operación viva)', money(cierre.ingreso_abonos)],
        ['Abonos de cursos cerrados', money(cierre.ingreso_abonos_archivados)],
        ['Ventas / comprobantes', money(cierre.ingreso_ventas)],
        ['Adicionales', money(cierre.ingreso_adicionales)],
        ['TOTAL', money(cierre.ingreso_total)],
    ]

    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#DDDDDD')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ])

    story.append(Table(resumen, colWidths=[300, 120], repeatRows=1, style=table_style))
    story.extend([Spacer(1, 12), Paragraph('Ingresos congelados', styles['Heading2'])])
    story.append(Table(ingresos, colWidths=[300, 120], repeatRows=1, style=table_style))

    story.extend([Spacer(1, 12), Paragraph('Egresos por categoría', styles['Heading2'])])
    egresos_data = [['Categoría', 'Total']]
    for e in egresos_detalle:
        egresos_data.append([f'{e.get("icono", "")} {e.get("categoria", "")}'.strip(), money(e.get('total') or 0)])
    if len(egresos_data) == 1:
        egresos_data.append(['Sin egresos en este corte', '$0.00'])
    egresos_data.append(['TOTAL', money(cierre.egreso_total)])
    story.append(Table(egresos_data, colWidths=[300, 120], repeatRows=1, style=table_style))

    if cierres_curso['count']:
        story.extend([Spacer(1, 12), Paragraph('Cierres de curso del periodo', styles['Heading2'])])
        cursos_data = [['Curso', 'Matrículas', 'Cobrado']]
        for c in cierres_curso['cierres']:
            cursos_data.append([c.curso_nombre, str(c.total_matriculas), money(c.total_cobrado)])
        story.append(Table(cursos_data, colWidths=[300, 70, 100], repeatRows=1, style=table_style))

    doc.build(story)
    out.seek(0)
    response = HttpResponse(out.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{_cierre_admin_filename(cierre, "pdf")}"'
    return response


@admin_requerido
@require_POST
def cierre_admin_eliminar(request, pk):
    """Elimina un cierre administrativo."""
    cierre = get_object_or_404(CierreAdministrativo, pk=pk)
    nombre = cierre.encabezado
    cierre.delete()
    messages.success(request, f'Cierre administrativo eliminado: {nombre}.')
    return redirect('academia:cierre_admin_historial')

@admin_requerido
def control_registro(request):
    """Auditoria de los registros realizados por los asesores/admin."""
    matriculas = Matricula.objects.select_related(
        'curso', 'estudiante', 'vendedora', 'registrado_por', 'jornada'
    ).order_by('-fecha_matricula', '-creado')[:300]
    
    pagos = Abono.objects.select_related(
        'matricula__estudiante', 'matricula__curso', 'registrado_por'
    ).order_by('-creado')[:300]
    
    return render(request, 'admin_panel/control_registro.html', {
        'matriculas': matriculas,
        'pagos': pagos,
        'titulo': 'Control de Registro',
    })
