from datetime import date
from decimal import Decimal
from io import BytesIO
import json

from django.contrib.auth.models import User
from django.test import TestCase, RequestFactory
from django.urls import reverse
from openpyxl import load_workbook

from .models import Curso, JornadaCurso, Estudiante, Matricula, Abono, SeleccionJornadasRecaudacion
from .views_pagos import _hojas_recaudacion_data, _paginas_recaudacion_combinada


class JornadasCombinadasTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_superuser('prueba_combinadas', password='prueba')
        self.client.force_login(self.usuario)
        self.curso = Curso.objects.create(nombre='Curso combinado', activo=True, ofrece_presencial=True,
                                         valor_presencial=110, numero_modulos=4)
        self.jornadas = [JornadaCurso.objects.create(curso=self.curso, modalidad='presencial',
                          descripcion='sabados_intensivos', fecha_inicio=date(2026, 7, dia),
                          ciudad='Guayaquil') for dia in (4, 11)]
        for i in range(11):
            est = Estudiante.objects.create(cedula=f'090000{i:04d}', nombres=f'Estudiante {10-i:02d}')
            m = Matricula.objects.create(estudiante=est, curso=self.curso, jornada=self.jornadas[i % 2],
                modalidad='presencial', tipo_matricula='reserva_abono', forma_pago='abono',
                valor_curso=110, valor_pagado=0, fecha_matricula=date(2026, 7, 1), tipo_registro='central_ia')
            Abono.objects.create(matricula=m, fecha=date(2026, 7, 1), monto=Decimal('35'),
                                 tipo_pago='abono', metodo='efectivo')
        self.datos = dict(curso=str(self.curso.pk), jornadas=[str(j.pk) for j in self.jornadas],
                          varias='1', fecha_desde='2026-07-01', fecha_hasta='2026-07-12', ciudad='', modalidad='')

    def test_combina_sin_duplicar_y_conserva_totales_planes_y_jornada(self):
        request = RequestFactory().get('/', self.datos)
        hojas, filtros = _hojas_recaudacion_data(request)
        self.assertEqual(len(hojas), 1)
        hoja = hojas[0]
        self.assertEqual(len(hoja['items']), 11)
        self.assertEqual(len({i['matricula_id'] for i in hoja['items']}), 11)
        self.assertEqual(hoja['total_recaudado'], Decimal('385'))
        self.assertEqual(hoja['total_efectivo'], Decimal('385'))
        nombres = [i['estudiante'].nombre_completo for i in hoja['items']]
        self.assertEqual(nombres, sorted(nombres))
        self.assertTrue(all(i['jornada_label'] for i in hoja['items']))
        self.assertIn('varias=1', filtros['querystring'])
        separadas, _ = _hojas_recaudacion_data(RequestFactory().get('/', dict(self.datos, varias='0')))
        originales = {i['matricula_id']: i for h in separadas for i in h['items']}
        for item in hoja['items']:
            for campo in ('cuota_sugerida', 'saldo_modulo', 'pago_pendiente_fecha', 'recaudado'):
                self.assertEqual(item[campo], originales[item['matricula_id']][campo])
        paginas = _paginas_recaudacion_combinada(hojas)
        self.assertEqual([len(p['items']) for p in paginas], [8, 3])
        self.assertFalse(paginas[0]['ultima'])
        self.assertTrue(paginas[1]['ultima'])

    def test_export_y_pantalla_mismos_estudiantes_y_resumen_unico(self):
        response = self.client.get(reverse('academia:hoja_recaudacion'), self.datos)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['hojas']), 1)
        self.assertContains(response, 'btn-varias-jornadas')
        self.assertContains(response, 'Página 2 de 2')
        self.assertEqual(SeleccionJornadasRecaudacion.objects.count(), 0)
        response = self.client.get(reverse('academia:hoja_recaudacion_export_excel'), self.datos)
        ws = load_workbook(BytesIO(response.content)).active
        nombres = [r[1] for r in ws.iter_rows(values_only=True) if r[1] and str(r[1]).startswith('Estudiante')]
        self.assertEqual(len(nombres), 11)
        self.assertTrue(all('\n' in nombre for nombre in nombres))
        self.assertEqual(str(ws.print_area).count('!'), 2)
        self.assertEqual(sum(r[1] == 'TOTAL RECAUDADO' for r in ws.iter_rows(values_only=True)), 1)

    def test_no_admite_jornada_de_otro_curso_ni_seleccion_vacia(self):
        otro = Curso.objects.create(nombre='Otro curso', activo=True)
        ajena = JornadaCurso.objects.create(curso=otro, modalidad='presencial', fecha_inicio=date(2026, 7, 1))
        for ids in ([], [str(ajena.pk)], ['no-numero']):
            datos = dict(self.datos, jornadas=ids)
            hojas, _ = _hojas_recaudacion_data(RequestFactory().get('/', datos))
            self.assertEqual(hojas, [])
            response = self.client.post(reverse('academia:hoja_recaudacion_historial'),
                                        data=json.dumps(datos), content_type='application/json')
            self.assertEqual(response.status_code, 400)
        self.assertEqual(SeleccionJornadasRecaudacion.objects.count(), 0)

    def test_historial_persistente_por_usuario_curso_y_fechas(self):
        url = reverse('academia:hoja_recaudacion_historial')
        response = self.client.post(url, data=json.dumps(self.datos), content_type='application/json')
        self.assertEqual(response.status_code, 200)
        self.assertIn('varias=1', response.json()['url'])
        registros = self.client.get(url, {'curso': self.curso.pk}).json()['registros']
        self.assertEqual(len(registros), 1)
        self.assertEqual(registros[0]['filtros']['fecha_desde'], '2026-07-01')
        self.assertEqual(len(registros[0]['jornadas']), 2)
        self.assertEqual(self.client.get(url, {'curso': self.curso.pk + 99}).json()['registros'], [])
        otro = User.objects.create_superuser('otro_historial', password='prueba')
        self.client.force_login(otro)
        self.assertEqual(self.client.get(url, {'curso': self.curso.pk}).json()['registros'], [])

    def test_historial_devuelve_mas_de_diez_y_elimina_solo_registros_propios(self):
        url = reverse('academia:hoja_recaudacion_historial')
        etiquetas = [dict(id=str(j.pk), nombre=j.etiqueta) for j in self.jornadas]
        for dia in range(1, 13):
            filtros = dict(self.datos, fecha_desde=f'2026-07-{dia:02d}', fecha_hasta=f'2026-07-{dia:02d}')
            SeleccionJornadasRecaudacion.objects.create(
                usuario=self.usuario, curso=self.curso, filtros=filtros, jornadas=etiquetas,
            )
        otro = User.objects.create_superuser('otro_historial_delete', password='prueba')
        ajeno = SeleccionJornadasRecaudacion.objects.create(
            usuario=otro, curso=self.curso, filtros=self.datos, jornadas=etiquetas,
        )

        registros = self.client.get(url, {'curso': self.curso.pk}).json()['registros']
        self.assertEqual(len(registros), 12)

        response = self.client.delete(
            url, data=json.dumps({'id': ajeno.pk}), content_type='application/json',
        )
        self.assertEqual(response.status_code, 404)
        self.assertTrue(SeleccionJornadasRecaudacion.objects.filter(pk=ajeno.pk).exists())

        response = self.client.delete(
            url, data=json.dumps({'id': registros[0]['id']}), content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(self.client.get(url, {'curso': self.curso.pk}).json()['registros']), 11)

    def test_duplicados_no_duplican_estudiantes_y_filtros_no_mezclan_ciudades(self):
        datos = dict(self.datos, jornadas=self.datos['jornadas'] * 2)
        hojas, _ = _hojas_recaudacion_data(RequestFactory().get('/', datos))
        self.assertEqual(len(hojas[0]['items']), 11)
        hojas, _ = _hojas_recaudacion_data(RequestFactory().get('/', dict(datos, ciudad='Quito')))
        self.assertEqual(hojas, [])

    def test_lista_larga_continua_en_tercera_pagina(self):
        for i in range(6):
            est = Estudiante.objects.create(cedula=f'099999{i:04d}', nombres=f'Nuevo {i}')
            Matricula.objects.create(estudiante=est, curso=self.curso, jornada=self.jornadas[0],
                modalidad='presencial', tipo_matricula='reserva_abono', forma_pago='abono',
                valor_curso=110, valor_pagado=0, fecha_matricula=date(2026, 7, 1), tipo_registro='central_ia')
        hojas, _ = _hojas_recaudacion_data(RequestFactory().get('/', self.datos))
        paginas = _paginas_recaudacion_combinada(hojas)
        self.assertEqual([len(p['items']) for p in paginas], [8, 8, 1])
        self.assertEqual([p['ultima'] for p in paginas], [False, False, True])
        response = self.client.get(reverse('academia:hoja_recaudacion_export_excel'), self.datos)
        ws = load_workbook(BytesIO(response.content)).active
        self.assertEqual(str(ws.print_area).count('!'), 3)

    def test_impresion_individual_conserva_todas_las_filas_y_cierre_vacio(self):
        datos = dict(self.datos, varias='0')
        response = self.client.get(reverse('academia:hoja_recaudacion'), datos)
        paginas = response.context['paginas_combinadas']
        self.assertEqual(sum(len(p['items']) for p in paginas), 11)
        self.assertEqual(sum(p['ultima'] for p in paginas), 2)
        self.assertContains(response, '<th>Recaudación en efectivo</th><td></td>')

    def test_pdf_descarga_individual_y_combinada(self):
        for varias in ('0', '1'):
            response = self.client.get(reverse('academia:hoja_recaudacion_export_pdf'),
                                       dict(self.datos, varias=varias))
            self.assertEqual(response.status_code, 200)
            self.assertEqual(response['Content-Type'], 'application/pdf')
            self.assertTrue(response.content.startswith(b'%PDF'))
