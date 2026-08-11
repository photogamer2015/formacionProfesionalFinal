"""
Vistas adicionales para Pagos, Historial de matriculados y Estudiantes.

Diseño:
- Todas usan el decorador @matricula_requerida (admin + asesor pueden ver).
- Las exportaciones a Excel usan openpyxl y devuelven un HttpResponse con el archivo.
- Filtros por GET querystring (q, curso, modalidad, estado, año, mes).
"""

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
import re
from urllib.parse import urlencode

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Prefetch, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_POST

from .forms import AbonoForm, RecuperacionPendienteForm
from .models import (
    Abono, Curso, Estudiante, JornadaCurso, Matricula, RecuperacionPendiente,
    distribuir_monto_en_cuotas_enteras,
)
from .permisos import matricula_requerida, admin_requerido
from .busqueda import filtrar_queryset_busqueda


# ═════════════════════════════════════════════════════════════════
# Constantes
# ═════════════════════════════════════════════════════════════════

MESES_ES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]

CENTAVO = Decimal('0.01')


def _redirect_despues_recuperacion(request, matricula=None):
    """Vuelve al origen si es una URL local; si no, vuelve a la matrícula o al listado."""
    next_url = (
        request.POST.get('next')
        or request.GET.get('next')
        or request.META.get('HTTP_REFERER')
        or ''
    )
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return redirect(next_url)
    if matricula is not None:
        return redirect('academia:matricula_abonos', pk=matricula.pk)
    return redirect('academia:recuperaciones_lista')


# ═════════════════════════════════════════════════════════════════
# Helpers de Excel
# ═════════════════════════════════════════════════════════════════

def _safe_excel_table_name(base, used_names=None):
    """Devuelve un nombre válido y único para una tabla de Excel."""
    used_names = used_names or set()
    name = re.sub(r'[^A-Za-z0-9_]', '_', str(base or 'Tabla'))
    name = re.sub(r'_+', '_', name).strip('_') or 'Tabla'
    if not re.match(r'^[A-Za-z_]', name):
        name = f'Tabla_{name}'
    name = f'Tabla_{name}'[:240]

    candidate = name
    suffix = 1
    while candidate in used_names:
        extra = f'_{suffix}'
        candidate = f'{name[:255 - len(extra)]}{extra}'
        suffix += 1
    return candidate


def _add_excel_table(ws, header_row, first_col, last_row, last_col, table_name):
    """
    Convierte un rango rectangular en tabla real de Excel.
    Si no hay filas de datos, deja al menos el autofiltro en los encabezados.
    """
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if last_col < first_col:
        return

    header_ref = (
        f'{get_column_letter(first_col)}{header_row}:'
        f'{get_column_letter(last_col)}{header_row}'
    )
    if last_row <= header_row:
        ws.auto_filter.ref = header_ref
        return

    ref = (
        f'{get_column_letter(first_col)}{header_row}:'
        f'{get_column_letter(last_col)}{last_row}'
    )
    used_names = set()
    for sheet in ws.parent.worksheets:
        used_names.update(sheet.tables.keys())
    display_name = _safe_excel_table_name(table_name, used_names)
    table = Table(displayName=display_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _build_excel_response(
    filename, sheet_name, headers, rows, totals=None,
    column_formats=None, text_columns=None, explicit_widths=None,
):
    """
    Genera un .xlsx en memoria y lo devuelve como HttpResponse para descarga.

    headers: lista de strings (encabezados de columna)
    rows: lista de listas (datos)
    totals: dict opcional {col_idx_0based: total} para fila final
    column_formats: dict opcional {col_idx_0based: formato Excel}
    text_columns: iterable opcional de col_idx_0based que se fuerzan como texto
    explicit_widths: dict opcional {col_idx_0based: ancho}
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.page import PageMargins

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel limita a 31 chars

    # ── Estilos ──
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    total_font = Font(bold=True, color='1A237E', size=11)
    total_fill = PatternFill('solid', fgColor='FFF8E1')
    data_align_wrap = Alignment(vertical='center', wrap_text=True)
    column_formats = column_formats or {}
    text_columns = set(text_columns or [])
    explicit_widths = explicit_widths or {}

    # ── Título de la hoja en fila 1 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=sheet_name)
    title_cell.font = Font(bold=True, size=14, color='1A237E')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # ── Encabezados en fila 2 ──
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    # ── Datos a partir de fila 3 ──
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, _header in enumerate(headers, start=1):
            value = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else ''
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_align_wrap
            if col_idx - 1 in text_columns:
                cell.number_format = '@'
            elif col_idx - 1 in column_formats:
                cell.number_format = column_formats[col_idx - 1]

    # ── Fila de totales ──
    if totals:
        total_row_idx = len(rows) + 3
        # Etiqueta "TOTAL" en la primera columna
        cell = ws.cell(row=total_row_idx, column=1, value='TOTAL')
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        for col_idx_0, value in totals.items():
            cell = ws.cell(row=total_row_idx, column=col_idx_0 + 1, value=value)
            cell.font = total_font
            cell.fill = total_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if col_idx_0 in column_formats:
                cell.number_format = column_formats[col_idx_0]

    # ── Auto-ancho aproximado por columna (con tope y mínimo razonables) ──
    for col_idx, header in enumerate(headers, start=1):
        if col_idx - 1 in explicit_widths:
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = explicit_widths[col_idx - 1]
            continue
        max_length = len(str(header))
        for row_data in rows:
            if col_idx - 1 < len(row_data):
                value = row_data[col_idx - 1]
                if value is not None:
                    # Considera la línea más larga si hay saltos de línea
                    longest = max((len(s) for s in str(value).split('\n')), default=0)
                    max_length = max(max_length, longest)
        # Un poco más amplio para que no se corte el texto al imprimir
        width = min(max(max_length + 3, 10), 38)
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

    # ── Congelar encabezados ──
    ws.freeze_panes = 'A3'
    last_row = len(rows) + (3 if totals else 2)
    _add_excel_table(ws, 2, 1, last_row, len(headers), sheet_name)

    # ── Configuración de impresión profesional ──
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A3 if len(headers) > 10 else ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # Tantas páginas como sean necesarias en alto
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.4,
                                   header=0.2, footer=0.2)
    # Repetir las dos primeras filas (título + encabezado) en cada página al imprimir
    ws.print_title_rows = '1:2'
    # Encabezado y pie de página al imprimir
    ws.oddHeader.center.text = f"&B{sheet_name}"
    ws.oddHeader.center.size = 11
    ws.oddHeader.center.color = "1A237E"
    ws.oddFooter.left.text = "Formación Profesional EC"
    ws.oddFooter.right.text = "Página &P de &N"
    ws.oddFooter.left.size = 9
    ws.oddFooter.right.size = 9

    # ── Devolver como HttpResponse ──
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _rango_fecha_matricula_desde_request(request):
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()
    fecha_desde_date = parse_date(fecha_desde) if fecha_desde else None
    fecha_hasta_date = parse_date(fecha_hasta) if fecha_hasta else None

    if not fecha_desde_date:
        fecha_desde = ''
    if not fecha_hasta_date:
        fecha_hasta = ''

    if fecha_desde_date and fecha_hasta_date and fecha_desde_date > fecha_hasta_date:
        fecha_desde_date, fecha_hasta_date = fecha_hasta_date, fecha_desde_date
        fecha_desde, fecha_hasta = fecha_desde_date.isoformat(), fecha_hasta_date.isoformat()

    return fecha_desde, fecha_hasta, fecha_desde_date, fecha_hasta_date


def _filtrar_matriculas(request):
    """
    Aplica filtros comunes a una queryset de Matricula según los GET params.
    Devuelve (queryset filtrada, dict de filtros aplicados).
    """
    qs = Matricula.objects.select_related(
        'estudiante', 'curso', 'curso__categoria', 'jornada', 'registrado_por'
    )

    estado = request.GET.get('estado', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()
    q = request.GET.get('q', '').strip()
    descuento_str = request.GET.get('descuento', '').strip()
    modulo_curso_id = request.GET.get('modulo_curso', '').strip()
    modulo_modalidad = request.GET.get('modulo_modalidad', '').strip().lower()
    modulo_campus = request.GET.get('modulo_campus', '').strip()
    modulo_numero = request.GET.get('modulo_numero', '').strip()
    modulo_estado = request.GET.get('modulo_estado', '').strip()
    fecha_desde, fecha_hasta, fecha_desde_date, fecha_hasta_date = (
        _rango_fecha_matricula_desde_request(request)
    )

    if curso_id:
        qs = qs.filter(curso_id=curso_id)
    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)
    if anio.isdigit():
        qs = qs.filter(fecha_matricula__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(fecha_matricula__month=int(mes))
    if fecha_desde_date:
        qs = qs.filter(fecha_matricula__gte=fecha_desde_date)
    if fecha_hasta_date:
        qs = qs.filter(fecha_matricula__lte=fecha_hasta_date)
    if q:
        qs = filtrar_queryset_busqueda(qs, q, [
            'estudiante__cedula',
            'estudiante__nombres',
            'estudiante__correo',
            'estudiante__celular',
            'curso__nombre',
            'fact_cedula',
            'fact_nombres',
        ])
        
    if descuento_str == 'si':
        qs = qs.filter(tiene_descuento=True)
    elif descuento_str == 'no':
        qs = qs.filter(tiene_descuento=False)

    # Filtro por estado financiero y por falta de pago de módulos.
    if estado == 'Retiro':
        qs = qs.filter(estado='retiro_voluntario')
    elif estado == 'Pagado':
        qs = qs.filter(valor_pagado__gte=models_F('valor_curso')).exclude(estado='retiro_voluntario')
    elif estado == 'Parcial':
        qs = qs.filter(valor_pagado__gt=0, valor_pagado__lt=models_F('valor_curso')).exclude(estado='retiro_voluntario')
    elif estado == 'Pendiente':
        qs = qs.filter(Q(valor_pagado=0) | Q(valor_pagado__isnull=True)).exclude(estado='retiro_voluntario')
    elif estado == 'ModuloPendiente':
        ids = [
            m.pk for m in qs.exclude(estado='retiro_voluntario').select_related('curso')
            if _tiene_modulo_pendiente(m)
        ]
        qs = qs.filter(pk__in=ids)
    elif estado == 'Recuperacion':
        # Recibos actuales de Recuperación y asociaciones históricas ya pagadas.
        qs = qs.filter(
            Q(abonos__tipo_pago='recuperacion')
            | Q(
                recuperaciones_pendientes__pagada=True,
                recuperaciones_pendientes__abono__isnull=False,
            )
        ).distinct()

    # Filtro independiente por curso + módulo + estado del módulo.
    # "Pendiente de pago" incluye módulos sin pago o con pago parcial.
    modulo_curso = None
    if modulo_curso_id.isdigit():
        modulo_curso = Curso.objects.filter(pk=int(modulo_curso_id)).first()
        qs = qs.filter(curso_id=int(modulo_curso_id))
    else:
        modulo_curso_id = ''

    if (
        modulo_curso
        and modulo_modalidad in ('presencial', 'online')
        and modulo_curso.ofrece(modulo_modalidad)
    ):
        qs = qs.filter(modalidad=modulo_modalidad)
    else:
        modulo_modalidad = ''

    campus_validos = {
        opcion['value']
        for opcion in _opciones_campus_curso_para_filtro(
            modulo_curso,
            modulo_modalidad,
        )
    }
    if modulo_campus in campus_validos:
        qs = _filtrar_matriculas_por_campus_modulo(qs, modulo_campus)
    else:
        modulo_campus = ''

    numeros_validos = {
        str(opcion['numero'])
        for opcion in _opciones_modulos_curso_para_filtro(
            modulo_curso,
            modulo_modalidad,
        )
    }
    if modulo_numero not in numeros_validos:
        modulo_numero = ''

    if (
        modulo_curso_id
        and modulo_numero
        and modulo_estado in ('pendiente', 'pagado')
    ):
        numero = int(modulo_numero)
        ids = [
            m.pk for m in qs.exclude(estado='retiro_voluntario').select_related('curso')
            if _matricula_coincide_filtro_modulo(m, numero, modulo_estado)
        ]
        qs = qs.filter(pk__in=ids)
    else:
        modulo_numero = '' if not modulo_numero.isdigit() else modulo_numero
        modulo_estado = '' if modulo_estado not in ('pendiente', 'pagado') else modulo_estado

    return qs, {
        'estado': estado,
        'curso': curso_id,
        'modalidad': modalidad,
        'anio': anio,
        'mes': mes,
        'q': q,
        'descuento': descuento_str,
        'modulo_curso': modulo_curso_id,
        'modulo_modalidad': modulo_modalidad,
        'modulo_campus': modulo_campus,
        'modulo_numero': modulo_numero,
        'modulo_estado': modulo_estado,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }


def _tiene_modulo_pendiente(matricula):
    """
    Devuelve True si a la matrícula le falta al menos un módulo por pagar.
    Un abono general no cuenta como pago de módulo; solo los pagos por módulo
    o recuperaciones aplicadas a módulo alimentan `desglose_pagos_por_modulo`.
    Si el curso ya está totalmente pagado, el modelo marca todos los módulos
    como pagados.
    """
    if matricula.estado == 'retiro_voluntario':
        return False
    return any(
        modulo['estado'] != 'Pagado'
        for modulo in matricula.desglose_pagos_por_modulo()
    )


def _detalle_modulo_pago(matricula, numero_modulo):
    """Devuelve el detalle de pago de un módulo concreto de una matrícula."""
    try:
        numero = int(numero_modulo)
    except (TypeError, ValueError):
        return None

    for modulo in matricula.desglose_pagos_por_modulo():
        if modulo['numero'] != numero:
            continue

        estado_real = modulo['estado']
        saldo_modulo = max(
            (modulo.get('esperado') or Decimal('0.00')) - (modulo.get('pagado') or Decimal('0.00')),
            Decimal('0.00'),
        )
        estado_codigo = 'pagado' if estado_real == 'Pagado' else 'pendiente'
        return {
            'numero': numero,
            'label': modulo.get('label') or f'Módulo {numero}',
            'estado': estado_real,
            'estado_codigo': estado_codigo,
            'pagado': modulo.get('pagado') or Decimal('0.00'),
            'esperado': modulo.get('esperado') or Decimal('0.00'),
            'saldo': saldo_modulo,
            'fecha_ultimo_pago': modulo.get('fecha_ultimo_pago'),
        }
    return None


def _matricula_coincide_filtro_modulo(matricula, numero_modulo, estado_filtro):
    detalle = _detalle_modulo_pago(matricula, numero_modulo)
    if not detalle:
        return False
    if estado_filtro == 'pagado':
        return detalle['estado'] == 'Pagado'
    if estado_filtro == 'pendiente':
        return detalle['estado'] in ('Pendiente', 'Parcial')
    return False


def _adjuntar_detalle_modulo_filtrado(matriculas, filtros):
    if not (
        filtros.get('modulo_curso')
        and filtros.get('modulo_numero')
        and filtros.get('modulo_estado') in ('pendiente', 'pagado')
    ):
        return matriculas

    for matricula in matriculas:
        matricula.modulo_filtro_detalle = _detalle_modulo_pago(
            matricula,
            filtros['modulo_numero'],
        )
    return matriculas


def _max_modulos_curso_para_filtro(curso):
    cantidades = []
    if curso.ofrece_presencial:
        cantidades.append(curso.numero_modulos or 1)
    if curso.ofrece_online:
        cantidades.append(
            1 if curso.usa_pago_unico_recaudacion('online')
            else (curso.numero_modulos_online or 1)
        )
    return max(cantidades or [1])


def _nombre_modulo_curso_para_filtro(curso, numero_modulo, modalidad=''):
    if not curso.nombrar_modulos or not isinstance(curso.nombres_modulos, dict):
        return ''

    nombres = []
    modalidades = (
        (modalidad,)
        if modalidad in ('presencial', 'online')
        else ('presencial', 'online')
    )
    for modalidad_actual in modalidades:
        if not curso.ofrece(modalidad_actual):
            continue
        total = (
            1 if curso.usa_pago_unico_recaudacion(modalidad_actual)
            else curso.get_numero_modulos(modalidad_actual)
        )
        if numero_modulo > total:
            continue
        nombres_modalidad = curso.nombres_modulos.get(modalidad_actual, [])
        if 1 <= numero_modulo <= len(nombres_modalidad):
            nombre = (nombres_modalidad[numero_modulo - 1] or '').strip()
            if nombre:
                nombres.append(nombre)

    nombres_unicos = sorted(set(nombres))
    return nombres_unicos[0] if len(nombres_unicos) == 1 else ''


def _opciones_modulos_curso_para_filtro(curso, modalidad=''):
    if not curso:
        return []
    if modalidad in ('presencial', 'online'):
        if not curso.ofrece(modalidad):
            return []
        total = (
            1 if curso.usa_pago_unico_recaudacion(modalidad)
            else curso.get_numero_modulos(modalidad)
        )
    else:
        total = _max_modulos_curso_para_filtro(curso)
    opciones = []
    for numero in range(1, total + 1):
        nombre = _nombre_modulo_curso_para_filtro(
            curso,
            numero,
            modalidad,
        )
        opciones.append({
            'numero': numero,
            'label': f'Módulo {numero}' + (f' - {nombre}' if nombre else ''),
        })
    return opciones


def _opciones_modalidades_curso_para_filtro(curso):
    if not curso:
        return []
    opciones = []
    if curso.ofrece_presencial:
        opciones.append({'value': 'presencial', 'label': 'Presencial'})
    if curso.ofrece_online:
        opciones.append({'value': 'online', 'label': 'Virtual / Online'})
    return opciones


def _opciones_campus_curso_para_filtro(
    curso,
    modalidad='presencial',
    campus_por_curso=None,
):
    """Campus con matrículas del curso; no aplica a la modalidad online."""
    if not curso or modalidad != 'presencial' or not curso.ofrece_presencial:
        return []
    if campus_por_curso is not None:
        return campus_por_curso.get(str(curso.pk), [])

    return _mapa_campus_cursos_para_filtro([curso]).get(str(curso.pk), [])


def _mapa_campus_cursos_para_filtro(cursos):
    """Carga en una consulta los campus usados en matrículas presenciales."""
    curso_ids = [curso.pk for curso in cursos if curso.ofrece_presencial]
    if not curso_ids:
        return {}

    matriculas = Matricula.objects.filter(
        curso_id__in=curso_ids,
        modalidad='presencial',
        jornada__isnull=False,
    ).select_related('jornada__sede').order_by(
        'jornada__sede__pais',
        'jornada__sede__orden',
        'jornada__sede__nombre',
        'jornada__ciudad',
    )
    opciones_por_curso = defaultdict(dict)
    for matricula in matriculas:
        jornada = matricula.jornada
        if jornada.sede_id:
            value = f'sede:{jornada.sede_id}'
            label = jornada.sede.etiqueta
        else:
            ciudad = (jornada.ciudad or '').strip()
            if not ciudad:
                continue
            value = f'ciudad:{ciudad}'
            label = ciudad
        opciones_por_curso[str(matricula.curso_id)].setdefault(
            value,
            {'value': value, 'label': label},
        )
    return {
        curso_id: list(opciones.values())
        for curso_id, opciones in opciones_por_curso.items()
    }


def _filtrar_matriculas_por_campus_modulo(qs, campus):
    tipo, separador, valor = campus.partition(':')
    if not separador or not valor:
        return qs.none()
    if tipo == 'sede' and valor.isdigit():
        return qs.filter(jornada__sede_id=int(valor))
    if tipo == 'ciudad':
        return qs.filter(jornada__ciudad__iexact=valor)
    return qs.none()


def _payload_cursos_modulo_filtro(cursos, campus_por_curso=None):
    payload = {}
    for curso in cursos:
        modalidades = []
        for modalidad in _opciones_modalidades_curso_para_filtro(curso):
            codigo = modalidad['value']
            modalidades.append({
                **modalidad,
                'modulos': _opciones_modulos_curso_para_filtro(curso, codigo),
                'campus': _opciones_campus_curso_para_filtro(
                    curso,
                    codigo,
                    campus_por_curso,
                ),
            })
        payload[str(curso.pk)] = {
            'nombre': curso.nombre,
            # `modulos` conserva compatibilidad con enlaces antiguos que no
            # incluían modalidad. La interfaz nueva usa los de cada modalidad.
            'modulos': _opciones_modulos_curso_para_filtro(curso),
            'modalidades': modalidades,
        }
    return payload


def _modalidad_filtro_label(modalidad):
    if modalidad == 'presencial':
        return 'Presencial'
    if modalidad == 'online':
        return 'Virtual / Online'
    return ''


def _modulo_filtro_estado_label(estado):
    if estado == 'pagado':
        return 'Pagado'
    if estado == 'pendiente':
        return 'Pendiente de pago'
    return ''


def _resumen_abonos(abonos):
    """
    Agrupa los pagos de una matrícula por tipo de pago y por método.
    Devuelve estructuras simples para poder pintarlas en tablas y exportarlas.
    """
    tipos = {}
    metodos = {}

    for abono in abonos:
        monto = abono.monto or Decimal('0.00')
        monto_2 = abono.monto_2 or Decimal('0.00')
        monto_1 = monto - monto_2 if monto_2 > 0 else monto

        tipo_label = abono.get_tipo_pago_display()
        if not abono.cuenta_para_saldo and abono.tipo_pago == 'recuperacion':
            tipo_label = f'{tipo_label} (aparte)'
        tipo = tipos.setdefault(tipo_label, {'label': tipo_label, 'total': Decimal('0.00'), 'count': 0})
        tipo['total'] += monto
        tipo['count'] += 1

        # Primer método
        metodo_label = abono.get_metodo_display()
        if abono.metodo in ('transferencia', 'tarjeta') and abono.banco:
            metodo_label = f'{metodo_label} · {abono.get_banco_display()}'
        metodo = metodos.setdefault(metodo_label, {'label': metodo_label, 'total': Decimal('0.00'), 'count': 0})
        metodo['total'] += monto_1
        metodo['count'] += 1

        # Segundo método (si es mixto)
        if monto_2 > 0 and abono.metodo_2:
            metodo2_label = dict(Abono.METODOS).get(abono.metodo_2, abono.metodo_2)
            if abono.metodo_2 in ('transferencia', 'tarjeta') and abono.banco_2:
                banco2_label = dict(Abono.BANCOS).get(abono.banco_2, abono.banco_2)
                metodo2_label = f'{metodo2_label} · {banco2_label}'
            metodo2 = metodos.setdefault(metodo2_label, {'label': metodo2_label, 'total': Decimal('0.00'), 'count': 0})
            metodo2['total'] += monto_2
            metodo2['count'] += 1

    return {
        'tipos': list(tipos.values()),
        'metodos': list(metodos.values()),
        'total_movimientos': sum((x['count'] for x in tipos.values()), 0),
    }


def _partes_pago_abono(abono):
    """Devuelve las partes reales de un abono, separando pagos mixtos."""
    monto = abono.monto or Decimal('0.00')
    monto_2 = abono.monto_2 or Decimal('0.00')
    monto_1 = monto - monto_2 if monto_2 > 0 else monto
    partes = []
    if monto_1 > 0:
        partes.append({
            'monto': monto_1,
            'metodo': abono.metodo,
            'metodo_display': abono.get_metodo_display(),
            'banco': abono.banco,
            'banco_display': abono.get_banco_display(),
        })
    if monto_2 > 0 and abono.metodo_2:
        bancos_map = dict(Abono.BANCOS)
        metodos_map = dict(Abono.METODOS)
        partes.append({
            'monto': monto_2,
            'metodo': abono.metodo_2,
            'metodo_display': metodos_map.get(abono.metodo_2, abono.metodo_2),
            'banco': abono.banco_2,
            'banco_display': bancos_map.get(abono.banco_2, abono.banco_2) if abono.banco_2 else '',
        })
    return partes


def _abono_corresponde_a_recuperacion(abono):
    """Reconoce recibos actuales y asociaciones históricas de recuperación."""
    if abono.tipo_pago == 'recuperacion':
        return True
    return any(
        recuperacion.pagada
        for recuperacion in abono.recuperaciones.all()
    )


def _adjuntar_resumen_abonos(matriculas):
    """Agrega a cada matrícula el resumen de sus abonos ya prefetched."""
    for matricula in matriculas:
        abonos = getattr(matricula, 'abonos_para_resumen', None)
        if abonos is None:
            abonos = list(matricula.abonos.all())
        matricula.resumen_abonos = _resumen_abonos(abonos)
        pagos_recuperacion = [
            abono for abono in abonos
            if _abono_corresponde_a_recuperacion(abono)
        ]
        matricula.tiene_pago_recuperacion = bool(pagos_recuperacion)
        matricula.total_pago_recuperacion = sum(
            (abono.monto or Decimal('0.00') for abono in pagos_recuperacion),
            Decimal('0.00'),
        )
    return matriculas


# Importación tardía para evitar circular imports en algunos casos
from django.db.models import F as models_F


# ═════════════════════════════════════════════════════════════════
# Pagos
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
def pagos_lista(request):
    """
    Vista centrada en lo financiero. Permite filtrar matrículas por:
    - Estado de pago: Pagado, Parcial, Pendiente
    - Curso (buscador)
    - Modalidad
    """
    qs, filtros = _filtrar_matriculas(request)
    qs = qs.prefetch_related(
        Prefetch(
            'abonos',
            queryset=Abono.objects.prefetch_related('recuperaciones').order_by('fecha', 'id'),
            to_attr='abonos_para_resumen',
        )
    ).order_by('-creado', '-id')
    matriculas = _adjuntar_resumen_abonos(list(qs))
    matriculas = _adjuntar_detalle_modulo_filtrado(matriculas, filtros)

    # Estadísticas globales (con los filtros aplicados, excepto el de estado)
    qs_sin_estado = Matricula.objects.select_related('curso').all()
    if filtros['curso']:
        qs_sin_estado = qs_sin_estado.filter(curso_id=filtros['curso'])
    if filtros['modalidad']:
        qs_sin_estado = qs_sin_estado.filter(modalidad=filtros['modalidad'])
    if filtros['anio'].isdigit():
        qs_sin_estado = qs_sin_estado.filter(fecha_matricula__year=int(filtros['anio']))
    if filtros['mes'].isdigit() and 1 <= int(filtros['mes']) <= 12:
        qs_sin_estado = qs_sin_estado.filter(fecha_matricula__month=int(filtros['mes']))
    fecha_desde_date = parse_date(filtros['fecha_desde']) if filtros['fecha_desde'] else None
    fecha_hasta_date = parse_date(filtros['fecha_hasta']) if filtros['fecha_hasta'] else None
    if fecha_desde_date:
        qs_sin_estado = qs_sin_estado.filter(fecha_matricula__gte=fecha_desde_date)
    if fecha_hasta_date:
        qs_sin_estado = qs_sin_estado.filter(fecha_matricula__lte=fecha_hasta_date)
    if filtros['q']:
        qs_sin_estado = filtrar_queryset_busqueda(qs_sin_estado.select_related('estudiante', 'curso'), filtros['q'], [
            'estudiante__cedula',
            'estudiante__nombres',
            'estudiante__correo',
            'estudiante__celular',
            'curso__nombre',
            'fact_cedula',
            'fact_nombres',
        ])

    totales = {
        'total_matriculas': qs_sin_estado.count(),
        'total_facturado': Decimal('0.00'),
        'total_cobrado': qs_sin_estado.aggregate(s=Sum('valor_pagado'))['s'] or Decimal('0.00'),
        'total_pendiente': Decimal('0.00'),
        'total_retiro': Decimal('0.00'),
    }

    # Conteo por estado y recálculo de saldo pendiente sin retiros
    todos_los_pagos = list(qs_sin_estado.values('valor_curso', 'descuento', 'valor_pagado', 'estado'))
    conteo_estado = {'Pagado': 0, 'Parcial': 0, 'Pendiente': 0, 'Retiro': 0}
    conteo_modulos_pendientes = 0

    for matricula in qs_sin_estado.select_related('curso').exclude(estado='retiro_voluntario'):
        if _tiene_modulo_pendiente(matricula):
            conteo_modulos_pendientes += 1
    
    for p in todos_los_pagos:
        vc = p['valor_curso'] or Decimal('0.00')
        desc = p['descuento'] or Decimal('0.00')
        vn = max(Decimal('0.00'), vc - desc)
        vp = p['valor_pagado'] or Decimal('0.00')
        st = p['estado']
        
        totales['total_facturado'] += vn
        
        if st == 'retiro_voluntario':
            conteo_estado['Retiro'] += 1
            totales['total_retiro'] += max(Decimal('0.00'), vn - vp)
        else:
            totales['total_pendiente'] += max(Decimal('0.00'), vn - vp)
            if vp >= vn:
                conteo_estado['Pagado'] += 1
            elif vp > 0:
                conteo_estado['Parcial'] += 1
            else:
                conteo_estado['Pendiente'] += 1

    cursos = list(Curso.objects.filter(activo=True).order_by('nombre'))
    anios = sorted(
        set(Matricula.objects.dates('fecha_matricula', 'year').values_list('fecha_matricula__year', flat=True)),
        reverse=True
    )
    filtros_query = urlencode({
        key: value for key, value in filtros.items() if value
    })
    filtros_query_sin_modulo = urlencode({
        key: value for key, value in filtros.items()
        if value and key not in (
            'modulo_curso', 'modulo_modalidad', 'modulo_campus',
            'modulo_numero', 'modulo_estado',
        )
    })

    modulo_curso_sel = next(
        (curso for curso in cursos if str(curso.pk) == filtros.get('modulo_curso')),
        None,
    )
    modulo_filtro_modalidades = _opciones_modalidades_curso_para_filtro(
        modulo_curso_sel
    )
    campus_modulo_por_curso = _mapa_campus_cursos_para_filtro(cursos)
    modulo_filtro_campus = _opciones_campus_curso_para_filtro(
        modulo_curso_sel,
        filtros.get('modulo_modalidad'),
        campus_modulo_por_curso,
    )
    modulo_filtro_opciones = _opciones_modulos_curso_para_filtro(
        modulo_curso_sel,
        filtros.get('modulo_modalidad'),
    )
    modulo_filtro_activo = bool(
        modulo_curso_sel
        and filtros.get('modulo_numero')
        and filtros.get('modulo_estado') in ('pendiente', 'pagado')
    )
    modulo_filtro_label = ''
    if filtros.get('modulo_numero'):
        modulo_filtro_label = next(
            (
                opcion['label'] for opcion in modulo_filtro_opciones
                if str(opcion['numero']) == filtros['modulo_numero']
            ),
            f"Módulo {filtros['modulo_numero']}",
        )

    return render(request, 'pagos/lista.html', {
        'matriculas': matriculas,
        'cursos': cursos,
        'anios': anios,
        'filtros': filtros,
        'filtros_query': filtros_query,
        'filtros_query_sin_modulo': filtros_query_sin_modulo,
        'hay_filtros': bool(filtros_query),
        'totales': totales,
        'conteo_estado': conteo_estado,
        'conteo_modulos_pendientes': conteo_modulos_pendientes,
        'cursos_modulo_filtro': _payload_cursos_modulo_filtro(
            cursos,
            campus_modulo_por_curso,
        ),
        'modulo_filtro_modalidades': modulo_filtro_modalidades,
        'modulo_filtro_campus': modulo_filtro_campus,
        'modulo_filtro_opciones': modulo_filtro_opciones,
        'modulo_filtro_activo': modulo_filtro_activo,
        'modulo_filtro_resumen': {
            'curso': modulo_curso_sel.nombre if modulo_curso_sel else '',
            'modalidad': _modalidad_filtro_label(
                filtros.get('modulo_modalidad')
            ),
            'campus': next(
                (
                    opcion['label'] for opcion in modulo_filtro_campus
                    if opcion['value'] == filtros.get('modulo_campus')
                ),
                '',
            ),
            'modulo': modulo_filtro_label,
            'estado': _modulo_filtro_estado_label(filtros.get('modulo_estado')),
        },
    })


@matricula_requerida
def pagos_export(request):
    """Descarga los pagos filtrados como Excel."""
    qs, filtros = _filtrar_matriculas(request)
    qs = qs.prefetch_related(
        Prefetch('abonos', queryset=Abono.objects.order_by('fecha', 'id'), to_attr='abonos_para_resumen')
    ).order_by('-creado', '-id')

    headers = [
        'Fecha matrícula', 'Cédula', 'Estudiante',
        'Curso', 'Categoría', 'Modalidad', 'Sede / Plataforma',
        'Jornada', 'Día (inicio jornada)', 'Horario',
        'Valor curso', 'Valor pagado', 'Saldo', 'Estado',
        'Tipos de pago', 'Métodos de pago',
        'Asistencia',
    ]

    rows = []
    total_curso = Decimal('0.00')
    total_pagado = Decimal('0.00')
    total_saldo = Decimal('0.00')

    for m in _adjuntar_resumen_abonos(list(qs)):
        tipos_pago = '; '.join(
            f"{x['label']}: ${x['total']:.2f} ({x['count']})"
            for x in m.resumen_abonos['tipos']
        ) or 'Sin pagos'
        metodos_pago = '; '.join(
            f"{x['label']}: ${x['total']:.2f} ({x['count']})"
            for x in m.resumen_abonos['metodos']
        ) or 'Sin pagos'

        # ── Datos de jornada ──
        if m.jornada:
            jornada_txt = m.jornada.descripcion_legible
            dia_inicio = m.jornada.fecha_inicio if m.jornada.fecha_inicio else ''
            if m.jornada.hora_inicio and m.jornada.hora_fin:
                horario_txt = f"{m.jornada.hora_inicio.strftime('%H:%M')} – {m.jornada.hora_fin.strftime('%H:%M')}"
            else:
                horario_txt = '—'
        else:
            jornada_txt = '—'
            dia_inicio = '—'
            horario_txt = '—'

        rows.append([
            m.fecha_matricula if m.fecha_matricula else '',
            m.estudiante.cedula,
            m.estudiante.nombres,
            m.curso.nombre,
            m.curso.categoria.nombre if m.curso.categoria else '—',
            m.get_modalidad_display(),
            m.sede,
            jornada_txt,
            dia_inicio,
            horario_txt,
            float(m.valor_curso or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
            m.estado_pago,
            tipos_pago,
            metodos_pago,
            '',  # Asistencia: columna en blanco para firma
        ])
        total_curso += m.valor_curso or Decimal('0.00')
        total_pagado += m.valor_pagado or Decimal('0.00')
        total_saldo += m.saldo or Decimal('0.00')

    totals = {
        10: float(total_curso),
        11: float(total_pagado),
        12: float(total_saldo),
    }

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = ''
    if filtros['estado']:
        sufijo += f"_{filtros['estado'].lower()}"
    if filtros['anio']:
        sufijo += f"_{filtros['anio']}"
    filename = f'pagos{sufijo}_{fecha_str}.xlsx'

    return _build_excel_response(
        filename=filename,
        sheet_name='Reporte de Pagos',
        headers=headers,
        rows=rows,
        totals=totals,
        column_formats={
            0: 'dd/mm/yyyy',
            8: 'dd/mm/yyyy',
            10: '"$"#,##0.00',
            11: '"$"#,##0.00',
            12: '"$"#,##0.00',
        },
        text_columns={1},
        explicit_widths={
            0: 15, 1: 15, 2: 26, 3: 28, 4: 18, 5: 14,
            6: 18, 7: 24, 8: 18, 9: 16, 10: 14, 11: 14,
            12: 12, 13: 14, 14: 34, 15: 34, 16: 16,
        },
    )


@matricula_requerida
def pagos_export_pdf(request):
    """Descarga los pagos filtrados como PDF horizontal con columna de asistencia."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A3
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    qs, filtros = _filtrar_matriculas(request)
    qs = qs.prefetch_related(
        Prefetch('abonos', queryset=Abono.objects.order_by('fecha', 'id'), to_attr='abonos_para_resumen')
    ).order_by('-creado', '-id')
    matriculas = _adjuntar_resumen_abonos(list(qs))

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A3),
        leftMargin=0.7*cm, rightMargin=0.7*cm, topMargin=1.0*cm, bottomMargin=0.8*cm,
        title='Estado de Pagos',
    )
    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle(
        'titulo_pagos', parent=styles['Title'],
        textColor=colors.HexColor('#1A237E'), fontSize=15,
        alignment=1, spaceAfter=4,
    )
    sub_st = ParagraphStyle(
        'sub_pagos', parent=styles['Normal'],
        textColor=colors.HexColor('#666666'), fontSize=9,
        alignment=1, spaceAfter=10,
    )
    cell_st = ParagraphStyle(
        'cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7, leading=8.5,
    )
    cell_bold_st = ParagraphStyle(
        'cell_bold', parent=cell_st, fontName='Helvetica-Bold',
    )

    elementos = [
        Paragraph('Estado de Pagos — Formación Profesional EC', titulo_st),
        Paragraph(
            f'Generado el {date.today().strftime("%d/%m/%Y")} · '
            f'{len(matriculas)} registro(s)',
            sub_st,
        ),
    ]

    headers = [
        'Fecha', 'Cédula', 'Estudiante', 'Curso', 'Modalidad',
        'Jornada', 'Día', 'Valor', 'Pagado', 'Tipo de pago',
        'Método', 'Saldo', 'Estado', 'Asistencia',
    ]
    # Header como Paragraphs (los wraps se hacen automáticamente)
    header_st = ParagraphStyle(
        'header_st', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8,
        textColor=colors.whitesmoke, alignment=1, leading=9,
    )
    data = [[Paragraph(h, header_st) for h in headers]]
    total_curso = total_pagado = total_saldo = 0.0

    for m in matriculas:
        if m.jornada:
            jornada_txt = m.jornada.descripcion_legible or '—'
            dia_inicio = m.jornada.fecha_inicio.strftime('%d/%m/%Y') if m.jornada.fecha_inicio else '—'
        else:
            jornada_txt = '—'
            dia_inicio = '—'

        tipos_pago = '<br/>'.join(
            f"{x['label']}: ${x['total']:.2f}"
            for x in m.resumen_abonos['tipos']
        ) or 'Sin pagos'
        metodos_pago = '<br/>'.join(
            f"{x['label']}: ${x['total']:.2f}"
            for x in m.resumen_abonos['metodos']
        ) or 'Sin método'

        estado_txt = m.estado_pago or '—'

        data.append([
            Paragraph(m.fecha_matricula.strftime('%d/%m/%Y') if m.fecha_matricula else '', cell_st),
            Paragraph(m.estudiante.cedula or '', cell_bold_st),
            Paragraph(f'{m.estudiante.nombres}'.strip(), cell_st),
            Paragraph(m.curso.nombre or '', cell_st),
            Paragraph(m.get_modalidad_display() or '', cell_st),
            Paragraph(jornada_txt, cell_st),
            Paragraph(dia_inicio, cell_st),
            Paragraph(f'${float(m.valor_curso or 0):.2f}', cell_st),
            Paragraph(f'<font color="#2e7d32"><b>${float(m.valor_pagado or 0):.2f}</b></font>', cell_st),
            Paragraph(tipos_pago, cell_st),
            Paragraph(metodos_pago, cell_st),
            Paragraph(f'<font color="{"#c62828" if (m.saldo or 0) > 0 else "#2e7d32"}"><b>${float(m.saldo or 0):.2f}</b></font>', cell_st),
            Paragraph(estado_txt, cell_st),
            '',  # Asistencia (firma) — vacío para llenar a mano
        ])
        total_curso += float(m.valor_curso or 0)
        total_pagado += float(m.valor_pagado or 0)
        total_saldo += float(m.saldo or 0)

    # Fila de totales
    data.append([
        Paragraph('', cell_st), Paragraph('', cell_st), Paragraph('', cell_st),
        Paragraph('', cell_st), Paragraph('', cell_st), Paragraph('', cell_st),
        Paragraph('<b>TOTAL</b>', cell_bold_st),
        Paragraph(f'<b>${total_curso:.2f}</b>', cell_bold_st),
        Paragraph(f'<b>${total_pagado:.2f}</b>', cell_bold_st),
        Paragraph('', cell_st), Paragraph('', cell_st),
        Paragraph(f'<b>${total_saldo:.2f}</b>', cell_bold_st),
        Paragraph('', cell_st), '',
    ])

    # ── Anchos de columna explícitos para A3 horizontal (≈ 41 cm de ancho útil) ──
    # Suman ~39.5 cm dejando margen
    col_widths = [
        1.7*cm,  # Fecha
        2.0*cm,  # Cédula
        4.5*cm,  # Estudiante
        3.5*cm,  # Curso
        1.8*cm,  # Modalidad
        3.5*cm,  # Jornada
        1.8*cm,  # Día
        1.8*cm,  # Valor
        1.8*cm,  # Pagado
        4.8*cm,  # Tipo de pago
        4.8*cm,  # Método
        1.8*cm,  # Saldo
        2.0*cm,  # Estado
        3.7*cm,  # Asistencia
    ]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#BBBBBB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FB')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E1')),
        # Línea de firma en columna Asistencia
        ('LINEBELOW', (-1, 1), (-1, -2), 0.5, colors.HexColor('#888888')),
    ]))
    elementos.append(table)
    doc.build(elementos)

    pdf_bytes = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    fecha_str = date.today().strftime('%Y%m%d')
    sufijo = ''
    if filtros['estado']:
        sufijo += f"_{filtros['estado'].lower()}"
    filename = f'pagos{sufijo}_{fecha_str}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Historial de matriculados (por año / mes)
# ═════════════════════════════════════════════════════════════════

class _HistorialItemArchivado:
    """
    Adaptador que envuelve una MatriculaArchivada y expone la MISMA interfaz
    que una Matricula viva, para que el template del historial la pinte sin
    cambios. Marca `es_archivada=True` para que el template pueda mostrar una
    etiqueta distintiva.
    """

    def __init__(self, archivada):
        self._a = archivada
        self.es_archivada = True
        self.fecha_matricula = archivada.fecha_matricula
        self.modalidad = archivada.modalidad
        self.valor_curso = archivada.valor_curso
        self.descuento = archivada.descuento
        self.valor_neto = archivada.valor_neto
        self.valor_pagado = archivada.valor_pagado or Decimal('0.00')
        self.saldo = archivada.saldo or Decimal('0.00')
        self.estado_pago = archivada.estado_pago
        self.sede = archivada.sede or '—'
        self._orden_id = archivada.pk
        # Datos para el cierre asociado (por si el template quiere enlazar)
        self.cierre_id = archivada.cierre_id

    @property
    def tiene_descuento(self):
        return self.descuento > 0

    def get_modalidad_display(self):
        return 'Online' if self.modalidad == 'online' else 'Presencial'

    @property
    def estudiante(self):
        """Devuelve un objeto ligero con cedula/nombres/pk."""
        a = self._a
        # Si el estudiante original sigue vivo, usar su pk real para el enlace.
        pk = a.estudiante_id
        return _EstudianteArchivadoProxy(
            cedula=a.cedula,
            
            nombres=a.nombres,
            pk=pk,
        )

    @property
    def curso(self):
        return _CursoProxy(
            nombre=self._a.curso_nombre,
            categoria_nombre=self._a.curso_categoria,
        )


class _EstudianteArchivadoProxy:
    def __init__(self, cedula, nombres, pk):
        self.cedula = cedula
        
        self.nombres = nombres
        self.pk = pk

    @property
    def nombre_completo(self):
        return f'{self.nombres}'.strip()


class _CursoProxy:
    def __init__(self, nombre, categoria_nombre=''):
        self.nombre = nombre
        self.categoria = _CategoriaProxy(categoria_nombre) if categoria_nombre else None


class _CategoriaProxy:
    def __init__(self, nombre):
        self.nombre = nombre


@matricula_requerida
def historial_lista(request):
    """
    Historial de matrículas agrupado por año y mes.
    Permite filtrar por año, mes, curso y modalidad.

    IMPORTANTE: combina matrículas VIVAS + matrículas ARCHIVADAS (de los cierres),
    para que el historial mensual NUNCA se pierda al ejecutar un cierre de curso.
    Las archivadas se muestran con una etiqueta "archivada" pero conservan su
    fecha, curso, modalidad, estado de pago, etc.
    """
    from .models import MatriculaArchivada

    qs, filtros = _filtrar_matriculas(request)
    qs = qs.order_by('-fecha_matricula', '-id')

    # ── También las matrículas archivadas (mismos filtros) ──
    arch_qs = MatriculaArchivada.objects.select_related('cierre', 'estudiante', 'curso')
    if filtros['curso']:
        arch_qs = arch_qs.filter(curso_id=filtros['curso'])
    if filtros['modalidad'] in ('presencial', 'online'):
        arch_qs = arch_qs.filter(modalidad=filtros['modalidad'])
    if filtros['anio'].isdigit():
        arch_qs = arch_qs.filter(fecha_matricula__year=int(filtros['anio']))
    if filtros['mes'].isdigit() and 1 <= int(filtros['mes']) <= 12:
        arch_qs = arch_qs.filter(fecha_matricula__month=int(filtros['mes']))
    if filtros['q']:
        arch_qs = filtrar_queryset_busqueda(arch_qs, filtros['q'], [
            'cedula',
            'nombres',
            'correo',
            'celular',
            'curso_nombre',
            'fact_cedula',
            'fact_nombres',
        ])
    if filtros['estado'] in ('Pagado', 'Parcial', 'Pendiente', 'Retiro'):
        arch_qs = arch_qs.filter(estado_pago=filtros['estado'])

    # Envolver las archivadas en un adaptador con la misma interfaz que Matricula
    items_archivados = [_HistorialItemArchivado(a) for a in arch_qs]

    # Combinar ambas fuentes
    todos = list(qs) + items_archivados

    # Agrupar por año → mes → matrículas
    grupos = defaultdict(lambda: defaultdict(list))
    totales_por_anio = defaultdict(lambda: {'count': 0, 'facturado': Decimal('0.00'), 'cobrado': Decimal('0.00')})
    totales_por_mes = defaultdict(lambda: {'count': 0, 'facturado': Decimal('0.00'), 'cobrado': Decimal('0.00')})

    for m in todos:
        anio = m.fecha_matricula.year
        mes = m.fecha_matricula.month
        grupos[anio][mes].append(m)

        totales_por_anio[anio]['count'] += 1
        totales_por_anio[anio]['facturado'] += m.valor_curso or Decimal('0.00')
        totales_por_anio[anio]['cobrado'] += m.valor_pagado or Decimal('0.00')

        key = (anio, mes)
        totales_por_mes[key]['count'] += 1
        totales_por_mes[key]['facturado'] += m.valor_curso or Decimal('0.00')
        totales_por_mes[key]['cobrado'] += m.valor_pagado or Decimal('0.00')

    # Convertir a lista ordenada para el template
    estructura = []
    for anio in sorted(grupos.keys(), reverse=True):
        meses_dict = grupos[anio]
        meses_lista = []
        for mes in sorted(meses_dict.keys(), reverse=True):
            # Ordenar las matrículas del mes por día descendente
            matriculas_mes = sorted(
                meses_dict[mes],
                key=lambda x: (x.fecha_matricula, getattr(x, '_orden_id', 0)),
                reverse=True
            )
            meses_lista.append({
                'numero': mes,
                'nombre': MESES_ES[mes],
                'matriculas': matriculas_mes,
                'totales': totales_por_mes[(anio, mes)],
            })
        estructura.append({
            'anio': anio,
            'meses': meses_lista,
            'totales': totales_por_anio[anio],
        })

    cursos = Curso.objects.filter(activo=True).order_by('nombre')

    # Años disponibles: vivos + archivados
    anios_vivos = set(Matricula.objects.dates('fecha_matricula', 'year').values_list('fecha_matricula__year', flat=True))
    anios_arch = set(MatriculaArchivada.objects.dates('fecha_matricula', 'year').values_list('fecha_matricula__year', flat=True))
    anios_disponibles = sorted(anios_vivos | anios_arch, reverse=True)

    return render(request, 'historial/lista.html', {
        'estructura': estructura,
        'cursos': cursos,
        'anios': anios_disponibles,
        'meses_es': MESES_ES,
        'filtros': filtros,
        'total_general': len(todos),
        'total_archivadas': len(items_archivados),
    })


@matricula_requerida
def historial_export(request):
    """
    Descarga del historial como Excel. El archivo tiene una hoja por año
    (o una sola si se filtró por año específico).
    Incluye matrículas vivas + archivadas (de cierres).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from .models import MatriculaArchivada

    qs, filtros = _filtrar_matriculas(request)
    qs = qs.order_by('-fecha_matricula', '-id')

    # ── Archivadas con los mismos filtros ──
    arch_qs = MatriculaArchivada.objects.select_related('cierre', 'estudiante', 'curso')
    if filtros['curso']:
        arch_qs = arch_qs.filter(curso_id=filtros['curso'])
    if filtros['modalidad'] in ('presencial', 'online'):
        arch_qs = arch_qs.filter(modalidad=filtros['modalidad'])
    if filtros['anio'].isdigit():
        arch_qs = arch_qs.filter(fecha_matricula__year=int(filtros['anio']))
    if filtros['mes'].isdigit() and 1 <= int(filtros['mes']) <= 12:
        arch_qs = arch_qs.filter(fecha_matricula__month=int(filtros['mes']))
    if filtros['q']:
        arch_qs = filtrar_queryset_busqueda(arch_qs, filtros['q'], [
            'cedula',
            'nombres',
            'correo',
            'celular',
            'curso_nombre',
            'fact_cedula',
            'fact_nombres',
        ])
    if filtros['estado'] in ('Pagado', 'Parcial', 'Pendiente', 'Retiro'):
        arch_qs = arch_qs.filter(estado_pago=filtros['estado'])

    todos = list(qs) + [_HistorialItemArchivado(a) for a in arch_qs]

    # Agrupar por año
    por_anio = defaultdict(list)
    for m in todos:
        por_anio[m.fecha_matricula.year].append(m)

    if not por_anio:
        # Excel vacío con mensaje
        return _build_excel_response(
            filename='historial_vacio.xlsx',
            sheet_name='Historial',
            headers=['Sin datos'],
            rows=[['No hay matrículas con los filtros aplicados.']],
        )

    # Construir el archivo manualmente con varias hojas
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    month_font = Font(bold=True, color='1A237E', size=12)
    month_fill = PatternFill('solid', fgColor='FFF8E1')
    total_font = Font(bold=True, color='2E7D32', size=10)
    total_fill = PatternFill('solid', fgColor='E8F5E9')
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = [
        'Fecha matrícula', 'Cédula', 'Apellidos y Nombres',
        'Curso', 'Modalidad', 'Categoría', 'Sede',
        'Valor curso', 'Pagado', 'Saldo', 'Estado',
    ]

    # Hoja limpia y filtrable: mantiene el historial agrupado en las hojas por año,
    # pero ofrece una tabla real para ordenar/filtrar sin mezclar subtotales.
    ws_data = wb.create_sheet(title='Tabla completa')
    ws_data.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws_data.cell(row=1, column=1, value='Historial de matrículas — Tabla completa')
    title.font = Font(bold=True, size=14, color='1A237E')
    title.alignment = Alignment(horizontal='center', vertical='center')
    ws_data.row_dimensions[1].height = 24
    for col_idx, h in enumerate(headers, start=1):
        cell = ws_data.cell(row=2, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws_data.row_dimensions[2].height = 30

    todos_ordenados = sorted(
        todos,
        key=lambda x: (x.fecha_matricula, getattr(x, '_orden_id', getattr(x, 'id', 0))),
        reverse=True,
    )
    for row_idx, m in enumerate(todos_ordenados, start=3):
        row_data = [
            m.fecha_matricula,
            m.estudiante.cedula,
            m.estudiante.nombre_completo,
            m.curso.nombre,
            m.get_modalidad_display(),
            m.curso.categoria.nombre if m.curso.categoria else '—',
            m.sede,
            float(m.valor_curso or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
            m.estado_pago,
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if col_idx == 1:
                cell.number_format = 'dd/mm/yyyy'
            elif col_idx == 2:
                cell.number_format = '@'
            elif col_idx in (8, 9, 10):
                cell.number_format = '"$"#,##0.00'
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(3, len(todos_ordenados) + 3):
            v = ws_data.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_length = max(max_length, len(str(v)))
        ws_data.column_dimensions[ws_data.cell(row=2, column=col_idx).column_letter].width = min(max_length + 3, 38)
    _add_excel_table(ws_data, 2, 1, len(todos_ordenados) + 2, len(headers), 'Historial Tabla Completa')
    ws_data.freeze_panes = 'A3'

    for anio in sorted(por_anio.keys(), reverse=True):
        ws = wb.create_sheet(title=f'Año {anio}')

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title = ws.cell(row=1, column=1, value=f'Historial de matrículas — {anio}')
        title.font = Font(bold=True, size=14, color='1A237E')
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        # Encabezados
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 30

        # Agrupar por mes dentro del año
        por_mes = defaultdict(list)
        for m in por_anio[anio]:
            por_mes[m.fecha_matricula.month].append(m)
        # Ordenar cada mes por fecha descendente (combina vivas + archivadas)
        for mes_key in por_mes:
            por_mes[mes_key].sort(
                key=lambda x: (x.fecha_matricula, getattr(x, '_orden_id', getattr(x, 'id', 0))),
                reverse=True
            )

        current_row = 3
        total_anio_facturado = Decimal('0.00')
        total_anio_pagado = Decimal('0.00')
        total_anio_saldo = Decimal('0.00')

        for mes in sorted(por_mes.keys(), reverse=True):
            # Fila separadora del mes
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(headers),
            )
            mes_cell = ws.cell(
                row=current_row, column=1,
                value=f'▸ {MESES_ES[mes]} {anio}  ({len(por_mes[mes])} matrícula(s))'
            )
            mes_cell.font = month_font
            mes_cell.fill = month_fill
            mes_cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1

            mes_facturado = Decimal('0.00')
            mes_pagado = Decimal('0.00')
            mes_saldo = Decimal('0.00')

            for m in por_mes[mes]:
                row_data = [
                    m.fecha_matricula.strftime('%d/%m/%Y'),
                    m.estudiante.cedula,
                    m.estudiante.nombre_completo,
                    m.curso.nombre,
                    m.get_modalidad_display(),
                    m.curso.categoria.nombre if m.curso.categoria else '—',
                    m.sede,
                    float(m.valor_curso or 0),
                    float(m.valor_pagado or 0),
                    float(m.saldo or 0),
                    m.estado_pago,
                ]
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                current_row += 1
                mes_facturado += m.valor_curso or Decimal('0.00')
                mes_pagado += m.valor_pagado or Decimal('0.00')
                mes_saldo += m.saldo or Decimal('0.00')

            # Subtotal del mes
            for col_idx in range(1, 8):
                cell = ws.cell(row=current_row, column=col_idx, value='')
                cell.fill = total_fill
                cell.border = thin_border
            ws.cell(row=current_row, column=7, value='Subtotal mes:').font = total_font
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right', vertical='center')

            for col_idx, val in [(8, float(mes_facturado)), (9, float(mes_pagado)),
                                 (10, float(mes_saldo))]:
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.font = total_font
                c.fill = total_fill
                c.border = thin_border
                c.alignment = Alignment(vertical='center')
            ws.cell(row=current_row, column=11, value='').fill = total_fill
            current_row += 2  # espacio extra antes del próximo mes

            total_anio_facturado += mes_facturado
            total_anio_pagado += mes_pagado
            total_anio_saldo += mes_saldo

        # Total del año
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_idx, value='')
            cell.fill = PatternFill('solid', fgColor='1A237E')
        ws.cell(row=current_row, column=7, value=f'TOTAL {anio}:').font = Font(bold=True, color='FFFFFF', size=11)
        ws.cell(row=current_row, column=7).fill = PatternFill('solid', fgColor='1A237E')
        ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right', vertical='center')

        for col_idx, val in [(8, float(total_anio_facturado)), (9, float(total_anio_pagado)),
                             (10, float(total_anio_saldo))]:
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font = Font(bold=True, color='FFFFFF', size=11)
            c.fill = PatternFill('solid', fgColor='1A237E')
            c.alignment = Alignment(vertical='center')

        # Auto-ancho
        for col_idx in range(1, len(headers) + 1):
            max_length = len(headers[col_idx - 1])
            for row_idx in range(3, current_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_length = max(max_length, len(str(v)))
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_length + 3, 38)

        ws.freeze_panes = 'A3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'historial_matriculados_{fecha_str}.xlsx'

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Estudiantes
# ═════════════════════════════════════════════════════════════════

def _sedes_con_matriculas():
    """Sedes usadas por matrículas activas, sin valores vacíos."""
    sedes = (
        Matricula.objects
        .filter(jornada__isnull=False)
        .exclude(jornada__ciudad__isnull=True)
        .exclude(jornada__ciudad__exact='')
        .values_list('jornada__ciudad', flat=True)
        .distinct()
        .order_by('jornada__ciudad')
    )
    return [s for s in sedes if s and s.strip() and s.strip() != '—']


def _filtrar_estudiantes_directorio(qs, q='', sede=''):
    if q:
        qs = filtrar_queryset_busqueda(qs, q, [
            'cedula',
            'nombres',
            'correo',
            'celular',
            'ciudad',
        ])
    if sede:
        qs = qs.filter(matriculas__jornada__ciudad__iexact=sede).distinct()
    return qs


def _matriculas_estudiante_para_resumen(estudiante, sede=''):
    matriculas = list(estudiante.matriculas.all())
    if not sede:
        return matriculas
    sede_normalizada = sede.strip().casefold()
    return [
        m for m in matriculas
        if (m.sede or '').strip().casefold() == sede_normalizada
    ]


@matricula_requerida
def estudiantes_lista(request):
    """
    Listado de estudiantes con búsqueda. Cada estudiante muestra el conteo
    de cursos matriculados, sus jornadas/sedes y un enlace al detalle.
    """
    q = request.GET.get('q', '').strip()
    sede = request.GET.get('sede', '').strip()
    qs = Estudiante.objects.annotate(
        num_matriculas=Count('matriculas', distinct=True)
    ).prefetch_related(
        'matriculas__jornada', 'matriculas__jornada__sede', 'matriculas__curso'
    ).order_by('nombres')

    sedes = _sedes_con_matriculas()
    if sede and sede not in sedes:
        sede = ''

    qs = _filtrar_estudiantes_directorio(qs, q=q, sede=sede)

    # Construir, por estudiante, el resumen de jornada(s) y sede(s) de sus matrículas.
    estudiantes = list(qs)
    for e in estudiantes:
        jornadas_set = []
        sedes_set = []
        matriculas_resumen = _matriculas_estudiante_para_resumen(e, sede=sede)
        if sede:
            e.num_matriculas = len(matriculas_resumen)
        for m in matriculas_resumen:
            if m.jornada_id:
                etiqueta_dia = m.jornada.descripcion_legible
                if etiqueta_dia and etiqueta_dia not in jornadas_set:
                    jornadas_set.append(etiqueta_dia)
            sede_nombre = (m.sede or '').strip()
            if sede_nombre and sede_nombre != '—' and sede_nombre not in sedes_set:
                sedes_set.append(sede_nombre)
        e.jornadas_resumen = ' · '.join(jornadas_set) if jornadas_set else '—'
        e.sedes_resumen = ' · '.join(sedes_set) if sedes_set else '—'

    return render(request, 'estudiantes/lista.html', {
        'estudiantes': estudiantes,
        'q': q,
        'sede': sede,
        'sedes': sedes,
        'filtros_query': urlencode({
            k: v for k, v in {'q': q, 'sede': sede}.items() if v
        }),
        'total': len(estudiantes),
    })


@matricula_requerida
def estudiantes_por_curso(request):
    """
    Estudiantes agrupados por curso. Útil cuando se quiere ver la nómina
    completa de un curso específico.
    """
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()

    cursos_qs = Curso.objects.filter(activo=True).order_by('nombre')

    grupos = []
    for curso in cursos_qs:
        if curso_id and str(curso.id) != curso_id:
            continue
        mat_qs = curso.matriculas.select_related('estudiante', 'jornada', 'jornada__sede').order_by(
            'estudiante__nombres'
        )
        if modalidad in ('presencial', 'online'):
            mat_qs = mat_qs.filter(modalidad=modalidad)
        if mat_qs.exists() or not curso_id:
            grupos.append({
                'curso': curso,
                'matriculas': mat_qs,
                'total': mat_qs.count(),
            })

    # Ocultar cursos sin matriculados (excepto si se filtró por curso)
    if not curso_id:
        grupos = [g for g in grupos if g['total'] > 0]

    return render(request, 'estudiantes/por_curso.html', {
        'grupos': grupos,
        'cursos': cursos_qs,
        'curso_seleccionado': curso_id,
        'modalidad': modalidad,
    })


@matricula_requerida
def estudiante_detalle(request, pk):
    """Detalle de un estudiante con todas sus matrículas."""
    estudiante = get_object_or_404(Estudiante, pk=pk)
    matriculas = estudiante.matriculas.select_related(
        'curso', 'curso__categoria', 'jornada'
    ).order_by('-fecha_matricula')

    # Agrupar por año para el "historial"
    por_anio = defaultdict(list)
    for m in matriculas:
        por_anio[m.fecha_matricula.year].append(m)

    historial = []
    for anio in sorted(por_anio.keys(), reverse=True):
        items = por_anio[anio]
        historial.append({
            'anio': anio,
            'matriculas': items,
            'total_facturado': sum((m.valor_curso or Decimal('0.00')) for m in items),
            'total_pagado': sum((m.valor_pagado or Decimal('0.00')) for m in items),
        })

    return render(request, 'estudiantes/detalle.html', {
        'estudiante': estudiante,
        'matriculas': matriculas,
        'historial': historial,
        'total_matriculas': matriculas.count(),
    })


@matricula_requerida
def estudiantes_export(request):
    """
    Descarga el directorio de estudiantes como Excel.
    Si se pasa ?por_curso=1, genera una hoja por curso.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    por_curso = request.GET.get('por_curso', '') == '1'
    q = request.GET.get('q', '').strip()
    sede = request.GET.get('sede', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    sedes = _sedes_con_matriculas()
    if sede and sede not in sedes:
        sede = ''

    if por_curso:
        # Una hoja por curso (solo cursos con matriculados)
        wb = Workbook()
        wb.remove(wb.active)

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill('solid', fgColor='1A237E')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'),
        )

        headers = [
            'Cédula', 'Estudiante', 'Edad',
            'Correo', 'Celular', 'Ciudad', 'Nivel',
            'Jornada', 'Sede',
            'Modalidad', 'Fecha matrícula', 'Valor', 'Pagado', 'Saldo', 'Estado',
        ]

        cursos_qs = Curso.objects.filter(activo=True).order_by('nombre')
        if curso_id and curso_id.isdigit():
            cursos_qs = cursos_qs.filter(id=int(curso_id))

        hojas_creadas = 0
        for curso in cursos_qs:
            mat_qs = curso.matriculas.select_related('estudiante', 'jornada', 'jornada__sede').order_by(
                'estudiante__nombres'
            )
            if modalidad in ('presencial', 'online'):
                mat_qs = mat_qs.filter(modalidad=modalidad)
            if sede:
                mat_qs = mat_qs.filter(jornada__ciudad__iexact=sede)
            if q:
                mat_qs = filtrar_queryset_busqueda(mat_qs, q, [
                    'estudiante__cedula',
                    'estudiante__nombres',
                    'estudiante__correo',
                    'estudiante__celular',
                    'curso__nombre',
                ])

            matriculas_curso = list(mat_qs)
            if not matriculas_curso:
                continue

            # Excel limita el nombre de hoja a 31 chars y prohíbe ciertos caracteres
            nombre_hoja = ''.join(c if c not in '\\/:*?[]' else '_' for c in curso.nombre)[:31]
            ws = wb.create_sheet(title=nombre_hoja)
            hojas_creadas += 1

            # Título
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            t = ws.cell(row=1, column=1, value=f'{curso.nombre} — {len(matriculas_curso)} estudiante(s)')
            t.font = Font(bold=True, size=14, color='1A237E')
            t.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 24

            # Encabezados
            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=2, column=col_idx, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = thin_border
            ws.row_dimensions[2].height = 30

            for row_idx, m in enumerate(matriculas_curso, start=3):
                e = m.estudiante
                row_data = [
                    e.cedula, e.nombres, e.edad or '',
                    e.correo or '', e.celular or '', e.ciudad or '',
                    e.get_nivel_formacion_display() if e.nivel_formacion else '',
                    m.jornada.descripcion_legible if m.jornada_id else '',
                    m.sede if m.sede != '—' else '',
                    m.get_modalidad_display(),
                    m.fecha_matricula if m.fecha_matricula else '',
                    float(m.valor_curso or 0),
                    float(m.valor_pagado or 0),
                    float(m.saldo or 0),
                    m.estado_pago,
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.border = thin_border
                    c.alignment = Alignment(vertical='center')
                    if col_idx in (1, 5):
                        c.number_format = '@'
                    elif col_idx == 11:
                        c.number_format = 'dd/mm/yyyy'
                    elif col_idx in (12, 13, 14):
                        c.number_format = '"$"#,##0.00'

            # Auto-ancho
            for col_idx in range(1, len(headers) + 1):
                max_length = len(headers[col_idx - 1])
                for row_idx in range(3, len(matriculas_curso) + 3):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is not None:
                        max_length = max(max_length, len(str(v)))
                ws.column_dimensions[
                    ws.cell(row=2, column=col_idx).column_letter
                ].width = min(max_length + 3, 35)

            _add_excel_table(ws, 2, 1, len(matriculas_curso) + 2, len(headers), nombre_hoja)
            ws.freeze_panes = 'A3'

        if hojas_creadas == 0:
            ws = wb.create_sheet(title='Sin datos')
            ws.cell(row=1, column=1, value='No hay estudiantes con los filtros aplicados.')

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'estudiantes_por_curso_{fecha_str}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # Modo plano: una sola hoja con todos los estudiantes
    estudiantes_qs = Estudiante.objects.annotate(
        num_matriculas=Count('matriculas', distinct=True)
    ).prefetch_related(
        'matriculas__jornada', 'matriculas__jornada__sede', 'matriculas__curso'
    ).order_by('nombres')

    estudiantes_qs = _filtrar_estudiantes_directorio(estudiantes_qs, q=q, sede=sede)

    headers = [
        'Cédula', 'Estudiante', 'Edad',
        'Correo', 'Celular', 'Ciudad', 'Nivel formación',
        'Título profesional', 'Jornada(s)', 'Sede(s)', '# Matrículas', 'Cursos',
    ]

    rows = []
    for e in estudiantes_qs:
        matriculas_resumen = _matriculas_estudiante_para_resumen(e, sede=sede)
        cursos_str = ', '.join(
            sorted({m.curso.nombre for m in matriculas_resumen})
        )
        # Resumen de jornadas y sedes (sin repetir)
        jornadas_set = []
        sedes_set = []
        for m in matriculas_resumen:
            if m.jornada_id:
                etiqueta_dia = m.jornada.descripcion_legible
                if etiqueta_dia and etiqueta_dia not in jornadas_set:
                    jornadas_set.append(etiqueta_dia)
            sede_nombre = (m.sede or '').strip()
            if sede_nombre and sede_nombre != '—' and sede_nombre not in sedes_set:
                sedes_set.append(sede_nombre)
        rows.append([
            e.cedula, e.nombres, e.edad or '',
            e.correo or '', e.celular or '', e.ciudad or '',
            e.get_nivel_formacion_display() if e.nivel_formacion else '',
            e.titulo_profesional or '',
            ' · '.join(jornadas_set),
            ' · '.join(sedes_set),
            len(matriculas_resumen) if sede else e.num_matriculas,
            cursos_str,
        ])

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'estudiantes_{fecha_str}.xlsx'
    return _build_excel_response(
        filename=filename,
        sheet_name='Directorio de Estudiantes',
        headers=headers,
        rows=rows,
        text_columns={0, 4},
        explicit_widths={
            0: 15, 1: 28, 2: 8, 3: 30, 4: 15, 5: 16,
            6: 24, 7: 24, 8: 32, 9: 22, 10: 14, 11: 36,
        },
    )


@matricula_requerida
def estudiante_export(request, pk):
    """Descarga el historial individual de un estudiante."""
    estudiante = get_object_or_404(Estudiante, pk=pk)
    matriculas = estudiante.matriculas.select_related(
        'curso', 'curso__categoria', 'jornada'
    ).order_by('-fecha_matricula')

    headers = [
        'Año', 'Mes', 'Fecha matrícula', 'Curso', 'Modalidad',
        'Categoría', 'Sede', 'Valor', 'Pagado', 'Saldo', 'Estado',
    ]

    rows = []
    total_facturado = Decimal('0.00')
    total_pagado = Decimal('0.00')

    for m in matriculas:
        rows.append([
            m.fecha_matricula.year if m.fecha_matricula else '',
            MESES_ES[m.fecha_matricula.month] if m.fecha_matricula else '',
            m.fecha_matricula if m.fecha_matricula else '',
            m.curso.nombre,
            m.get_modalidad_display(),
            m.curso.categoria.nombre if m.curso.categoria else '—',
            m.sede,
            float(m.valor_curso or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
            m.estado_pago,
        ])
        total_facturado += m.valor_curso or Decimal('0.00')
        total_pagado += m.valor_pagado or Decimal('0.00')

    totals = {
        7: float(total_facturado),
        8: float(total_pagado),
        9: float(total_facturado - total_pagado),
    }

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    filename = f'estudiante_{estudiante.cedula}_{fecha_str}.xlsx'
    return _build_excel_response(
        filename=filename,
        sheet_name=f'{estudiante.nombres}'[:31],
        headers=headers,
        rows=rows,
        totals=totals,
        column_formats={
            2: 'dd/mm/yyyy',
            7: '"$"#,##0.00',
            8: '"$"#,##0.00',
            9: '"$"#,##0.00',
        },
    )


# ═════════════════════════════════════════════════════════════════
# Gestión de abonos por matrícula
# ═════════════════════════════════════════════════════════════════

@matricula_requerida
@require_POST
def matricula_activar_retiro(request, pk):
    """Marca la matrícula como Retiro Voluntario, lo que ajusta su saldo a 0."""
    matricula = get_object_or_404(Matricula, pk=pk)
    if matricula.estado != 'retiro_voluntario':
        matricula.estado = 'retiro_voluntario'
        matricula.save(update_fields=['estado'])
        messages.success(request, 'La matrícula ha sido marcada como Retiro Voluntario. El saldo pendiente ahora es $0.00.')
    return redirect('academia:matricula_abonos', pk=matricula.pk)


@admin_requerido
@require_POST
@transaction.atomic
def matricula_revertir_retiro(request, pk):
    """
    Revierte un retiro voluntario sin modificar pagos ni valores.

    El saldo no se almacena: se calcula como valor neto menos lo pagado.
    Por eso, al regresar la matrícula a Activa, recupera automáticamente el
    saldo exacto que tenía antes del retiro.
    """
    matricula = get_object_or_404(
        Matricula.objects.select_for_update().select_related(
            'estudiante', 'curso',
        ),
        pk=pk,
    )
    if matricula.estado != 'retiro_voluntario':
        messages.info(
            request,
            'La matrícula ya está activa; no fue necesario revertirla.',
        )
        return redirect(
            'academia:matricula_retirados',
            modalidad=matricula.modalidad,
        )

    matricula.estado = 'activa'
    matricula.save(update_fields=['estado'])
    saldo_restaurado = matricula.saldo
    messages.success(
        request,
        (
            f'El retiro voluntario de {matricula.estudiante.nombre_completo} '
            f'fue revertido. La matrícula está activa y su saldo se restauró '
            f'a ${saldo_restaurado:.2f}.'
        ),
    )
    return redirect(
        'academia:matricula_retirados',
        modalidad=matricula.modalidad,
    )


@matricula_requerida
def matricula_abonos(request, pk):
    """
    Pantalla central de pagos de una matrícula:
    - Resumen (valor curso, pagado, saldo, estado)
    - Historial de abonos
    - Botón "Registrar abono" (modal)
    """
    matricula = get_object_or_404(
        Matricula.objects.select_related(
            'estudiante', 'curso', 'curso__categoria', 'jornada'
        ).prefetch_related(
            Prefetch(
                'recuperaciones_pendientes',
                queryset=RecuperacionPendiente.objects.select_related(
                    'abono',
                ).order_by('pagada', '-fecha_marcada', '-creado'),
            )
        ),
        pk=pk
    )
    abonos = matricula.abonos.select_related('registrado_por').order_by('-fecha', '-creado')

    # Saldo restante para el modal
    saldo_pendiente = matricula.saldo

    # Distribución por método (para mostrar resumen)
    dist_metodo = defaultdict(lambda: {'count': 0, 'total': Decimal('0.00')})
    for a in abonos:
        dist_metodo[a.get_metodo_display()]['count'] += 1
        dist_metodo[a.get_metodo_display()]['total'] += a.monto

    # Form pre-cargado para el modal (fecha=hoy, monto=0.00)
    form_inicial = AbonoForm(
        initial={
            'fecha': date.today(),
            'monto': Decimal('0.00'),
        },
        matricula=matricula,
    )

    return render(request, 'pagos/matricula_abonos.html', {
        'matricula': matricula,
        'pago_unico_online': matricula.tiene_pago_unico_online,
        'abonos': abonos,
        'saldo_pendiente': saldo_pendiente,
        'dist_metodo': dict(dist_metodo),
        'form': form_inicial,
        'siguiente_recibo': Abono.generar_numero_recibo(matricula),
    })


def _sincronizar_recuperacion_desde_pago(
    abono, *, fecha_marcada=None, fecha_programada=None
):
    """Vincula automáticamente un pago tipo Recuperación con su registro.

    Si el módulo ya tenía una recuperación pendiente, la marca como pagada.
    Si el usuario registró directamente el pago desde la matrícula, crea el
    registro pagado para que también aparezca en Clases en Recuperación y en
    la Hoja de Recaudación.
    """
    if (
        not abono
        or abono.tipo_pago != 'recuperacion'
        or not abono.matricula_id
        or not abono.numero_modulo
    ):
        return None

    matricula = abono.matricula
    recuperacion = (
        RecuperacionPendiente.objects.filter(abono=abono).first()
        or RecuperacionPendiente.objects.filter(
            matricula=matricula,
            numero_modulo=abono.numero_modulo,
            pagada=False,
        ).order_by('fecha_marcada', 'pk').first()
    )

    if recuperacion is None:
        monto_que_afecta_saldo = (
            abono.monto if abono.cuenta_para_saldo else Decimal('0.00')
        )
        saldo_antes_del_pago = min(
            matricula.valor_neto,
            max(matricula.saldo + monto_que_afecta_saldo, Decimal('0.00')),
        )
        recuperacion = RecuperacionPendiente(
            matricula=matricula,
            numero_modulo=abono.numero_modulo,
            fecha_marcada=fecha_marcada or abono.fecha,
            saldo_pendiente_al_marcar=saldo_antes_del_pago,
            observaciones=(
                'Registro generado automáticamente desde el pago de recuperación.'
            ),
        )

    recuperacion.numero_modulo = abono.numero_modulo
    if fecha_marcada:
        recuperacion.fecha_marcada = fecha_marcada
    if fecha_programada is not None or recuperacion.pk is None:
        recuperacion.fecha_programada = fecha_programada
    recuperacion.pagada = True
    # La fecha propia del abono es la fecha en que se realizó el pago.
    recuperacion.fecha_recuperacion = abono.fecha
    recuperacion.abono = abono
    recuperacion.save()
    return recuperacion


@matricula_requerida
@require_POST
@transaction.atomic
def abono_crear(request, matricula_pk):
    """Crear un abono nuevo. Llamado desde el modal."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    form = AbonoForm(request.POST, matricula=matricula)

    if form.is_valid():
        tipo_cobro = form.cleaned_data.get('tipo_cobro')
        if tipo_cobro == 'mixto':
            # Create a single abono with both payments
            abono = form.save(commit=False)
            abono.matricula = matricula
            abono.registrado_por = request.user
            monto_2 = form.cleaned_data.get('monto_pago_2') or Decimal('0.00')
            abono.monto = form.cleaned_data.get('monto')
            abono.metodo = form.cleaned_data.get('metodo_pago_1') or 'efectivo'
            abono.banco = form.cleaned_data.get('banco_1') or ''
            abono.monto_2 = monto_2
            abono.metodo_2 = form.cleaned_data.get('metodo_pago_2') or 'efectivo'
            abono.banco_2 = form.cleaned_data.get('banco_2') or ''
            abono.save()
            _sincronizar_recuperacion_desde_pago(
                abono,
                fecha_marcada=form.cleaned_data.get('fecha_marcada'),
                fecha_programada=form.cleaned_data.get('fecha_programada'),
            )
            messages.success(
                request,
                f'Pago mixto registrado: {abono.numero_recibo} (${abono.monto}). '
                f'Nuevo saldo: ${matricula.saldo}.'
            )
        else:
            abono = form.save(commit=False)
            abono.matricula = matricula
            abono.registrado_por = request.user
            abono.save()
            _sincronizar_recuperacion_desde_pago(
                abono,
                fecha_marcada=form.cleaned_data.get('fecha_marcada'),
                fecha_programada=form.cleaned_data.get('fecha_programada'),
            )
            messages.success(
                request,
                f'Abono registrado: {abono.numero_recibo} por ${abono.monto}. '
                f'Nuevo saldo: ${matricula.saldo}.'
            )
    else:
        # Recopilar errores legibles (sin __all__ ni nombres internos)
        errores = []
        for field, errs in form.errors.items():
            prefijo = '' if field == '__all__' else f'{form.fields[field].label or field}: '
            for err in errs:
                errores.append(f'{prefijo}{err}')
        messages.error(
            request,
            'No se pudo registrar el abono. ' + ' / '.join(errores)
        )

    return redirect('academia:matricula_abonos', pk=matricula_pk)


@admin_requerido
@transaction.atomic
def abono_editar(request, matricula_pk, abono_pk):
    """Editar un abono existente."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    abono = get_object_or_404(Abono, pk=abono_pk, matricula=matricula)

    if request.method == 'POST':
        form = AbonoForm(request.POST, instance=abono, matricula=matricula)
        if form.is_valid():
            tipo_cobro = form.cleaned_data.get('tipo_cobro')
            if tipo_cobro == 'mixto':
                abono = form.save(commit=False)
                monto_2 = form.cleaned_data.get('monto_pago_2') or Decimal('0.00')
                abono.monto = form.cleaned_data.get('monto')
                abono.metodo = form.cleaned_data.get('metodo_pago_1') or 'efectivo'
                abono.banco = form.cleaned_data.get('banco_1') or ''
                abono.monto_2 = monto_2
                abono.metodo_2 = form.cleaned_data.get('metodo_pago_2') or 'efectivo'
                abono.banco_2 = form.cleaned_data.get('banco_2') or ''
                abono.save()
                _sincronizar_recuperacion_desde_pago(
                    abono,
                    fecha_marcada=form.cleaned_data.get('fecha_marcada'),
                    fecha_programada=form.cleaned_data.get('fecha_programada'),
                )
            else:
                abono = form.save(commit=False)
                abono.monto_2 = None
                abono.metodo_2 = ''
                abono.banco_2 = ''
                abono.save()
                _sincronizar_recuperacion_desde_pago(
                    abono,
                    fecha_marcada=form.cleaned_data.get('fecha_marcada'),
                    fecha_programada=form.cleaned_data.get('fecha_programada'),
                )
            messages.success(request, f'Abono {abono.numero_recibo} actualizado.')
            return redirect('academia:matricula_abonos', pk=matricula_pk)
    else:
        form = AbonoForm(instance=abono, matricula=matricula)

    return render(request, 'pagos/abono_editar.html', {
        'form': form,
        'abono': abono,
        'matricula': matricula,
    })


@admin_requerido
@require_POST
def abono_eliminar(request, matricula_pk, abono_pk):
    """Eliminar un abono y recalcular el total."""
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    abono = get_object_or_404(Abono, pk=abono_pk, matricula=matricula)
    numero = abono.numero_recibo
    monto = abono.monto
    abono.delete()
    messages.success(
        request,
        f'Abono {numero} eliminado (${monto}). Saldo recalculado: ${matricula.saldo}.'
    )
    return redirect('academia:matricula_abonos', pk=matricula_pk)


@matricula_requerida
def abonos_export(request):
    """
    Reporte de abonos en Excel — todos los abonos del periodo,
    con filtros por mes, año, método.
    """
    qs = Abono.objects.select_related(
        'matricula', 'matricula__estudiante', 'matricula__curso',
        'registrado_por',
    ).order_by('-fecha', '-creado')

    anio = request.GET.get('anio', '').strip()
    mes = request.GET.get('mes', '').strip()
    metodo = request.GET.get('metodo', '').strip()

    if anio.isdigit():
        qs = qs.filter(fecha__year=int(anio))
    if mes.isdigit() and 1 <= int(mes) <= 12:
        qs = qs.filter(fecha__month=int(mes))
    if metodo in ('efectivo', 'transferencia', 'tarjeta'):
        qs = qs.filter(metodo=metodo)

    headers = [
        'Nº Recibo', 'Fecha', 'Cédula', 'Estudiante', 'Curso',
        'Modalidad', 'Método', 'Monto', 'Valor curso', 'Saldo restante',
        'Registrado por', 'Observaciones',
    ]

    rows = []
    total_monto = Decimal('0.00')
    total_efectivo = Decimal('0.00')
    total_transf = Decimal('0.00')
    total_tarjeta = Decimal('0.00')

    for a in qs:
        m = a.matricula
        rows.append([
            a.numero_recibo,
            a.fecha if a.fecha else '',
            m.estudiante.cedula,
            m.estudiante.nombre_completo,
            m.curso.nombre,
            m.get_modalidad_display(),
            a.get_metodo_display(),
            float(a.monto),
            float(m.valor_curso or 0),
            float(m.saldo or 0),
            (a.registrado_por.get_full_name() or a.registrado_por.username) if a.registrado_por else '—',
            a.observaciones or '',
        ])
        total_monto += a.monto
        if a.metodo == 'efectivo':
            total_efectivo += a.monto
        elif a.metodo == 'transferencia':
            total_transf += a.monto
        elif a.metodo == 'tarjeta':
            total_tarjeta += a.monto

    totals = {7: float(total_monto)}

    fecha_str = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = ''
    if anio:
        sufijo += f'_{anio}'
    if mes:
        sufijo += f'_{mes:0>2}' if not mes.startswith('0') else f'_{mes}'
    if metodo:
        sufijo += f'_{metodo}'
    filename = f'abonos{sufijo}_{fecha_str}.xlsx'

    # Construir el archivo con totales por método al final
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte de Abonos'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A237E')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )
    total_font = Font(bold=True, color='1A237E', size=11)
    total_fill = PatternFill('solid', fgColor='FFF8E1')
    method_font = Font(bold=True, color='2E7D32', size=11)
    method_fill = PatternFill('solid', fgColor='E8F5E9')

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    t = ws.cell(row=1, column=1, value='Reporte de Abonos')
    t.font = Font(bold=True, size=14, color='1A237E')
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    # Encabezados
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin
    ws.row_dimensions[2].height = 30

    # Datos
    for row_idx, row_data in enumerate(rows, start=3):
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin
            c.alignment = Alignment(vertical='center')
            if col_idx in (1, 3):
                c.number_format = '@'
            elif col_idx == 2:
                c.number_format = 'dd/mm/yyyy'
            elif col_idx in (8, 9, 10):
                c.number_format = '"$"#,##0.00'

    # Total general
    total_row = len(rows) + 3
    ws.cell(row=total_row, column=6, value='TOTAL GENERAL:').font = total_font
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=6).fill = total_fill
    ws.cell(row=total_row, column=6).border = thin
    c = ws.cell(row=total_row, column=8, value=float(total_monto))
    c.font = total_font
    c.fill = total_fill
    c.border = thin
    c.number_format = '"$"#,##0.00'

    # Desglose por método
    metodo_row = total_row + 2
    ws.cell(row=metodo_row, column=1, value='💵 Por método de pago:').font = method_font
    metodo_row += 1
    for label, total in [
        ('Efectivo', total_efectivo),
        ('Transferencia', total_transf),
        ('Tarjeta', total_tarjeta),
    ]:
        ws.cell(row=metodo_row, column=1, value=label).font = method_font
        ws.cell(row=metodo_row, column=1).fill = method_fill
        ws.cell(row=metodo_row, column=1).border = thin
        c = ws.cell(row=metodo_row, column=2, value=float(total))
        c.font = method_font
        c.fill = method_fill
        c.border = thin
        c.number_format = '"$"#,##0.00'
        metodo_row += 1

    # Auto-ancho
    for col_idx in range(1, len(headers) + 1):
        max_length = len(headers[col_idx - 1])
        for row_idx in range(3, len(rows) + 3):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                max_length = max(max_length, len(str(v)))
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = min(max_length + 3, 38)

    _add_excel_table(ws, 2, 1, len(rows) + 2, len(headers), 'Reporte de Abonos')
    ws.freeze_panes = 'A3'

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@matricula_requerida
def abono_recibo(request, abono_pk):
    """
    Vista del recibo individual (HTML imprimible).
    Cada abono tiene su comprobante.
    """
    abono = get_object_or_404(
        Abono.objects.select_related(
            'matricula', 'matricula__estudiante', 'matricula__curso',
            'registrado_por',
        ),
        pk=abono_pk
    )
    return render(request, 'pagos/recibo.html', {
        'abono': abono,
        'matricula': abono.matricula,
    })


# ═════════════════════════════════════════════════════════════════
# Pagos por Módulo (control semanal del avance del curso)
# ═════════════════════════════════════════════════════════════════

# Tipos de matrícula que SÍ implican una reserva inicial (los únicos que
# hacen sentido para el control de morosidad por módulo).
TIPOS_CON_RESERVA = ('reserva_abono', 'reserva_modulo_1')

TIPOS_MATRICULA_FILTRO_PAGOS_MODULO = (
    ('reserva_abono', 'Reserva / Abono'),
    ('programa_completo', 'Programa Completo'),
)


def _normalizar_matricula_filtro_pagos_modulo(matricula_id, curso_sel):
    if not curso_sel or not matricula_id or not str(matricula_id).isdigit():
        return ''
    existe = Matricula.objects.filter(
        pk=int(matricula_id),
        curso=curso_sel,
    ).exclude(estado='retiro_voluntario').exists()
    return str(matricula_id) if existe else ''


def _label_matricula_estudiante_jornada(matricula):
    estudiante = matricula.estudiante
    jornada = matricula.jornada
    nombre = (getattr(estudiante, 'nombre_completo', '') or estudiante.nombres or '').strip()
    nombre = nombre or 'Sin nombre'
    cedula = (estudiante.cedula or '').strip()
    if cedula:
        nombre = f'{nombre} ({cedula})'

    partes = [matricula.get_modalidad_display()]
    if jornada:
        partes.append(jornada.descripcion_legible)
        if jornada.fecha_inicio:
            partes.append(jornada.fecha_inicio.strftime('%d/%m/%Y'))
        if jornada.ciudad:
            partes.append(jornada.ciudad)
    else:
        partes.append('Sin jornada')

    return f"{nombre} - {' · '.join(p for p in partes if p)} - Mat. #{matricula.pk}"


def _opciones_matriculas_estudiantes_pagos_modulo(curso_sel, modalidad='',
                                                   ciudad='', tipo_matricula='',
                                                   fecha_jornada=''):
    if not curso_sel:
        return []

    qs = Matricula.objects.filter(
        curso=curso_sel,
    ).select_related(
        'estudiante', 'jornada'
    ).exclude(
        estado='retiro_voluntario'
    ).order_by(
        'estudiante__nombres', 'jornada__fecha_inicio', 'pk'
    )

    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)
    if ciudad:
        qs = qs.filter(jornada__ciudad__iexact=ciudad)
    if tipo_matricula:
        qs = qs.filter(tipo_matricula=tipo_matricula)
    if fecha_jornada:
        qs = qs.filter(jornada__fecha_inicio=fecha_jornada)

    return [
        {
            'id': str(m.pk),
            'label': _label_matricula_estudiante_jornada(m),
        }
        for m in qs
    ]


def _construir_matriz_pagos(curso_sel, modalidad='', ciudad='',
                            tipo_matricula='', filtro_modulo_estado='',
                            fecha_jornada='', matricula_id='',
                            fecha_modulo_desde=None, fecha_modulo_hasta=None):
    """
    Construye la matriz estudiantes x modulos para un curso.
    filtro_modulo_estado: cadena con formato "<num>_<estado>",
    por ejemplo "2_Parcial", "1_Pagado" o "3_Recuperacion".
    Filtra por el modulo y estado indicados. Cadena vacia = sin filtro.
    fecha_modulo_desde/fecha_modulo_hasta filtran por la fecha de pago que
    se muestra dentro de cada celda de módulo.
    """
    if modalidad in ('presencial', 'online'):
        n_mod = (
            1
            if curso_sel.usa_pago_unico_recaudacion(modalidad)
            else curso_sel.get_numero_modulos(modalidad)
        )
    else:
        cuotas_por_modalidad = []
        if curso_sel.ofrece_presencial:
            cuotas_por_modalidad.append(curso_sel.numero_modulos or 1)
        if curso_sel.ofrece_online:
            cuotas_por_modalidad.append(
                1 if curso_sel.usa_pago_unico_recaudacion('online')
                else (curso_sel.numero_modulos_online or 1)
            )
        n_mod = max(cuotas_por_modalidad or [1])
    modulos = list(range(1, n_mod + 1))
    modulos_visibles = list(modulos)

    qs = Matricula.objects.filter(
        curso=curso_sel
    ).select_related(
        'estudiante', 'jornada', 'registrado_por'
    ).prefetch_related(
        'abonos', 'recuperaciones_pendientes__abono'
    ).exclude(estado='retiro_voluntario')

    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)
    if ciudad:
        qs = qs.filter(jornada__ciudad__iexact=ciudad)
    if tipo_matricula:
        qs = qs.filter(tipo_matricula=tipo_matricula)
    if fecha_jornada:
        qs = qs.filter(jornada__fecha_inicio=fecha_jornada)
    if matricula_id and str(matricula_id).isdigit():
        qs = qs.filter(pk=int(matricula_id))

    matriculas = []
    for m in qs:
        jornada = m.jornada
        todos_abonos = list(m.abonos.all())
        todas_recuperaciones = list(m.recuperaciones_pendientes.all())

        # Desglose por módulo: SOLO cuenta los abonos asignados explícitamente
        # a un módulo (tipo_pago='por_modulo' o 'recuperacion' con
        # numero_modulo). La reserva y los abonos libres NO entran a la matriz
        # (suman al saldo total pero no se asignan a ningún módulo). Cada
        # módulo se cobra de forma independiente. Toda la lógica está
        # centralizada en `Matricula.desglose_pagos_por_modulo()` para que
        # haya un solo lugar de verdad.
        desglose = m.desglose_pagos_por_modulo()

        modulos_data = []
        for d in desglose:
            numero = d['numero']
            recuperaciones_modulo = [
                recuperacion for recuperacion in todas_recuperaciones
                if recuperacion.numero_modulo == numero
            ]
            pagos_recuperacion = {
                abono.pk: abono
                for abono in todos_abonos
                if (
                    abono.tipo_pago == 'recuperacion'
                    and abono.cuenta_para_saldo
                    and abono.numero_modulo == numero
                )
            }
            for recuperacion in recuperaciones_modulo:
                if (
                    recuperacion.pagada
                    and recuperacion.abono
                    and recuperacion.abono.cuenta_para_saldo
                ):
                    pagos_recuperacion[recuperacion.abono.pk] = recuperacion.abono

            modulos_data.append({
                'numero': numero,
                'estado': d['estado'],
                'pagado': d['pagado'],
                'esperado': d['esperado'],
                'fecha_pago': d['fecha_ultimo_pago'],
                'aplica': True,
                'es_recuperacion': bool(recuperaciones_modulo or pagos_recuperacion),
                'recuperacion_monto': sum(
                    (a.monto for a in pagos_recuperacion.values()),
                    Decimal('0.00'),
                ),
                'recuperacion_recibos': ', '.join(
                    a.numero_recibo for a in pagos_recuperacion.values()
                    if a.numero_recibo
                ),
            })
        # Si la matrícula tiene menos módulos que el máximo del curso
        # (por ejemplo, una matrícula online con 2 módulos en un curso que
        # también ofrece presencial con 4 módulos), rellenamos las celdas
        # extra con un placeholder "no aplica" para que la tabla quede
        # alineada y no falle al recorrer modulos_data.
        for n in range(len(modulos_data) + 1, n_mod + 1):
            modulos_data.append({
                'numero': n,
                'estado': 'No aplica',
                'pagado': Decimal('0.00'),
                'esperado': Decimal('0.00'),
                'fecha_pago': None,
                'aplica': False,
                'es_recuperacion': False,
                'recuperacion_monto': Decimal('0.00'),
                'recuperacion_recibos': '',
            })

        # Diccionario plano para los cálculos posteriores que sí lo necesitan
        pagos_efectivos = {d['numero']: d['pagado'] for d in desglose}

        # modulo_control: el primer módulo que aplica y no está pagado.
        # Si todos están pagados, usamos el último que aplica.
        aplicables = [x for x in modulos_data if x['aplica']]
        modulo_control = next(
            (x for x in aplicables if x['estado'] != 'Pagado'),
            aplicables[-1] if aplicables else None,
        )
        
        # Hoja de recaudacion: SOLO pagos tipo 'por_modulo' asignados
        # explicitamente al modulo de control.
        abonos_modulo = []
        recaudado_hoja = Decimal('0.00')
        if modulo_control:
            abonos_modulo = [
                a for a in todos_abonos
                if a.cuenta_para_saldo
                and a.tipo_pago in ('por_modulo', 'solo_modulo')
                and a.numero_modulo == modulo_control['numero']
            ]
            recaudado_hoja = sum(
                (a.monto for a in abonos_modulo), Decimal('0.00')
            )
            
        metodos = sorted({a.get_metodo_display() for a in abonos_modulo})
        tipos = sorted({a.get_tipo_pago_display() for a in abonos_modulo})
        bancos = sorted({a.get_banco_display() for a in abonos_modulo if a.banco})
        recuperaciones_pendientes = [
            r for r in todas_recuperaciones if not r.pagada
        ]
        recuperacion_txt = ', '.join(
            f"Mód. {r.numero_modulo}" for r in recuperaciones_pendientes
        )
        
        # Reserva / abono libre = abonos que NO son por_modulo. Estos no
        # entran a la matriz pero sí suman al saldo total. Se muestran como
        # un texto pequeño "Reservado: $X" debajo del valor pagado.
        reserva_total = sum(
            (a.monto for a in m.abonos.filter(cuenta_para_saldo=True).exclude(tipo_pago__in=('por_modulo', 'solo_modulo'))),
            Decimal('0.00'),
        )

        matriculas.append({
            'matricula': m,
            'pago_unico_online': m.tiene_pago_unico_online,
            'estudiante': m.estudiante,
            'curso_nombre': m.curso.nombre,
            'modulos_data': modulos_data,
            'valor_modulo_sugerido': (
                modulo_control['esperado']
                if modulo_control else Decimal('0.00')
            ),
            'tipo_matricula_codigo': m.tipo_matricula,
            'tipo_matricula_label': m.get_tipo_matricula_display(),
            'reserva_total': reserva_total,
            'jornada_inicio': jornada.fecha_inicio if jornada and jornada.fecha_inicio else None,
            'jornada_dia': jornada.descripcion_legible if jornada else '—',
            'jornada_horario': m.horario,
            'jornada_sede': m.sede,
            'jornada_resumen': (
                f"{jornada.descripcion_legible} · "
                f"{jornada.fecha_inicio.strftime('%d/%m/%Y') if jornada and jornada.fecha_inicio else 'Sin fecha'}"
                f"{' · ' + m.horario if m.horario != '—' else ''}"
                f"{' · ' + m.sede if m.sede != '—' else ''}"
            ) if jornada else '—',
            'modulo_control': modulo_control['numero'] if modulo_control else '—',
            'recaudar_control': (
                max(
                    modulo_control['esperado'] - recaudado_hoja,
                    Decimal('0.00'),
                )
                if modulo_control else Decimal('0.00')
            ),
            'recaudado_control': recaudado_hoja,
            'forma_pago_control': ', '.join(metodos) if metodos else 'Sin pagar',
            'tipo_pago_control': ', '.join(tipos) if tipos else 'Sin pagar',
            'banco_control': ', '.join(bancos) if bancos else '—',
            'asistencia_control': '—',
            'recuperacion_control': recuperacion_txt or '—',
        })

    # ── Filtro por módulo + estado (e.g. "2_Pagado", "2_Pendiente") ──
    # Bajo la regla visual binaria, "Pagado" agrupa cualquier módulo que
    # haya recibido al menos un pago directo (estado interno Pagado o
    # Parcial). "Pendiente" sigue siendo solo los módulos sin pagos.
    # "Recuperacion" filtra matrículas con recuperación marcada o cobrada
    # en ese módulo, aunque el pago del módulo siga pendiente.
    if filtro_modulo_estado:
        partes = filtro_modulo_estado.split('_', 1)
        if len(partes) == 2 and partes[0].isdigit() and partes[1] in ('Pagado', 'Parcial', 'Pendiente', 'Recuperacion'):
            num_filtro = int(partes[0])
            est_filtro = partes[1]
            if num_filtro in modulos:
                modulos_visibles = [num_filtro]
            if est_filtro == 'Recuperacion':
                matriculas = [
                    x for x in matriculas
                    if any(
                        mod['numero'] == num_filtro
                        and mod.get('aplica', True)
                        and mod.get('es_recuperacion')
                        for mod in x['modulos_data']
                    )
                ]
            else:
                if est_filtro == 'Pagado':
                    estados_match = ('Pagado', 'Parcial')
                else:
                    estados_match = (est_filtro,)
                matriculas = [
                    x for x in matriculas
                    if any(
                        mod['numero'] == num_filtro
                        and mod.get('aplica', True)
                        and mod['estado'] in estados_match
                        for mod in x['modulos_data']
                    )
                ]

    if fecha_modulo_desde or fecha_modulo_hasta:
        if fecha_modulo_desde and fecha_modulo_hasta and fecha_modulo_desde > fecha_modulo_hasta:
            fecha_modulo_desde, fecha_modulo_hasta = fecha_modulo_hasta, fecha_modulo_desde

        def _modulo_en_rango_fecha(mod):
            fecha_pago = mod.get('fecha_pago')
            if not mod.get('aplica', True) or not fecha_pago:
                return False
            if fecha_modulo_desde and fecha_pago < fecha_modulo_desde:
                return False
            if fecha_modulo_hasta and fecha_pago > fecha_modulo_hasta:
                return False
            return True

        matriculas = [
            x for x in matriculas
            if any(
                mod['numero'] in modulos_visibles and _modulo_en_rango_fecha(mod)
                for mod in x['modulos_data']
            )
        ]
        modulos_con_fecha = []
        for x in matriculas:
            for mod in x['modulos_data']:
                coincide_fecha = (
                    mod['numero'] in modulos_visibles
                    and _modulo_en_rango_fecha(mod)
                )
                mod['coincide_fecha_modulo'] = coincide_fecha
                if coincide_fecha and mod['numero'] not in modulos_con_fecha:
                    modulos_con_fecha.append(mod['numero'])
        modulos_visibles = [
            n for n in modulos_visibles
            if n in modulos_con_fecha
        ]

    for x in matriculas:
        x['modulos_visibles_data'] = [
            mod for mod in x['modulos_data']
            if mod['numero'] in modulos_visibles
        ]

    # ── Resumen por módulo ──
    # Importante: cada matrícula tiene `modulos_data` con n_mod elementos,
    # pero algunos pueden estar marcados con 'aplica': False (cuando la
    # modalidad de esa matrícula tiene menos módulos que el máximo del
    # curso). Esos NO se cuentan en el resumen.
    resumen_lista = []
    for n in modulos:
        def _modulo_n(x, num=n):
            return next((m for m in x['modulos_data']
                         if m['numero'] == num and m.get('aplica', True)), None)

        modulos_n = [_modulo_n(x) for x in matriculas]
        # Solo las matrículas que realmente tienen ese módulo.
        modulos_n = [m for m in modulos_n if m is not None]

        pagados = sum(1 for m in modulos_n if m['estado'] == 'Pagado')
        parciales = sum(1 for m in modulos_n if m['estado'] == 'Parcial')
        pendientes = sum(1 for m in modulos_n if m['estado'] == 'Pendiente')
        recaudado = sum(
            (m['pagado'] for m in modulos_n if m['estado'] == 'Pagado'),
            Decimal('0.00')
        )
        resumen_lista.append({
            'numero': n,
            'pagados': pagados,
            'parciales': parciales,
            'pendientes': pendientes,
            'recaudado': recaudado,
            'total_estudiantes': len(modulos_n),
        })

    return matriculas, modulos, resumen_lista, modulos_visibles


@matricula_requerida
def pagos_por_modulo(request):
    """
    Vista MATRIZ: por cada matrícula del curso filtrado, muestra el estado
    de pago de CADA módulo (Pagado / Parcial / Pendiente).

    Filtros: curso (obligatorio para ver detalle), modalidad, ciudad,
             tipo de matrícula, estado por módulo.

    Reglas del control por módulo:
    - En la matriz SOLO se cuentan los abonos asignados explícitamente a
      un módulo (tipo "Por Módulo" o "Recuperación" con número de módulo).
      Cada módulo se paga de forma independiente.
    - La reserva y los abonos libres (tipo "Abono" o "Pago Completo" sin
      número de módulo) NO aparecen en la matriz, pero sí suman al valor
      pagado total y al saldo del curso. Quedan visibles como
      "Reservado: $X" debajo del valor pagado.
    - Caso especial: si el tipo de matrícula es "Programa Completo" y el
      saldo está en $0, los módulos se muestran como "Programa Completo"
      en verde, porque el estudiante canceló el curso de una sola vez
      (no por módulos individuales).
    - Filtro "tipo de matrícula" para enfocar reservas (los que SÍ tienen
      pendientes que cobrar mes a mes).
    - Filtro "estado por módulo" para ver de un vistazo morosos.
    """
    cursos = Curso.objects.filter(activo=True).order_by('nombre')

    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    ciudad = request.GET.get('ciudad', '').strip()
    tipo_matricula = request.GET.get('tipo_matricula', '').strip()
    if tipo_matricula not in dict(TIPOS_MATRICULA_FILTRO_PAGOS_MODULO):
        tipo_matricula = ''
    filtro_modulo_estado = request.GET.get('filtro_modulo_estado', '').strip()
    fecha_jornada = request.GET.get('fecha_jornada', '').strip()
    matricula_id = request.GET.get('matricula', '').strip()
    (
        fecha_modulo_desde, fecha_modulo_hasta,
        fecha_modulo_desde_date, fecha_modulo_hasta_date,
    ) = _rango_fecha_modulo_desde_request(request)

    curso_sel = None
    matriculas = []
    modulos = []
    modulos_visibles = []
    resumen_por_modulo = []

    if curso_id and curso_id.isdigit():
        try:
            curso_sel = Curso.objects.get(pk=int(curso_id), activo=True)
        except Curso.DoesNotExist:
            curso_sel = None
    matricula_id = _normalizar_matricula_filtro_pagos_modulo(
        matricula_id, curso_sel
    )
    if not curso_sel:
        fecha_modulo_desde = fecha_modulo_hasta = ''
        fecha_modulo_desde_date = fecha_modulo_hasta_date = None

    if curso_sel:
        matriculas, modulos, resumen_por_modulo, modulos_visibles = _construir_matriz_pagos(
            curso_sel,
            modalidad=modalidad,
            ciudad=ciudad,
            tipo_matricula=tipo_matricula,
            filtro_modulo_estado=filtro_modulo_estado,
            fecha_jornada=fecha_jornada,
            matricula_id=matricula_id,
            fecha_modulo_desde=fecha_modulo_desde_date,
            fecha_modulo_hasta=fecha_modulo_hasta_date,
        )
    # Fechas de jornada disponibles para el curso seleccionado
    fechas_jornada = []
    estudiantes_jornada = []
    if curso_sel:
        from academia.models import JornadaCurso
        fechas_jornada = (
            JornadaCurso.objects.filter(curso=curso_sel, activo=True)
            .order_by('fecha_inicio')
            .values_list('fecha_inicio', flat=True)
            .distinct()
        )
        estudiantes_jornada = _opciones_matriculas_estudiantes_pagos_modulo(
            curso_sel,
            modalidad=modalidad,
            ciudad=ciudad,
            tipo_matricula=tipo_matricula,
            fecha_jornada=fecha_jornada,
        )

    export_querystring = urlencode({
        'curso': curso_id,
        'modalidad': modalidad,
        'ciudad': ciudad,
        'tipo_matricula': tipo_matricula,
        'filtro_modulo_estado': filtro_modulo_estado,
        'fecha_jornada': fecha_jornada,
        'matricula': matricula_id,
        'fecha_modulo_desde': fecha_modulo_desde,
        'fecha_modulo_hasta': fecha_modulo_hasta,
    })

    return render(request, 'pagos/por_modulo.html', {
        'cursos': cursos,
        'curso_sel': curso_sel,
        'modulos': modulos,
        'modulos_visibles': modulos_visibles,
        'matriculas_data': matriculas,
        'resumen_por_modulo': resumen_por_modulo,
        'fechas_jornada': fechas_jornada,
        'estudiantes_jornada': estudiantes_jornada,
        'tipos_matricula': TIPOS_MATRICULA_FILTRO_PAGOS_MODULO,
        'filtro_pago_unico_online': bool(
            curso_sel
            and modalidad == 'online'
            and curso_sel.usa_pago_unico_recaudacion('online')
        ),
        'export_querystring': export_querystring,
        'filtros': {
            'curso': curso_id,
            'modalidad': modalidad,
            'ciudad': ciudad,
            'tipo_matricula': tipo_matricula,
            'filtro_modulo_estado': filtro_modulo_estado,
            'fecha_jornada': fecha_jornada,
            'matricula': matricula_id,
            'fecha_modulo_desde': fecha_modulo_desde,
            'fecha_modulo_hasta': fecha_modulo_hasta,
            'fecha_modulo_label': _periodo_fecha_modulo_label(
                fecha_modulo_desde_date,
                fecha_modulo_hasta_date,
            ),
        },
    })


# ═════════════════════════════════════════════════════════════════
# Clases en Recuperación
# ═════════════════════════════════════════════════════════════════

def _rango_fecha_get(request, desde_param, hasta_param):
    """Lee un rango de fechas GET, descarta valores inválidos y ordena extremos."""
    fecha_desde = request.GET.get(desde_param, '').strip()
    fecha_hasta = request.GET.get(hasta_param, '').strip()

    try:
        fecha_desde_date = parse_date(fecha_desde) if fecha_desde else None
    except ValueError:
        fecha_desde_date = None
    try:
        fecha_hasta_date = parse_date(fecha_hasta) if fecha_hasta else None
    except ValueError:
        fecha_hasta_date = None

    if not fecha_desde_date:
        fecha_desde = ''
    if not fecha_hasta_date:
        fecha_hasta = ''

    if fecha_desde_date and fecha_hasta_date and fecha_desde_date > fecha_hasta_date:
        fecha_desde_date, fecha_hasta_date = fecha_hasta_date, fecha_desde_date
        fecha_desde, fecha_hasta = fecha_desde_date.isoformat(), fecha_hasta_date.isoformat()

    return fecha_desde, fecha_hasta, fecha_desde_date, fecha_hasta_date


def _rango_fecha_modulo_desde_request(request):
    """
    Lee el rango de fecha de módulo. Si el usuario llena un solo extremo,
    se trata como búsqueda exacta de ese día para evitar rangos abiertos.
    """
    fecha_desde, fecha_hasta, fecha_desde_date, fecha_hasta_date = _rango_fecha_get(
        request,
        'fecha_modulo_desde',
        'fecha_modulo_hasta',
    )

    if fecha_desde_date and not fecha_hasta_date:
        fecha_hasta_date = fecha_desde_date
        fecha_hasta = fecha_desde
    elif fecha_hasta_date and not fecha_desde_date:
        fecha_desde_date = fecha_hasta_date
        fecha_desde = fecha_hasta

    return fecha_desde, fecha_hasta, fecha_desde_date, fecha_hasta_date


def _periodo_fecha_modulo_label(fecha_desde, fecha_hasta):
    if not fecha_desde or not fecha_hasta:
        return ''
    if fecha_desde == fecha_hasta:
        return fecha_desde.strftime('%d/%m/%Y')
    return f'{fecha_desde:%d/%m/%Y} - {fecha_hasta:%d/%m/%Y}'


def _filtrar_recuperaciones(request):
    """Aplica los filtros de la tabla de recuperaciones y devuelve queryset + filtros."""
    estado = request.GET.get('estado', 'pendientes').strip() or 'pendientes'
    curso_id = request.GET.get('curso', '').strip()
    estudiante_id = request.GET.get('estudiante', '').strip()
    q = request.GET.get('q', '').strip()
    (
        fecha_falta_desde,
        fecha_falta_hasta,
        fecha_falta_desde_date,
        fecha_falta_hasta_date,
    ) = _rango_fecha_get(request, 'fecha_falta_desde', 'fecha_falta_hasta')
    (
        fecha_programada_desde,
        fecha_programada_hasta,
        fecha_programada_desde_date,
        fecha_programada_hasta_date,
    ) = _rango_fecha_get(
        request,
        'fecha_programada_desde',
        'fecha_programada_hasta',
    )

    if estado not in ('pendientes', 'pagadas', 'todas'):
        estado = 'pendientes'
    if curso_id and not curso_id.isdigit():
        curso_id = ''
    if estudiante_id and not estudiante_id.isdigit():
        estudiante_id = ''
    if estudiante_id and not curso_id:
        estudiante_id = ''
    if estudiante_id and curso_id:
        estudiante_scope = RecuperacionPendiente.objects.filter(
            matricula__curso_id=int(curso_id),
            matricula__estudiante_id=int(estudiante_id),
        )
        if estado == 'pendientes':
            estudiante_scope = estudiante_scope.filter(pagada=False)
        elif estado == 'pagadas':
            estudiante_scope = estudiante_scope.filter(pagada=True)
        if not estudiante_scope.exists():
            estudiante_id = ''

    qs = RecuperacionPendiente.objects.select_related(
        'matricula', 'matricula__estudiante', 'matricula__curso',
        'matricula__jornada', 'abono',
    )

    if estado == 'pendientes':
        qs = qs.filter(pagada=False)
    elif estado == 'pagadas':
        qs = qs.filter(pagada=True)

    if curso_id:
        qs = qs.filter(matricula__curso_id=int(curso_id))
    if estudiante_id:
        qs = qs.filter(matricula__estudiante_id=int(estudiante_id))

    if q:
        qs = filtrar_queryset_busqueda(qs, q, [
            'matricula__estudiante__cedula',
            'matricula__estudiante__nombres',
            'matricula__curso__nombre',
        ])

    if fecha_falta_desde_date:
        qs = qs.filter(fecha_marcada__gte=fecha_falta_desde_date)
    if fecha_falta_hasta_date:
        qs = qs.filter(fecha_marcada__lte=fecha_falta_hasta_date)
    if fecha_programada_desde_date:
        qs = qs.filter(fecha_programada__gte=fecha_programada_desde_date)
    if fecha_programada_hasta_date:
        qs = qs.filter(fecha_programada__lte=fecha_programada_hasta_date)

    filtros = {
        'curso': curso_id,
        'estudiante': estudiante_id,
        'q': q,
        'estado': estado,
        'fecha_falta_desde': fecha_falta_desde,
        'fecha_falta_hasta': fecha_falta_hasta,
        'fecha_programada_desde': fecha_programada_desde,
        'fecha_programada_hasta': fecha_programada_hasta,
    }
    filtros['query'] = urlencode({
        key: value for key, value in filtros.items()
        if key != 'query' and value
    })

    return qs.order_by('pagada', '-fecha_marcada', '-creado'), {
        **filtros,
    }


def _estudiantes_para_filtro_recuperaciones():
    """Opciones del filtro dependiente Estado + Curso -> Estudiante."""
    estudiantes = {}
    rows = (
        RecuperacionPendiente.objects
        .select_related('matricula__estudiante', 'matricula__curso')
        .values_list(
            'matricula__estudiante_id',
            'matricula__estudiante__nombres',
            'matricula__estudiante__cedula',
            'matricula__curso_id',
            'pagada',
        )
        .distinct()
    )
    for estudiante_id, nombres, cedula, curso_id, pagada in rows:
        if not estudiante_id or not curso_id:
            continue
        data = estudiantes.setdefault(estudiante_id, {
            'id': estudiante_id,
            'nombre': nombres or 'Sin nombre',
            'cedula': cedula or '',
            'curso_ids': set(),
            'curso_ids_pendientes': set(),
            'curso_ids_pagadas': set(),
        })
        curso_id_str = str(curso_id)
        data['curso_ids'].add(curso_id_str)
        if pagada:
            data['curso_ids_pagadas'].add(curso_id_str)
        else:
            data['curso_ids_pendientes'].add(curso_id_str)

    opciones = []
    for data in estudiantes.values():
        opciones.append({
            **data,
            'curso_ids': sorted(data['curso_ids']),
            'curso_ids_pendientes': sorted(data['curso_ids_pendientes']),
            'curso_ids_pagadas': sorted(data['curso_ids_pagadas']),
        })
    return sorted(opciones, key=lambda item: (item['nombre'].lower(), item['cedula']))


@matricula_requerida
def recuperaciones_lista(request):
    """
    Listado central de clases en recuperación.
    Muestra: pendientes (sin cobrar) y resueltas (ya cobradas).
    Cada pendiente trae el saldo previo del estudiante.
    """
    qs, filtros = _filtrar_recuperaciones(request)
    cursos = Curso.objects.filter(activo=True).order_by('nombre')
    estudiantes_filtro = _estudiantes_para_filtro_recuperaciones()

    # Conteos para tarjetas
    total_pendientes = RecuperacionPendiente.objects.filter(pagada=False).count()
    total_pagadas = RecuperacionPendiente.objects.filter(pagada=True).count()

    return render(request, 'pagos/recuperaciones.html', {
        'recuperaciones': qs,
        'cursos': cursos,
        'estudiantes_filtro': estudiantes_filtro,
        'estado': filtros['estado'],
        'filtros': filtros,
        'total_pendientes': total_pendientes,
        'total_pagadas': total_pagadas,
    })


@matricula_requerida
def recuperaciones_export_excel(request):
    """Exporta la tabla filtrada de clases en recuperación a Excel."""
    recuperaciones, filtros = _filtrar_recuperaciones(request)

    headers = [
        'Estado', 'Fecha falta', 'Cédula', 'Estudiante', 'Curso',
        'Modalidad', 'Módulo', 'Saldo al marcar', 'Fecha para recuperar',
        'Fecha recuperación real', 'Recibo', 'Tipo de pago', 'Monto pagado',
        'Método', 'Banco / app', 'Cuenta para saldo', 'Observaciones',
        'Asistencia',
    ]

    rows = []
    total_saldo = Decimal('0.00')
    total_pagado = Decimal('0.00')

    for r in recuperaciones:
        abono = r.abono
        estado_label = 'Pagada' if r.pagada else 'Pendiente'
        estudiante = r.matricula.estudiante
        banco = abono.get_banco_display() if (abono and abono.banco) else '—'
        monto = abono.monto if abono else Decimal('0.00')
        tipo_pago_label = abono.get_tipo_pago_display() if abono else '—'
        if abono and abono.tipo_pago != 'recuperacion':
            tipo_pago_label = f'Descontada con {tipo_pago_label}'

        rows.append([
            estado_label,
            r.fecha_marcada if r.fecha_marcada else '',
            estudiante.cedula,
            estudiante.nombre_completo,
            r.matricula.curso.nombre,
            r.matricula.get_modalidad_display(),
            r.numero_modulo,
            float(r.saldo_pendiente_al_marcar or 0),
            r.fecha_programada if r.fecha_programada else '',
            r.fecha_recuperacion if r.fecha_recuperacion else '',
            abono.numero_recibo if abono else '—',
            tipo_pago_label,
            float(monto or 0),
            abono.get_metodo_display() if abono else '—',
            banco,
            'Sí' if (abono and abono.cuenta_para_saldo) else ('No' if abono else '—'),
            r.observaciones or '—',
            '',  # Asistencia: en blanco para firma
        ])
        total_saldo += r.saldo_pendiente_al_marcar or Decimal('0.00')
        if abono and abono.tipo_pago == 'recuperacion':
            total_pagado += monto or Decimal('0.00')

    totals = {
        7: float(total_saldo),
        12: float(total_pagado),
    }
    filename = f'recuperaciones_{filtros["estado"]}_{date.today().strftime("%Y%m%d")}.xlsx'
    return _build_excel_response(
        filename=filename,
        sheet_name='Clases en Recuperación',
        headers=headers,
        rows=rows,
        totals=totals,
        column_formats={
            1: 'dd/mm/yyyy',
            7: '"$"#,##0.00',
            8: 'dd/mm/yyyy',
            9: 'dd/mm/yyyy',
            12: '"$"#,##0.00',
        },
        text_columns={2},
    )


@matricula_requerida
def recuperaciones_export_pdf(request):
    """Exporta la tabla filtrada de clases en recuperación a PDF horizontal."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    recuperaciones, filtros = _filtrar_recuperaciones(request)
    recuperaciones = list(recuperaciones)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=0.9*cm, rightMargin=0.9*cm, topMargin=1.1*cm, bottomMargin=0.9*cm,
        title='Clases en Recuperación',
    )
    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle(
        'titulo_recuperaciones', parent=styles['Title'],
        textColor=colors.HexColor('#1A237E'), fontSize=15,
        alignment=1, spaceAfter=4,
    )
    sub_st = ParagraphStyle(
        'sub_recuperaciones', parent=styles['Normal'],
        textColor=colors.HexColor('#666666'), fontSize=9,
        alignment=1, spaceAfter=10,
    )

    elementos = [
        Paragraph('Clases en Recuperación', titulo_st),
        Paragraph(
            f'Estado: {filtros["estado"].title()} · Generado el '
            f'{date.today().strftime("%d/%m/%Y")} · {len(recuperaciones)} registro(s)',
            sub_st,
        ),
    ]

    data = [[
        'Estado', 'Fecha falta', 'Fecha pactada', 'Estudiante', 'Cédula',
        'Curso', 'Mód.', 'Saldo', 'Pago recuperación', 'Método', 'Obs.',
        'Asistencia',
    ]]
    total_saldo = Decimal('0.00')
    total_pagado = Decimal('0.00')

    for r in recuperaciones:
        abono = r.abono
        estudiante = r.matricula.estudiante
        monto = abono.monto if abono else Decimal('0.00')
        metodo = abono.get_metodo_display() if abono else '—'
        if abono and abono.banco:
            metodo = f'{metodo} · {abono.get_banco_display()}'
        if abono and abono.tipo_pago != 'recuperacion':
            pago_recuperacion = f'Descontada con {abono.numero_recibo} · ${float(monto or 0):.2f}'
        else:
            pago_recuperacion = f'{abono.numero_recibo} · ${float(monto or 0):.2f}' if abono else 'Por cobrar'

        data.append([
            'Pagada' if r.pagada else 'Pendiente',
            r.fecha_marcada.strftime('%d/%m/%Y') if r.fecha_marcada else '',
            r.fecha_programada.strftime('%d/%m/%Y') if r.fecha_programada else '—',
            estudiante.nombre_completo,
            estudiante.cedula,
            r.matricula.curso.nombre,
            f'M{r.numero_modulo}',
            f'${float(r.saldo_pendiente_al_marcar or 0):.2f}',
            pago_recuperacion,
            metodo,
            (r.observaciones or '—')[:60],
            '',  # Asistencia: vacía para firma a mano
        ])
        total_saldo += r.saldo_pendiente_al_marcar or Decimal('0.00')
        if abono and abono.tipo_pago == 'recuperacion':
            total_pagado += monto or Decimal('0.00')

    data.append([
        '', '', '', '', '', 'TOTAL', '',
        f'${float(total_saldo):.2f}',
        f'${float(total_pagado):.2f}',
        '', '', '',
    ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('FONTSIZE', (0, 1), (-1, -2), 7),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F7F7F7')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1A237E')),
        # Línea para firma en la columna Asistencia (última, solo filas de datos)
        ('LINEBELOW', (-1, 1), (-1, -2), 0.5, colors.HexColor('#888888')),
    ]))
    elementos.append(table)
    elementos.append(Spacer(1, 0.2*cm))
    doc.build(elementos)

    pdf_bytes = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    filename = f'recuperaciones_{filtros["estado"]}_{date.today().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@matricula_requerida
@transaction.atomic
def recuperacion_marcar(request, matricula_pk):
    """
    Marcar una clase como pendiente de recuperación para una matrícula.
    Guarda automáticamente el saldo pendiente al momento.
    """
    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    if request.method == 'POST':
        form = RecuperacionPendienteForm(request.POST, matricula=matricula)
        if form.is_valid():
            recup = form.save(commit=False)
            recup.matricula = matricula
            recup.saldo_pendiente_al_marcar = matricula.saldo
            recup.save()
            messages.success(
                request,
                f'Clase de Módulo {recup.numero_modulo} marcada para recuperación. '
                f'Saldo arrastrado: ${recup.saldo_pendiente_al_marcar:.2f}.'
            )
            return redirect('academia:matricula_abonos', pk=matricula.pk)
    else:
        form = RecuperacionPendienteForm(
            initial={'fecha_marcada': date.today()},
            matricula=matricula,
        )

    return render(request, 'pagos/recuperacion_marcar.html', {
        'form': form,
        'matricula': matricula,
        'modo_edicion': False,
    })


@matricula_requerida
@transaction.atomic
def recuperacion_editar(request, recup_pk):
    """Editar una marca de recuperación sin borrar pagos asociados."""
    recup = get_object_or_404(
        RecuperacionPendiente.objects.select_related(
            'matricula', 'matricula__estudiante', 'matricula__curso', 'abono',
        ),
        pk=recup_pk,
    )
    matricula = recup.matricula
    abono_original = recup.abono
    tiene_pago_asociado = bool(abono_original)

    if request.method == 'POST':
        form = RecuperacionPendienteForm(
            request.POST, instance=recup, matricula=matricula
        )
        if form.is_valid():
            recup_editada = form.save(commit=False)
            recup_editada.matricula = matricula

            # Los pagos ya asociados, incluidos registros históricos, se
            # conservan. Ya no se pueden crear nuevos descuentos con un pago
            # "Abono + Módulo" desde esta pantalla.
            if tiene_pago_asociado:
                recup_editada.pagada = True
                recup_editada.fecha_recuperacion = (
                    recup.fecha_recuperacion or abono_original.fecha
                )
                recup_editada.abono = abono_original
                recup_editada.save()
                if (
                    abono_original.tipo_pago == 'recuperacion'
                    and abono_original.numero_modulo != recup_editada.numero_modulo
                ):
                    abono_original.numero_modulo = recup_editada.numero_modulo
                    abono_original.save(update_fields=['numero_modulo', 'actualizado'])
                messages.success(request, 'Recuperación actualizada. El pago asociado se conservó.')
                return _redirect_despues_recuperacion(request, matricula)

            recup_editada.pagada = False
            recup_editada.fecha_recuperacion = None
            recup_editada.abono = None
            recup_editada.save()
            messages.success(request, 'Recuperación actualizada y dejada como pendiente.')
            return _redirect_despues_recuperacion(request, matricula)
    else:
        form = RecuperacionPendienteForm(
            instance=recup,
            matricula=matricula,
        )

    return render(request, 'pagos/recuperacion_marcar.html', {
        'form': form,
        'matricula': matricula,
        'recuperacion': recup,
        'modo_edicion': True,
        'next_url': request.GET.get('next', ''),
    })


@matricula_requerida
@transaction.atomic
def recuperacion_cobrar(request, recup_pk):
    """
    Cobra la clase de recuperación: crea un Abono con tipo='recuperacion'
    y deja la marca como pagada=True.
    El usuario decide si el cobro suma al saldo del curso o se cobra aparte.
    """
    recup = get_object_or_404(
        RecuperacionPendiente.objects.select_related(
            'matricula', 'matricula__estudiante', 'matricula__curso',
        ),
        pk=recup_pk,
    )

    if recup.pagada:
        messages.info(request, 'Esta clase de recuperación ya fue cobrada.')
        return redirect('academia:recuperaciones_lista')

    matricula = recup.matricula

    if request.method == 'POST':
        # Forzar tipo_pago='recuperacion' y modulo=el de la recuperación
        post = request.POST.copy()
        post['tipo_pago'] = 'recuperacion'
        post['numero_modulo'] = recup.numero_modulo
        post['fecha_marcada'] = recup.fecha_marcada.isoformat()
        post['fecha_programada'] = (
            recup.fecha_programada.isoformat()
            if recup.fecha_programada else ''
        )
        form = AbonoForm(post, matricula=matricula)
        if form.is_valid():
            abono = form.save(commit=False)
            abono.matricula = matricula
            abono.tipo_pago = 'recuperacion'
            abono.numero_modulo = recup.numero_modulo
            abono.registrado_por = request.user
            if form.cleaned_data.get('tipo_cobro') == 'mixto':
                # `monto` conserva el total. En el modelo `monto_2` guarda
                # la segunda parte y la primera se obtiene como total - monto_2.
                abono.monto = form.cleaned_data['monto']
                abono.metodo = (
                    form.cleaned_data.get('metodo_pago_1') or 'efectivo'
                )
                abono.banco = form.cleaned_data.get('banco_1') or ''
                abono.monto_2 = (
                    form.cleaned_data.get('monto_pago_2')
                    or Decimal('0.00')
                )
                abono.metodo_2 = (
                    form.cleaned_data.get('metodo_pago_2') or 'efectivo'
                )
                abono.banco_2 = form.cleaned_data.get('banco_2') or ''
            else:
                abono.monto_2 = None
                abono.metodo_2 = ''
                abono.banco_2 = ''
            abono.save()
            # Marcar recuperación como pagada
            recup.pagada = True
            recup.fecha_recuperacion = abono.fecha
            recup.abono = abono
            recup.save()
            messages.success(
                request,
                f'Recuperación cobrada: {abono.numero_recibo} por ${abono.monto}. '
                f'{"(Sumó al saldo del curso)" if abono.cuenta_para_saldo else "(Cobrada aparte, no afecta saldo)"}.'
            )
            return redirect('academia:recuperaciones_lista')
    else:
        form = AbonoForm(
            initial={
                'fecha': date.today(),
                'monto': Decimal('25.00'),
                'tipo_pago': 'recuperacion',
                'numero_modulo': recup.numero_modulo,
                'cuenta_para_saldo': True,
                'fecha_marcada': recup.fecha_marcada,
                'fecha_programada': recup.fecha_programada,
            },
            matricula=matricula,
        )

    return render(request, 'pagos/recuperacion_cobrar.html', {
        'form': form,
        'recuperacion': recup,
        'matricula': matricula,
    })


@matricula_requerida
@require_POST
def recuperacion_eliminar(request, recup_pk):
    """Eliminar la marca de recuperación sin eliminar el abono asociado."""
    recup = get_object_or_404(
        RecuperacionPendiente.objects.select_related('matricula', 'abono'),
        pk=recup_pk,
    )
    matricula = recup.matricula
    abono = recup.abono
    mensaje_extra = ''
    if abono:
        mensaje_extra = f' El recibo {abono.numero_recibo} se conservó en el historial de pagos.'
    recup.delete()
    messages.success(request, f'Marca de recuperación eliminada.{mensaje_extra}')
    return _redirect_despues_recuperacion(request, matricula)


# ═════════════════════════════════════════════════════════════════
# Hoja de Recaudación imprimible (formato del PDF de Glenda/Kimberly)
# ═════════════════════════════════════════════════════════════════

DIAS_SEMANA_ES = ['LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO']


def _rango_recaudacion_desde_request(request):
    """Normaliza fecha única o rango de fechas de la hoja de recaudación."""
    fecha_legacy = (request.GET.get('fecha') or '').strip()
    fecha_desde = (request.GET.get('fecha_desde') or '').strip()
    fecha_hasta = (request.GET.get('fecha_hasta') or '').strip()

    fecha_desde_date = parse_date(fecha_desde) if fecha_desde else None
    fecha_hasta_date = parse_date(fecha_hasta) if fecha_hasta else None

    if fecha_desde and not fecha_desde_date:
        fecha_desde = ''
    if fecha_hasta and not fecha_hasta_date:
        fecha_hasta = ''

    if not fecha_desde_date and not fecha_hasta_date:
        fecha_legacy_date = parse_date(fecha_legacy) if fecha_legacy else None
        if fecha_legacy_date:
            fecha_desde_date = fecha_legacy_date
            fecha_hasta_date = fecha_legacy_date
            fecha_desde = fecha_hasta = fecha_legacy_date.isoformat()
        else:
            fecha_legacy = ''
    elif fecha_desde_date and not fecha_hasta_date:
        fecha_hasta_date = fecha_desde_date
        fecha_hasta = fecha_desde
    elif fecha_hasta_date and not fecha_desde_date:
        fecha_desde_date = fecha_hasta_date
        fecha_desde = fecha_hasta

    if (
        fecha_desde_date and fecha_hasta_date
        and fecha_desde_date > fecha_hasta_date
    ):
        fecha_desde_date, fecha_hasta_date = fecha_hasta_date, fecha_desde_date
        fecha_desde, fecha_hasta = (
            fecha_desde_date.isoformat(), fecha_hasta_date.isoformat()
        )

    fecha = (
        fecha_desde
        if fecha_desde_date and fecha_hasta_date and fecha_desde_date == fecha_hasta_date
        else ''
    )
    return fecha, fecha_desde, fecha_hasta, fecha_desde_date, fecha_hasta_date


def _recaudacion_periodo_label(fecha_desde, fecha_hasta):
    if not fecha_desde or not fecha_hasta:
        return ''
    if fecha_desde == fecha_hasta:
        return fecha_desde.strftime('%d/%m/%Y')
    return f'{fecha_desde:%d/%m/%Y} - {fecha_hasta:%d/%m/%Y}'


def _recaudacion_periodo_slug(fecha_desde, fecha_hasta):
    if not fecha_desde or not fecha_hasta:
        return 'sin_fecha'
    if fecha_desde == fecha_hasta:
        return fecha_desde.strftime('%Y%m%d')
    return f'{fecha_desde:%Y%m%d}_{fecha_hasta:%Y%m%d}'


def _recaudacion_filtros(fecha, fecha_desde, fecha_hasta, ciudad,
                         curso_id, jornada_id, modalidad):
    fecha_desde_date = parse_date(fecha_desde) if fecha_desde else None
    fecha_hasta_date = parse_date(fecha_hasta) if fecha_hasta else None
    params = {
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'ciudad': ciudad,
        'curso': curso_id,
        'jornada': jornada_id,
        'modalidad': modalidad,
    }
    if fecha:
        params['fecha'] = fecha

    querystring = urlencode({
        key: value for key, value in params.items() if value
    })
    return {
        'fecha': fecha,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'ciudad': ciudad,
        'curso': curso_id,
        'jornada': jornada_id,
        'modalidad': modalidad,
        'es_rango': bool(
            fecha_desde_date and fecha_hasta_date
            and fecha_desde_date != fecha_hasta_date
        ),
        'periodo_label': _recaudacion_periodo_label(
            fecha_desde_date, fecha_hasta_date,
        ),
        'periodo_slug': _recaudacion_periodo_slug(
            fecha_desde_date, fecha_hasta_date,
        ),
        'querystring': querystring,
    }


def _modulo_recaudacion_label(matricula, numero_modulo):
    """Etiqueta real del modulo vigente para pantalla, impresion y exportaciones."""
    try:
        numero = max(int(numero_modulo or 1), 1)
    except (TypeError, ValueError):
        numero = 1

    curso = getattr(matricula, 'curso', None)
    nombre = ''
    if curso and curso.nombrar_modulos and curso.nombres_modulos:
        nombres_por_modalidad = (
            curso.nombres_modulos if isinstance(curso.nombres_modulos, dict) else {}
        )
        nombres = nombres_por_modalidad.get(matricula.modalidad, []) or []
        if isinstance(nombres, list) and 1 <= numero <= len(nombres):
            nombre = str(nombres[numero - 1] or '').strip()

    return f'Módulo {numero} - {nombre}' if nombre else f'Módulo {numero}'


def _recuperacion_recaudacion_label(matricula, abonos_dia):
    """Estado de recuperación visible en pantalla, impresión y Excel."""
    recuperaciones = list(matricula.recuperaciones_pendientes.all())
    abonos_dia_ids = {a.pk for a in abonos_dia if a.pk}
    abonos_vinculados = set()
    etiquetas = []

    for recuperacion in recuperaciones:
        if (
            recuperacion.pagada
            and recuperacion.abono_id in abonos_dia_ids
        ):
            etiquetas.append(
                f'✱ Pagada · Módulo {recuperacion.numero_modulo}'
            )
            abonos_vinculados.add(recuperacion.abono_id)

    # Protege también pagos de recuperación antiguos o importados que aún no
    # tenían vínculo con RecuperacionPendiente.
    for abono in abonos_dia:
        if (
            abono.tipo_pago == 'recuperacion'
            and abono.pk not in abonos_vinculados
        ):
            etiquetas.append(
                f'✱ Pagada · Módulo {abono.numero_modulo or 1}'
            )

    for recuperacion in recuperaciones:
        if not recuperacion.pagada:
            etiquetas.append(
                f'✱ Pendiente · Módulo {recuperacion.numero_modulo}'
            )

    # Conserva el orden y evita duplicados si un dato histórico coincidiera.
    return ' · '.join(dict.fromkeys(etiquetas))


def _jornadas_recaudacion_queryset(curso_id, ciudad='', modalidad=''):
    """Jornadas disponibles para el selector de la hoja de recaudación."""
    if not (curso_id and str(curso_id).isdigit()):
        return JornadaCurso.objects.none()

    qs = JornadaCurso.objects.filter(curso_id=int(curso_id)).select_related(
        'curso', 'sede',
    ).order_by('fecha_inicio', 'modalidad', 'hora_inicio', 'id')
    if ciudad:
        qs = qs.filter(ciudad__iexact=ciudad)
    if modalidad in ('presencial', 'online'):
        qs = qs.filter(modalidad=modalidad)
    return qs


def _construir_hoja_recaudacion(curso, matriculas, fecha_obj, ciudad='',
                                jornada=None, fecha_hasta_obj=None):
    """Construye una hoja ya separada por jornada y fecha o periodo."""
    fecha_hasta_obj = fecha_hasta_obj or fecha_obj
    if fecha_hasta_obj < fecha_obj:
        fecha_obj, fecha_hasta_obj = fecha_hasta_obj, fecha_obj
    es_rango = fecha_obj != fecha_hasta_obj

    items = []
    total_efectivo = Decimal('0.00')
    total_transferencia = Decimal('0.00')
    total_recaudado = Decimal('0.00')
    total_cuotas = Decimal('0.00')

    for m in matriculas:
        # Abonos del estudiante registrados en la fecha o periodo seleccionado.
        if es_rango:
            abonos_dia = list(
                m.abonos.filter(fecha__gte=fecha_obj, fecha__lte=fecha_hasta_obj)
            )
        else:
            abonos_dia = list(m.abonos.filter(fecha=fecha_obj))
        pagado_dia = sum((a.monto for a in abonos_dia), Decimal('0.00'))

        partes_dia = [
            parte
            for abono in abonos_dia
            for parte in _partes_pago_abono(abono)
        ]

        # Forma de pago del día (concatenadas si hay varias)
        metodos = sorted({p['metodo_display'] for p in partes_dia})
        bancos = sorted({p['banco_display'] for p in partes_dia if p['banco']})
        forma = ', '.join(metodos) if metodos else '—'
        banco_str = ', '.join(bancos) if bancos else '—'

        plan_recaudacion = _plan_recaudacion_matricula(m, fecha_hasta_obj)
        modulo_actual = plan_recaudacion['modulo']

        recup_str = _recuperacion_recaudacion_label(m, abonos_dia)

        cuota_sugerida = plan_recaudacion['cuota_sugerida']

        # Suma a totales por método
        for parte in partes_dia:
            if parte['metodo'] == 'efectivo':
                total_efectivo += parte['monto']
            elif parte['metodo'] in ('transferencia', 'tarjeta'):
                total_transferencia += parte['monto']

        saldo_modulo = plan_recaudacion['saldo_modulo']
        total_cuotas += cuota_sugerida
        total_recaudado += pagado_dia

        items.append({
            'estudiante': m.estudiante,
            'matricula_id': m.pk,
            'modulo': modulo_actual,
            'modulo_label': _modulo_recaudacion_label(m, modulo_actual),
            'saldo_modulo': saldo_modulo,
            'cuota_sugerida': cuota_sugerida,
            'cuota_manual': plan_recaudacion.get('cuota_manual', False),
            'recaudado': pagado_dia,
            'forma_pago': forma,
            'banco': banco_str,
            'asistencia': '—',  # no tenemos campo asistencia, queda manual
            'recuperacion': recup_str,
            'talla': m.talla_camiseta or '',
            'jornada_inicio': (
                m.jornada.fecha_inicio
                if (m.jornada and m.jornada.fecha_inicio)
                else None
            ),
            'jornada_descripcion': (
                m.jornada.descripcion_legible if m.jornada else '—'
            ),
        })

    # Responsable: usuario que más matrículas registró en esta jornada.
    responsables = {}
    for m in matriculas:
        if m.registrado_por_id:
            nombre = (
                f'{m.registrado_por.first_name} {m.registrado_por.last_name}'.strip()
                or m.registrado_por.username
            )
            responsables[nombre] = responsables.get(nombre, 0) + 1
    responsable = max(responsables.items(), key=lambda x: x[1])[0] if responsables else '—'

    dia_semana = 'PERÍODO' if es_rango else DIAS_SEMANA_ES[fecha_obj.weekday()]
    ciudad_hoja = (
        jornada.ciudad if jornada and jornada.ciudad else ciudad or '—'
    )
    periodo_label = _recaudacion_periodo_label(fecha_obj, fecha_hasta_obj)

    return {
        'curso': curso,
        'jornada': jornada,
        'jornada_id': jornada.pk if jornada else '',
        'jornada_label': jornada.etiqueta if jornada else 'Sin jornada asignada',
        'jornada_inicio': jornada.fecha_inicio if jornada else None,
        'jornada_modalidad': jornada.get_modalidad_display() if jornada else '—',
        'fecha': fecha_obj,
        'fecha_desde': fecha_obj,
        'fecha_hasta': fecha_hasta_obj,
        'es_rango': es_rango,
        'periodo_label': periodo_label,
        'dia_semana': dia_semana,
        'ciudad': ciudad_hoja,
        'responsable': responsable,
        'items': items,
        'total_efectivo': total_efectivo,
        'total_transferencia': total_transferencia,
        'total_cuotas': total_cuotas,
        'total_recaudado': total_recaudado,
    }


def _construir_hojas_recaudacion(fecha_obj, curso_id, ciudad='',
                                 modalidad='', jornada_id='',
                                 fecha_hasta_obj=None):
    """
    Construye hojas de recaudación desde los filtros.
    Requiere curso; si no hay jornada específica, separa una hoja por jornada.
    """
    if not fecha_obj or not (curso_id and str(curso_id).isdigit()):
        return []
    fecha_hasta_obj = fecha_hasta_obj or fecha_obj
    if fecha_hasta_obj < fecha_obj:
        fecha_obj, fecha_hasta_obj = fecha_hasta_obj, fecha_obj

    curso = Curso.objects.filter(activo=True, pk=int(curso_id)).first()
    if not curso:
        return []

    mat_qs = Matricula.objects.filter(curso=curso).exclude(
        estado='retiro_voluntario',
    ).select_related(
        'estudiante', 'jornada', 'jornada__sede', 'registrado_por',
    ).prefetch_related(
        'abonos', 'recuperaciones_pendientes__abono',
    ).order_by(
        'jornada__fecha_inicio', 'jornada__modalidad', 'jornada__hora_inicio',
        'jornada__id', 'estudiante__nombres', 'id',
    )
    if ciudad:
        mat_qs = mat_qs.filter(jornada__ciudad__iexact=ciudad)
    if modalidad in ('presencial', 'online'):
        mat_qs = mat_qs.filter(jornada__modalidad=modalidad)
    if jornada_id and str(jornada_id).isdigit():
        mat_qs = mat_qs.filter(jornada_id=int(jornada_id))

    grupos = {}
    for matricula in mat_qs:
        key = matricula.jornada_id or 0
        if key not in grupos:
            grupos[key] = {
                'jornada': matricula.jornada if matricula.jornada_id else None,
                'matriculas': [],
            }
        grupos[key]['matriculas'].append(matricula)

    hojas = []
    for grupo in grupos.values():
        hojas.append(_construir_hoja_recaudacion(
            curso=curso,
            matriculas=grupo['matriculas'],
            fecha_obj=fecha_obj,
            ciudad=ciudad,
            jornada=grupo['jornada'],
            fecha_hasta_obj=fecha_hasta_obj,
        ))
    return hojas


def _distribuir_saldo_en_cuotas_enteras(saldo, cantidad):
    return distribuir_monto_en_cuotas_enteras(saldo, cantidad)


# ─────────────────────────────────────────────────────────────────
# ALGORITMO ÚNICO DE CUOTA DE RECAUDACIÓN
#
#   cuota_de_hoy = saldo pendiente del módulo vigente
#
# La Hoja de Recaudación cobra exclusivamente el módulo vigente. El valor
# objetivo de cada módulo sale del valor neto posterior a la reserva dividido
# entre la cantidad de módulos configurada. Los pagos se aplican en orden:
# primero completan el módulo actual y luego avanzan a los siguientes.
#
# Ejemplo: curso $110, reserva $10 y 4 módulos de $25. Si el estudiante
# pagó $40 en total, cubrió la reserva, el módulo 1 y $5 del módulo 2. La
# hoja muestra únicamente $20 como saldo del módulo 2; no redistribuye los
# $70 pendientes de todo el curso.
# ─────────────────────────────────────────────────────────────────


def _semanas_recaudacion_matricula(matricula):
    """Cantidad de cuotas configuradas para la hoja de recaudación."""
    if matricula.curso_id:
        if matricula.curso.usa_pago_unico_recaudacion(matricula.modalidad):
            return 1
        return max(int(matricula.curso.get_numero_modulos(matricula.modalidad) or 1), 1)
    return 1


def _reserva_base_recaudacion(matricula):
    """La reserva fija es independiente de las cuotas semanales."""
    return matricula.reserva_inicial_plan


def _cuotas_objetivo_recaudacion(matricula, total_semanas):
    """Cuotas originales del curso después de descontar la reserva fija."""
    cuotas = matricula.cuotas_modulos_objetivo()
    if len(cuotas) == total_semanas:
        return cuotas
    saldo_cuotas = max(
        matricula.valor_neto - _reserva_base_recaudacion(matricula),
        Decimal('0.00'),
    )
    return _distribuir_saldo_en_cuotas_enteras(saldo_cuotas, total_semanas)


def _pago_aplicable_a_cuotas(matricula):
    """Pago acumulado que corresponde a módulos, excluyendo la reserva."""
    valor_neto = matricula.valor_neto or Decimal('0.00')
    saldo = max(matricula.saldo, Decimal('0.00'))
    pagado_total = max(valor_neto - saldo, Decimal('0.00'))
    return max(
        pagado_total - _reserva_base_recaudacion(matricula),
        Decimal('0.00'),
    )


def _semanas_cubiertas_por_pago(matricula, total_semanas):
    """
    Semanas cubiertas por pagos posteriores a la reserva. Para matrículas
    nuevas de Reserva/Abono, los primeros $10 nunca cuentan como módulo:
    las cuotas se calculan sobre (valor neto - $10).

    El flujo antiguo "Reserva + Módulo" no separaba la reserva, por eso
    conserva el comportamiento acumulado previo.
    """
    total_semanas = max(int(total_semanas or 1), 1)
    valor_neto = matricula.valor_neto or Decimal('0.00')
    if valor_neto <= 0:
        return total_semanas

    cuotas_objetivo = _cuotas_objetivo_recaudacion(matricula, total_semanas)
    pagado = _pago_aplicable_a_cuotas(matricula)
    tolerancia = Decimal('0.01')  # absorbe diferencias de redondeo

    cubiertas = 0
    requerido = Decimal('0.00')
    for k, cuota in enumerate(cuotas_objetivo, start=1):
        requerido += cuota
        if pagado + tolerancia >= requerido:
            cubiertas = k
        else:
            break
    return cubiertas


def _semanas_calendario_restantes(matricula, fecha_recaudacion, total_semanas):
    """
    Semanas de pago que aún NO vencieron a la fecha de la hoja, según el
    calendario de la modalidad (el mismo del panel de alertas: presencial
    semanal desde el inicio de jornada; online un día antes del inicio y el
    saldo a los 13 días; ciclo corto online con cobro semanal o pago único,
    según la configuración del curso).

    Esto sostiene la regla invariable: aunque el estudiante venga atrasado,
    su saldo se reparte solo entre las semanas que realmente quedan, de modo
    que llega a $0 en la última semana del curso.
    """
    if not fecha_recaudacion or not matricula.jornada_id:
        return total_semanas
    if not getattr(matricula.jornada, 'fecha_inicio', None):
        return total_semanas
    try:
        calendario = _calendario_vencimientos(matricula)
    except Exception:
        return total_semanas

    # Se cuentan FECHAS DISTINTAS, no módulos: varios módulos pueden vencer
    # el mismo día y entonces son UN solo cobro. Es lo que ocurre con el
    # ciclo corto online configurado con pago único y con el
    # segundo cobro del online normal (todo el saldo restante a los 13 días).
    fechas = {
        venc for (venc, _hito) in calendario.values()
        if venc >= fecha_recaudacion
    }
    return len(fechas)


def _plan_recaudacion_matricula(matricula, fecha_recaudacion=None,
                                aplicar_manual=True):
    """
    Calcula el plan de cuotas de la Hoja de Recaudación para una matrícula.

    Devuelve dict con:
      modulo            → módulo/semana en curso (el primero no cubierto)
      saldo_pendiente   → saldo total del curso (solo para compatibilidad)
      saldo_modulo      → saldo pendiente del módulo vigente
      cuota_sugerida    → cuánto cobrar HOY; nunca supera saldo_modulo
      cuotas_pendientes → cantidad de módulos con saldo
      cuotas            → saldos del módulo vigente y módulos posteriores
      cuota_manual      → True si la cuota de hoy fue fijada a mano por el
                          usuario para esta fecha (ver CuotaManualRecaudacion)

    Si existe una cuota manual guardada para (matrícula, fecha) y
    aplicar_manual=True, esa cuota reemplaza a la automática en
    cuota_sugerida (acotada a [0, saldo_modulo]). HTML, impresión y exports
    pasan por esta misma función, así que el valor manual nunca puede cobrar
    parte de un módulo futuro.
    """
    saldo = matricula.saldo if matricula.saldo > 0 else Decimal('0.00')
    saldo = saldo.quantize(CENTAVO, rounding=ROUND_HALF_UP)
    total_semanas = _semanas_recaudacion_matricula(matricula)
    cubiertas_pago = _semanas_cubiertas_por_pago(matricula, total_semanas)

    # Sin saldo no hay nada que recaudar.
    if saldo <= 0:
        return {
            'modulo': total_semanas,
            'saldo_pendiente': Decimal('0.00'),
            'saldo_modulo': Decimal('0.00'),
            'cuota_sugerida': Decimal('0.00'),
            'cuotas_pendientes': 0,
            'cuotas': [],
            'cuota_estandar': Decimal('0.00'),
            'cuota_manual': False,
        }

    cuotas_objetivo = _cuotas_objetivo_recaudacion(
        matricula, total_semanas
    )
    pagado_cuotas = _pago_aplicable_a_cuotas(matricula)

    if cubiertas_pago < total_semanas:
        modulo_actual = cubiertas_pago + 1
        requerido_anterior = sum(
            cuotas_objetivo[:cubiertas_pago], Decimal('0.00')
        )
        pagado_modulo = max(
            pagado_cuotas - requerido_anterior,
            Decimal('0.00'),
        )
        saldo_modulo = max(
            cuotas_objetivo[cubiertas_pago] - pagado_modulo,
            Decimal('0.00'),
        ).quantize(CENTAVO, rounding=ROUND_HALF_UP)
        cuotas = [saldo_modulo] + cuotas_objetivo[cubiertas_pago + 1:]
        cuotas = [cuota for cuota in cuotas if cuota > 0]
    else:
        # Puede existir saldo total por una reserva incompleta, pero nunca se
        # disfraza ese valor como saldo de un módulo.
        modulo_actual = total_semanas
        saldo_modulo = Decimal('0.00')
        cuotas = []

    # ── Cuota manual guardada para esta fecha (si existe) ──
    cuota_sugerida = saldo_modulo
    cuota_manual = False
    if aplicar_manual and fecha_recaudacion is not None:
        from .models import CuotaManualRecaudacion
        override = CuotaManualRecaudacion.objects.filter(
            matricula=matricula, fecha=fecha_recaudacion,
        ).only('monto').first()
        if override is not None:
            # Nunca permite cobrar módulos futuros: el límite es el saldo
            # pendiente del módulo vigente, no el saldo total del curso.
            cuota_sugerida = max(
                min(override.monto, saldo_modulo), Decimal('0.00')
            ).quantize(CENTAVO, rounding=ROUND_HALF_UP)
            cuota_manual = True

    return {
        'modulo': modulo_actual,
        'saldo_pendiente': saldo,
        'saldo_modulo': saldo_modulo,
        'cuota_sugerida': cuota_sugerida,
        'cuotas_pendientes': len(cuotas),
        'cuotas': cuotas,
        'cuota_estandar': saldo_modulo,
        'cuota_manual': cuota_manual,
    }


@matricula_requerida
def matricula_comprobante_pdf(request, pk):
    """
    PDF con el resumen completo de una matrícula: los mismos datos de la
    pantalla "Confirmar matrícula" (Estudiante, Matrícula, Pago inicial y
    Comprobante). Solo lee datos, no modifica nada.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    from .views import _ids_abonos_pago_inicial

    matricula = get_object_or_404(
        Matricula.objects.select_related(
            'estudiante', 'curso', 'jornada', 'vendedora', 'registrado_por',
        ),
        pk=pk,
    )
    est = matricula.estudiante

    AZUL = colors.HexColor('#1A237E')
    GRIS = colors.HexColor('#666666')
    LINEA = colors.HexColor('#E3E6EF')
    FONDO = colors.HexColor('#FAFBFF')

    def _t(valor):
        """
        Texto seguro para la celda: '—' cuando no hay dato.
        Se eliminan los emojis (🏫, 💻, ...) porque las fuentes base de
        ReportLab no los tienen y se dibujarían como un cuadrito negro.
        """
        if valor is None:
            return '—'
        texto = str(valor).strip()
        if not texto:
            return '—'
        limpio = ''.join(ch for ch in texto if ord(ch) < 0x2190)
        limpio = ' '.join(limpio.split())  # normaliza espacios sobrantes
        return limpio if limpio else '—'

    def _money(valor):
        return f'${(valor or Decimal("0.00")):.2f}'

    # ── Pago inicial: los abonos del bloque de la matrícula ──
    ids_ini = _ids_abonos_pago_inicial(matricula)
    abonos_ini = list(
        Abono.objects.filter(id__in=ids_ini).order_by('creado', 'id')
    )
    monto_inicial = sum((a.monto for a in abonos_ini), Decimal('0.00'))

    filas_pago = [
        ('Valor curso', _money(matricula.valor_curso)),
        ('Descuento', _money(matricula.descuento)),
        ('Valor a pagar', _money(matricula.valor_neto)),
        ('Valor pagado (inicial)', _money(monto_inicial)),
        ('Forma pago', _t(matricula.get_forma_pago_display() if matricula.forma_pago else '')),
    ]
    if abonos_ini:
        a0 = abonos_ini[0]
        monto2 = getattr(a0, 'monto_2', None) or Decimal('0.00')
        if monto2 > 0:
            filas_pago.append(('Distribución', 'Pago Mixto'))
            filas_pago.append(('Monto 1', _money(a0.monto - monto2)))
            filas_pago.append(('Método 1', _t(a0.get_metodo_display())))
            if a0.banco:
                filas_pago.append(('Banco 1', _t(a0.get_banco_display())))
            filas_pago.append(('Monto 2', _money(monto2)))
            filas_pago.append(('Método 2', _t(a0.get_metodo_2_display() if hasattr(a0, 'get_metodo_2_display') else a0.metodo_2)))
            if getattr(a0, 'banco_2', ''):
                filas_pago.append(('Banco 2', _t(a0.get_banco_2_display())))
        else:
            filas_pago.append(('Distribución', 'Un solo método'))
            filas_pago.append(('Método', _t(a0.get_metodo_display())))
            if a0.banco:
                filas_pago.append(('Banco', _t(a0.get_banco_display())))
    filas_pago.append(('Saldo pendiente', _money(matricula.saldo)))

    filas_est = [
        ('Cédula/RUC', _t(est.cedula)),
        ('Nombres', _t(est.nombre_completo)),
        ('Celular', _t(est.celular)),
        ('Correo', _t(est.correo)),
        ('Ciudad', _t(est.ciudad)),
    ]

    filas_mat = [
        ('Estado', _t(matricula.get_estado_display())),
        ('Curso', _t(matricula.curso.nombre if matricula.curso_id else '')),
        ('Modalidad', _t(matricula.get_modalidad_display())),
        ('Tipo', _t(matricula.get_tipo_matricula_display())),
        ('Fecha', matricula.fecha_matricula.strftime('%d/%m/%Y') if matricula.fecha_matricula else '—'),
        ('Jornada', _t(matricula.jornada.etiqueta if matricula.jornada_id else '')),
    ]

    filas_comp = [
        ('Tipo registro', _t(matricula.get_tipo_registro_display() if matricula.tipo_registro else '')),
        ('Asesor', _t(matricula.vendedora.get_full_name() or matricula.vendedora.username if matricula.vendedora_id else '')),
        ('Factura', 'Sí' if matricula.factura_realizada == 'si' else 'No'),
    ]
    if matricula.factura_realizada == 'si':
        filas_comp.append(('Titular factura', _t(matricula.fact_nombres)))
        filas_comp.append(('Cédula/RUC factura', _t(matricula.fact_cedula)))
        filas_comp.append(('Correo factura', _t(matricula.fact_correo)))

    # ── Construcción del documento ──
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=1.6*cm, rightMargin=1.6*cm, topMargin=1.5*cm, bottomMargin=1.4*cm,
        title=f'Matrícula #{matricula.pk} — {est.nombre_completo}',
    )
    styles = getSampleStyleSheet()
    st_titulo = ParagraphStyle('t', parent=styles['Title'], textColor=AZUL,
                               fontSize=17, alignment=0, spaceAfter=2)
    st_sub = ParagraphStyle('s', parent=styles['Normal'], textColor=GRIS,
                            fontSize=9, spaceAfter=10)
    st_pie = ParagraphStyle('p', parent=styles['Normal'], textColor=GRIS,
                            fontSize=7.5, alignment=1)

    def bloque(titulo, filas):
        """Tabla de una sección (título + filas etiqueta/valor)."""
        st_lbl = ParagraphStyle('l', parent=styles['Normal'], fontSize=8.5,
                                textColor=colors.HexColor('#555555'))
        st_val = ParagraphStyle('v', parent=styles['Normal'], fontSize=8.5,
                                fontName='Helvetica-Bold')
        # Valores largos (correos, jornadas): fuente menor para que no se
        # corten a mitad de palabra dentro de la celda.
        st_val_sm = ParagraphStyle('vs', parent=st_val, fontSize=7,
                                   leading=8.6)
        st_tit = ParagraphStyle('bt', parent=styles['Normal'], fontSize=9,
                                textColor=AZUL, fontName='Helvetica-Bold')
        data = [[Paragraph(titulo.upper(), st_tit), '']]
        for etiqueta, valor in filas:
            estilo_valor = st_val_sm if len(valor) > 20 else st_val
            data.append([Paragraph(etiqueta, st_lbl), Paragraph(valor, estilo_valor)])
        tabla = Table(data, colWidths=[3.3*cm, 4.7*cm])
        estilo = [
            ('SPAN', (0, 0), (1, 0)),
            ('BACKGROUND', (0, 0), (-1, -1), FONDO),
            ('BOX', (0, 0), (-1, -1), 0.6, LINEA),
            ('LINEBELOW', (0, 1), (-1, -2), 0.4, LINEA),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 3.5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
        ]
        tabla.setStyle(TableStyle(estilo))
        return tabla

    elementos = [
        Paragraph('Resumen de Matrícula', st_titulo),
        Paragraph(
            f'Matrícula #{matricula.pk} · {est.nombre_completo} · '
            f'Generado el {date.today().strftime("%d/%m/%Y")}',
            st_sub,
        ),
    ]

    # Dos columnas: (Estudiante | Matrícula) y (Pago inicial | Comprobante)
    for izq, der in (
        (bloque('Estudiante', filas_est), bloque('Matrícula', filas_mat)),
        (bloque('Pago inicial', filas_pago), bloque('Comprobante', filas_comp)),
    ):
        grid = Table([[izq, der]], colWidths=[8.6*cm, 8.6*cm])
        grid.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elementos.append(grid)

    if matricula.observaciones:
        st_obs = ParagraphStyle('o', parent=styles['Normal'], fontSize=8.5)
        elementos.append(Spacer(1, 2))
        elementos.append(bloque('Observaciones', [('Detalle', _t(matricula.observaciones))]))

    elementos.append(Spacer(1, 14))
    elementos.append(Paragraph(
        'Formación Técnica Profesional · Documento informativo generado por el sistema.',
        st_pie,
    ))

    doc.build(elementos)
    buf.seek(0)

    nombre_archivo = f'matricula_{matricula.pk}_{est.cedula or "sin_cedula"}.pdf'
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return resp


@matricula_requerida
def hoja_recaudacion(request):
    """
    Vista imprimible: una hoja por jornada para una fecha o rango y curso dados.
    Replica el formato de las hojas físicas (Recaudaciones GYE/QUITO).

    Filtros: fecha/rango y curso obligatorios; ciudad, modalidad y jornada opcionales.
    Si no se filtra una jornada específica, genera hojas separadas por jornada
    del curso para evitar mezclar inicios distintos en la misma tabla.
    """
    (
        fecha_str, fecha_desde, fecha_hasta,
        fecha_desde_obj, fecha_hasta_obj,
    ) = _rango_recaudacion_desde_request(request)
    ciudad = request.GET.get('ciudad', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    jornada_id = request.GET.get('jornada', '').strip()
    modalidad = request.GET.get('modalidad', '').strip().lower()
    if modalidad not in ('presencial', 'online'):
        modalidad = ''  # vacío = todas las modalidades

    cursos_disponibles = Curso.objects.filter(activo=True).order_by('nombre')
    jornadas_disponibles = _jornadas_recaudacion_queryset(
        curso_id, ciudad=ciudad, modalidad=modalidad
    )
    jornadas_todas = JornadaCurso.objects.filter(
        curso__activo=True,
    ).select_related('curso', 'sede').order_by(
        'curso__nombre', 'fecha_inicio', 'modalidad', 'hora_inicio', 'id',
    )
    hojas = _construir_hojas_recaudacion(
        fecha_desde_obj, curso_id, ciudad=ciudad,
        modalidad=modalidad, jornada_id=jornada_id,
        fecha_hasta_obj=fecha_hasta_obj,
    )

    return render(request, 'pagos/hoja_recaudacion.html', {
        'cursos_disponibles': cursos_disponibles,
        'jornadas_disponibles': jornadas_disponibles,
        'jornadas_todas': jornadas_todas,
        'hojas': hojas,
        'filtros': _recaudacion_filtros(
            fecha_str, fecha_desde, fecha_hasta, ciudad,
            curso_id, jornada_id, modalidad,
        ),
    })


# ═════════════════════════════════════════════════════════════════
# Exportación de "Pagos por Módulo" (Excel y PDF)
# ═════════════════════════════════════════════════════════════════

def _export_pagos_modulo_filtros(request):
    """
    Resuelve el curso seleccionado y otros filtros para las exportaciones
    de Pagos por Módulo. Devuelve (curso_sel | None, dict_filtros).

    NOTA: el formulario de la pantalla y los enlaces de exportación
    arman el querystring con `filtro_modulo_estado` (mismo nombre que
    espera `_construir_matriz_pagos`). Mantener este nombre alineado.
    """
    curso_id = request.GET.get('curso', '').strip()
    modalidad = request.GET.get('modalidad', '').strip()
    ciudad = request.GET.get('ciudad', '').strip()
    tipo_matricula = request.GET.get('tipo_matricula', '').strip()
    if tipo_matricula not in dict(TIPOS_MATRICULA_FILTRO_PAGOS_MODULO):
        tipo_matricula = ''
    filtro_modulo_estado = request.GET.get('filtro_modulo_estado', '').strip()
    fecha_jornada = request.GET.get('fecha_jornada', '').strip()
    matricula_id = request.GET.get('matricula', '').strip()
    (
        fecha_modulo_desde, fecha_modulo_hasta,
        fecha_modulo_desde_date, fecha_modulo_hasta_date,
    ) = _rango_fecha_modulo_desde_request(request)

    curso_sel = None
    if curso_id and curso_id.isdigit():
        try:
            curso_sel = Curso.objects.get(pk=int(curso_id), activo=True)
        except Curso.DoesNotExist:
            curso_sel = None
    matricula_id = _normalizar_matricula_filtro_pagos_modulo(
        matricula_id, curso_sel
    )

    return curso_sel, {
        'modalidad': modalidad,
        'ciudad': ciudad,
        'tipo_matricula': tipo_matricula,
        'filtro_modulo_estado': filtro_modulo_estado,
        'fecha_jornada': fecha_jornada,
        'matricula': matricula_id,
        'fecha_modulo_desde': fecha_modulo_desde,
        'fecha_modulo_hasta': fecha_modulo_hasta,
        'fecha_modulo_desde_date': fecha_modulo_desde_date,
        'fecha_modulo_hasta_date': fecha_modulo_hasta_date,
    }


@matricula_requerida
def pagos_por_modulo_export_excel(request):
    """Exporta la matriz de pagos por módulo del curso filtrado a Excel."""
    curso_sel, filtros = _export_pagos_modulo_filtros(request)
    if not curso_sel:
        messages.error(request, 'Selecciona un curso para exportar la matriz.')
        return redirect('academia:pagos_por_modulo')

    matriculas, _modulos, _resumen, modulos_visibles = _construir_matriz_pagos(
        curso_sel,
        modalidad=filtros['modalidad'],
        ciudad=filtros['ciudad'],
        tipo_matricula=filtros['tipo_matricula'],
        filtro_modulo_estado=filtros['filtro_modulo_estado'],
        fecha_jornada=filtros['fecha_jornada'],
        matricula_id=filtros['matricula'],
        fecha_modulo_desde=filtros['fecha_modulo_desde_date'],
        fecha_modulo_hasta=filtros['fecha_modulo_hasta_date'],
    )

    # Encabezados base + 1 columna por módulo + Asistencia
    headers = [
        'Cédula', 'Estudiante', 'Curso', 'Jornada', 'Día (inicio jornada)',
        'Tipo matrícula', 'Horario', 'Sede',
        'Valor neto', 'Pagado', 'Saldo',
    ]
    for n in modulos_visibles:
        headers.append(f'Mód. {n} (estado / pagado)')
    headers.append('Asistencia')

    rows = []
    total_neto = total_pagado = total_saldo = Decimal('0.00')
    for x in matriculas:
        m = x['matricula']
        estu = x['estudiante']
        j = m.jornada
        if j and j.hora_inicio and j.hora_fin:
            horario_txt = f"{j.hora_inicio.strftime('%H:%M')} – {j.hora_fin.strftime('%H:%M')}"
        else:
            horario_txt = '—'
        fila = [
            estu.cedula,
            f'{estu.nombres}'.strip(),
            x['curso_nombre'],
            x['jornada_dia'],
            x['jornada_inicio'] if x['jornada_inicio'] else '',
            x['tipo_matricula_label'],
            horario_txt,
            (j.ciudad if (j and j.ciudad) else '—'),
            float(m.valor_neto or 0),
            float(m.valor_pagado or 0),
            float(m.saldo or 0),
        ]
        for mod in x['modulos_visibles_data']:
            if (
                (filtros['fecha_modulo_desde_date'] or filtros['fecha_modulo_hasta_date'])
                and not mod.get('coincide_fecha_modulo')
            ):
                fila.append('—')
                continue
            if mod.get('aplica', True):
                detalle = (
                    f"{mod['estado']} – ${float(mod['pagado']):.2f} / "
                    f"${float(mod['esperado']):.2f}"
                )
                if mod.get('es_recuperacion'):
                    detalle += ' – Recuperación'
                fila.append(detalle)
            else:
                fila.append('— (no aplica)')
        fila.append('')  # Asistencia: en blanco para firmar
        rows.append(fila)
        total_neto += m.valor_neto or Decimal('0.00')
        total_pagado += m.valor_pagado or Decimal('0.00')
        total_saldo += m.saldo or Decimal('0.00')

    totals = {
        8: round(float(total_neto), 2),
        9: round(float(total_pagado), 2),
        10: round(float(total_saldo), 2),
    }
    filename = f'pagos_modulo_{curso_sel.pk}_{date.today().strftime("%Y%m%d")}.xlsx'
    sheet_name = f'Pagos por Módulo - {curso_sel.nombre}'[:31]
    return _build_excel_response(
        filename,
        sheet_name,
        headers,
        rows,
        totals=totals,
        column_formats={
            4: 'dd/mm/yyyy',
            8: '"$"#,##0.00',
            9: '"$"#,##0.00',
            10: '"$"#,##0.00',
        },
        text_columns={0},
    )


@matricula_requerida
def pagos_por_modulo_export_pdf(request):
    """Exporta la matriz de pagos por módulo a un PDF horizontal."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A3, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        )
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    curso_sel, filtros = _export_pagos_modulo_filtros(request)
    if not curso_sel:
        messages.error(request, 'Selecciona un curso para exportar la matriz.')
        return redirect('academia:pagos_por_modulo')

    matriculas, _modulos, _resumen, modulos_visibles = _construir_matriz_pagos(
        curso_sel,
        modalidad=filtros['modalidad'],
        ciudad=filtros['ciudad'],
        tipo_matricula=filtros['tipo_matricula'],
        filtro_modulo_estado=filtros['filtro_modulo_estado'],
        fecha_jornada=filtros['fecha_jornada'],
        matricula_id=filtros['matricula'],
        fecha_modulo_desde=filtros['fecha_modulo_desde_date'],
        fecha_modulo_hasta=filtros['fecha_modulo_hasta_date'],
    )

    # ── Elegir tamaño de página según número de módulos ──
    n_mod = len(modulos_visibles)
    # 9 columnas fijas + n_mod + 1 (asistencia)
    n_cols_total = 9 + n_mod + 1
    if n_cols_total > 13:
        page_size = landscape(A3)  # ancho útil ≈ 41 cm
        page_width_cm = 41.0
    else:
        page_size = landscape(A4)  # ancho útil ≈ 28 cm
        page_width_cm = 28.0

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=0.7*cm, rightMargin=0.7*cm, topMargin=1.0*cm, bottomMargin=0.8*cm,
        title=f'Pagos por Módulo — {curso_sel.nombre}',
    )
    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle('titulo', parent=styles['Title'],
                               textColor=colors.HexColor('#1A237E'),
                               fontSize=14, alignment=1, spaceAfter=4)
    sub_st = ParagraphStyle('sub', parent=styles['Normal'],
                            textColor=colors.HexColor('#666666'),
                            fontSize=9, alignment=1, spaceAfter=10)
    cell_st = ParagraphStyle(
        'cell', parent=styles['Normal'],
        fontName='Helvetica', fontSize=7, leading=8.5,
    )
    cell_bold_st = ParagraphStyle('cell_b', parent=cell_st, fontName='Helvetica-Bold')
    header_st = ParagraphStyle(
        'h_st', parent=styles['Normal'],
        fontName='Helvetica-Bold', fontSize=8,
        textColor=colors.whitesmoke, alignment=1, leading=9,
    )

    elementos = [
        Paragraph(f'Pagos por Módulo — {curso_sel.nombre}', titulo_st),
        Paragraph(
            f'Formación Técnica y Profesional EC · Generado el '
            f'{date.today().strftime("%d/%m/%Y")} · {len(matriculas)} matrícula(s)',
            sub_st,
        ),
    ]

    headers = [
        'Cédula', 'Estudiante', 'Curso', 'Jornada', 'Día',
        'Tipo matrícula', 'Valor', 'Pagado', 'Saldo',
    ] + [f'Mód. {n}' for n in modulos_visibles] + ['Asistencia']

    data = [[Paragraph(h, header_st) for h in headers]]
    total_neto = total_pagado = total_saldo = 0.0
    for x in matriculas:
        m = x['matricula']
        e = x['estudiante']
        fila = [
            Paragraph(e.cedula or '', cell_bold_st),
            Paragraph(f'{e.nombres}'.strip(), cell_st),
            Paragraph(x['curso_nombre'] or '', cell_st),
            Paragraph(x['jornada_dia'] or '', cell_st),
            Paragraph(x['jornada_inicio'].strftime('%d/%m/%Y') if x['jornada_inicio'] else '—', cell_st),
            Paragraph(x['tipo_matricula_label'] or '', cell_st),
            Paragraph(f'${float(m.valor_neto or 0):.2f}', cell_st),
            Paragraph(f'<font color="#2e7d32"><b>${float(m.valor_pagado or 0):.2f}</b></font>', cell_st),
            Paragraph(f'<font color="{"#c62828" if (m.saldo or 0) > 0 else "#2e7d32"}"><b>${float(m.saldo or 0):.2f}</b></font>', cell_st),
        ]
        for mod in x['modulos_visibles_data']:
            if (
                (filtros['fecha_modulo_desde_date'] or filtros['fecha_modulo_hasta_date'])
                and not mod.get('coincide_fecha_modulo')
            ):
                fila.append(Paragraph('<font color="#bbbbbb">—</font>', cell_st))
                continue
            if not mod.get('aplica', True):
                fila.append(Paragraph('<font color="#bbbbbb">—</font>', cell_st))
                continue
            estado = mod['estado']
            simbolo = {'Pagado': '✓', 'Parcial': '◐', 'Pendiente': '○'}.get(estado, '')
            color = {'Pagado': '#2e7d32', 'Parcial': '#f0ad4e', 'Pendiente': '#c62828'}.get(estado, '#000')
            marca_recuperacion = (
                '<br/><font color="#8a5a00"><b>Recuperación</b></font>'
                if mod.get('es_recuperacion') else ''
            )
            fila.append(Paragraph(
                f'<font color="{color}"><b>{simbolo} ${float(mod["pagado"]):.2f}</b></font>'
                f'{marca_recuperacion}',
                cell_st,
            ))
        fila.append('')  # Asistencia: vacío para firma a mano
        data.append(fila)
        total_neto += float(m.valor_neto or 0)
        total_pagado += float(m.valor_pagado or 0)
        total_saldo += float(m.saldo or 0)

    # Fila de totales
    fila_total = [Paragraph('', cell_st)] * 5 + [Paragraph('<b>TOTAL</b>', cell_bold_st)]
    fila_total += [
        Paragraph(f'<b>${total_neto:.2f}</b>', cell_bold_st),
        Paragraph(f'<b>${total_pagado:.2f}</b>', cell_bold_st),
        Paragraph(f'<b>${total_saldo:.2f}</b>', cell_bold_st),
    ]
    fila_total += [Paragraph('', cell_st)] * len(modulos)
    fila_total.append('')
    data.append(fila_total)

    # ── Anchos de columna explícitos: distribuir el ancho útil ──
    # 9 columnas fijas con anchos predefinidos + módulos + asistencia
    fixed_widths_cm = {
        'cedula': 1.9, 'estudiante': 4.2, 'curso': 3.0, 'jornada': 2.6,
        'dia': 1.7, 'tipo': 2.6, 'valor': 1.7, 'pagado': 1.7, 'saldo': 1.7,
    }
    fixed_total = sum(fixed_widths_cm.values())  # ≈ 21.1 cm
    asistencia_cm = 2.5
    remaining = page_width_cm - fixed_total - asistencia_cm
    mod_width_cm = max(1.4, remaining / max(n_mod, 1)) if n_mod else 0

    col_widths = [
        fixed_widths_cm['cedula']*cm,
        fixed_widths_cm['estudiante']*cm,
        fixed_widths_cm['curso']*cm,
        fixed_widths_cm['jornada']*cm,
        fixed_widths_cm['dia']*cm,
        fixed_widths_cm['tipo']*cm,
        fixed_widths_cm['valor']*cm,
        fixed_widths_cm['pagado']*cm,
        fixed_widths_cm['saldo']*cm,
    ] + [mod_width_cm*cm] * n_mod + [asistencia_cm*cm]

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN',     (0, 0), (-1, 0), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('VALIGN',     (0, 1), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#BBBBBB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8F9FB')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E1')),
        # Línea de firma en la columna de Asistencia
        ('LINEBELOW', (-1, 1), (-1, -2), 0.5, colors.HexColor('#888888')),
    ]))
    elementos.append(table)
    doc.build(elementos)

    pdf_bytes = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    filename = f'pagos_modulo_{curso_sel.pk}_{date.today().strftime("%Y%m%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Exportación de la "Hoja de Recaudación" (Excel y PDF)
# ═════════════════════════════════════════════════════════════════

def _hojas_recaudacion_data(request):
    """
    Re-construye la data que hoja_recaudacion() entrega al template,
    usando los mismos filtros GET. Devuelve (hojas, filtros).
    """
    (
        fecha_str, fecha_desde, fecha_hasta,
        fecha_desde_obj, fecha_hasta_obj,
    ) = _rango_recaudacion_desde_request(request)
    ciudad = request.GET.get('ciudad', '').strip()
    curso_id = request.GET.get('curso', '').strip()
    jornada_id = request.GET.get('jornada', '').strip()
    modalidad = request.GET.get('modalidad', '').strip().lower()
    if modalidad not in ('presencial', 'online'):
        modalidad = ''

    hojas = _construir_hojas_recaudacion(
        fecha_desde_obj, curso_id, ciudad=ciudad,
        modalidad=modalidad, jornada_id=jornada_id,
        fecha_hasta_obj=fecha_hasta_obj,
    )
    return hojas, _recaudacion_filtros(
        fecha_str, fecha_desde, fecha_hasta, ciudad,
        curso_id, jornada_id, modalidad,
    )


@matricula_requerida
@require_POST
def hoja_recaudacion_guardar_cuotas(request):
    """
    Guarda las cuotas manuales de la Hoja de Recaudación para UNA fecha.

    Recibe JSON: {"fecha": "YYYY-MM-DD",
                  "cuotas": [{"matricula_id": 1, "monto": "15.00"}, ...]}

    Reglas:
      • Cada monto se acota a [0, saldo pendiente del módulo vigente].
      • Si el monto queda IGUAL a la cuota automática del sistema, el
        registro manual se elimina: esa fila vuelve al cálculo dinámico.
      • Si difiere, se guarda/actualiza como CuotaManualRecaudacion y a
        partir de ahí la hoja de esa fecha (HTML, Excel y PDF) lo usa.
    """
    import json as _json
    from datetime import datetime as _dt
    from .models import CuotaManualRecaudacion

    try:
        payload = _json.loads(request.body.decode('utf-8'))
        fecha = _dt.strptime(str(payload.get('fecha', '')), '%Y-%m-%d').date()
        cuotas = payload.get('cuotas', [])
        assert isinstance(cuotas, list)
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Datos inválidos.'}, status=400)

    guardadas, restauradas, sin_cambio = 0, 0, 0
    for fila in cuotas:
        try:
            matricula = Matricula.objects.get(pk=int(fila.get('matricula_id')))
            monto = Decimal(str(fila.get('monto'))).quantize(
                CENTAVO, rounding=ROUND_HALF_UP
            )
        except Exception:
            continue

        # Cuota automática (sin overrides) para saber si hay diferencia real.
        plan_auto = _plan_recaudacion_matricula(
            matricula, fecha, aplicar_manual=False
        )
        saldo_modulo = plan_auto['saldo_modulo']
        monto = max(min(monto, saldo_modulo), Decimal('0.00'))
        if abs(monto - plan_auto['cuota_sugerida']) < Decimal('0.005'):
            borradas, _ = CuotaManualRecaudacion.objects.filter(
                matricula=matricula, fecha=fecha,
            ).delete()
            if borradas:
                restauradas += 1
            else:
                sin_cambio += 1
        else:
            CuotaManualRecaudacion.objects.update_or_create(
                matricula=matricula, fecha=fecha,
                defaults={'monto': monto, 'registrado_por': request.user},
            )
            guardadas += 1

    return JsonResponse({
        'ok': True,
        'guardadas': guardadas,
        'restauradas': restauradas,
        'sin_cambio': sin_cambio,
    })


@matricula_requerida
def hoja_recaudacion_export_excel(request):
    """Exporta las hojas de recaudación a Excel (todos los cursos en una sola hoja)."""
    hojas, filtros = _hojas_recaudacion_data(request)
    if not hojas:
        messages.error(
            request,
            'No hay hojas para exportar. Verifica que la fecha tenga matrículas.'
        )
        return redirect('academia:hoja_recaudacion')

    headers = [
        'Curso', 'Jornada', 'Fecha', 'Día', 'Ciudad', 'Responsable', '#',
        'Estudiante', 'Inicio jornada', 'Mód.',
        'Recaudado', 'Forma de pago', 'Banco', 'Recuperación', 'Talla',
    ]
    rows = []
    for h in hojas:
        for idx, item in enumerate(h['items'], start=1):
            rows.append([
                h['curso'].nombre,
                h.get('jornada_label') or '—',
                h['periodo_label'] if h.get('es_rango') else h['fecha'],
                h['dia_semana'],
                h['ciudad'],
                h['responsable'],
                idx,
                item['estudiante'].nombre_completo if hasattr(item['estudiante'], 'nombre_completo')
                else f"{item['estudiante'].nombres}".strip(),
                item['jornada_inicio'] if item['jornada_inicio'] else '',
                item['modulo_label'],
                float(item['recaudado'] or 0),
                item['forma_pago'],
                item['banco'],
                item['recuperacion'],
                item['talla'],
            ])

    filename = f'hoja_recaudacion_{filtros["periodo_slug"]}.xlsx'
    sheet_name = f'Recaudación {filtros["periodo_slug"]}'[:31]
    return _build_excel_response(
        filename,
        sheet_name,
        headers,
        rows,
        column_formats={
            2: 'dd/mm/yyyy',
            8: 'dd/mm/yyyy',
            10: '"$"#,##0.00',
        },
    )


@matricula_requerida
def hoja_recaudacion_export_pdf(request):
    """Exporta las hojas de recaudación a un PDF (una página por curso)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
        )
    except ImportError:
        return HttpResponse(
            'Para exportar a PDF instala reportlab: pip install reportlab',
            status=500, content_type='text/plain; charset=utf-8',
        )

    hojas, filtros = _hojas_recaudacion_data(request)
    if not hojas:
        messages.error(
            request,
            'No hay hojas para exportar. Verifica que la fecha tenga matrículas.'
        )
        return redirect('academia:hoja_recaudacion')

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=1*cm, rightMargin=1*cm, topMargin=1.2*cm, bottomMargin=1*cm,
        title=f'Hoja de Recaudación — {filtros["periodo_label"]}',
    )
    styles = getSampleStyleSheet()
    titulo_st = ParagraphStyle('titulo', parent=styles['Title'],
                               textColor=colors.HexColor('#1A237E'),
                               fontSize=14, alignment=1, spaceAfter=4)
    sub_st = ParagraphStyle('sub', parent=styles['Normal'],
                            textColor=colors.HexColor('#666666'),
                            fontSize=9, alignment=1, spaceAfter=10)
    meta_st = ParagraphStyle('meta', parent=styles['Normal'],
                             fontSize=9, spaceAfter=6)

    elementos = []

    for idx_hoja, h in enumerate(hojas):
        elementos.append(Paragraph(
            f'Recaudación — {h["curso"].nombre} — {h.get("jornada_label") or "—"}',
            titulo_st,
        ))
        elementos.append(Paragraph(
            f'<b>{"Período" if h.get("es_rango") else "Fecha"}:</b> '
            f'{h["dia_semana"]} {h["periodo_label"]} · '
            f'<b>Ciudad:</b> {h["ciudad"]} · '
            f'<b>Jornada:</b> {h.get("jornada_label") or "—"} · '
            f'<b>Responsable:</b> {h["responsable"]} · '
            f'<b>Estudiantes:</b> {len(h["items"])}',
            meta_st,
        ))

        headers = [
            '#', 'Estudiante', 'Inicio jornada', 'Mód.',
            'Recaudado', 'Forma', 'Banco', 'Recuperación',
        ]
        data = [headers]
        for i, item in enumerate(h['items'], start=1):
            est = item['estudiante']
            nombre = (est.nombre_completo if hasattr(est, 'nombre_completo')
                      else f'{est.nombres}'.strip())
            data.append([
                str(i),
                nombre,
                item['jornada_inicio'].strftime('%d/%m/%Y') if item['jornada_inicio'] else '—',
                item['modulo_label'],
                f"${float(item['recaudado']):.2f}",
                item['forma_pago'],
                item['banco'],
                item['recuperacion'],
            ])
        # Fila de totales
        data.append([
            '', 'TOTAL', '', '',
            f"${float(h['total_recaudado']):.2f}",
            f"Efectivo: ${float(h['total_efectivo']):.2f}",
            f"Transf.: ${float(h['total_transferencia']):.2f}",
            '',
        ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F0AD4E')),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, 0), 8),
            ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
            ('FONTSIZE',   (0, 1), (-1, -2), 7),
            ('GRID',       (0, 0), (-1, -1), 0.3, colors.HexColor('#CCCCCC')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF8E1')),
            ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR',  (0, -1), (-1, -1), colors.HexColor('#1A237E')),
            ('FONTSIZE',   (0, -1), (-1, -1), 8),
        ]))
        elementos.append(table)

        if idx_hoja < len(hojas) - 1:
            elementos.append(PageBreak())

    doc.build(elementos)
    pdf_bytes = buf.getvalue()
    buf.close()
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    filename = f'hoja_recaudacion_{filtros["periodo_slug"]}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ═════════════════════════════════════════════════════════════════
# Alertas de pago pendiente (calendario según modalidad y tipo de curso)
# ═════════════════════════════════════════════════════════════════

def _calendario_vencimientos(matricula):
    """
    Construye el calendario de vencimiento de cada módulo según la modalidad
    de la matrícula y si el curso es de ciclo corto.

    La única fecha base es `jornada.fecha_inicio`. `fecha_matricula` no
    interviene: indica cuándo se inscribió el estudiante, no cuándo empiezan
    sus semanas de pago.

    Reglas de negocio:
      • PRESENCIAL (cualquier cantidad de módulos):
        cada módulo representa una semana. El módulo k vence al iniciar
        la semana k, es decir
        inicio_jornada + (k-1)*7 días. Primer pago el mismo día de inicio,
        el siguiente a los 7 días, y así sucesivamente hasta completar
        exactamente la cantidad de módulos configurada para el curso.
      • CURSO CON PAGOS CADA DOS SEMANAS: conserva la fecha normal del
        primer pago y programa cada pago posterior cada 14 días desde el
        inicio de la jornada. La configuración es exclusiva de cada curso.
      • ONLINE normal: el módulo 1 vence UN DÍA ANTES del inicio
        de la jornada (ej. inicia 10/07 → vence 09/07). El resto del valor
        pendiente (saldo restante) vence a los 13 días del inicio (un día
        antes de que arranque la segunda mitad del curso).
      • ONLINE ciclo corto: por defecto, un cobro por módulo separado por 7
        días. Si tiene activado "pago único online", todo el saldo posterior
        a la reserva vence en un solo cobro un día antes del inicio.
      • ONLINE de 1 solo módulo: un solo pago un día antes del inicio.

    Devuelve un dict {numero_modulo: (fecha_vencimiento, hito)} donde hito es:
      - 'modulo'         → pago de módulo (semanal presencial / módulo 1 online)
      - 'pago_unico'     → un solo pago por todo el curso
      - 'saldo_restante' → segundo pago online: todo el saldo que quede
    """
    from datetime import timedelta

    inicio = matricula.jornada.fecha_inicio
    n_mod = (matricula.curso.get_numero_modulos(matricula.modalidad) or 1)
    es_corto = bool(getattr(matricula.curso, 'es_ciclo_corto', False))
    pago_unico = matricula.curso.usa_pago_unico_recaudacion(
        matricula.modalidad
    )

    calendario = {}
    if matricula.modalidad == 'online':
        if n_mod == 1 or pago_unico:
            calendario[1] = (inicio - timedelta(days=1), 'pago_unico')
        elif matricula.curso.pagos_cada_dos_semanas:
            calendario[1] = (inicio - timedelta(days=1), 'modulo')
            for k in range(2, n_mod + 1):
                calendario[k] = (
                    inicio + timedelta(days=(k - 1) * 14),
                    'modulo',
                )
        elif es_corto:
            # Ciclo corto: conserva exactamente una cuota por módulo/semana.
            for k in range(1, n_mod + 1):
                calendario[k] = (
                    inicio - timedelta(days=1) + timedelta(days=(k - 1) * 7),
                    'modulo',
                )
        else:
            # Online normal: módulo 1 un día antes; saldo restante a los 13 días.
            calendario[1] = (inicio - timedelta(days=1), 'modulo')
            for k in range(2, n_mod + 1):
                calendario[k] = (inicio + timedelta(days=13), 'saldo_restante')
    else:
        # Presencial: un módulo equivale a una semana. La cantidad de fechas
        # sale del número de módulos configurado en el curso, sin valores fijos.
        intervalo_dias = (
            14 if matricula.curso.pagos_cada_dos_semanas else 7
        )
        for k in range(1, n_mod + 1):
            calendario[k] = (
                inicio + timedelta(days=(k - 1) * intervalo_dias),
                'modulo',
            )

    return calendario


def _calendario_alertas_pago(matricula):
    """
    Calendario exclusivo del panel "Gestión de Matrículas".

    La primera obligación online se habilita un día antes del inicio de la
    jornada; la presencial conserva exactamente la fecha de inicio. Las
    obligaciones siguientes mantienen sus fechas actuales: cada siete días,
    o cada catorce cuando el curso tiene activada esa configuración
    independiente. El panel conserva el primer módulo impago: nunca salta al
    módulo de la semana actual. Los cursos con pago único mantienen una sola
    obligación, aunque académicamente tengan más de un módulo.
    """
    from datetime import timedelta

    inicio = matricula.jornada.fecha_inicio
    total_cuotas = _semanas_recaudacion_matricula(matricula)
    pago_unico = matricula.curso.usa_pago_unico_recaudacion(
        matricula.modalidad
    )
    primera_fecha = (
        inicio - timedelta(days=1)
        if matricula.modalidad == 'online'
        else inicio
    )

    if pago_unico:
        return {1: (primera_fecha, 'pago_unico')}

    intervalo_dias = (
        14 if matricula.curso.pagos_cada_dos_semanas else 7
    )

    return {
        numero: (
            primera_fecha if numero == 1 else (
                inicio + timedelta(days=(numero - 1) * intervalo_dias)
            ),
            'modulo',
        )
        for numero in range(1, total_cuotas + 1)
    }


def _calcular_alertas_pago(usuario_actual=None):
    """
    Devuelve la lista de alertas activas: matrículas tipo "Reserva/Abono" o
    "Reserva + Módulo 1" con saldo pendiente cuyo hito de pago YA venció
    según el calendario semanal del panel (ver _calendario_alertas_pago).

    Siempre muestra el primer módulo que todavía no está cubierto. Cuando se
    paga, la alerta se oculta hasta que llegue la fecha del siguiente módulo.
    Con saldo cero desaparece definitivamente. Excluye también las alertas que
    ya fueron marcadas como "revisadas hoy".
    """
    from .models import AlertaPagoRevisada
    from datetime import timedelta

    hoy = date.today()

    # Online entra al panel un día antes; presencial conserva el día exacto.
    qs = Matricula.objects.filter(
        tipo_matricula__in=TIPOS_CON_RESERVA,
    ).filter(
        Q(
            modalidad='online',
            jornada__fecha_inicio__lte=hoy + timedelta(days=1),
        )
        | Q(
            modalidad='presencial',
            jornada__fecha_inicio__lte=hoy,
        )
    ).exclude(estado='retiro_voluntario').select_related(
        'estudiante', 'curso', 'jornada', 'jornada__sede'
    ).prefetch_related(
        Prefetch('abonos', queryset=Abono.objects.prefetch_related('recuperaciones'))
    )

    # Set de (matricula_id, modulo) ya revisadas hoy → para excluir
    revisadas_hoy = set(
        AlertaPagoRevisada.objects.filter(fecha=hoy)
        .values_list('matricula_id', 'numero_modulo')
    )

    alertas = []
    for m in qs:
        saldo_total = m.saldo
        if saldo_total <= 0:
            continue  # curso totalmente pagado: nunca alertar

        n_mod = m.curso.get_numero_modulos(m.modalidad) or 1
        total_cuotas = _semanas_recaudacion_matricula(m)
        cuotas_objetivo = _cuotas_objetivo_recaudacion(m, total_cuotas)

        calendario = _calendario_alertas_pago(m)

        # Primer módulo no cubierto por los pagos posteriores a la reserva.
        pagado_cuotas = _pago_aplicable_a_cuotas(m)
        tolerancia = Decimal('0.01')  # cubre diferencias de redondeo

        primer_modulo_pendiente = None
        requerido = Decimal('0.00')
        for k, cuota in enumerate(cuotas_objetivo, start=1):
            requerido += cuota
            if pagado_cuotas + tolerancia < requerido:
                primer_modulo_pendiente = k
                break

        if primer_modulo_pendiente is None:
            continue  # está al día

        numero_modulo = primer_modulo_pendiente

        fecha_venc, hito = calendario.get(
            numero_modulo, (m.jornada.fecha_inicio, 'modulo')
        )

        if hoy < fecha_venc:
            continue  # su próximo pago todavía no vence: no molestar

        if (m.pk, numero_modulo) in revisadas_hoy:
            continue  # ya revisada hoy para ese módulo

        cuota_modulo = cuotas_objetivo[numero_modulo - 1]
        requerido_anterior = sum(
            cuotas_objetivo[:numero_modulo - 1], Decimal('0.00')
        )
        # Lo aportado que corresponde a este módulo.
        pagado_mod = min(
            max(pagado_cuotas - requerido_anterior, Decimal('0.00')),
            cuota_modulo,
        )

        # Días de seguimiento desde que venció ESE hito de pago.
        dias_atraso = max((hoy - fecha_venc).days, 0)

        # Monto que corresponde reclamar en este hito:
        #  - pago único / saldo restante → todo el saldo pendiente
        #  - módulo → lo que falta de ese módulo
        if hito in ('pago_unico', 'saldo_restante'):
            monto_hito = saldo_total
        else:
            monto_hito = max(cuota_modulo - pagado_mod, Decimal('0.00'))

        if hito == 'pago_unico':
            hito_label = 'Pago único'
        elif hito == 'saldo_restante':
            hito_label = 'Saldo restante'
        else:
            hito_label = f'Módulo {numero_modulo}'

        es_ciclo_corto = bool(getattr(m.curso, 'es_ciclo_corto', False))
        if es_ciclo_corto:
            if m.curso.usa_pago_unico_recaudacion(m.modalidad):
                ciclo_corto_label = 'Ciclo corto · Un solo pago'
            else:
                ciclo_corto_label = (
                    f'Ciclo corto · {total_cuotas} pagos por módulo'
                )
        else:
            ciclo_corto_label = ''

        celular = (m.estudiante.celular or '').strip()
        # Limpieza básica del celular para WhatsApp (solo dígitos, agregamos 593 si parece local)
        digitos = ''.join(c for c in celular if c.isdigit())
        if digitos.startswith('0') and len(digitos) == 10:
            celular_wa = '593' + digitos[1:]
        elif digitos.startswith('593'):
            celular_wa = digitos
        else:
            celular_wa = digitos  # asumimos que ya viene en formato internacional

        sede_label = ''
        if m.jornada:
            sede_label = m.jornada.sede_nombre or m.jornada.ciudad or ''
        if not sede_label:
            sede_label = 'Online' if m.modalidad == 'online' else '—'

        pagos_recuperacion = sorted(
            (
                abono for abono in m.abonos.all()
                if _abono_corresponde_a_recuperacion(abono)
            ),
            key=lambda abono: (abono.fecha, abono.pk),
            reverse=True,
        )
        modulos_recuperacion = sorted({
            abono.numero_modulo
            for abono in pagos_recuperacion
            if abono.numero_modulo
        })
        recuperacion_modulos_label = (
            ', '.join(f'Módulo {numero}' for numero in modulos_recuperacion)
            if modulos_recuperacion else 'Sin módulo indicado'
        )
        ultimo_pago_recuperacion = (
            pagos_recuperacion[0] if pagos_recuperacion else None
        )

        alertas.append({
            'matricula': m,
            'estudiante': m.estudiante,
            'curso': m.curso,
            'jornada': m.jornada,
            'sede_label': sede_label,
            'modalidad': m.modalidad,
            'modalidad_label': m.get_modalidad_display(),
            'fecha_inicio_jornada': m.jornada.fecha_inicio,
            'fecha_vencimiento': fecha_venc,
            'hito': hito,
            'hito_label': hito_label,
            'dias_atraso': dias_atraso,
            'numero_modulo': numero_modulo,
            'total_modulos': n_mod,
            'tipo_matricula_label': m.get_tipo_matricula_display(),
            'pagado_m1': pagado_mod,
            'valor_m1': cuota_modulo,
            'saldo_m1': monto_hito,
            'saldo_total': saldo_total,
            'es_ciclo_corto': es_ciclo_corto,
            'ciclo_corto_label': ciclo_corto_label,
            'pagos_cada_dos_semanas': m.curso.pagos_cada_dos_semanas,
            'celular': celular,
            'celular_wa': celular_wa,
            'recuperacion_pagada': bool(pagos_recuperacion),
            'recuperacion_modulos_label': recuperacion_modulos_label,
            'recuperacion_fecha': (
                ultimo_pago_recuperacion.fecha
                if ultimo_pago_recuperacion else None
            ),
            'recuperacion_recibo': (
                ultimo_pago_recuperacion.numero_recibo
                if ultimo_pago_recuperacion else ''
            ),
        })

    # Ordenar de mayor atraso a menor
    alertas.sort(key=lambda x: (-x['dias_atraso'], x['matricula'].pk))
    return alertas


@matricula_requerida
@require_POST
def alerta_marcar_revisada(request, matricula_pk):
    """
    Marca como "revisada hoy" la alerta de pago pendiente del módulo 1
    para la matrícula indicada. Esto la oculta del dashboard hasta mañana.
    """
    from .models import AlertaPagoRevisada

    matricula = get_object_or_404(Matricula, pk=matricula_pk)
    numero_modulo = int(request.POST.get('numero_modulo', '1') or 1)
    notas = (request.POST.get('notas', '') or '').strip()

    AlertaPagoRevisada.objects.update_or_create(
        matricula=matricula,
        numero_modulo=numero_modulo,
        fecha=date.today(),
        defaults={
            'revisada_por': request.user,
            'notas': notas,
        },
    )
    messages.success(
        request,
        f'Alerta de {matricula.estudiante} marcada como revisada por hoy. '
        'Si mañana sigue pendiente, volverá a aparecer.'
    )

    redirect_to = request.POST.get('next') or 'academia:bienvenida'
    return redirect(redirect_to)
