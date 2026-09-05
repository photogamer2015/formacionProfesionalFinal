from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re
from types import SimpleNamespace
from unittest.mock import patch

from django.conf import settings
from django.contrib.auth.hashers import check_password
from django.contrib.auth.models import Group, User
from django.test import RequestFactory, SimpleTestCase, TestCase
from django.urls import reverse
from django.utils import timezone

from .authentication import (
    LOGIN_CAPTCHA_ANSWER_SESSION_KEY, LOGIN_MFA_CODE_HASH_SESSION_KEY,
    LOGIN_MFA_EMAIL_SESSION_KEY, LOGIN_MFA_EMAIL_TO_SAVE_SESSION_KEY,
    LOGIN_MFA_USER_ID_SESSION_KEY,
)
from .forms import (
    AbonoForm, AdicionalSupletorioRapidoForm, CursoForm, EstudianteForm,
    MatriculaForm, es_cedula_ruc_ecuador_valido, es_ruc_ecuador,
)
from .models import (
    Abono, ActividadUsuario, Adicional, AdicionalArchivado, AmistadUsuario,
    CierreCurso, Comprobante, Curso, CuotaManualRecaudacion, Estudiante,
    EstudianteArchivado, JornadaCurso, Matricula, MatriculaArchivada,
    MeGustaPerfil, PerfilUsuario, PersonaExterna, Recordatorio,
    RecuperacionPendiente, Sede,
)
from .permisos import puede_gestionar_jornadas, puede_ver_jornadas
from .views import _registrar_pago_inicial
from .views_pagos import (
    _calcular_alertas_pago, _calendario_vencimientos,
    _construir_matriz_pagos, _hojas_recaudacion_data,
    _plan_recaudacion_matricula,
)


class GlobalTablePaginationTests(SimpleTestCase):
    def _read_project_file(self, relative_path):
        return (Path(settings.BASE_DIR) / relative_path).read_text(
            encoding='utf-8'
        )

    def test_base_carga_paginacion_global_con_version_actual(self):
        base = self._read_project_file('templates/base.html')

        self.assertIn("responsive.css' %}?v=20260822-3", base)
        self.assertIn("responsive.js' %}?v=20260822-3", base)

    def test_paginacion_global_usa_diez_registros_y_controles_accesibles(self):
        script = self._read_project_file('static/responsive.js')

        self.assertIn('var TABLE_PAGE_SIZE = 10;', script)
        self.assertIn('Página <strong data-current-page>', script)
        self.assertIn('Ir a la página anterior', script)
        self.assertIn('Ir a la página siguiente', script)
        self.assertIn('if (!table.tHead || !table.tBodies.length) return;', script)

    def test_estilos_incluyen_movil_modo_oscuro_e_impresion_completa(self):
        styles = self._read_project_file('static/responsive.css')

        self.assertIn('.fp-table-pagination__controls', styles)
        self.assertIn('body.dark-mode .fp-table-pagination', styles)
        self.assertIn('@media screen and (max-width: 760px)', styles)
        self.assertIn('display: table-row !important;', styles)
        self.assertIn(
            '.fp-table-pagination,\n    .fp-table-navigator {\n        display: none !important;',
            styles,
        )

    def test_tablas_anchas_tienen_navegacion_horizontal_global(self):
        template = self._read_project_file('templates/matricula/lista.html')
        script = self._read_project_file('static/responsive.js')
        styles = self._read_project_file('static/responsive.css')

        self.assertIn('data-table-navigation="true"', template)
        self.assertIn('class="matricula-col-student"', template)
        self.assertIn('class="matricula-phone-link"', template)
        self.assertIn('C.I. {{ m.estudiante.cedula }} · {{ m.curso.nombre }}', template)
        self.assertIn('Ir al final de la tabla', script)
        self.assertIn('Ir al inicio de la tabla', script)
        self.assertIn('if (table.dataset.tableNavigation === "off") return;', script)
        self.assertIn('state.navigator.hidden = !overflowing;', script)
        self.assertIn('left: destination,', script)
        self.assertIn('shell.scrollLeft = Number(topScroll.value);', script)
        self.assertIn('topScroll.value = String(Math.round(shell.scrollLeft));', script)
        self.assertIn('shell.classList.toggle("is-scrolled-x"', script)
        self.assertNotIn('target.scrollIntoView({', script)
        self.assertIn('.fp-table-navigator__scroll', styles)
        self.assertIn('.fp-table-navigator__jump', styles)
        self.assertIn('.fp-table-navigator[hidden]', styles)

    def test_tablas_conservan_identidad_y_datos_de_contexto_al_desplazarse(self):
        template = self._read_project_file('templates/matricula/lista.html')
        script = self._read_project_file('static/responsive.js')
        styles = self._read_project_file('static/responsive.css')

        self.assertIn('data-table-context="custom"', template)
        self.assertIn('function prepareTableIdentity(table)', script)
        self.assertIn('function identityColumnIndex(headers)', script)
        self.assertIn('C.I./RUC', script)
        self.assertIn('prepareTableIdentity(changedTable);', script)
        self.assertIn('.fp-table-identity-column', styles)
        self.assertIn(
            '.responsive-table-shell.is-scrolled-x .fp-table-identity-context',
            styles,
        )
        self.assertIn('position: static !important;', styles)

    def test_alertas_del_panel_paginan_todos_los_casos_y_conservan_filtros(self):
        template = self._read_project_file('templates/bienvenida.html')

        self.assertIn('const GM_PAGE_SIZE = 10;', template)
        self.assertIn('aria-label="Paginación de estudiantes con saldo pendiente"', template)
        self.assertIn('data-gm-current-page', template)
        self.assertIn('data-gm-total-pages', template)
        self.assertIn("card.dataset.gmFilterMatch = matches ? 'true' : 'false';", template)
        self.assertIn("const matchingCards = cards.filter", template)
        self.assertNotIn('<details class="gm-vermas">', template)

    def test_paginacion_de_alertas_es_responsive_accesible_e_imprimible(self):
        template = self._read_project_file('templates/bienvenida.html')

        self.assertIn('aria-controls="gm-case-list"', template)
        self.assertIn('aria-live="polite"', template)
        self.assertIn('.gm-card[hidden] { display: none; }', template)
        self.assertIn('.gm-card[hidden] { display: grid !important; }', template)
        self.assertIn("window.matchMedia('(prefers-reduced-motion: reduce)')", template)

    def test_cursos_tienen_buscador_normalizado_por_nombre(self):
        template = self._read_project_file('templates/cursos/lista.html')

        self.assertIn('data-course-search', template)
        self.assertIn('data-course-card', template)
        self.assertIn('data-course-name="{{ curso.nombre }}"', template)
        self.assertIn(".normalize('NFD')", template)
        self.assertIn("replace(/[\\u0300-\\u036f]/g, '')", template)
        self.assertIn('No hay cursos que coincidan con la búsqueda.', template)

    def test_centro_ayuda_actualizado_sin_video(self):
        template = self._read_project_file('templates/ayuda.html')

        self.assertNotIn('<video', template)
        self.assertNotIn('El Plan Maestro.mp4', template)
        self.assertNotIn('Video Tutorial', template)
        self.assertIn('Última revisión: agosto 2026.', template)
        self.assertIn('📚 Cursos y Categorías', template)
        self.assertIn('Filtro separado por módulo', template)
        self.assertIn('mayúsculas, minúsculas, tildes', template)

    def test_historial_meses_tienen_flecha_desplegable_funcional(self):
        template = self._read_project_file('templates/historial/lista.html')

        self.assertIn('class="hist-month"', template)
        self.assertIn('class="hist-month-summary"', template)
        self.assertIn('class="hist-month-arrow"', template)
        self.assertIn('.hist-month summary::-webkit-details-marker', template)
        self.assertIn('.hist-month[open] .hist-month-arrow', template)
        self.assertIn("summary.setAttribute('aria-expanded'", template)


class ActividadUsuarioTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username='admin_actividad',
            password='clave12345',
            first_name='Ana',
            last_name='Administradora',
        )
        self.asesor = User.objects.create_user(
            username='asesor_actividad',
            password='clave12345',
            first_name='Carlos',
            last_name='Asesor',
        )

    def test_middleware_registra_navegacion_autenticada_con_hora(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse('academia:bienvenida'))

        self.assertEqual(response.status_code, 200)
        actividad = ActividadUsuario.objects.get(usuario=self.asesor)
        self.assertEqual(actividad.usuario_nombre, 'Carlos Asesor')
        self.assertEqual(actividad.accion, 'Consultó el panel de inicio')
        self.assertEqual(actividad.categoria, 'consulta')
        self.assertEqual(actividad.ruta, reverse('academia:bienvenida'))
        self.assertIsNotNone(actividad.creado)

    def test_middleware_no_registra_keepalive_tecnico(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse('academia:session_keepalive'))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(ActividadUsuario.objects.filter(usuario=self.asesor).exists())


    def test_registro_diario_filtra_usuario_y_muestra_mensaje_vacio(self):
        self.client.force_login(self.admin)
        fecha_sin_actividad = timezone.localdate() + timedelta(days=5)

        response = self.client.get(reverse('academia:actividad_usuarios'), {
            'fecha': fecha_sin_actividad.isoformat(),
            'usuario': self.asesor.pk,
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Aún este usuario no ha hecho nada en esta fecha')
        self.assertContains(response, 'Carlos Asesor')

    def test_registro_diario_limpia_actividad_mayor_a_48_horas(self):
        vieja = ActividadUsuario.objects.create(
            usuario=self.asesor,
            usuario_nombre='Carlos Asesor',
            categoria='consulta',
            accion='Consultó información antigua',
        )
        reciente = ActividadUsuario.objects.create(
            usuario=self.asesor,
            usuario_nombre='Carlos Asesor',
            categoria='consulta',
            accion='Consultó información reciente',
        )
        ActividadUsuario.objects.filter(pk=vieja.pk).update(
            creado=timezone.now() - timedelta(hours=49),
        )
        ActividadUsuario.objects.filter(pk=reciente.pk).update(
            creado=timezone.now() - timedelta(hours=47),
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse('academia:actividad_usuarios'))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(ActividadUsuario.objects.filter(pk=vieja.pk).exists())
        self.assertTrue(ActividadUsuario.objects.filter(pk=reciente.pk).exists())
        self.assertEqual(response.context['actividad_limpieza_eliminados'], 1)
        self.assertContains(response, 'Se borró automáticamente 1 registro vencido.')
        self.assertContains(response, 'a partir de las 48 horas')

    def test_registro_diario_es_exclusivo_de_administradores(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse('academia:actividad_usuarios'))

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('academia:bienvenida'))

    def test_pdf_de_actividad_respeta_usuario_y_fecha(self):
        ActividadUsuario.objects.create(
            usuario=self.asesor,
            usuario_nombre='Carlos Asesor',
            categoria='pago',
            accion='Registró un pago de estudiante',
            detalle='Matrícula: 49',
            ruta='/matricula/49/abonos/crear/',
            metodo_http='POST',
            estado_http=302,
        )
        self.client.force_login(self.admin)

        response = self.client.get(reverse('academia:actividad_usuarios_pdf'), {
            'fecha': timezone.localdate().isoformat(),
            'usuario': self.asesor.pk,
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))
        self.assertIn('registro-actividad-', response['Content-Disposition'])


class InicioRankingVendedorasTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username='admin_ranking_inicio',
            password='clave12345',
        )
        self.asesor = User.objects.create_user(
            username='asesor_ranking_inicio',
            password='clave12345',
        )
        grupo = Group.objects.create(name='Asesores')
        self.asesor.groups.add(grupo)
        self.curso = Curso.objects.create(
            nombre='Curso para ranking por rol',
            valor_presencial=Decimal('100.00'),
        )
        self._crear_comprobante(
            self.asesor, date(2026, 8, 10), Decimal('80.00'), Decimal('20.00'),
        )
        self._crear_comprobante(
            self.admin, date(2026, 8, 30), Decimal('50.00'), Decimal('0.00'),
        )

    def _crear_comprobante(self, vendedora, fecha, pago, diferencia):
        return Comprobante.objects.create(
            curso=self.curso,
            modalidad='presencial',
            fecha_inscripcion=fecha,
            jornada='Sábados',
            inicio_curso=date(2026, 9, 5),
            nombre_persona=f'Estudiante {vendedora.username}',
            celular='0999999999',
            pago_abono=pago,
            diferencia=diferencia,
            vendedora=vendedora,
            fact_nombres=f'Factura {vendedora.username}',
            fact_cedula='1200000000',
            fact_correo=f'{vendedora.username}@example.com',
        )

    def test_admin_ve_ranking_de_vendedoras_al_inicio(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse('academia:bienvenida'))
        contenido = response.content.decode()

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Ranking de Vendedoras')
        self.assertContains(response, reverse('academia:comprobante_totales'))
        self.assertContains(response, 'card-ranking')
        self.assertLess(
            contenido.index('card-title">Ranking de Vendedoras'),
            contenido.index('card-title">Ayuda'),
        )

    def test_asesor_ve_ranking_de_vendedoras_en_inicio_sin_montos(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse('academia:bienvenida'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Ranking de Vendedoras (Asesora)')
        self.assertContains(response, reverse('academia:comprobante_totales'))
        self.assertContains(response, 'sin valores monetarios')

    def test_asesor_puede_abrir_ranking_y_no_recibe_datos_monetarios(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse('academia:comprobante_totales'))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['mostrar_montos'])
        self.assertEqual(response.context['total_ventas'], 2)
        self.assertIsNone(response.context['total_cobrado'])
        self.assertIsNone(response.context['total_pendiente'])
        self.assertIsNone(response.context['total_general'])
        self.assertNotIn('total_general', response.context['ranking'][0])
        self.assertContains(response, 'Ranking de Vendedoras (Asesora)')
        self.assertContains(response, 'Vista asesora · solo cantidades')
        self.assertContains(response, 'Ver por rango')
        self.assertContains(response, 'data-ranking-date-picker')
        self.assertNotContains(response, 'TOTAL GENERAL')
        self.assertNotContains(response, 'cobrados')
        contenido_visible = re.sub(
            r'<(?:script|style)\b.*?</(?:script|style)>',
            '',
            response.content.decode(),
            flags=re.IGNORECASE | re.DOTALL,
        )
        self.assertNotIn('$', contenido_visible)
        self.assertContains(response, 'Ver ventas', count=2)

    def test_admin_conserva_montos_y_usa_el_mismo_filtro_profesional(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse('academia:comprobante_totales'))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['mostrar_montos'])
        self.assertEqual(response.context['total_general'], Decimal('150.00'))
        self.assertIn('total_general', response.context['ranking'][0])
        self.assertContains(response, 'TOTAL GENERAL')
        self.assertContains(response, '$')
        self.assertContains(response, 'Ver por rango')
        self.assertContains(response, 'data-ranking-date-picker')

    def test_boton_ver_ventas_conserva_el_rango_del_ranking(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse('academia:comprobante_totales'), {
            'desde': '2026-08-01',
            'hasta': '2026-08-31',
        })
        ventas_url = reverse(
            'academia:comprobante_asesor_ventas',
            args=[self.admin.pk],
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            f'{ventas_url}?desde=2026-08-01&amp;hasta=2026-08-31',
            response.content.decode(),
        )

    def test_asesor_ve_ventas_de_una_persona_sin_montos(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse(
            'academia:comprobante_asesor_ventas',
            args=[self.admin.pk],
        ))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['mostrar_montos'])
        self.assertEqual(response.context['total_ventas'], 1)
        self.assertIsNone(response.context['total_cobrado'])
        self.assertIsNone(response.context['total_pendiente'])
        self.assertIsNone(response.context['total_general'])
        self.assertContains(response, 'Ventas de admin_ranking_inicio')
        self.assertContains(response, 'Estudiante admin_ranking_inicio')
        self.assertNotContains(response, 'Estudiante asesor_ranking_inicio')
        self.assertNotContains(response, '<th>Pago</th>')
        self.assertNotContains(response, '<th>Diferencia</th>')
        self.assertNotContains(response, '<th>Total</th>')
        contenido_visible = re.sub(
            r'<(?:script|style)\b.*?</(?:script|style)>',
            '',
            response.content.decode(),
            flags=re.IGNORECASE | re.DOTALL,
        )
        self.assertNotIn('$', contenido_visible)

    def test_admin_ve_ventas_de_una_persona_con_montos(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse(
            'academia:comprobante_asesor_ventas',
            args=[self.asesor.pk],
        ))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['mostrar_montos'])
        self.assertEqual(response.context['total_ventas'], 1)
        self.assertEqual(response.context['total_cobrado'], Decimal('80.00'))
        self.assertEqual(response.context['total_pendiente'], Decimal('20.00'))
        self.assertEqual(response.context['total_general'], Decimal('100.00'))
        self.assertContains(response, 'Estudiante asesor_ranking_inicio')
        self.assertNotContains(response, 'Estudiante admin_ranking_inicio')
        self.assertContains(response, '<th>Pago</th>')
        self.assertContains(response, '<th>Diferencia</th>')
        self.assertContains(response, '<th>Total</th>')

    def test_ventas_de_asesor_filtra_por_rango(self):
        self.client.force_login(self.admin)

        response = self.client.get(reverse(
            'academia:comprobante_asesor_ventas',
            args=[self.asesor.pk],
        ), {
            'desde': '2026-08-15',
            'hasta': '2026-08-31',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_ventas'], 0)
        self.assertContains(response, 'No hay ventas de esta vendedora')

    def test_ranking_normaliza_un_rango_de_fechas_invertido(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse('academia:comprobante_totales'), {
            'desde': '2026-08-31',
            'hasta': '2026-08-01',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['filtros']['desde'], '2026-08-01')
        self.assertEqual(response.context['filtros']['hasta'], '2026-08-31')

    def test_ranking_filtra_las_ventas_por_el_rango_seleccionado(self):
        self.client.force_login(self.asesor)

        response = self.client.get(reverse('academia:comprobante_totales'), {
            'desde': '2026-08-15',
            'hasta': '2026-08-31',
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_ventas'], 1)
        self.assertEqual(len(response.context['ranking']), 1)
        self.assertEqual(response.context['ranking'][0]['nombre'], self.admin.username)


class LoginCaptchaTests(TestCase):
    def test_login_muestra_captcha_matematico(self):
        response = self.client.get(reverse('login'))

        self.assertEqual(response.status_code, 200)
        self.assertRegex(response.context['captcha_question'], r'^\d+ (\+|x) \d+$')
        self.assertIn(LOGIN_CAPTCHA_ANSWER_SESSION_KEY, self.client.session)
        self.assertContains(response, 'Captcha de seguridad')

    def test_login_rechaza_clave_valida_si_captcha_incorrecto(self):
        User.objects.create_user(username='admin_test', password='clave12345')
        self.client.get(reverse('login'))

        response = self.client.post(reverse('login'), {
            'username': 'admin_test',
            'password': 'clave12345',
            'captcha': '999',
        })

        self.assertEqual(response.status_code, 200)
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertContains(response, 'El resultado del captcha no es correcto.')

    @patch('academia.authentication.enviar_codigo_login')
    def test_login_con_clave_valida_envia_codigo_y_no_autentica_aun(self, enviar_codigo):
        user = User.objects.create_user(
            username='admin_test',
            password='clave12345',
            email='admin@example.com',
        )
        self.client.get(reverse('login'))
        captcha_answer = self.client.session[LOGIN_CAPTCHA_ANSWER_SESSION_KEY]

        response = self.client.post(reverse('login'), {
            'username': user.username,
            'password': 'clave12345',
            'captcha': captcha_answer,
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('login_code'))
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertEqual(self.client.session[LOGIN_MFA_USER_ID_SESSION_KEY], str(user.pk))
        self.assertIn(LOGIN_MFA_CODE_HASH_SESSION_KEY, self.client.session)
        self.assertIn(LOGIN_MFA_EMAIL_SESSION_KEY, self.client.session)
        self.assertNotIn(LOGIN_CAPTCHA_ANSWER_SESSION_KEY, self.client.session)
        enviar_codigo.assert_called_once()

        code_response = self.client.get(reverse('login_code'))
        self.assertContains(code_response, '¿Está seguro que desea reenviar el código?')

    @patch('academia.authentication.generar_codigo_verificacion', return_value='123456')
    @patch('academia.authentication.enviar_codigo_login')
    def test_codigo_correcto_completa_inicio_de_sesion(self, enviar_codigo, generar_codigo):
        user = User.objects.create_user(
            username='admin_test',
            password='clave12345',
            email='admin@example.com',
        )
        self.client.get(reverse('login'))
        captcha_answer = self.client.session[LOGIN_CAPTCHA_ANSWER_SESSION_KEY]
        self.client.post(reverse('login'), {
            'username': user.username,
            'password': 'clave12345',
            'captcha': captcha_answer,
        })

        response = self.client.post(reverse('login_code'), {'code': '123456'})

        self.assertEqual(response.status_code, 302)
        self.assertIn('_auth_user_id', self.client.session)
        self.assertNotIn(LOGIN_MFA_USER_ID_SESSION_KEY, self.client.session)
        self.assertNotIn(LOGIN_MFA_CODE_HASH_SESSION_KEY, self.client.session)

    @patch('academia.authentication.generar_codigo_verificacion', return_value='123456')
    @patch('academia.authentication.enviar_codigo_login')
    def test_codigo_incorrecto_no_autentica(self, enviar_codigo, generar_codigo):
        user = User.objects.create_user(
            username='admin_test',
            password='clave12345',
            email='admin@example.com',
        )
        self.client.get(reverse('login'))
        captcha_answer = self.client.session[LOGIN_CAPTCHA_ANSWER_SESSION_KEY]
        self.client.post(reverse('login'), {
            'username': user.username,
            'password': 'clave12345',
            'captcha': captcha_answer,
        })

        response = self.client.post(reverse('login_code'), {'code': '000000'})

        self.assertEqual(response.status_code, 200)
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertContains(response, 'Codigo incorrecto.')

    @patch('academia.authentication.generar_codigo_verificacion', return_value='123456')
    @patch('academia.authentication.enviar_codigo_login')
    def test_usuario_sin_correo_lo_registra_una_vez_y_luego_solo_pide_codigo(
        self,
        enviar_codigo,
        generar_codigo,
    ):
        user = User.objects.create_user(username='admin_test', password='clave12345')
        self.client.get(reverse('login'))
        captcha_answer = self.client.session[LOGIN_CAPTCHA_ANSWER_SESSION_KEY]

        response = self.client.post(reverse('login'), {
            'username': user.username,
            'password': 'clave12345',
            'captcha': captcha_answer,
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('login_email'))
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertEqual(self.client.session[LOGIN_MFA_USER_ID_SESSION_KEY], str(user.pk))
        self.assertNotIn(LOGIN_MFA_CODE_HASH_SESSION_KEY, self.client.session)

        response = self.client.post(reverse('login_email'), {
            'email': 'Nuevo@Example.COM',
            'email_confirm': 'nuevo@example.com',
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('login_code'))
        self.assertEqual(
            self.client.session[LOGIN_MFA_EMAIL_TO_SAVE_SESSION_KEY],
            'nuevo@example.com',
        )
        self.assertIn(LOGIN_MFA_CODE_HASH_SESSION_KEY, self.client.session)
        self.assertNotIn('_auth_user_id', self.client.session)
        self.assertEqual(enviar_codigo.call_args.args[2], 'nuevo@example.com')

        response = self.client.post(reverse('login_code'), {'code': '123456'})

        self.assertEqual(response.status_code, 302)
        self.assertIn('_auth_user_id', self.client.session)
        user.refresh_from_db()
        self.assertEqual(user.email, 'nuevo@example.com')

        self.client.logout()
        self.client.get(reverse('login'))
        captcha_answer = self.client.session[LOGIN_CAPTCHA_ANSWER_SESSION_KEY]

        response = self.client.post(reverse('login'), {
            'username': user.username,
            'password': 'clave12345',
            'captcha': captcha_answer,
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response['Location'], reverse('login_code'))

    @patch('academia.authentication.enviar_codigo_login')
    def test_registro_correo_requiere_confirmacion_igual(self, enviar_codigo):
        user = User.objects.create_user(username='admin_test', password='clave12345')
        self.client.get(reverse('login'))
        captcha_answer = self.client.session[LOGIN_CAPTCHA_ANSWER_SESSION_KEY]
        self.client.post(reverse('login'), {
            'username': user.username,
            'password': 'clave12345',
            'captcha': captcha_answer,
        })

        response = self.client.post(reverse('login_email'), {
            'email': 'uno@example.com',
            'email_confirm': 'dos@example.com',
        })

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Los correos no coinciden.')
        self.assertNotIn(LOGIN_MFA_CODE_HASH_SESSION_KEY, self.client.session)
        enviar_codigo.assert_not_called()

    @patch('academia.authentication.generar_codigo_verificacion', side_effect=['111111', '222222'])
    @patch('academia.authentication.enviar_codigo_login')
    def test_reenviar_codigo_invalida_el_anterior_y_permite_el_ultimo(
        self,
        enviar_codigo,
        generar_codigo,
    ):
        user = User.objects.create_user(
            username='admin_test',
            password='clave12345',
            email='admin@example.com',
        )
        self.client.get(reverse('login'))
        captcha_answer = self.client.session[LOGIN_CAPTCHA_ANSWER_SESSION_KEY]
        self.client.post(reverse('login'), {
            'username': user.username,
            'password': 'clave12345',
            'captcha': captcha_answer,
        })
        self.assertTrue(
            check_password('111111', self.client.session[LOGIN_MFA_CODE_HASH_SESSION_KEY])
        )

        response = self.client.post(reverse('login_code'), {'action': 'resend'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Te enviamos un nuevo codigo.')
        self.assertEqual(enviar_codigo.call_count, 2)
        self.assertTrue(
            check_password('222222', self.client.session[LOGIN_MFA_CODE_HASH_SESSION_KEY])
        )

        response = self.client.post(reverse('login_code'), {'code': '111111'})

        self.assertEqual(response.status_code, 200)
        self.assertNotIn('_auth_user_id', self.client.session)

        response = self.client.post(reverse('login_code'), {'code': '222222'})

        self.assertEqual(response.status_code, 302)
        self.assertIn('_auth_user_id', self.client.session)


class PerfilUsuarioTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_superuser(
            username='yandri',
            password='clave12345',
            email='yandri@example.com',
            first_name='Yandri',
        )
        self.client.force_login(self.user)
        self.url = reverse(
            'academia:comprobante_asesor_detalle',
            args=[self.user.pk],
        )

    def test_perfil_muestra_selector_y_avatar_predeterminado(self):
        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Elige tu avatar')
        self.assertContains(response, 'avatars/corona.svg')
        self.assertContains(response, 'Princesa Peach')
        self.assertContains(response, 'Mujer pelirroja clara')
        self.assertContains(response, 'avatars/pelirroja_clara.svg')
        self.assertContains(response, 'Mujer morena pelirroja')
        self.assertContains(response, 'avatars/pelirroja_morena.svg')
        self.assertContains(response, 'Cabello rojo con corona')
        self.assertContains(response, 'avatars/pelirroja_corona.svg')
        self.assertContains(response, 'Morena con corona')
        self.assertContains(response, 'avatars/morena_corona.svg')
        self.assertContains(response, 'Gorra de Luigi')
        self.assertContains(response, 'avatars/gorra_luigi.svg')
        self.assertContains(response, 'Caballo')
        self.assertContains(response, 'avatars/caballo.svg')
        self.assertContains(response, 'Dinosaurio')
        self.assertContains(response, 'avatars/dinosaurio.svg')
        self.assertContains(response, 'Tigre')
        self.assertContains(response, 'avatars/tigre.svg')
        self.assertContains(response, 'Perro')
        self.assertContains(response, 'avatars/perro.svg')
        self.assertContains(response, 'Gato')
        self.assertContains(response, 'avatars/gato.svg')
        self.assertContains(response, 'Logo Formación Profesional')
        self.assertContains(response, 'Elige tu portada')
        self.assertContains(response, 'portadas/mariposas.jpg')
        self.assertContains(response, 'portadas/castillo.jpg')
        self.assertContains(response, 'portadas/paisaje.jpg')
        self.assertContains(response, 'Mundo mágico en la nieve')
        self.assertContains(response, 'portadas/mundo_magico_nieve.jpg')
        self.assertContains(response, 'Mundo champiñón')
        self.assertContains(response, 'portadas/mundo_champinon.jpg')

    def test_usuario_puede_guardar_avatar_y_el_header_lo_refleja(self):
        response = self.client.post(
            self.url,
            {'avatar': 'leon'},
            follow=True,
        )

        self.assertRedirects(response, self.url)
        self.assertEqual(
            PerfilUsuario.objects.get(user=self.user).avatar,
            'leon',
        )
        self.assertContains(response, 'avatars/leon.svg')
        self.assertContains(
            response,
            'Tu avatar se actualizó correctamente en todo el sistema.',
        )

    def test_avatar_invalido_no_se_guarda(self):
        response = self.client.post(self.url, {'avatar': 'desconocido'})

        self.assertRedirects(response, self.url)
        self.assertFalse(PerfilUsuario.objects.filter(user=self.user).exists())

    def test_usuario_puede_elegir_logo_como_avatar(self):
        response = self.client.post(
            self.url,
            {'avatar': 'logo_formacion'},
            follow=True,
        )

        self.assertRedirects(response, self.url)
        self.assertEqual(
            PerfilUsuario.objects.get(user=self.user).avatar,
            'logo_formacion',
        )
        self.assertContains(response, 'Logo Formación Profesional')

    def test_usuario_puede_guardar_portada_sin_perder_avatar(self):
        PerfilUsuario.objects.create(user=self.user, avatar='leon')

        response = self.client.post(
            self.url,
            {'preferencia': 'portada', 'portada': 'castillo'},
            follow=True,
        )

        self.assertRedirects(response, self.url)
        perfil = PerfilUsuario.objects.get(user=self.user)
        self.assertEqual(perfil.portada, 'castillo')
        self.assertEqual(perfil.avatar, 'leon')
        self.assertEqual(response.context['portada_seleccionada'], 'castillo')
        self.assertContains(response, 'Tu portada se actualizó correctamente.')
        self.assertContains(response, 'profile-cover-custom')

    def test_portada_invalida_no_se_guarda(self):
        response = self.client.post(
            self.url,
            {'preferencia': 'portada', 'portada': 'desconocida'},
        )

        self.assertRedirects(response, self.url)
        self.assertFalse(PerfilUsuario.objects.filter(user=self.user).exists())

    def test_usuario_puede_guardar_su_mural_sin_perder_avatar_ni_portada(self):
        PerfilUsuario.objects.create(
            user=self.user,
            avatar='leon',
            portada='castillo',
        )

        response = self.client.post(
            self.url,
            {
                'preferencia': 'mural',
                'descripcion_personal': 'Me define la creatividad y ayudar al equipo.',
                'musica': ['latina', 'rock', 'reggaeton'],
                'hobbies': ['lectura', 'viajes'],
                'peliculas': ['comedia', 'ciencia_ficcion'],
                'intereses': ['tecnologia', 'educacion'],
            },
            follow=True,
        )

        self.assertRedirects(response, self.url)
        perfil = PerfilUsuario.objects.get(user=self.user)
        self.assertEqual(perfil.avatar, 'leon')
        self.assertEqual(perfil.portada, 'castillo')
        self.assertEqual(
            perfil.descripcion_personal,
            'Me define la creatividad y ayudar al equipo.',
        )
        self.assertEqual(perfil.musica_favorita, ['latina', 'rock', 'reggaeton'])
        self.assertEqual(perfil.hobbies_favoritos, ['lectura', 'viajes'])
        self.assertEqual(
            perfil.peliculas_favoritas,
            ['comedia', 'ciencia_ficcion'],
        )
        self.assertEqual(
            perfil.intereses_personales,
            ['tecnologia', 'educacion'],
        )
        self.assertContains(
            response,
            'Tu mural de Formación EC se actualizó correctamente.',
        )
        self.assertContains(response, 'Reggaetón')
        self.assertContains(response, 'Editar mural')
        self.assertNotContains(response, 'id="profile-mural-form"')

        response_edicion = self.client.get(self.url, {'editar_mural': '1'})
        self.assertContains(response_edicion, 'id="profile-mural-form"')
        self.assertContains(response_edicion, 'value="reggaeton" checked')

    def test_mural_rechaza_opciones_manipuladas(self):
        response = self.client.post(
            self.url,
            {
                'preferencia': 'mural',
                'descripcion_personal': 'Presentación válida',
                'musica': ['opcion_inexistente'],
            },
            follow=True,
        )

        self.assertRedirects(response, self.url)
        self.assertFalse(PerfilUsuario.objects.filter(user=self.user).exists())
        self.assertContains(
            response,
            'Selecciona hasta 6 opciones válidas en cada categoría.',
        )

    def test_mural_se_muestra_al_consultar_otro_perfil(self):
        otro_usuario = User.objects.create_user(
            username='perfil_publico',
            password='clave12345',
            first_name='Ana',
        )
        PerfilUsuario.objects.create(
            user=otro_usuario,
            descripcion_personal='Soy curiosa, organizada y disfruto aprender.',
            musica_favorita=['salsa'],
            hobbies_favoritos=['fotografia'],
            peliculas_favoritas=['documentales'],
            intereses_personales=['educacion'],
        )

        response = self.client.get(reverse(
            'academia:comprobante_asesor_detalle',
            args=[otro_usuario.pk],
        ))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Soy curiosa, organizada y disfruto aprender.')
        self.assertContains(response, 'Salsa')
        self.assertContains(response, 'Fotografía')
        self.assertContains(response, 'Documentales')
        self.assertContains(response, 'Educación')
        self.assertNotContains(response, 'id="profile-mural-form"')

    def test_resumen_muestra_conteos_no_montos(self):
        curso = Curso.objects.create(
            nombre='Curso para indicadores del perfil',
            valor_presencial=Decimal('100.00'),
        )

        otro_usuario = User.objects.create_user(
            username='otro_asesor',
            password='clave12345',
            first_name='Otro',
        )

        def crear_matricula(cedula, pagado, estado='activa', vendedora=None):
            vendedora = vendedora or self.user
            estudiante = Estudiante.objects.create(
                cedula=cedula,
                nombres=f'Estudiante {cedula}',
            )
            return Matricula.objects.create(
                estudiante=estudiante,
                curso=curso,
                modalidad='presencial',
                estado=estado,
                tipo_matricula='reserva_abono',
                fecha_matricula=date(2026, 8, 1),
                valor_curso=Decimal('100.00'),
                valor_pagado=pagado,
                registrado_por=vendedora,
                vendedora=vendedora,
            )

        pendiente = crear_matricula('0900000001', Decimal('10.00'))
        crear_matricula('0900000002', Decimal('10.00'), 'retiro_voluntario')
        crear_matricula('0900000003', Decimal('100.00'))
        for numero_modulo in (1, 2):
            RecuperacionPendiente.objects.create(
                matricula=pendiente,
                numero_modulo=numero_modulo,
                fecha_marcada=date(2026, 8, 1),
                saldo_pendiente_al_marcar=Decimal('90.00'),
            )
        pendiente_otro_usuario = crear_matricula(
            '0900000004',
            Decimal('20.00'),
            vendedora=otro_usuario,
        )
        crear_matricula(
            '0900000005',
            Decimal('20.00'),
            estado='retiro_voluntario',
            vendedora=otro_usuario,
        )
        RecuperacionPendiente.objects.create(
            matricula=pendiente_otro_usuario,
            numero_modulo=1,
            fecha_marcada=date(2026, 8, 1),
            saldo_pendiente_al_marcar=Decimal('80.00'),
        )

        response = self.client.get(self.url)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_ventas'], 3)
        self.assertEqual(response.context['total_activas'], 2)
        self.assertEqual(response.context['total_saldos_pendientes'], 2)
        self.assertEqual(response.context['total_retiros'], 2)
        self.assertEqual(response.context['total_recuperaciones'], 3)
        self.assertContains(response, 'Recuperación')
        self.assertContains(response, 'Estudiantes por recaudar')
        self.assertNotContains(response, 'Total cobrado')
        self.assertContains(response, 'data-summary-target="summary-ventas"')
        self.assertContains(response, 'data-summary-target="summary-recuperaciones"')
        self.assertContains(response, 'data-summary-target="summary-pendientes"')
        self.assertContains(response, 'data-summary-target="summary-retiros"')
        self.assertContains(response, 'id="profile-summary-dialog"')
        self.assertContains(response, 'Detalle general de las clases marcadas para recuperación.')
        self.assertContains(response, 'Saldo pendiente: $90,00')
        self.assertContains(response, 'Saldo pendiente: $80,00')
        self.assertContains(response, 'Retiro voluntario')


class PerfilSocialTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_user(
            username='carlos_social',
            password='clave12345',
            first_name='Carlos',
            last_name='Mora',
        )
        self.otro = User.objects.create_user(
            username='ana_social',
            password='clave12345',
            email='ana.social@example.com',
            first_name='Ana',
            last_name='Paredes',
        )
        self.tercero = User.objects.create_superuser(
            username='super_social',
            password='clave12345',
            first_name='Lucía',
            last_name='Admin',
        )
        PerfilUsuario.objects.create(user=self.usuario, avatar='leon')
        PerfilUsuario.objects.create(user=self.otro, avatar='latina')
        self.client.force_login(self.usuario)

    def perfil_url(self, user):
        return reverse(
            'academia:comprobante_asesor_detalle',
            args=[user.pk],
        )

    def test_busqueda_ignora_mayusculas_y_no_filtra_por_rol(self):
        response = self.client.get(reverse('academia:buscar_amigos'), {
            'q': 'aNa',
        })

        self.assertEqual(response.status_code, 200)
        resultados = response.json()['resultados']
        self.assertEqual([item['id'] for item in resultados], [self.otro.pk])
        self.assertEqual(resultados[0]['estado'], 'sin_relacion')

        response_todos = self.client.get(reverse('academia:buscar_amigos'))
        ids = {item['id'] for item in response_todos.json()['resultados']}
        self.assertIn(self.otro.pk, ids)
        self.assertIn(self.tercero.pk, ids)

    def test_solicitud_aparece_en_campana_y_destinatario_puede_aceptar(self):
        response = self.client.post(reverse(
            'academia:amistad_solicitar',
            args=[self.otro.pk],
        ))

        self.assertRedirects(response, self.perfil_url(self.otro))
        relacion = AmistadUsuario.objects.get()
        self.assertEqual(relacion.estado, 'pendiente')
        self.assertEqual(relacion.solicitada_por, self.usuario)
        self.assertLess(relacion.usuario_a_id, relacion.usuario_b_id)

        self.client.force_login(self.otro)
        response_notificacion = self.client.get(self.perfil_url(self.otro))
        self.assertEqual(
            response_notificacion.context['solicitudes_amistad_pendientes_n'],
            1,
        )
        self.assertEqual(
            response_notificacion.context['notificaciones_no_leidas_n'],
            1,
        )
        self.assertContains(response_notificacion, 'Te envió una solicitud de amistad.')
        self.assertContains(response_notificacion, 'value="aceptar"')

        response_aceptar = self.client.post(
            reverse('academia:amistad_accion', args=[relacion.pk]),
            {'accion': 'aceptar'},
        )

        self.assertRedirects(response_aceptar, self.perfil_url(self.usuario))
        relacion.refresh_from_db()
        self.assertEqual(relacion.estado, 'aceptada')

        response_perfil = self.client.get(self.perfil_url(self.otro))
        self.assertEqual(response_perfil.context['total_amigos'], 1)
        self.assertContains(response_perfil, 'Carlos Mora')
        self.assertContains(response_perfil, 'id="friends-dialog"')

    def test_solicitud_puede_rechazarse_y_cancelarse(self):
        self.client.post(reverse(
            'academia:amistad_solicitar',
            args=[self.otro.pk],
        ))
        relacion = AmistadUsuario.objects.get()

        response_cancelar = self.client.post(
            reverse('academia:amistad_accion', args=[relacion.pk]),
            {'accion': 'cancelar'},
        )
        self.assertEqual(response_cancelar.status_code, 302)
        self.assertFalse(AmistadUsuario.objects.exists())

        self.client.post(reverse(
            'academia:amistad_solicitar',
            args=[self.otro.pk],
        ))
        relacion = AmistadUsuario.objects.get()
        self.client.force_login(self.otro)
        response_rechazar = self.client.post(
            reverse('academia:amistad_accion', args=[relacion.pk]),
            {'accion': 'rechazar'},
        )
        self.assertEqual(response_rechazar.status_code, 302)
        self.assertFalse(AmistadUsuario.objects.exists())

    def test_me_gusta_se_puede_activar_y_quitar(self):
        url = reverse('academia:perfil_me_gusta', args=[self.otro.pk])

        response = self.client.post(url)

        self.assertRedirects(response, self.perfil_url(self.otro))
        self.assertTrue(MeGustaPerfil.objects.filter(
            usuario=self.usuario,
            perfil=self.otro,
        ).exists())
        response_perfil = self.client.get(self.perfil_url(self.otro))
        self.assertEqual(response_perfil.context['total_me_gusta'], 1)
        self.assertTrue(response_perfil.context['dio_me_gusta'])
        self.assertContains(response_perfil, 'Te gusta')

        self.client.post(url)
        self.assertFalse(MeGustaPerfil.objects.exists())

    def test_no_permite_amistad_ni_me_gusta_con_el_perfil_propio(self):
        self.client.post(reverse(
            'academia:amistad_solicitar',
            args=[self.usuario.pk],
        ))
        self.client.post(reverse(
            'academia:perfil_me_gusta',
            args=[self.usuario.pk],
        ))

        self.assertFalse(AmistadUsuario.objects.exists())
        self.assertFalse(MeGustaPerfil.objects.exists())

    def test_otro_usuario_ve_perfil_social_pero_no_datos_operativos(self):
        PerfilUsuario.objects.filter(user=self.otro).update(
            descripcion_personal='Me gusta aprender y compartir ideas.',
            musica_favorita=['rock'],
        )

        response = self.client.get(self.perfil_url(self.otro))

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context['puede_ver_actividad_perfil'])
        self.assertContains(response, 'Me gusta aprender y compartir ideas.')
        self.assertContains(response, 'Rock')
        self.assertContains(response, 'Añadir amigo')
        self.assertContains(response, 'Me gusta')
        self.assertNotContains(response, 'Ventas registradas')
        self.assertNotContains(response, 'Estudiantes matriculados')
        self.assertNotContains(response, self.otro.email)

    def test_admin_ve_perfil_ajeno_solo_con_tarjeta_publica(self):
        PerfilUsuario.objects.filter(user=self.otro).update(
            descripcion_personal='Perfil visible para la comunidad.',
            hobbies_favoritos=['viajes'],
        )
        self.client.force_login(self.tercero)

        response = self.client.get(self.perfil_url(self.otro))

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['puede_ver_actividad_perfil'])
        self.assertFalse(response.context['mostrar_detalle_operativo_perfil'])
        self.assertContains(response, 'Perfil visible para la comunidad.')
        self.assertContains(response, 'Viajar')
        self.assertNotContains(response, 'class="profile-tabs"')
        self.assertNotContains(response, 'Ventas registradas')
        self.assertNotContains(response, 'Estudiantes matriculados')
        self.assertNotContains(response, 'Actividad en el sistema')


class RecordatorioAvatarTests(TestCase):
    def setUp(self):
        self.creador = User.objects.create_superuser(
            username='creador_recordatorio',
            password='clave12345',
            first_name='Yandri',
        )
        self.destinatario = User.objects.create_user(
            username='destino_recordatorio',
            password='clave12345',
            first_name='Admin',
        )
        grupo_asesores = Group.objects.create(name='Asesores')
        self.destinatario.groups.add(grupo_asesores)
        PerfilUsuario.objects.create(user=self.creador, avatar='leon')
        PerfilUsuario.objects.create(user=self.destinatario, avatar='latina')

    def test_formulario_incluye_avatar_para_preview_del_destinatario(self):
        self.client.force_login(self.creador)

        response = self.client.get(reverse('academia:recordatorio_crear'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'recipient-profile-preview')
        self.assertContains(response, 'recordatorio-destinatarios-avatar-data')
        self.assertContains(response, 'avatars/latina.svg')

    def test_lista_muestra_avatar_segun_contexto_del_mensaje(self):
        Recordatorio.objects.create(
            titulo='Seguimiento',
            contenido='Revisar pago pendiente.',
            prioridad='media',
            creado_por=self.creador,
            destinatario=self.destinatario,
            fecha=timezone.localdate(),
            fecha_vencimiento=timezone.localdate() + timedelta(days=3),
        )

        self.client.force_login(self.destinatario)
        response = self.client.get(reverse('academia:recordatorio_lista'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'avatars/leon.svg')

        self.client.force_login(self.creador)
        response = self.client.get(reverse('academia:recordatorio_lista'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'avatars/latina.svg')


class SessionKeepaliveTests(TestCase):
    def test_keepalive_requiere_login(self):
        response = self.client.get(reverse('academia:session_keepalive'))
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login/', response['Location'])

    def test_keepalive_refresca_sesion_autenticada(self):
        usuario = User.objects.create_user(username='soporte', password='clave12345')
        self.client.force_login(usuario)

        response = self.client.get(reverse('academia:session_keepalive'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {'ok': True})


class JornadaMatriculasAccessTests(TestCase):
    def setUp(self):
        self.asesor = User.objects.create_user(username='asesor')
        grupo = Group.objects.create(name='Asesores')
        self.asesor.groups.add(grupo)
        self.curso = Curso.objects.create(
            nombre='Curso Jornadas Test',
            ofrece_presencial=True,
            valor_presencial=Decimal('100.00'),
        )
        self.jornada_1 = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='lun_mie_vie',
            fecha_inicio=date(2026, 7, 5),
        )
        self.jornada_2 = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='mar_jue',
            fecha_inicio=date(2026, 7, 6),
        )
        self.estudiante_1 = Estudiante.objects.create(
            cedula='1207342716',
            nombres='Estudiante Jornada Uno',
        )
        self.estudiante_2 = Estudiante.objects.create(
            cedula='1207342717',
            nombres='Estudiante Jornada Dos',
        )
        Matricula.objects.create(
            estudiante=self.estudiante_1,
            curso=self.curso,
            jornada=self.jornada_1,
            modalidad='presencial',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('100.00'),
            tipo_registro='central_ia',
            registrado_por=self.asesor,
        )
        Matricula.objects.create(
            estudiante=self.estudiante_2,
            curso=self.curso,
            jornada=self.jornada_2,
            modalidad='presencial',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('100.00'),
            tipo_registro='central_ia',
            registrado_por=self.asesor,
        )

    def test_asesor_puede_ver_panel_de_jornadas(self):
        self.assertTrue(puede_ver_jornadas(self.asesor))
        self.assertTrue(puede_gestionar_jornadas(self.asesor))

    def test_recuperaciones_es_acceso_principal_y_no_boton_de_pagos(self):
        self.client.force_login(self.asesor)
        recuperaciones_url = reverse('academia:recuperaciones_lista')

        bienvenida = self.client.get(reverse('academia:bienvenida'))
        self.assertEqual(bienvenida.status_code, 200)
        self.assertContains(bienvenida, recuperaciones_url)
        self.assertContains(bienvenida, 'Recuperaciones')

        pagos = self.client.get(reverse('academia:pagos_lista'))
        self.assertEqual(pagos.status_code, 200)
        self.assertNotContains(pagos, recuperaciones_url)

    def test_lista_matriculas_filtra_por_jornada(self):
        self.client.force_login(self.asesor)

        response = self.client.get(
            reverse('academia:matricula_lista', kwargs={'modalidad': 'presencial'}),
            {'jornada': str(self.jornada_1.id)},
        )

        self.assertEqual(response.status_code, 200)
        matriculas = list(response.context['matriculas'])
        self.assertEqual(len(matriculas), 1)
        self.assertEqual(matriculas[0].estudiante, self.estudiante_1)
        self.assertEqual(response.context['jornada_filtrada'], self.jornada_1)

    def test_lista_matriculas_busqueda_ignora_tildes(self):
        self.estudiante_1.nombres = 'Osmár Jornada Uno'
        self.estudiante_1.save(update_fields=['nombres'])
        self.client.force_login(self.asesor)

        response = self.client.get(
            reverse('academia:matricula_lista', kwargs={'modalidad': 'presencial'}),
            {'q': 'Osmar'},
        )

        self.assertEqual(response.status_code, 200)
        matriculas = list(response.context['matriculas'])
        self.assertEqual(len(matriculas), 1)
        self.assertEqual(matriculas[0].estudiante, self.estudiante_1)

    def test_lista_matriculas_filtra_por_estado_de_pago(self):
        pendiente = Matricula.objects.get(estudiante=self.estudiante_1)
        pagada = Matricula.objects.get(estudiante=self.estudiante_2)
        pendiente.valor_pagado = Decimal('30.00')
        pendiente.save(update_fields=['valor_pagado'])
        self.client.force_login(self.asesor)

        response_pendiente = self.client.get(
            reverse('academia:matricula_lista', kwargs={'modalidad': 'presencial'}),
            {'estado_pago': 'pendiente'},
        )
        response_pagado = self.client.get(
            reverse('academia:matricula_lista', kwargs={'modalidad': 'presencial'}),
            {'estado_pago': 'pagado'},
        )

        self.assertEqual(response_pendiente.status_code, 200)
        self.assertEqual(response_pendiente.context['estado_pago_seleccionado'], 'pendiente')
        self.assertEqual(list(response_pendiente.context['matriculas']), [pendiente])
        self.assertContains(response_pendiente, 'Saldo pendiente')

        self.assertEqual(response_pagado.status_code, 200)
        self.assertEqual(response_pagado.context['estado_pago_seleccionado'], 'pagado')
        self.assertEqual(list(response_pagado.context['matriculas']), [pagada])

    def test_apartado_retirados_muestra_solo_retiros_voluntarios(self):
        retirada = Matricula.objects.get(estudiante=self.estudiante_1)
        retirada.estado = 'retiro_voluntario'
        retirada.save(update_fields=['estado'])
        activa = Matricula.objects.get(estudiante=self.estudiante_2)
        self.client.force_login(self.asesor)

        response = self.client.get(
            reverse(
                'academia:matricula_retirados',
                kwargs={'modalidad': 'presencial'},
            )
        )

        self.assertEqual(response.status_code, 200)
        matriculas = list(response.context['matriculas'])
        self.assertTrue(response.context['solo_retirados'])
        self.assertEqual([m.pk for m in matriculas], [retirada.pk])
        self.assertNotIn(activa.pk, [m.pk for m in matriculas])
        self.assertContains(response, 'Estudiantes retirados')
        self.assertContains(response, 'Retiro')

    def test_menu_matricula_incluye_acceso_y_total_de_retirados(self):
        retirada = Matricula.objects.get(estudiante=self.estudiante_1)
        retirada.estado = 'retiro_voluntario'
        retirada.save(update_fields=['estado'])
        self.client.force_login(self.asesor)

        response = self.client.get(
            reverse(
                'academia:matricula_menu',
                kwargs={'modalidad': 'presencial'},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_retirados'], 1)
        self.assertContains(response, 'Estudiantes retirados')
        self.assertContains(
            response,
            reverse(
                'academia:matricula_retirados',
                kwargs={'modalidad': 'todos'},
            ),
        )

    def test_menu_abre_listas_en_pestana_todos(self):
        self.client.force_login(self.asesor)

        response = self.client.get(
            reverse(
                'academia:matricula_menu',
                kwargs={'modalidad': 'presencial'},
            )
        )

        self.assertContains(
            response,
            reverse(
                'academia:matricula_lista',
                kwargs={'modalidad': 'todos'},
            ),
        )
        self.assertContains(
            response,
            reverse(
                'academia:matricula_retirados',
                kwargs={'modalidad': 'todos'},
            ),
        )

    def test_administrador_revierte_retiro_y_restaura_saldo(self):
        administrador = User.objects.create_user(username='administrador')
        grupo_admin, _ = Group.objects.get_or_create(name='Administradores')
        administrador.groups.add(grupo_admin)
        retirada = Matricula.objects.get(estudiante=self.estudiante_1)
        retirada.estado = 'retiro_voluntario'
        retirada.valor_pagado = Decimal('70.00')
        retirada.save(update_fields=['estado', 'valor_pagado'])
        self.assertEqual(retirada.saldo, Decimal('0.00'))
        self.client.force_login(administrador)

        response = self.client.post(
            reverse(
                'academia:matricula_revertir_retiro',
                kwargs={'pk': retirada.pk},
            )
        )

        retirada.refresh_from_db()
        self.assertRedirects(
            response,
            reverse(
                'academia:matricula_retirados',
                kwargs={'modalidad': 'presencial'},
            ),
        )
        self.assertEqual(retirada.estado, 'activa')
        self.assertEqual(retirada.valor_pagado, Decimal('70.00'))
        self.assertEqual(retirada.saldo, Decimal('30.00'))

    def test_asesor_no_puede_revertir_retiro(self):
        retirada = Matricula.objects.get(estudiante=self.estudiante_1)
        retirada.estado = 'retiro_voluntario'
        retirada.valor_pagado = Decimal('70.00')
        retirada.save(update_fields=['estado', 'valor_pagado'])
        self.client.force_login(self.asesor)

        response = self.client.post(
            reverse(
                'academia:matricula_revertir_retiro',
                kwargs={'pk': retirada.pk},
            )
        )

        retirada.refresh_from_db()
        self.assertRedirects(response, reverse('academia:bienvenida'))
        self.assertEqual(retirada.estado, 'retiro_voluntario')
        self.assertEqual(retirada.saldo, Decimal('0.00'))

    def test_boton_revertir_solo_se_muestra_al_administrador(self):
        administrador = User.objects.create_superuser(
            username='admin_reversion',
        )
        retirada = Matricula.objects.get(estudiante=self.estudiante_1)
        retirada.estado = 'retiro_voluntario'
        retirada.save(update_fields=['estado'])
        url = reverse(
            'academia:matricula_retirados',
            kwargs={'modalidad': 'presencial'},
        )

        self.client.force_login(self.asesor)
        response_asesor = self.client.get(url)
        self.assertNotContains(response_asesor, 'REVERTIR')
        edicion_url = reverse(
            'academia:matricula_editar',
            kwargs={'modalidad': 'presencial', 'pk': retirada.pk},
        )
        edicion_asesor = self.client.get(edicion_url)
        self.assertTrue(edicion_asesor.context['mat_form'].fields['estado'].disabled)

        self.client.force_login(administrador)
        response_admin = self.client.get(url)
        self.assertContains(response_admin, 'REVERTIR')
        edicion_admin = self.client.get(edicion_url)
        self.assertFalse(edicion_admin.context['mat_form'].fields['estado'].disabled)

    def test_lista_matriculas_todos_mezcla_presencial_y_online(self):
        sede = Sede.objects.create(nombre='Guayaquil', orden=1)
        self.jornada_1.sede = sede
        self.jornada_1.save(update_fields=['sede', 'ciudad'])
        curso_online = Curso.objects.create(
            nombre='Curso Online Jornadas Test',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('80.00'),
        )
        jornada_online = JornadaCurso.objects.create(
            curso=curso_online,
            modalidad='online',
            descripcion='sabados_intensivos',
            fecha_inicio=date(2026, 7, 7),
        )
        estudiante_online = Estudiante.objects.create(
            cedula='1207342718',
            nombres='Estudiante Online',
        )
        Matricula.objects.create(
            estudiante=estudiante_online,
            curso=curso_online,
            jornada=jornada_online,
            modalidad='online',
            fecha_matricula=date(2026, 7, 7),
            valor_curso=Decimal('80.00'),
            valor_pagado=Decimal('80.00'),
            tipo_registro='central_ia',
            registrado_por=self.asesor,
        )
        self.client.force_login(self.asesor)

        response = self.client.get(
            reverse('academia:matricula_lista', kwargs={'modalidad': 'todos'})
        )

        self.assertEqual(response.status_code, 200)
        modalidades = {m.modalidad for m in response.context['matriculas']}
        self.assertEqual(response.context['modalidad'], 'todos')
        self.assertEqual(response.context['modalidad_registro'], 'presencial')
        self.assertIn('presencial', modalidades)
        self.assertIn('online', modalidades)

    def test_lista_matriculas_todos_filtra_modalidad_y_sede(self):
        sede_guayaquil = Sede.objects.create(nombre='Guayaquil', orden=1)
        sede_quito = Sede.objects.create(nombre='Quito', orden=2)
        self.jornada_1.sede = sede_guayaquil
        self.jornada_1.save(update_fields=['sede', 'ciudad'])
        self.jornada_2.sede = sede_quito
        self.jornada_2.save(update_fields=['sede', 'ciudad'])
        curso_online = Curso.objects.create(
            nombre='Curso Online Filtro',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('80.00'),
        )
        jornada_online = JornadaCurso.objects.create(
            curso=curso_online,
            modalidad='online',
            descripcion='sabados_intensivos',
            fecha_inicio=date(2026, 7, 7),
        )
        estudiante_online = Estudiante.objects.create(
            cedula='1207342718',
            nombres='Estudiante Online Filtro',
        )
        Matricula.objects.create(
            estudiante=estudiante_online,
            curso=curso_online,
            jornada=jornada_online,
            modalidad='online',
            fecha_matricula=date(2026, 7, 7),
            valor_curso=Decimal('80.00'),
            valor_pagado=Decimal('80.00'),
            tipo_registro='central_ia',
            registrado_por=self.asesor,
        )
        self.client.force_login(self.asesor)

        response = self.client.get(
            reverse('academia:matricula_lista', kwargs={'modalidad': 'todos'}),
            {
                'modalidad_filtro': 'presencial',
                'campus': f'sede:{sede_guayaquil.pk}',
            },
        )

        self.assertEqual(response.status_code, 200)
        matriculas = list(response.context['matriculas'])
        self.assertEqual(response.context['modalidad_filtro'], 'presencial')
        self.assertEqual(response.context['campus_seleccionado'], f'sede:{sede_guayaquil.pk}')
        self.assertTrue(matriculas)
        self.assertTrue(all(m.modalidad == 'presencial' for m in matriculas))
        self.assertTrue(all(m.jornada.sede_id == sede_guayaquil.pk for m in matriculas))

    def test_listas_muestran_ultimo_registro_primero_aunque_fecha_sea_anterior(self):
        self.client.force_login(self.asesor)
        vieja = Matricula.objects.get(estudiante=self.estudiante_1)
        nueva = Matricula.objects.get(estudiante=self.estudiante_2)
        ahora = timezone.now()
        Matricula.objects.filter(pk=vieja.pk).update(
            fecha_matricula=date(2026, 7, 6),
            creado=ahora - timezone.timedelta(minutes=10),
        )
        Matricula.objects.filter(pk=nueva.pk).update(
            fecha_matricula=date(2026, 7, 1),
            creado=ahora,
        )

        matricula_response = self.client.get(
            reverse('academia:matricula_lista', kwargs={'modalidad': 'presencial'})
        )
        pagos_response = self.client.get(reverse('academia:pagos_lista'))

        self.assertEqual(matricula_response.status_code, 200)
        self.assertEqual(pagos_response.status_code, 200)
        self.assertEqual(
            list(matricula_response.context['matriculas'])[0].estudiante,
            self.estudiante_2,
        )
        self.assertEqual(pagos_response.context['matriculas'][0].estudiante, self.estudiante_2)

    def test_control_registro_ordena_matriculas_por_fecha_de_inscripcion(self):
        admin = User.objects.create_superuser(username='admin_control')
        reciente = Matricula.objects.get(estudiante=self.estudiante_1)
        antigua = Matricula.objects.get(estudiante=self.estudiante_2)
        estudiante_futura = Estudiante.objects.create(
            cedula='1207342718',
            nombres='Estudiante Fecha Futura',
        )
        futura = Matricula.objects.create(
            estudiante=estudiante_futura,
            curso=self.curso,
            jornada=self.jornada_1,
            modalidad='presencial',
            fecha_matricula=date(2026, 7, 16),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=self.asesor,
        )
        ahora = timezone.now()
        Matricula.objects.filter(pk=reciente.pk).update(
            fecha_matricula=date(2026, 7, 15),
            creado=ahora - timezone.timedelta(minutes=10),
        )
        Matricula.objects.filter(pk=antigua.pk).update(
            fecha_matricula=date(2026, 7, 1),
            creado=ahora,
        )
        Matricula.objects.filter(pk=futura.pk).update(creado=ahora - timezone.timedelta(minutes=20))
        self.client.force_login(admin)

        response = self.client.get(reverse('academia:control_registro'))

        self.assertEqual(response.status_code, 200)
        matriculas = list(response.context['matriculas'])
        self.assertEqual(
            [m.estudiante for m in matriculas[:3]],
            [estudiante_futura, self.estudiante_1, self.estudiante_2],
        )

    def test_control_registro_ordena_pagos_por_fecha_reciente(self):
        admin = User.objects.create_superuser(username='admin_control_orden_pagos')
        matricula_1 = Matricula.objects.get(estudiante=self.estudiante_1)
        matricula_2 = Matricula.objects.get(estudiante=self.estudiante_2)
        pago_15 = Abono.objects.create(
            matricula=matricula_1,
            fecha=date(2026, 7, 15),
            monto=Decimal('15.00'),
            metodo='efectivo',
            registrado_por=self.asesor,
        )
        pago_01 = Abono.objects.create(
            matricula=matricula_2,
            fecha=date(2026, 7, 1),
            monto=Decimal('10.00'),
            metodo='efectivo',
            registrado_por=self.asesor,
        )
        pago_16 = Abono.objects.create(
            matricula=matricula_1,
            fecha=date(2026, 7, 16),
            monto=Decimal('16.00'),
            metodo='transferencia',
            registrado_por=self.asesor,
        )
        Abono.objects.filter(pk=pago_15.pk).update(
            creado=timezone.make_aware(datetime(2026, 7, 15, 12, 0))
        )
        Abono.objects.filter(pk=pago_01.pk).update(
            creado=timezone.make_aware(datetime(2026, 7, 1, 12, 0))
        )
        Abono.objects.filter(pk=pago_16.pk).update(
            creado=timezone.make_aware(datetime(2026, 7, 16, 12, 0))
        )
        self.client.force_login(admin)

        response = self.client.get(reverse('academia:control_registro'), {'tab': 'pagos'})

        self.assertEqual(response.status_code, 200)
        pagos = list(response.context['pagos'])
        self.assertEqual(pagos[:3], [pago_16, pago_15, pago_01])

    def test_control_registro_filtra_estudiante_sin_tildes_ni_mayusculas(self):
        admin = User.objects.create_superuser(username='admin_control_busqueda')
        self.estudiante_1.nombres = 'Osmár Jornada Uno'
        self.estudiante_1.save(update_fields=['nombres'])
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:control_registro'),
            {'q': 'OSMAR jornada'},
        )

        self.assertEqual(response.status_code, 200)
        matriculas = list(response.context['matriculas'])
        self.assertEqual(len(matriculas), 1)
        self.assertEqual(matriculas[0].estudiante, self.estudiante_1)

    def test_control_registro_filtra_por_curso_campus_y_fecha_matricula(self):
        admin = User.objects.create_superuser(username='admin_control_filtros')
        sede_guayaquil = Sede.objects.create(nombre='Guayaquil', orden=1)
        sede_quito = Sede.objects.create(nombre='Quito', orden=2)
        sede_cuenca = Sede.objects.create(nombre='Cuenca', orden=3)
        self.jornada_1.sede = sede_guayaquil
        self.jornada_1.save(update_fields=['sede', 'ciudad'])
        self.jornada_2.sede = sede_quito
        self.jornada_2.save(update_fields=['sede', 'ciudad'])
        otra_curso = Curso.objects.create(
            nombre='Curso Otro Control',
            ofrece_presencial=True,
            valor_presencial=Decimal('100.00'),
        )
        otra_jornada = JornadaCurso.objects.create(
            curso=otra_curso,
            modalidad='presencial',
            descripcion='sabados_intensivos',
            fecha_inicio=date(2026, 7, 9),
            sede=sede_cuenca,
        )
        otro_estudiante = Estudiante.objects.create(
            cedula='1207342718',
            nombres='Estudiante Otro Curso',
        )
        Matricula.objects.create(
            estudiante=otro_estudiante,
            curso=otra_curso,
            jornada=otra_jornada,
            modalidad='presencial',
            fecha_matricula=date(2026, 7, 9),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=self.asesor,
        )
        Matricula.objects.filter(estudiante=self.estudiante_1).update(
            fecha_matricula=date(2026, 7, 5),
        )
        Matricula.objects.filter(estudiante=self.estudiante_2).update(
            fecha_matricula=date(2026, 7, 15),
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:control_registro'),
            {
                'curso': str(self.curso.pk),
                'campus': f'sede:{sede_guayaquil.pk}',
                'fecha_desde': '2026-07-01',
                'fecha_hasta': '2026-07-10',
            },
        )

        self.assertEqual(response.status_code, 200)
        matriculas = list(response.context['matriculas'])
        self.assertEqual(len(matriculas), 1)
        self.assertEqual(matriculas[0].estudiante, self.estudiante_1)
        campus_labels = [opcion['label'] for opcion in response.context['campus_opciones']]
        self.assertIn('Guayaquil', campus_labels)
        self.assertIn('Quito', campus_labels)
        self.assertNotIn('Cuenca', campus_labels)

    def test_control_registro_filtra_pagos_por_fecha_curso_metodo_y_monto(self):
        admin = User.objects.create_superuser(username='admin_control_pagos')
        matricula_1 = Matricula.objects.get(estudiante=self.estudiante_1)
        matricula_2 = Matricula.objects.get(estudiante=self.estudiante_2)
        otro_curso = Curso.objects.create(
            nombre='Curso Pagos Otro Control',
            ofrece_presencial=True,
            valor_presencial=Decimal('100.00'),
        )
        otra_jornada = JornadaCurso.objects.create(
            curso=otro_curso,
            modalidad='presencial',
            descripcion='sabados_intensivos',
            fecha_inicio=date(2026, 7, 9),
        )
        otro_estudiante = Estudiante.objects.create(
            cedula='1207342718',
            nombres='Estudiante Pago Otro',
        )
        otra_matricula = Matricula.objects.create(
            estudiante=otro_estudiante,
            curso=otro_curso,
            jornada=otra_jornada,
            modalidad='presencial',
            fecha_matricula=date(2026, 7, 9),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=self.asesor,
        )
        pago_objetivo = Abono.objects.create(
            matricula=matricula_1,
            fecha=date(2026, 7, 9),
            monto=Decimal('25.00'),
            metodo='transferencia',
            registrado_por=self.asesor,
        )
        pago_otro_monto = Abono.objects.create(
            matricula=matricula_2,
            fecha=date(2026, 7, 9),
            monto=Decimal('10.00'),
            metodo='transferencia',
            registrado_por=self.asesor,
        )
        pago_otro_curso = Abono.objects.create(
            matricula=otra_matricula,
            fecha=date(2026, 7, 9),
            monto=Decimal('25.00'),
            metodo='transferencia',
            registrado_por=self.asesor,
        )
        pago_otro_metodo = Abono.objects.create(
            matricula=matricula_2,
            fecha=date(2026, 7, 9),
            monto=Decimal('25.00'),
            metodo='efectivo',
            registrado_por=self.asesor,
        )
        fecha_registro = timezone.make_aware(datetime(2026, 7, 9, 17, 25))
        Abono.objects.filter(
            pk__in=[
                pago_objetivo.pk, pago_otro_monto.pk,
                pago_otro_curso.pk, pago_otro_metodo.pk,
            ]
        ).update(creado=fecha_registro)
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:control_registro'),
            {
                'tab': 'pagos',
                'pago_fecha': '2026-07-09',
                'pago_curso': str(self.curso.pk),
                'pago_metodo': 'transferencia',
                'pago_monto': '25,00',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['active_tab'], 'pagos')
        pagos = list(response.context['pagos'])
        self.assertEqual(pagos, [pago_objetivo])

    def test_admin_dashboard_valida_recuperaciones_descontadas_con_abono_modulo(self):
        admin = User.objects.create_superuser(username='admin_control_recup')
        matricula = Matricula.objects.get(estudiante=self.estudiante_1)
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 9),
            monto=Decimal('70.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 10),
            saldo_pendiente_al_marcar=matricula.saldo,
            pagada=True,
            fecha_recuperacion=date(2026, 7, 10),
            abono=abono,
        )

        self.client.force_login(admin)
        response = self.client.get(reverse('academia:admin_dashboard'), {'anio': '2026', 'mes': '7'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['recuperaciones_mes']['total'], Decimal('0.00'))
        self.assertEqual(response.context['recuperaciones_abono_modulo_mes']['count'], 1)
        self.assertEqual(
            response.context['recuperaciones_abono_modulo_mes']['total_recibos'],
            Decimal('70.00'),
        )
        self.assertContains(response, 'Recuperación con Abono + Módulo')
        self.assertContains(response, 'No suma ingreso nuevo')

    def test_control_registro_muestra_tipo_abono_modulo_en_pagos(self):
        admin = User.objects.create_superuser(username='admin_control_tipo_pago')
        matricula = Matricula.objects.get(estudiante=self.estudiante_1)
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 9),
            monto=Decimal('70.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
            metodo='efectivo',
            registrado_por=admin,
        )

        self.client.force_login(admin)
        response = self.client.get(reverse('academia:control_registro'), {'tab': 'pagos'})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Abono + Módulo')
        self.assertContains(response, 'Mód. 1')


class BusquedaSinTildesTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_user(username='asesor_busqueda')
        grupo = Group.objects.create(name='Asesores')
        self.usuario.groups.add(grupo)
        self.client.force_login(self.usuario)

        self.curso = Curso.objects.create(
            nombre='Técnico Contable',
            ofrece_presencial=True,
            valor_presencial=Decimal('100.00'),
        )
        self.estudiante = Estudiante.objects.create(
            cedula='1207342716',
            nombres='Osmár Dahmér',
            correo='osmar@example.com',
            celular='0999999999',
        )
        self.matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            modalidad='presencial',
            fecha_matricula=date(2026, 7, 10),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('20.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        self.persona_externa = PersonaExterna.objects.create(
            cedula='0912345678',
            nombres='Jeffrey Dahmér',
        )
        self.adicional = Adicional.objects.create(
            tipo_adicional='cert_antiguo',
            persona_externa=self.persona_externa,
            curso=self.curso,
            modalidad='presencial',
            fecha=date(2026, 7, 10),
            valor=Decimal('10.00'),
        )
        self.cierre = CierreCurso.objects.create(
            curso=self.curso,
            curso_nombre='Técnico Contable',
            jornada_modalidad='presencial',
            alcance='curso',
            total_matriculas=1,
        )
        self.estudiante_archivado = EstudianteArchivado.objects.create(
            cierre=self.cierre,
            estudiante_original_id=self.estudiante.pk,
            cedula='1207342717',
            nombres='Jazzyel Kleinér',
            correo='jazzy@example.com',
            celular='0888888888',
        )
        self.adicional_archivado = AdicionalArchivado.objects.create(
            cierre=self.cierre,
            tipo_adicional='cert_antiguo',
            tipo_adicional_label='Certificado antiguo',
            persona_nombre='Jazzyel Kleinér',
            persona_cedula='1207342717',
            curso_nombre='Técnico Contable',
            fecha=date(2026, 7, 10),
            valor=Decimal('10.00'),
            metodo_pago='efectivo',
        )

    def test_estudiantes_filtra_nombre_sin_tildes(self):
        response = self.client.get(reverse('academia:estudiantes_lista'), {'q': 'Osmar Dahmer'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context['estudiantes']), [self.estudiante])

    def test_pagos_filtra_curso_sin_tildes(self):
        response = self.client.get(reverse('academia:pagos_lista'), {'q': 'Tecnico'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['matriculas'][0].pk, self.matricula.pk)

    def test_historial_filtra_nombre_sin_tildes(self):
        response = self.client.get(reverse('academia:historial_lista'), {'q': 'Osmar'})

        self.assertEqual(response.status_code, 200)
        matriculas = response.context['estructura'][0]['meses'][0]['matriculas']
        self.assertEqual([m.pk for m in matriculas], [self.matricula.pk])

    def test_estudiantes_archivados_filtra_nombre_sin_tildes(self):
        response = self.client.get(reverse('academia:estudiantes_archivados_lista'), {'q': 'Kleiner'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context['estudiantes']), [self.estudiante_archivado])

    def test_adicional_filtra_persona_externa_sin_tildes(self):
        response = self.client.get(reverse('academia:adicional_lista'), {'q': 'Jeffrey Dahmer'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context['adicionales']), [self.adicional])

    def test_adicional_archivado_filtra_nombre_sin_tildes(self):
        response = self.client.get(reverse('academia:adicionales_archivados_lista'), {'q': 'Kleiner'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context['adicionales']), [self.adicional_archivado])


class CierreCursoManualTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create_superuser(
            username='admin_cierre',
            email='admin@example.com',
            password='clave-admin-123',
        )
        self.client.force_login(self.admin)
        self.curso = Curso.objects.create(
            nombre='Asistente Contable Cierre Manual Test',
            ofrece_presencial=True,
            valor_presencial=Decimal('100.00'),
        )
        self.jornada_1 = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='lun_mie_vie',
            fecha_inicio=date(2026, 8, 1),
            ciudad='Guayaquil',
        )
        self.jornada_2 = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='mar_jue',
            fecha_inicio=date(2026, 8, 2),
            ciudad='Guayaquil',
        )
        self.estudiante_1 = Estudiante.objects.create(
            cedula='000001',
            nombres='Osmár Manual',
            celular='0991111111',
        )
        self.estudiante_2 = Estudiante.objects.create(
            cedula='000002',
            nombres='Estudiante Otra Jornada',
            celular='0992222222',
        )
        self.matricula_1 = Matricula.objects.create(
            estudiante=self.estudiante_1,
            curso=self.curso,
            jornada=self.jornada_1,
            modalidad='presencial',
            fecha_matricula=date(2026, 8, 10),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('30.00'),
            tipo_registro='central_ia',
            registrado_por=self.admin,
        )
        self.matricula_2 = Matricula.objects.create(
            estudiante=self.estudiante_2,
            curso=self.curso,
            jornada=self.jornada_2,
            modalidad='presencial',
            fecha_matricula=date(2026, 8, 10),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('100.00'),
            tipo_registro='central_ia',
            registrado_por=self.admin,
        )

    def test_preview_busca_matricula_manual_por_nombre_sin_tilde(self):
        response = self.client.get(
            reverse('academia:cierre_preview', kwargs={'curso_pk': self.curso.pk}),
            {'archivo_mes': '8', 'archivo_anio': '2026', 'manual_q': 'Osmar'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Cierre manual por estudiante')
        self.assertEqual(list(response.context['manual_matriculas']), [self.matricula_1])

    def test_preview_manual_busca_matricula_de_otro_mes(self):
        response = self.client.get(
            reverse('academia:cierre_preview', kwargs={'curso_pk': self.curso.pk}),
            {'archivo_mes': '7', 'archivo_anio': '2026', 'manual_q': 'Osmar'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(list(response.context['manual_matriculas']), [self.matricula_1])

    def test_cierre_por_jornada_no_archiva_otras_jornadas(self):
        response = self.client.post(
            reverse('academia:cierre_ejecutar', kwargs={'curso_pk': self.curso.pk}),
            {
                'jornada_id': str(self.jornada_1.pk),
                'archivo_mes': '8',
                'archivo_anio': '2026',
                'admin_password': 'clave-admin-123',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Matricula.objects.filter(pk=self.matricula_1.pk).exists())
        self.assertTrue(Matricula.objects.filter(pk=self.matricula_2.pk).exists())

        cierre = CierreCurso.objects.get()
        self.assertEqual(cierre.alcance, 'jornada')
        self.assertEqual(cierre.jornada, self.jornada_1)
        self.assertEqual(cierre.total_matriculas, 1)
        self.assertEqual(cierre.matriculas_archivadas.count(), 1)

    def test_cierre_por_jornada_cierra_agosto_pero_guarda_en_julio(self):
        response = self.client.post(
            reverse('academia:cierre_ejecutar', kwargs={'curso_pk': self.curso.pk}),
            {
                'jornada_id': str(self.jornada_1.pk),
                'periodo_mes': '8',
                'periodo_anio': '2026',
                'archivo_mes': '7',
                'archivo_anio': '2026',
                'archivo_dia': '15',
                'aplicar_mes_archivo': 'on',
                'admin_password': 'clave-admin-123',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Matricula.objects.filter(pk=self.matricula_1.pk).exists())
        self.assertTrue(Matricula.objects.filter(pk=self.matricula_2.pk).exists())

        cierre = CierreCurso.objects.get()
        self.assertEqual(timezone.localtime(cierre.fecha_cierre).date(), date(2026, 7, 15))
        archivada = MatriculaArchivada.objects.get(cierre=cierre)
        self.assertEqual(timezone.localtime(archivada.archivado_en).date(), date(2026, 7, 15))

    def test_cierre_manual_estudiante_archiva_solo_esa_matricula(self):
        Abono.objects.create(
            matricula=self.matricula_1,
            fecha=date(2026, 8, 10),
            monto=Decimal('30.00'),
            tipo_pago='abono',
            metodo='efectivo',
            registrado_por=self.admin,
        )

        response = self.client.post(
            reverse(
                'academia:cierre_manual_estudiante_ejecutar',
                kwargs={'curso_pk': self.curso.pk, 'matricula_pk': self.matricula_1.pk},
            ),
            {
                'archivo_mes': '8',
                'archivo_anio': '2026',
                'admin_password': 'clave-admin-123',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Matricula.objects.filter(pk=self.matricula_1.pk).exists())
        self.assertTrue(Matricula.objects.filter(pk=self.matricula_2.pk).exists())

        cierre = CierreCurso.objects.get()
        self.assertEqual(cierre.alcance, 'manual')
        self.assertEqual(cierre.total_matriculas, 1)
        self.assertEqual(cierre.total_estudiantes_archivados, 1)
        self.assertFalse(cierre.limpio_directorio)

        archivada = MatriculaArchivada.objects.get(cierre=cierre)
        self.assertEqual(archivada.matricula_original_id, self.matricula_1.pk)
        self.assertEqual(archivada.nombre_completo, 'Osmár Manual')
        self.assertEqual(archivada.abonos_archivados.count(), 1)

        estudiante_archivado = EstudianteArchivado.objects.get(cierre=cierre)
        self.assertEqual(estudiante_archivado.estudiante_original_id, self.estudiante_1.pk)
        self.assertEqual(estudiante_archivado.nombre_completo, 'Osmár Manual')
        self.assertTrue(Estudiante.objects.filter(pk=self.estudiante_1.pk).exists())

    def test_cierre_manual_guarda_en_julio_si_se_confirma_mes_distinto(self):
        response = self.client.post(
            reverse(
                'academia:cierre_manual_estudiante_ejecutar',
                kwargs={'curso_pk': self.curso.pk, 'matricula_pk': self.matricula_1.pk},
            ),
            {
                'archivo_mes': '7',
                'archivo_anio': '2026',
                'archivo_dia': '20',
                'aplicar_mes_archivo': 'on',
                'admin_password': 'clave-admin-123',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Matricula.objects.filter(pk=self.matricula_1.pk).exists())

        cierre = CierreCurso.objects.get()
        self.assertEqual(cierre.alcance, 'manual')
        self.assertEqual(timezone.localtime(cierre.fecha_cierre).date(), date(2026, 7, 20))

        archivada = MatriculaArchivada.objects.get(cierre=cierre)
        self.assertEqual(timezone.localtime(archivada.archivado_en).date(), date(2026, 7, 20))
        estudiante_archivado = EstudianteArchivado.objects.get(cierre=cierre)
        self.assertEqual(timezone.localtime(estudiante_archivado.archivado_en).date(), date(2026, 7, 20))

    def test_cierre_manual_bloquea_mes_distinto_sin_confirmacion(self):
        response = self.client.post(
            reverse(
                'academia:cierre_manual_estudiante_ejecutar',
                kwargs={'curso_pk': self.curso.pk, 'matricula_pk': self.matricula_1.pk},
            ),
            {
                'archivo_mes': '7',
                'archivo_anio': '2026',
                'archivo_dia': '20',
                'admin_password': 'clave-admin-123',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertTrue(Matricula.objects.filter(pk=self.matricula_1.pk).exists())
        self.assertFalse(CierreCurso.objects.exists())
        self.assertFalse(MatriculaArchivada.objects.exists())

    def test_cierre_manual_con_limpieza_quita_estudiante_sin_matriculas_vivas(self):
        response = self.client.post(
            reverse(
                'academia:cierre_manual_estudiante_ejecutar',
                kwargs={'curso_pk': self.curso.pk, 'matricula_pk': self.matricula_1.pk},
            ),
            {
                'archivo_mes': '8',
                'archivo_anio': '2026',
                'admin_password': 'clave-admin-123',
                'limpiar_directorio': 'on',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Matricula.objects.filter(pk=self.matricula_1.pk).exists())
        self.assertFalse(Estudiante.objects.filter(pk=self.estudiante_1.pk).exists())
        self.assertTrue(Matricula.objects.filter(pk=self.matricula_2.pk).exists())

        cierre = CierreCurso.objects.get()
        self.assertEqual(cierre.alcance, 'manual')
        self.assertEqual(cierre.total_estudiantes_archivados, 1)
        self.assertTrue(cierre.limpio_directorio)

        estudiante_archivado = EstudianteArchivado.objects.get(cierre=cierre)
        self.assertEqual(estudiante_archivado.nombre_completo, 'Osmár Manual')


class CamposNumericosMatriculaTests(TestCase):
    def _estudiante_data(self, **overrides):
        data = {
            'cedula': '0102030405',
            'nombres': 'Estudiante Numérico',
            'edad': '24',
            'correo': '',
            'celular': '0991234567',
            'nivel_formacion': '',
            'titulo_profesional': '',
            'ciudad': '',
        }
        data.update(overrides)
        return data

    def test_campos_personales_numericos_conservan_ceros_iniciales(self):
        form = EstudianteForm(self._estudiante_data())

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['cedula'], '0102030405')
        self.assertEqual(form.cleaned_data['celular'], '0991234567')

    def test_factura_si_no_exige_correo_del_estudiante(self):
        form = EstudianteForm(
            self._estudiante_data(correo='', ciudad='Guayaquil'),
            factura_si=True,
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['correo'], '')

    def test_cedula_ruc_rechaza_letras_y_signos(self):
        form = EstudianteForm(
            self._estudiante_data(cedula='01020A040-5')
        )

        self.assertFalse(form.is_valid())
        self.assertIn('cedula', form.errors)
        self.assertIn('únicamente números', form.errors['cedula'][0])

    def test_cedula_rechaza_menos_de_diez_digitos(self):
        form = EstudianteForm(self._estudiante_data(cedula='010203040'))

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors['cedula'][0],
            'La cédula debe tener 10 dígitos.',
        )

    def test_ruc_valido_acepta_trece_digitos_terminados_en_001(self):
        ruc = '1207342716001'
        form = EstudianteForm(self._estudiante_data(cedula=ruc))

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['cedula'], ruc)
        self.assertTrue(es_cedula_ruc_ecuador_valido(ruc))

    def test_ruc_requiere_exactamente_trece_digitos_y_terminacion_001(self):
        self.assertFalse(es_ruc_ecuador('983001'))
        self.assertFalse(es_ruc_ecuador('1207342716123'))
        self.assertTrue(es_ruc_ecuador('1207342716001'))

    def test_registro_matricula_acepta_documentos_de_longitud_flexible(self):
        for documento in ('123', '12345678901', '123456789012345'):
            with self.subTest(documento=documento):
                form = EstudianteForm(
                    self._estudiante_data(cedula=documento),
                    documento_flexible=True,
                )

                self.assertTrue(form.is_valid(), form.errors)
                self.assertEqual(form.cleaned_data['cedula'], documento)
                self.assertTrue(
                    es_cedula_ruc_ecuador_valido(
                        documento,
                        permitir_longitud_flexible=True,
                    )
                )

    def test_registro_matricula_flexible_sigue_rechazando_letras(self):
        form = EstudianteForm(
            self._estudiante_data(cedula='123ABC001'),
            documento_flexible=True,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('únicamente números', form.errors['cedula'][0])

    def test_widget_flexible_quita_patron_de_longitud(self):
        form = EstudianteForm(documento_flexible=True)
        attrs = form.fields['cedula'].widget.attrs

        self.assertNotIn('pattern', attrs)
        self.assertEqual(attrs['maxlength'], '20')
        self.assertIn('exactamente 13 dígitos', attrs['title'])

    def test_cedula_pegada_con_separadores_se_normaliza(self):
        form = EstudianteForm(self._estudiante_data(cedula='010-203-0405'))

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['cedula'], '0102030405')

    def test_ruc_rechaza_longitud_incompleta(self):
        form = EstudianteForm(self._estudiante_data(cedula='120734271602'))

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors['cedula'][0],
            'El RUC debe tener 13 dígitos y terminar en 001.',
        )

    def test_ruc_rechaza_trece_digitos_sin_terminacion_001(self):
        form = EstudianteForm(self._estudiante_data(cedula='1207342716123'))

        self.assertFalse(form.is_valid())
        self.assertEqual(
            form.errors['cedula'][0],
            'El RUC debe terminar en 001.',
        )

    def test_celular_rechaza_letras_aunque_contenga_diez_digitos(self):
        form = EstudianteForm(
            self._estudiante_data(celular='0991234567abc')
        )

        self.assertFalse(form.is_valid())
        self.assertIn('celular', form.errors)
        self.assertIn('únicamente números', form.errors['celular'][0])

    def test_celular_pegado_desde_whatsapp_se_normaliza(self):
        form = EstudianteForm(
            self._estudiante_data(celular='+593 99 759 6744')
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['celular'], '0997596744')

    def test_edad_rechaza_letras(self):
        form = EstudianteForm(self._estudiante_data(edad='2a'))

        self.assertFalse(form.is_valid())
        self.assertIn('edad', form.errors)

    def test_widgets_numericos_exponen_teclado_y_restriccion_correctos(self):
        estudiante_form = EstudianteForm()
        matricula_form = MatriculaForm()

        for campo in ('edad', 'celular'):
            attrs = estudiante_form.fields[campo].widget.attrs
            self.assertEqual(attrs['inputmode'], 'numeric')
            self.assertEqual(attrs['pattern'], '[0-9]*')
            self.assertEqual(attrs['data-digits-only'], 'true')
        cedula_attrs = estudiante_form.fields['cedula'].widget.attrs
        self.assertEqual(cedula_attrs['inputmode'], 'numeric')
        self.assertEqual(cedula_attrs['data-digits-only'], 'true')
        self.assertEqual(
            estudiante_form.fields['celular'].widget.attrs['maxlength'],
            '10',
        )
        self.assertEqual(
            estudiante_form.fields['celular'].widget.attrs['data-phone-ecuador'],
            'true',
        )
        self.assertEqual(
            estudiante_form.fields['cedula'].widget.attrs['maxlength'],
            '13',
        )
        self.assertEqual(
            estudiante_form.fields['cedula'].widget.attrs['pattern'],
            '(?:[0-9]{10}|[0-9]{10}001)',
        )

        fact_attrs = matricula_form.fields['fact_cedula'].widget.attrs
        self.assertEqual(fact_attrs['inputmode'], 'numeric')
        self.assertEqual(fact_attrs['data-digits-only'], 'true')

        for campo in (
            'valor_curso', 'descuento', 'valor_pagado',
            'monto_pago_1', 'monto_pago_2',
        ):
            attrs = matricula_form.fields[campo].widget.attrs
            self.assertEqual(attrs['inputmode'], 'decimal')
            self.assertEqual(attrs['data-decimal-only'], 'true')


class PagoInicialMatriculaTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_user(username='soporte')
        self.sede = Sede.objects.create(nombre='Guayaquil', orden=1)
        self.curso = Curso.objects.create(
            nombre='Curso Técnico',
            ofrece_presencial=True,
            valor_presencial=Decimal('115.00'),
            numero_modulos=4,
        )
        self.jornada = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='lun_mie_vie',
            fecha_inicio=date(2026, 7, 5),
            sede=self.sede,
        )
        self.estudiante = Estudiante.objects.create(
            cedula='1207342716',
            nombres='Gianny Guevara',
        )

    def _matricula_form_data(self, **overrides):
        data = {
            'mat-curso': str(self.curso.pk),
            'mat-jornada': str(self.jornada.pk),
            'mat-estado': 'activa',
            'mat-tipo_matricula': 'reserva_abono',
            'mat-forma_pago': 'abono',
            'mat-fecha_matricula': '2026-07-05',
            'mat-valor_curso': '115.00',
            'mat-descuento': '0.00',
            'mat-valor_pagado': '10.00',
            'mat-observaciones': '',
            'mat-tipo_registro': 'central_ia',
            'mat-link_comprobante': '',
            'mat-factura_realizada': 'no',
            'mat-fact_nombres': '',
            'mat-fact_cedula': '',
            'mat-fact_correo': '',
            'mat-tipo_cobro': 'un_solo_metodo',
            'mat-metodo_pago': 'efectivo',
            'mat-banco': '',
            'mat-monto_pago_1': '',
            'mat-metodo_pago_1': '',
            'mat-banco_1': '',
            'mat-monto_pago_2': '',
            'mat-metodo_pago_2': '',
            'mat-banco_2': '',
            'mat-modulos_a_pagar': '1',
        }
        data.update(overrides)
        return data

    def _estudiante_post_data(self, **overrides):
        data = {
            'est-cedula': self.estudiante.cedula,
            'est-nombres': self.estudiante.nombres,
            'est-edad': '',
            'est-correo': '',
            'est-celular': '0991234567',
            'est-nivel_formacion': '',
            'est-titulo_profesional': '',
            'est-ciudad': 'Guayaquil',
        }
        data.update(overrides)
        return data

    def test_matricula_metodo_pago_muestra_seleccione_primero(self):
        form = MatriculaForm(prefix='mat')

        choices = list(form.fields['metodo_pago'].choices)

        self.assertEqual(choices[0], ('', 'Seleccione'))

    def test_matricula_nueva_no_ofrece_reserva_mas_modulo(self):
        form = MatriculaForm(prefix='mat')

        tipos = dict(form.fields['tipo_matricula'].choices)
        formas = dict(form.fields['forma_pago'].choices)

        self.assertNotIn('reserva_modulo_1', tipos)
        self.assertNotIn('abono_modulo', formas)
        self.assertIn('reserva_abono', tipos)
        self.assertIn('abono', formas)

    def test_matricula_nueva_acepta_reserva_fija_de_diez_dolares(self):
        form = MatriculaForm(self._matricula_form_data(), prefix='mat')

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['valor_pagado'], Decimal('10.00'))

    def test_matricula_nueva_acepta_abono_mayor_a_la_reserva(self):
        form = MatriculaForm(
            self._matricula_form_data(**{'mat-valor_pagado': '15.00'}),
            prefix='mat',
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['valor_pagado'], Decimal('15.00'))

    def test_matricula_nueva_rechaza_abono_menor_a_la_reserva(self):
        form = MatriculaForm(
            self._matricula_form_data(**{'mat-valor_pagado': '5.00'}),
            prefix='mat',
        )

        self.assertFalse(form.is_valid())
        self.assertIn('valor_pagado', form.errors)

    def test_matricula_antigua_conserva_reserva_mas_modulo_al_editar(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('0.00'),
        )

        form = MatriculaForm(prefix='mat', instance=matricula)

        self.assertIn('reserva_modulo_1', dict(form.fields['tipo_matricula'].choices))
        self.assertIn('abono_modulo', dict(form.fields['forma_pago'].choices))

    def test_matricula_rechaza_metodo_pago_vacio(self):
        form = MatriculaForm(
            self._matricula_form_data(**{'mat-metodo_pago': ''}),
            prefix='mat',
        )

        self.assertFalse(form.is_valid())
        self.assertIn('metodo_pago', form.errors)

    def test_matricula_rechaza_letras_en_cedula_ruc_de_factura(self):
        form = MatriculaForm(
            self._matricula_form_data(
                **{
                    'mat-factura_realizada': 'si',
                    'mat-fact_nombres': 'Cliente Factura',
                    'mat-fact_cedula': '0102030405A',
                    'mat-fact_correo': 'cliente@example.com',
                }
            ),
            prefix='mat',
        )

        self.assertFalse(form.is_valid())
        self.assertIn('fact_cedula', form.errors)
        self.assertIn('únicamente números', form.errors['fact_cedula'][0])

    def test_matricula_normaliza_cedula_ruc_de_factura_pegada(self):
        form = MatriculaForm(
            self._matricula_form_data(
                **{
                    'mat-factura_realizada': 'si',
                    'mat-fact_nombres': 'Cliente Factura',
                    'mat-fact_cedula': '010 203 0405',
                    'mat-fact_correo': 'cliente@example.com',
                }
            ),
            prefix='mat',
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['fact_cedula'], '0102030405')

    def test_matricula_factura_si_no_exige_correo_de_factura(self):
        form = MatriculaForm(
            self._matricula_form_data(
                **{
                    'mat-factura_realizada': 'si',
                    'mat-fact_nombres': 'Cliente Factura',
                    'mat-fact_cedula': '0102030405',
                    'mat-fact_correo': '',
                }
            ),
            prefix='mat',
        )

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['fact_correo'], '')

    def test_registro_con_estudiante_existente_actualiza_correo_para_confirmacion(self):
        asesor = User.objects.create_superuser(username='admin_matricula_view')
        self.client.force_login(asesor)

        response = self.client.post(
            reverse(
                'academia:matricula_registrar',
                kwargs={'modalidad': 'presencial'},
            ),
            {
                **self._estudiante_post_data(
                    **{'est-correo': 'nuevo.estudiante@example.com'}
                ),
                **self._matricula_form_data(),
                'vendedora_id': str(asesor.pk),
            },
        )

        self.assertEqual(response.status_code, 302)
        self.estudiante.refresh_from_db()
        self.assertEqual(
            self.estudiante.correo,
            'nuevo.estudiante@example.com',
        )
        matricula = Matricula.objects.get(estudiante=self.estudiante)
        self.assertEqual(matricula.vendedora, asesor)

    def test_matricula_mixta_rechaza_metodos_vacios(self):
        form = MatriculaForm(
            self._matricula_form_data(
                **{
                    'mat-tipo_cobro': 'mixto',
                    'mat-valor_pagado': '10.00',
                    'mat-monto_pago_1': '5.00',
                    'mat-metodo_pago_1': '',
                    'mat-monto_pago_2': '5.00',
                    'mat-metodo_pago_2': '',
                }
            ),
            prefix='mat',
        )

        self.assertFalse(form.is_valid())
        self.assertIn('metodo_pago_1', form.errors)
        self.assertIn('metodo_pago_2', form.errors)

    def test_matricula_mixta_rechaza_suma_distinta_al_valor_pagado(self):
        form = MatriculaForm(
            self._matricula_form_data(
                **{
                    'mat-tipo_cobro': 'mixto',
                    'mat-valor_pagado': '10.00',
                    'mat-monto_pago_1': '4.00',
                    'mat-metodo_pago_1': 'efectivo',
                    'mat-monto_pago_2': '4.00',
                    'mat-metodo_pago_2': 'efectivo',
                }
            ),
            prefix='mat',
        )

        self.assertFalse(form.is_valid())
        self.assertIn('monto_pago_2', form.errors)

    def test_matricula_rechaza_jornada_presencial_sin_sede(self):
        jornada_sin_sede = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='mar_jue',
            fecha_inicio=date(2026, 7, 6),
        )
        form = MatriculaForm(
            self._matricula_form_data(**{'mat-jornada': str(jornada_sin_sede.pk)}),
            prefix='mat',
        )

        self.assertFalse(form.is_valid())
        self.assertIn('jornada', form.errors)

    def _mat_form(self, modulos_a_pagar=1, **overrides):
        data = {
            'tipo_cobro': 'un_solo_metodo',
            'metodo_pago': 'efectivo',
            'banco': '',
            'monto_pago_1': Decimal('0.00'),
            'metodo_pago_1': 'efectivo',
            'banco_1': '',
            'monto_pago_2': Decimal('0.00'),
            'metodo_pago_2': 'efectivo',
            'banco_2': '',
            'modulos_a_pagar': modulos_a_pagar,
        }
        data.update(overrides)
        return SimpleNamespace(cleaned_data=data)

    def _crear_recuperacion_para_filtro(
        self,
        usuario,
        estudiante,
        fecha_falta,
        fecha_programada,
    ):
        matricula = Matricula.objects.create(
            estudiante=estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('25.00'),
            tipo_registro='central_ia',
            registrado_por=usuario,
        )
        return RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=fecha_falta,
            fecha_programada=fecha_programada,
            saldo_pendiente_al_marcar=Decimal('90.00'),
        )

    def _crear_recuperacion_cobrable(self, usuario):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=usuario,
        )
        recuperacion = RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=2,
            fecha_marcada=date(2026, 8, 1),
            saldo_pendiente_al_marcar=Decimal('25.00'),
        )
        return matricula, recuperacion

    def test_recuperaciones_lista_filtra_por_fecha_de_falta(self):
        admin = User.objects.create_superuser(
            username='admin_filtro_falta',
            password='clave12345',
        )
        otro_estudiante = Estudiante.objects.create(
            cedula='0102030405',
            nombres='Jazzyel Kleiner',
        )
        recuperacion_en_rango = self._crear_recuperacion_para_filtro(
            admin,
            self.estudiante,
            date(2026, 7, 10),
            date(2026, 7, 15),
        )
        recuperacion_fuera = self._crear_recuperacion_para_filtro(
            admin,
            otro_estudiante,
            date(2026, 7, 25),
            date(2026, 7, 30),
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:recuperaciones_lista'),
            {
                'estado': 'todas',
                'fecha_falta_desde': '2026-07-09',
                'fecha_falta_hasta': '2026-07-11',
            },
        )

        recuperaciones = set(response.context['recuperaciones'].values_list('id', flat=True))
        self.assertEqual(response.status_code, 200)
        self.assertIn(recuperacion_en_rango.id, recuperaciones)
        self.assertNotIn(recuperacion_fuera.id, recuperaciones)
        self.assertEqual(response.context['filtros']['fecha_falta_desde'], '2026-07-09')
        self.assertEqual(response.context['filtros']['fecha_falta_hasta'], '2026-07-11')

    def test_recuperaciones_lista_filtra_por_fecha_programada(self):
        admin = User.objects.create_superuser(
            username='admin_filtro_programada',
            password='clave12345',
        )
        otro_estudiante = Estudiante.objects.create(
            cedula='0102030406',
            nombres='Daniela Mora',
        )
        recuperacion_fuera = self._crear_recuperacion_para_filtro(
            admin,
            self.estudiante,
            date(2026, 7, 10),
            date(2026, 7, 15),
        )
        recuperacion_en_rango = self._crear_recuperacion_para_filtro(
            admin,
            otro_estudiante,
            date(2026, 7, 25),
            date(2026, 7, 30),
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:recuperaciones_lista'),
            {
                'estado': 'todas',
                'fecha_programada_desde': '2026-07-31',
                'fecha_programada_hasta': '2026-07-29',
            },
        )

        recuperaciones = set(response.context['recuperaciones'].values_list('id', flat=True))
        self.assertEqual(response.status_code, 200)
        self.assertIn(recuperacion_en_rango.id, recuperaciones)
        self.assertNotIn(recuperacion_fuera.id, recuperaciones)
        self.assertEqual(response.context['filtros']['fecha_programada_desde'], '2026-07-29')
        self.assertEqual(response.context['filtros']['fecha_programada_hasta'], '2026-07-31')
        self.assertIn(
            'fecha_programada_desde=2026-07-29',
            response.context['filtros']['query'],
        )

    def test_recuperaciones_lista_filtra_estudiantes_del_curso(self):
        admin = User.objects.create_superuser(
            username='admin_filtro_estudiante_recuperacion',
            password='clave12345',
        )
        curso_extra = Curso.objects.create(
            nombre='Excel Administrativo',
            ofrece_presencial=True,
            valor_presencial=Decimal('90.00'),
            numero_modulos=3,
        )
        jornada_extra = JornadaCurso.objects.create(
            curso=curso_extra,
            modalidad='presencial',
            descripcion='sabados',
            fecha_inicio=date(2026, 7, 12),
            sede=self.sede,
        )
        otro_estudiante = Estudiante.objects.create(
            cedula='0102030407',
            nombres='Daniela Mora',
        )
        recuperacion_curso = self._crear_recuperacion_para_filtro(
            admin,
            self.estudiante,
            date(2026, 7, 10),
            date(2026, 7, 15),
        )
        matricula_extra = Matricula.objects.create(
            estudiante=otro_estudiante,
            curso=curso_extra,
            jornada=jornada_extra,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 12),
            valor_curso=Decimal('90.00'),
            valor_pagado=Decimal('20.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        recuperacion_otro_curso = RecuperacionPendiente.objects.create(
            matricula=matricula_extra,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 12),
            fecha_programada=date(2026, 7, 20),
            saldo_pendiente_al_marcar=Decimal('70.00'),
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:recuperaciones_lista'),
            {
                'estado': 'todas',
                'curso': str(self.curso.pk),
                'estudiante': str(self.estudiante.pk),
            },
        )

        recuperaciones = set(response.context['recuperaciones'].values_list('id', flat=True))
        self.assertEqual(response.status_code, 200)
        self.assertIn(recuperacion_curso.id, recuperaciones)
        self.assertNotIn(recuperacion_otro_curso.id, recuperaciones)
        self.assertEqual(response.context['filtros']['estudiante'], str(self.estudiante.pk))
        self.assertIn(f'estudiante={self.estudiante.pk}', response.context['filtros']['query'])
        self.assertContains(response, 'id="recup-estudiante-filter"')
        self.assertContains(response, self.estudiante.nombres)

        opciones = response.context['estudiantes_filtro']
        opcion_actual = next(
            opcion for opcion in opciones
            if opcion['id'] == self.estudiante.pk
        )
        self.assertIn(str(self.curso.pk), opcion_actual['curso_ids'])

        response_sin_curso = self.client.get(
            reverse('academia:recuperaciones_lista'),
            {
                'estado': 'todas',
                'estudiante': str(self.estudiante.pk),
            },
        )
        recuperaciones_sin_curso = set(
            response_sin_curso.context['recuperaciones'].values_list('id', flat=True),
        )
        self.assertEqual(response_sin_curso.context['filtros']['estudiante'], '')
        self.assertIn(recuperacion_curso.id, recuperaciones_sin_curso)
        self.assertIn(recuperacion_otro_curso.id, recuperaciones_sin_curso)

    def test_recuperaciones_lista_estudiantes_respetan_estado_de_pago(self):
        admin = User.objects.create_superuser(
            username='admin_filtro_estudiante_estado_recuperacion',
            password='clave12345',
        )
        estudiante_pagado = Estudiante.objects.create(
            cedula='0102030410',
            nombres='Estudiante Recuperacion Pagada',
        )
        recuperacion_pendiente = self._crear_recuperacion_para_filtro(
            admin,
            self.estudiante,
            date(2026, 7, 10),
            date(2026, 7, 15),
        )
        matricula_pagada = Matricula.objects.create(
            estudiante=estudiante_pagado,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('25.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        abono = Abono.objects.create(
            matricula=matricula_pagada,
            fecha=date(2026, 7, 20),
            monto=Decimal('10.00'),
            tipo_pago='recuperacion',
            numero_modulo=1,
            cuenta_para_saldo=True,
        )
        recuperacion_pagada = RecuperacionPendiente.objects.create(
            matricula=matricula_pagada,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 10),
            fecha_programada=date(2026, 7, 15),
            fecha_recuperacion=date(2026, 7, 20),
            saldo_pendiente_al_marcar=Decimal('90.00'),
            pagada=True,
            abono=abono,
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:recuperaciones_lista'),
            {
                'estado': 'pendientes',
                'curso': str(self.curso.pk),
                'estudiante': str(estudiante_pagado.pk),
            },
        )

        recuperaciones = set(response.context['recuperaciones'].values_list('id', flat=True))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['filtros']['estudiante'], '')
        self.assertIn(recuperacion_pendiente.id, recuperaciones)
        self.assertNotIn(recuperacion_pagada.id, recuperaciones)

        opciones = response.context['estudiantes_filtro']
        opcion_pagada = next(
            opcion for opcion in opciones
            if opcion['id'] == estudiante_pagado.pk
        )
        self.assertNotIn(str(self.curso.pk), opcion_pagada['curso_ids_pendientes'])
        self.assertIn(str(self.curso.pk), opcion_pagada['curso_ids_pagadas'])
        self.assertContains(response, 'data-course-ids-pendientes=""')

        response_pagadas = self.client.get(
            reverse('academia:recuperaciones_lista'),
            {
                'estado': 'pagadas',
                'curso': str(self.curso.pk),
                'estudiante': str(estudiante_pagado.pk),
            },
        )

        recuperaciones_pagadas = set(
            response_pagadas.context['recuperaciones'].values_list('id', flat=True),
        )
        self.assertEqual(
            response_pagadas.context['filtros']['estudiante'],
            str(estudiante_pagado.pk),
        )
        self.assertIn(recuperacion_pagada.id, recuperaciones_pagadas)
        self.assertNotIn(recuperacion_pendiente.id, recuperaciones_pagadas)

    def test_reserva_modulo_respeta_valor_pagado_digitado(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('50.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        _registrar_pago_inicial(matricula, self.usuario, self._mat_form())

        matricula.refresh_from_db()
        abono = Abono.objects.get(matricula=matricula)
        self.assertEqual(matricula.valor_pagado, Decimal('50.00'))
        self.assertEqual(matricula.saldo, Decimal('65.00'))
        self.assertEqual(abono.monto, Decimal('50.00'))
        self.assertEqual(abono.numero_modulo, 1)
        self.assertEqual(abono.tipo_pago, 'por_modulo')

    def test_recuperacion_no_reutiliza_abono_modulo_existente(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 5),
            monto=Decimal('60.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        matricula.refresh_from_db()

        self.client.force_login(admin)
        response = self.client.post(
            reverse('academia:recuperacion_marcar', kwargs={'matricula_pk': matricula.pk}),
            {
                'numero_modulo': '2',
                'modo_registro': 'descontar_abono_modulo',
                'fecha_marcada': '2026-07-10',
                'observaciones': 'Faltó al módulo 2.',
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(Abono.objects.filter(matricula=matricula).count(), 1)
        recup = RecuperacionPendiente.objects.get(matricula=matricula)
        self.assertEqual(recup.numero_modulo, 2)
        self.assertFalse(recup.pagada)
        self.assertIsNone(recup.abono)
        self.assertIsNone(recup.fecha_recuperacion)
        abono.refresh_from_db()
        self.assertEqual(abono.tipo_pago, 'por_modulo')
        matricula.refresh_from_db()
        self.assertEqual(matricula.valor_pagado, Decimal('60.00'))

    def test_marcar_recuperacion_no_muestra_descuento_abono_modulo(self):
        admin = User.objects.create_superuser(
            username='admin_sin_descuento_recuperacion',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            registrado_por=admin,
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse(
                'academia:recuperacion_marcar',
                kwargs={'matricula_pk': matricula.pk},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotIn('modo_registro', response.context['form'].fields)
        self.assertNotContains(response, 'Descontar con Abono + Módulo')
        self.assertNotContains(response, 'Cómo usar el descuento')

    def test_cobrar_recuperacion_muestra_distribucion_de_pago(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_distribucion',
            password='clave12345',
        )
        _matricula, recuperacion = self._crear_recuperacion_cobrable(admin)
        self.client.force_login(admin)

        response = self.client.get(
            reverse(
                'academia:recuperacion_cobrar',
                kwargs={'recup_pk': recuperacion.pk},
            )
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Distribución de pago *')
        self.assertContains(response, 'Pago Mixto (Dividir el monto en dos)')
        self.assertContains(response, 'data-pago-mixto-resumen')
        self.assertContains(response, 'data-mixto-suma')

    def test_cobrar_recuperacion_guarda_pago_mixto(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_mixta',
            password='clave12345',
        )
        matricula, recuperacion = self._crear_recuperacion_cobrable(admin)
        self.client.force_login(admin)

        response = self.client.post(
            reverse(
                'academia:recuperacion_cobrar',
                kwargs={'recup_pk': recuperacion.pk},
            ),
            {
                'fecha': '2026-08-01',
                'monto': '25.00',
                'cuenta_para_saldo': 'True',
                'tipo_cobro': 'mixto',
                'metodo': '',
                'banco': '',
                'monto_pago_1': '10.00',
                'metodo_pago_1': 'efectivo',
                'banco_1': '',
                'monto_pago_2': '15.00',
                'metodo_pago_2': 'transferencia',
                'banco_2': 'pichincha',
                'numero_recibo': '',
                'observaciones': 'Recuperación con pago combinado.',
            },
        )

        recuperacion.refresh_from_db()
        abono = Abono.objects.get(matricula=matricula)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(recuperacion.pagada)
        self.assertEqual(recuperacion.abono, abono)
        self.assertEqual(abono.tipo_pago, 'recuperacion')
        self.assertEqual(abono.monto, Decimal('25.00'))
        self.assertEqual(abono.metodo, 'efectivo')
        self.assertEqual(abono.monto_2, Decimal('15.00'))
        self.assertEqual(abono.metodo_2, 'transferencia')
        self.assertEqual(abono.banco_2, 'pichincha')

    def test_matricula_abonos_muestra_pago_en_tabla_de_recuperaciones(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_pago_tabla',
            password='clave12345',
        )
        matricula, recuperacion = self._crear_recuperacion_cobrable(admin)
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 2),
            monto=Decimal('25.00'),
            tipo_pago='recuperacion',
            numero_modulo=recuperacion.numero_modulo,
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        recuperacion.pagada = True
        recuperacion.fecha_recuperacion = abono.fecha
        recuperacion.abono = abono
        recuperacion.save()
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:matricula_abonos', kwargs={'pk': matricula.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Pago recuperación')
        self.assertContains(response, '$25,00')
        self.assertContains(response, abono.numero_recibo)
        self.assertContains(response, 'Efectivo')

    def test_cobrar_recuperacion_rechaza_suma_mixta_incorrecta(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_suma_invalida',
            password='clave12345',
        )
        matricula, recuperacion = self._crear_recuperacion_cobrable(admin)
        self.client.force_login(admin)

        response = self.client.post(
            reverse(
                'academia:recuperacion_cobrar',
                kwargs={'recup_pk': recuperacion.pk},
            ),
            {
                'fecha': '2026-08-01',
                'monto': '25.00',
                'cuenta_para_saldo': 'True',
                'tipo_cobro': 'mixto',
                'metodo': '',
                'banco': '',
                'monto_pago_1': '10.00',
                'metodo_pago_1': 'efectivo',
                'banco_1': '',
                'monto_pago_2': '10.00',
                'metodo_pago_2': 'transferencia',
                'banco_2': 'pichincha',
                'numero_recibo': '',
                'observaciones': '',
            },
        )

        recuperacion.refresh_from_db()
        self.assertEqual(response.status_code, 200)
        self.assertIn('monto_pago_2', response.context['form'].errors)
        self.assertFalse(recuperacion.pagada)
        self.assertFalse(Abono.objects.filter(matricula=matricula).exists())

    def test_recuperacion_solo_muestra_modulos_sin_registro_de_pago(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_pendientes',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        for numero_modulo, monto in (
            (1, Decimal('22.50')),
            (2, Decimal('22.50')),
            (3, Decimal('25.00')),
        ):
            Abono.objects.create(
                matricula=matricula,
                fecha=date(2026, 7, 20),
                monto=monto,
                tipo_pago='por_modulo',
                numero_modulo=numero_modulo,
                cuenta_para_saldo=True,
                metodo='efectivo',
                registrado_por=admin,
            )
        self.client.force_login(admin)

        response = self.client.get(
            reverse(
                'academia:recuperacion_marcar',
                kwargs={'matricula_pk': matricula.pk},
            )
        )

        form = response.context['form']
        opciones = list(form.fields['numero_modulo'].widget.choices)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(form.modulos_pendientes, [4])
        self.assertEqual([str(valor) for valor, _label in opciones], ['', '4'])
        self.assertContains(response, 'Solo aparecen los módulos')

    def test_recuperacion_rechaza_modulo_que_ya_tiene_pago(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_bloqueada',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 20),
            monto=Decimal('25.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        self.client.force_login(admin)

        response = self.client.post(
            reverse(
                'academia:recuperacion_marcar',
                kwargs={'matricula_pk': matricula.pk},
            ),
            {
                'numero_modulo': '1',
                'modo_registro': 'normal',
                'fecha_marcada': '2026-07-29',
                'observaciones': 'Intento de módulo pagado.',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('numero_modulo', response.context['form'].errors)
        self.assertFalse(
            RecuperacionPendiente.objects.filter(matricula=matricula).exists()
        )

    def test_editar_recuperacion_guarda_fecha_programada(self):
        admin = User.objects.create_superuser(
            username='admin_fecha_recuperacion',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('10.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        recuperacion = RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 10),
            saldo_pendiente_al_marcar=Decimal('90.00'),
        )
        self.client.force_login(admin)

        response = self.client.post(
            reverse(
                'academia:recuperacion_editar',
                kwargs={'recup_pk': recuperacion.pk},
            ),
            {
                'numero_modulo': '1',
                'modo_registro': 'normal',
                'fecha_marcada': '2026-07-10',
                'fecha_programada': '2026-07-15',
                'observaciones': 'Recuperación acordada.',
            },
        )

        recuperacion.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            recuperacion.fecha_programada,
            date(2026, 7, 15),
        )
        self.assertFalse(recuperacion.pagada)
        self.assertIsNone(recuperacion.fecha_recuperacion)

    def test_fecha_programada_no_puede_ser_anterior_a_la_falta(self):
        admin = User.objects.create_superuser(
            username='admin_fecha_recuperacion_invalida',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('100.00'),
            valor_pagado=Decimal('10.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        self.client.force_login(admin)

        response = self.client.post(
            reverse(
                'academia:recuperacion_marcar',
                kwargs={'matricula_pk': matricula.pk},
            ),
            {
                'numero_modulo': '1',
                'modo_registro': 'normal',
                'fecha_marcada': '2026-07-10',
                'fecha_programada': '2026-07-09',
                'observaciones': 'Fecha inválida.',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertIn('fecha_programada', response.context['form'].errors)
        self.assertFalse(
            RecuperacionPendiente.objects.filter(matricula=matricula).exists()
        )

    def test_matricula_muestra_acciones_para_recuperacion_descontada(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_acciones',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 5),
            monto=Decimal('50.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        recup = RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 10),
            saldo_pendiente_al_marcar=matricula.saldo,
            pagada=True,
            fecha_recuperacion=date(2026, 7, 10),
            abono=abono,
        )

        self.client.force_login(admin)
        response = self.client.get(reverse('academia:matricula_abonos', kwargs={'pk': matricula.pk}))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, reverse('academia:recuperacion_editar', kwargs={'recup_pk': recup.pk}))
        self.assertContains(response, reverse('academia:recuperacion_eliminar', kwargs={'recup_pk': recup.pk}))
        self.assertContains(response, 'Eliminar')

    def test_matricula_abonos_muestra_aviso_para_tipo_abono(self):
        admin = User.objects.create_superuser(
            username='admin_aviso_abono_inicial',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('10.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 5),
            monto=Decimal('10.00'),
            tipo_pago='abono',
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )

        self.client.force_login(admin)
        response = self.client.get(
            reverse('academia:matricula_abonos', kwargs={'pk': matricula.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            'El abono inicial ya fue pagado en el registro de matrícula',
        )

    def test_eliminar_recuperacion_descontada_no_borra_abono(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_eliminar',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 5),
            monto=Decimal('50.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        recup = RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 10),
            saldo_pendiente_al_marcar=matricula.saldo,
            pagada=True,
            fecha_recuperacion=date(2026, 7, 10),
            abono=abono,
        )

        self.client.force_login(admin)
        response = self.client.post(
            reverse('academia:recuperacion_eliminar', kwargs={'recup_pk': recup.pk}),
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(RecuperacionPendiente.objects.filter(pk=recup.pk).exists())
        self.assertTrue(Abono.objects.filter(pk=abono.pk).exists())
        matricula.refresh_from_db()
        self.assertEqual(matricula.valor_pagado, Decimal('50.00'))

    def test_recuperacion_posterior_no_reutiliza_un_mismo_abono(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_modulos',
            password='clave12345',
        )
        curso = Curso.objects.create(
            nombre='Curso Tres Modulos',
            ofrece_presencial=True,
            valor_presencial=Decimal('90.00'),
            numero_modulos=3,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='presencial',
            descripcion='lun_mie_vie',
            fecha_inicio=date(2026, 7, 5),
            sede=self.sede,
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('90.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 5),
            monto=Decimal('90.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 8),
            saldo_pendiente_al_marcar=matricula.saldo,
            pagada=True,
            fecha_recuperacion=date(2026, 7, 8),
            abono=abono,
        )

        self.client.force_login(admin)
        response = self.client.post(
            reverse('academia:recuperacion_marcar', kwargs={'matricula_pk': matricula.pk}),
            {
                'numero_modulo': '3',
                'modo_registro': 'descontar_abono_modulo',
                'fecha_marcada': '2026-07-10',
                'observaciones': 'Faltó al módulo 3.',
            },
        )

        self.assertEqual(response.status_code, 302)
        recup_modulo_3 = RecuperacionPendiente.objects.get(
            matricula=matricula,
            numero_modulo=3,
        )
        self.assertFalse(recup_modulo_3.pagada)
        self.assertIsNone(recup_modulo_3.abono)
        self.assertEqual(Abono.objects.filter(matricula=matricula).count(), 1)

    def test_modo_descuento_manipulado_se_ignora_y_crea_recuperacion_pendiente(self):
        admin = User.objects.create_superuser(
            username='admin_recuperacion_acumulado',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('90.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 5),
            monto=Decimal('10.00'),
            tipo_pago='abono',
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 5),
            monto=Decimal('20.00'),
            tipo_pago='solo_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 5),
            monto=Decimal('40.00'),
            tipo_pago='abono',
            cuenta_para_saldo=True,
            metodo='transferencia',
            banco='guayaquil',
            registrado_por=admin,
        )

        self.client.force_login(admin)
        response = self.client.post(
            reverse('academia:recuperacion_marcar', kwargs={'matricula_pk': matricula.pk}),
            {
                'numero_modulo': '3',
                'modo_registro': 'descontar_abono_modulo',
                'fecha_marcada': '2026-07-10',
                'observaciones': 'Faltó al módulo 3.',
            },
        )

        self.assertEqual(response.status_code, 302)
        recuperacion = RecuperacionPendiente.objects.get(matricula=matricula)
        self.assertFalse(recuperacion.pagada)
        self.assertIsNone(recuperacion.abono)
        self.assertEqual(Abono.objects.filter(matricula=matricula).count(), 3)

    def test_reserva_modulo_reparte_monto_real_si_hay_varios_modulos(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('50.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        _registrar_pago_inicial(matricula, self.usuario, self._mat_form(modulos_a_pagar=2))

        matricula.refresh_from_db()
        montos = list(
            Abono.objects.filter(matricula=matricula)
            .order_by('numero_modulo')
            .values_list('monto', flat=True)
        )
        self.assertEqual(matricula.valor_pagado, Decimal('50.00'))
        self.assertEqual(montos, [Decimal('26.25'), Decimal('23.75')])

    def test_reserva_modulo_mixto_conserva_montos_por_metodo(self):
        curso = Curso.objects.create(
            nombre='Curso Mixto',
            ofrece_presencial=True,
            valor_presencial=Decimal('110.00'),
            numero_modulos=4,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='presencial',
            descripcion='lun_mie_vie',
            fecha_inicio=date(2026, 7, 8),
        )
        estudiante = Estudiante.objects.create(
            cedula='1207342717',
            nombres='Randy Mixto',
        )
        matricula = Matricula.objects.create(
            estudiante=estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 8),
            valor_curso=Decimal('110.00'),
            valor_pagado=Decimal('40.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        _registrar_pago_inicial(
            matricula,
            self.usuario,
            self._mat_form(
                tipo_cobro='mixto',
                metodo_pago='efectivo',
                banco='',
                monto_pago_1=Decimal('20.00'),
                metodo_pago_1='efectivo',
                banco_1='',
                monto_pago_2=Decimal('20.00'),
                metodo_pago_2='transferencia',
                banco_2='guayaquil',
            ),
        )

        abono = Abono.objects.get(matricula=matricula)
        self.assertEqual(abono.monto, Decimal('40.00'))
        self.assertEqual(abono.monto_2, Decimal('20.00'))
        self.assertEqual(abono.metodo, 'efectivo')
        self.assertEqual(abono.metodo_2, 'transferencia')

        request = RequestFactory().get('/pagos/hoja-recaudacion/', {
            'fecha': '2026-07-08',
            'curso': str(curso.id),
        })
        hojas, _filtros = _hojas_recaudacion_data(request)

        self.assertEqual(hojas[0]['total_efectivo'], Decimal('20.00'))
        self.assertEqual(hojas[0]['total_transferencia'], Decimal('20.00'))

    def test_hoja_recaudacion_separa_jornadas_del_mismo_curso(self):
        segunda_jornada = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='mar_jue',
            fecha_inicio=date(2026, 8, 15),
        )
        segundo_estudiante = Estudiante.objects.create(
            cedula='1207342799',
            nombres='Estudiante Segunda Jornada',
        )
        Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        Matricula.objects.create(
            estudiante=segundo_estudiante,
            curso=self.curso,
            jornada=segunda_jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        request = RequestFactory().get('/pagos/hoja-recaudacion/', {
            'fecha': '2026-07-10',
            'curso': str(self.curso.id),
        })
        hojas, _filtros = _hojas_recaudacion_data(request)

        self.assertEqual(len(hojas), 2)
        self.assertEqual([h['jornada_id'] for h in hojas], [
            self.jornada.pk,
            segunda_jornada.pk,
        ])
        self.assertEqual([len(h['items']) for h in hojas], [1, 1])

        request_jornada = RequestFactory().get('/pagos/hoja-recaudacion/', {
            'fecha': '2026-07-10',
            'curso': str(self.curso.id),
            'jornada': str(segunda_jornada.id),
        })
        hojas_jornada, _filtros = _hojas_recaudacion_data(request_jornada)

        self.assertEqual(len(hojas_jornada), 1)
        self.assertEqual(hojas_jornada[0]['jornada_id'], segunda_jornada.pk)
        self.assertEqual(
            hojas_jornada[0]['items'][0]['estudiante'],
            segundo_estudiante,
        )

    def test_hoja_recaudacion_fecha_usa_selector_visual(self):
        admin = User.objects.create_superuser(
            username='admin_fecha_hoja_recaudacion',
            password='clave12345',
        )
        self.client.force_login(admin)
        Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('10.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )

        response = self.client.get(
            reverse('academia:hoja_recaudacion'),
            {
                'fecha': '2026-07-30',
                'curso': str(self.curso.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'data-single-date-picker')
        self.assertContains(response, 'data-date-range-picker')
        self.assertContains(
            response,
            '<input type="hidden" name="fecha_desde" value="2026-07-30" data-date-start>',
        )
        self.assertContains(
            response,
            '<input type="hidden" name="fecha_hasta" value="2026-07-30" data-date-end>',
        )
        self.assertContains(response, '.hoja-totales { display: none !important; }')
        self.assertContains(
            response,
            '.hoja-tabla .col-cuota { display: table-cell !important; }',
        )
        self.assertContains(response, '.screen-only { display: none !important; }')
        self.assertContains(response, '.print-only { display: inline !important; }')
        self.assertContains(response, '<span class="print-only">Recaudar</span>')
        self.assertNotContains(response, 'Saldo módulo')
        self.assertContains(response, 'Módulo 1')
        self.assertContains(response, 'Exportar a Excel')
        self.assertNotContains(response, 'Exportar a PDF')
        self.assertNotContains(response, 'type="date" name="fecha"')
        self.assertContains(response, 'hoja-tabla-wrap responsive-table-shell')
        self.assertContains(response, 'data-table-navigation="true"')
        self.assertContains(response, 'data-table-identity-column="1"')
        self.assertContains(response, 'id="btn-editar-cuotas"')
        self.assertContains(response, 'Editar A Recaudar')
        self.assertContains(response, 'fp-table-identity-context')
        self.assertContains(response, '<strong>Inicio:</strong>')
        self.assertContains(response, '<strong>Módulo:</strong> Módulo 1')
        self.assertContains(response, 'min-width: 1560px;')
        self.assertContains(response, 'min-width: 0 !important;')
        self.assertContains(response, 'data-fecha="2026-07-30"')

    def test_hoja_recaudacion_rango_suma_pagos_del_periodo(self):
        admin = User.objects.create_superuser(
            username='admin_rango_hoja_recaudacion',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('35.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 29),
            monto=Decimal('5.00'),
            tipo_pago='abono',
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 30),
            monto=Decimal('10.00'),
            tipo_pago='abono',
            cuenta_para_saldo=True,
            metodo='efectivo',
            registrado_por=admin,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 1),
            monto=Decimal('20.00'),
            tipo_pago='abono',
            cuenta_para_saldo=True,
            metodo='transferencia',
            banco='guayaquil',
            registrado_por=admin,
        )

        request = RequestFactory().get('/pagos/hoja-recaudacion/', {
            'fecha_desde': '2026-07-30',
            'fecha_hasta': '2026-08-01',
            'curso': str(self.curso.id),
        })
        hojas, filtros = _hojas_recaudacion_data(request)

        self.assertTrue(filtros['es_rango'])
        self.assertEqual(filtros['fecha'], '')
        self.assertEqual(filtros['periodo_label'], '30/07/2026 - 01/08/2026')
        self.assertEqual(filtros['fecha_guardado'], '2026-08-01')
        self.assertEqual(hojas[0]['items'][0]['recaudado'], Decimal('30.00'))
        self.assertEqual(hojas[0]['total_efectivo'], Decimal('10.00'))
        self.assertEqual(hojas[0]['total_transferencia'], Decimal('20.00'))

        self.client.force_login(admin)
        response = self.client.get(
            reverse('academia:hoja_recaudacion'),
            {
                'fecha_desde': '2026-07-30',
                'fecha_hasta': '2026-08-01',
                'curso': str(self.curso.pk),
            },
        )
        self.assertContains(response, 'id="btn-editar-cuotas"')
        self.assertContains(response, 'Editar A Recaudar')
        self.assertContains(response, 'data-fecha="2026-08-01"')
        self.assertContains(response, 'para el final del período')

    def test_hoja_recaudacion_excel_incluye_resumen_y_formato_impresion(self):
        admin = User.objects.create_superuser(
            username='admin_excel_hoja_recaudacion',
            password='clave12345',
        )
        Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('15.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:hoja_recaudacion_export_excel'),
            {
                'fecha': '2026-07-30',
                'curso': str(self.curso.pk),
            },
        )

        from openpyxl import load_workbook

        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        encabezados = list(next(ws.iter_rows(
            min_row=3,
            max_row=3,
            values_only=True,
        )))
        primera_columna = [
            row[0] for row in ws.iter_rows(
                min_col=1,
                max_col=1,
                values_only=True,
            )
            if row[0] not in (None, '')
        ]
        self.assertNotIn('Saldo Pend.', encabezados)
        self.assertNotIn('A Recaudar (Cuota)', encabezados)
        self.assertIn('RECAUDAR', encabezados)
        self.assertIn('RECAUDADO', encabezados)
        self.assertIn(self.curso.nombre.upper(), ws.cell(row=1, column=2).value)
        self.assertIn('FECHA: JUEVES 30 DE JULIO', ws.cell(row=2, column=2).value)
        self.assertEqual(ws.cell(row=3, column=2).value, 'NOMBRE DEL ESTUDIANTE')
        self.assertEqual(ws.cell(row=3, column=10).value, 'RECUPERACIÓN')
        self.assertEqual(ws.cell(row=4, column=3).value, 1)
        self.assertEqual(ws.cell(row=4, column=4).value, 22)
        self.assertEqual(ws.column_dimensions['B'].width, 32)
        self.assertEqual(ws.cell(row=2, column=2).fill.fgColor.rgb, 'FF6D9EEB')
        self.assertEqual(ws.cell(row=2, column=9).fill.fgColor.rgb, 'FF00FF00')
        self.assertEqual(ws.cell(row=3, column=2).fill.fgColor.rgb, 'FFE06666')
        self.assertEqual(ws.row_dimensions[4].height, 45)
        self.assertNotIn('TOTAL', primera_columna)
        self.assertIn('JORNADA:', ws['G2'].value)
        self.assertEqual(ws['B15'].value, 'RECAUDACIÓN EN EFECTIVO')
        self.assertEqual(ws['B16'].value, 'RECAUDACIÓN EN TRANSFERENCIA')
        self.assertEqual(ws['B17'].value, 'RECAUDACIÓN EN PAYPHONE')
        self.assertEqual(ws['B18'].value, 'TOTAL RECAUDADO')
        self.assertEqual(ws['B19'].value, 'OBSERVACIONES ADICIONALES')
        self.assertEqual(ws.page_setup.orientation, 'landscape')
        self.assertEqual(ws.page_setup.fitToWidth, 1)
        self.assertIn('$A$1:$J$22', str(ws.print_area))

    def test_hoja_recaudacion_excel_usa_valor_manual_guardado_en_rango(self):
        admin = User.objects.create_superuser(
            username='admin_excel_manual_rango',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('15.00'),
            tipo_registro='central_ia',
            registrado_por=admin,
        )
        self.client.force_login(admin)

        response = self.client.post(
            reverse('academia:hoja_recaudacion_guardar_cuotas'),
            data=(
                '{"fecha":"2026-08-01","cuotas":['
                f'{{"matricula_id":{matricula.pk},"monto":"7.50"}}'
                ']}'
            ),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['ok'])
        self.assertTrue(
            CuotaManualRecaudacion.objects.filter(
                matricula=matricula,
                fecha=date(2026, 8, 1),
                monto=Decimal('7.50'),
            ).exists()
        )

        response = self.client.get(
            reverse('academia:hoja_recaudacion_export_excel'),
            {
                'fecha_desde': '2026-07-30',
                'fecha_hasta': '2026-08-01',
                'curso': str(self.curso.pk),
            },
        )

        from openpyxl import load_workbook

        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content), read_only=True)
        ws = wb.active
        encabezados = list(next(ws.iter_rows(
            min_row=3,
            max_row=3,
            values_only=True,
        )))
        col_recaudar = encabezados.index('RECAUDAR') + 1
        self.assertEqual(ws.cell(row=4, column=col_recaudar).value, 7.5)

    def test_editar_solo_datos_oculta_campos_de_pago(self):
        admin = User.objects.create_superuser(
            username='admin_edicion_datos',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('30.00'),
            tipo_registro='central_ia',
            factura_realizada='no',
            registrado_por=admin,
            vendedora=self.usuario,
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse(
                'academia:matricula_editar',
                kwargs={'modalidad': 'presencial', 'pk': matricula.pk},
            )
        )

        self.assertEqual(response.status_code, 200)
        html = response.content.decode('utf-8')
        self.assertIn('Seleccione la jornada', html)
        self.assertIn('¿Factura con datos?', html)
        self.assertIn('Selecciona el asesor', html)
        self.assertNotIn('Valor pagado (USD)', html)
        self.assertNotIn('Forma de pago *', html)
        self.assertNotIn('Distribución de pago', html)
        self.assertNotIn('Método de pago *', html)
        self.assertNotIn('Valor del curso (USD)', html)

    def test_editar_pago_inicial_mantiene_campos_de_pago_visibles(self):
        admin = User.objects.create_superuser(
            username='admin_edicion_pago',
            password='clave12345',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('30.00'),
            tipo_registro='central_ia',
            factura_realizada='no',
            registrado_por=admin,
            vendedora=self.usuario,
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse(
                'academia:matricula_editar',
                kwargs={'modalidad': 'presencial', 'pk': matricula.pk},
            ),
            {'editar_pago': '1'},
        )

        self.assertEqual(response.status_code, 200)
        html = response.content.decode('utf-8')
        self.assertIn('Valor pagado (USD)', html)
        self.assertIn('Forma de pago *', html)
        self.assertIn('Distribución de pago', html)

    def test_comprobante_usa_vendedora_de_matricula(self):
        registrador = User.objects.create_user(username='registrador')
        vendedora_1 = User.objects.create_user(
            username='asesora_uno',
            first_name='Asesora',
            last_name='Uno',
        )
        vendedora_2 = User.objects.create_user(
            username='asesora_dos',
            first_name='Asesora',
            last_name='Dos',
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('30.00'),
            tipo_registro='central_ia',
            factura_realizada='no',
            registrado_por=registrador,
            vendedora=vendedora_1,
        )

        comprobante = Comprobante.objects.get(matricula=matricula)
        self.assertEqual(comprobante.vendedora, vendedora_1)
        self.assertEqual(comprobante.vendedora_nombre, 'Asesora Uno')

        matricula.vendedora = vendedora_2
        matricula.save()
        comprobante.refresh_from_db()

        self.assertEqual(comprobante.vendedora, vendedora_2)
        self.assertEqual(comprobante.vendedora_nombre, 'Asesora Dos')

    def _abono_data(self, **overrides):
        data = {
            'fecha': '2026-07-06',
            'monto': '25.00',
            'tipo_pago': 'solo_modulo',
            'numero_modulo': '1',
            'cuenta_para_saldo': 'True',
            'metodo': 'efectivo',
            'banco': '',
            'numero_recibo': '',
            'observaciones': '',
            'tipo_cobro': 'mixto',
            'monto_pago_1': '10.00',
            'metodo_pago_1': 'efectivo',
            'banco_1': '',
            'monto_pago_2': '15.00',
            'metodo_pago_2': 'efectivo',
            'banco_2': '',
        }
        data.update(overrides)
        return data

    def test_abono_metodo_pago_muestra_seleccione_primero(self):
        form = AbonoForm(matricula=None)

        self.assertEqual(list(form.fields['metodo'].choices)[0], ('', 'Seleccione'))
        self.assertEqual(list(form.fields['metodo_pago_1'].choices)[0], ('', 'Seleccione'))
        self.assertEqual(list(form.fields['metodo_pago_2'].choices)[0], ('', 'Seleccione'))
        self.assertEqual(form['metodo'].value(), '')

    def test_registrar_pago_no_ofrece_abono_mas_modulo(self):
        form = AbonoForm(matricula=None)

        tipos = dict(form.fields['tipo_pago'].choices)

        self.assertNotIn('por_modulo', tipos)
        self.assertIn('abono', tipos)
        self.assertIn('solo_modulo', tipos)
        self.assertIn('pago_completo', tipos)
        self.assertIn('recuperacion', tipos)

    def test_registrar_pago_rechaza_abono_mas_modulo_manipulado(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('10.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        form = AbonoForm(
            self._abono_data(tipo_pago='por_modulo'),
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('tipo_pago', form.errors)

    def _crear_matricula_pago_unico_online(self):
        curso = Curso.objects.create(
            nombre='Ciclo Corto Online Pago Único',
            ofrece_online=True,
            valor_online=Decimal('35.00'),
            numero_modulos_online=2,
            es_ciclo_corto=True,
            pago_unico_online=True,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 8, 4),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 8, 1),
            valor_curso=Decimal('35.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 1),
            monto=Decimal('10.00'),
            tipo_pago='abono',
            metodo='tarjeta',
        )
        matricula.refresh_from_db()
        return matricula

    def test_pago_unico_online_mantiene_dos_modulos_pero_una_obligacion(self):
        matricula = self._crear_matricula_pago_unico_online()

        self.assertEqual(matricula.curso.get_numero_modulos('online'), 2)
        self.assertEqual(matricula.numero_cuotas_pago, 1)
        self.assertEqual(matricula.cuotas_modulos_objetivo(), [Decimal('25.00')])
        self.assertEqual(len(matricula.desglose_pagos_por_modulo()), 1)
        self.assertEqual(
            matricula.desglose_pagos_por_modulo()[0]['label'],
            'Un solo pago',
        )

        pago = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 2),
            monto=Decimal('25.00'),
            tipo_pago='solo_modulo',
            numero_modulo=1,
        )
        self.assertEqual(pago.get_modulo_display, 'Módulo 1')

    def test_formulario_pago_unico_online_no_permite_cuota_del_modulo_dos(self):
        matricula = self._crear_matricula_pago_unico_online()
        form = AbonoForm(
            self._abono_data(numero_modulo='2'),
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('numero_modulo', form.errors)
        self.assertIn('un solo pago', form.errors['numero_modulo'][0].lower())

        form_valido = AbonoForm(
            self._abono_data(numero_modulo='1'),
            matricula=matricula,
        )
        self.assertTrue(form_valido.is_valid(), form_valido.errors)
        self.assertEqual(
            dict(form_valido.fields['tipo_pago'].choices)['solo_modulo'],
            'Un solo pago',
        )

    def test_detalle_pago_unico_online_muestra_plan_y_selector_protegido(self):
        matricula = self._crear_matricula_pago_unico_online()
        admin = User.objects.create_superuser(
            username='admin_pago_unico_online',
            password='clave12345',
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:matricula_abonos', kwargs={'pk': matricula.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['pago_unico_online'])
        self.assertContains(response, 'Plan financiero: un solo pago de $25,00')
        self.assertContains(response, 'data-pago-unico-online="1"')
        self.assertContains(response, '>Un solo pago</option>')

    def test_registrar_pago_rechaza_modulo_ya_registrado_en_historial(self):
        matricula = self._crear_matricula_pago_unico_online()
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 3),
            monto=Decimal('20.00'),
            tipo_pago='solo_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
        )

        form = AbonoForm(
            self._abono_data(
                fecha='2026-08-10',
                monto='10.00',
                tipo_pago='recuperacion',
                numero_modulo='1',
                cuenta_para_saldo='False',
                fecha_marcada='2026-08-02',
                fecha_programada='2026-08-10',
                tipo_cobro='un_solo_metodo',
                metodo='efectivo',
                monto_pago_1='',
                metodo_pago_1='',
                monto_pago_2='',
                metodo_pago_2='',
            ),
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('numero_modulo', form.errors)
        self.assertIn(
            'ya se encuentra registrado',
            form.errors['numero_modulo'][0],
        )

    def test_registrar_pago_normal_bloquea_modulo_en_recuperacion_pendiente(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('10.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 12),
            saldo_pendiente_al_marcar=Decimal('105.00'),
        )

        form = AbonoForm(
            self._abono_data(tipo_pago='solo_modulo', numero_modulo='1'),
            matricula=matricula,
        )

        self.assertEqual(form.modulos_con_recuperacion_pendiente, [1])
        self.assertIn(
            (1, 'Módulo 1 - recuperación'),
            list(form.fields['numero_modulo'].widget.choices),
        )
        self.assertFalse(form.is_valid())
        self.assertIn('numero_modulo', form.errors)
        self.assertIn('recuperación', form.errors['numero_modulo'][0])

    def test_pago_recuperacion_permite_modulo_en_recuperacion_pendiente(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('10.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 12),
            saldo_pendiente_al_marcar=Decimal('105.00'),
        )

        form = AbonoForm(
            self._abono_data(
                tipo_pago='recuperacion',
                numero_modulo='1',
                fecha_marcada='2026-07-12',
                fecha_programada='2026-07-19',
            ),
            matricula=matricula,
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_detalle_pago_expone_modulos_registrados_para_el_selector(self):
        matricula = self._crear_matricula_pago_unico_online()
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 3),
            monto=Decimal('20.00'),
            tipo_pago='solo_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
        )
        admin = User.objects.create_superuser(
            username='admin_modulo_registrado_selector',
            password='clave12345',
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:matricula_abonos', kwargs={'pk': matricula.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['form'].modulos_con_pago_registrado,
            [1],
        )
        self.assertContains(response, 'modulos-registrados-pago')
        self.assertContains(response, 'aviso-modulo-registrado')

    def test_detalle_pago_expone_modulos_en_recuperacion_para_el_selector(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('10.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 12),
            saldo_pendiente_al_marcar=Decimal('105.00'),
        )
        admin = User.objects.create_superuser(
            username='admin_modulo_recuperacion_selector',
            password='clave12345',
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:matricula_abonos', kwargs={'pk': matricula.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.context['form'].modulos_con_recuperacion_pendiente,
            [1],
        )
        self.assertContains(response, 'modulos-recuperacion-pendiente')
        self.assertContains(response, 'Módulo 1 - recuperación')

    def test_editar_pago_conserva_su_modulo_actual(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('25.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 6),
            monto=Decimal('25.00'),
            tipo_pago='solo_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
        )

        form = AbonoForm(
            self._abono_data(tipo_pago='solo_modulo', numero_modulo='1'),
            instance=abono,
            matricula=matricula,
        )

        self.assertTrue(form.is_valid(), form.errors)

    def test_editar_pago_rechaza_modulo_de_otro_recibo(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('50.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 6),
            monto=Decimal('25.00'),
            tipo_pago='solo_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
        )
        abono_modulo_2 = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 13),
            monto=Decimal('25.00'),
            tipo_pago='solo_modulo',
            numero_modulo=2,
            cuenta_para_saldo=True,
            metodo='efectivo',
        )

        form = AbonoForm(
            self._abono_data(tipo_pago='solo_modulo', numero_modulo='1'),
            instance=abono_modulo_2,
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('numero_modulo', form.errors)
        self.assertIn(
            'ya se encuentra registrado',
            form.errors['numero_modulo'][0],
        )

    def test_pago_recuperacion_vincula_automaticamente_la_marca_pendiente(self):
        matricula = self._crear_matricula_pago_unico_online()
        recuperacion = RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 8, 2),
            fecha_programada=date(2026, 8, 10),
            saldo_pendiente_al_marcar=Decimal('25.00'),
        )
        admin = User.objects.create_superuser(
            username='admin_vincula_pago_recuperacion',
            password='clave12345',
        )
        self.client.force_login(admin)

        response = self.client.post(
            reverse(
                'academia:abono_crear',
                kwargs={'matricula_pk': matricula.pk},
            ),
            self._abono_data(
                fecha='2026-08-12',
                tipo_pago='recuperacion',
                numero_modulo='1',
                fecha_marcada='2026-08-03',
                fecha_programada='2026-08-11',
                tipo_cobro='un_solo_metodo',
                metodo='efectivo',
                monto_pago_1='',
                metodo_pago_1='',
                monto_pago_2='',
                metodo_pago_2='',
            ),
        )

        recuperacion.refresh_from_db()
        self.assertEqual(response.status_code, 302)
        self.assertTrue(recuperacion.pagada)
        self.assertEqual(recuperacion.fecha_marcada, date(2026, 8, 3))
        self.assertEqual(recuperacion.fecha_programada, date(2026, 8, 11))
        self.assertEqual(recuperacion.fecha_recuperacion, date(2026, 8, 12))
        self.assertIsNotNone(recuperacion.abono)
        self.assertEqual(recuperacion.abono.fecha, date(2026, 8, 12))
        self.assertEqual(recuperacion.abono.tipo_pago, 'recuperacion')
        self.assertEqual(
            RecuperacionPendiente.objects.filter(matricula=matricula).count(),
            1,
        )

    def test_pago_recuperacion_exige_fecha_de_la_falta(self):
        matricula = self._crear_matricula_pago_unico_online()
        form = AbonoForm(
            self._abono_data(
                tipo_pago='recuperacion',
                numero_modulo='1',
                fecha_marcada='',
                fecha_programada='2026-08-10',
            ),
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('fecha_marcada', form.errors)

    def test_pago_recuperacion_directo_guarda_las_tres_fechas(self):
        matricula = self._crear_matricula_pago_unico_online()
        admin = User.objects.create_superuser(
            username='admin_pago_recuperacion_directo',
            password='clave12345',
        )
        self.client.force_login(admin)

        response = self.client.post(
            reverse(
                'academia:abono_crear',
                kwargs={'matricula_pk': matricula.pk},
            ),
            self._abono_data(
                fecha='2026-08-12',
                tipo_pago='recuperacion',
                numero_modulo='1',
                cuenta_para_saldo='False',
                fecha_marcada='2026-08-03',
                fecha_programada='2026-08-11',
                tipo_cobro='un_solo_metodo',
                metodo='efectivo',
                monto_pago_1='',
                metodo_pago_1='',
                monto_pago_2='',
                metodo_pago_2='',
            ),
        )

        recuperacion = RecuperacionPendiente.objects.get(matricula=matricula)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(recuperacion.fecha_marcada, date(2026, 8, 3))
        self.assertEqual(recuperacion.fecha_programada, date(2026, 8, 11))
        self.assertEqual(recuperacion.fecha_recuperacion, date(2026, 8, 12))
        self.assertEqual(recuperacion.abono.fecha, date(2026, 8, 12))
        self.assertTrue(recuperacion.pagada)

    def test_pago_recuperacion_rechaza_fecha_programada_anterior(self):
        matricula = self._crear_matricula_pago_unico_online()
        form = AbonoForm(
            self._abono_data(
                tipo_pago='recuperacion',
                numero_modulo='1',
                fecha_marcada='2026-08-10',
                fecha_programada='2026-08-09',
            ),
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('fecha_programada', form.errors)

    def test_editar_pago_recuperacion_precarga_fechas_academicas(self):
        matricula = self._crear_matricula_pago_unico_online()
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 12),
            monto=Decimal('10.00'),
            tipo_pago='recuperacion',
            numero_modulo=1,
            cuenta_para_saldo=False,
            metodo='efectivo',
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 8, 3),
            fecha_programada=date(2026, 8, 11),
            fecha_recuperacion=date(2026, 8, 12),
            pagada=True,
            abono=abono,
        )

        form = AbonoForm(instance=abono, matricula=matricula)

        self.assertEqual(form['fecha_marcada'].value(), date(2026, 8, 3))
        self.assertEqual(form['fecha_programada'].value(), date(2026, 8, 11))

    def test_editar_pago_antiguo_conserva_abono_mas_modulo(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('40.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 6),
            monto=Decimal('25.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
            registrado_por=self.usuario,
        )

        form = AbonoForm(instance=abono, matricula=matricula)

        self.assertIn('por_modulo', dict(form.fields['tipo_pago'].choices))

    def test_abono_rechaza_metodo_pago_vacio_en_un_solo_metodo(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('40.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        form = AbonoForm(
            self._abono_data(
                tipo_cobro='un_solo_metodo',
                metodo='',
                monto_pago_1='',
                metodo_pago_1='',
                monto_pago_2='',
                metodo_pago_2='',
            ),
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('metodo', form.errors)

    def test_abono_mixto_rechaza_suma_distinta_al_monto_principal(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('40.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        form = AbonoForm(
            self._abono_data(
                monto='25.00',
                monto_pago_1='15.00',
                monto_pago_2='15.00',
            ),
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('monto_pago_2', form.errors)

    def test_abono_mixto_acepta_suma_igual_al_monto_principal(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('40.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        form = AbonoForm(self._abono_data(), matricula=matricula)

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['monto'], Decimal('25.00'))
        self.assertEqual(form.cleaned_data['monto_pago_1'], Decimal('10.00'))
        self.assertEqual(form.cleaned_data['monto_pago_2'], Decimal('15.00'))

    def _supletorio_data(self, **overrides):
        data = {
            'numero_modulo': '1',
            'fecha': '2026-07-06',
            'valor': '20.00',
            'metodo_pago': 'efectivo',
            'banco': '',
            'tipo_cobro': 'mixto',
            'monto_pago_1': '10.00',
            'metodo_pago_1': 'efectivo',
            'banco_1': '',
            'monto_pago_2': '10.00',
            'metodo_pago_2': 'efectivo',
            'banco_2': '',
            'numero_recibo': '',
            'observaciones': '',
        }
        data.update(overrides)
        return data

    def test_supletorio_rapido_mixto_rechaza_suma_distinta_al_valor(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('40.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        form = AdicionalSupletorioRapidoForm(
            self._supletorio_data(
                valor='20.00',
                monto_pago_1='15.00',
                monto_pago_2='10.00',
            ),
            matricula=matricula,
        )

        self.assertFalse(form.is_valid())
        self.assertIn('monto_pago_2', form.errors)

    def test_supletorio_rapido_mixto_acepta_suma_igual_al_valor(self):
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('115.00'),
            valor_pagado=Decimal('40.00'),
            tipo_registro='central_ia',
            registrado_por=self.usuario,
        )

        form = AdicionalSupletorioRapidoForm(self._supletorio_data(), matricula=matricula)

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.cleaned_data['tipo_cobro'], 'mixto')
        self.assertEqual(form.cleaned_data['monto_pago_1'], Decimal('10.00'))
        self.assertEqual(form.cleaned_data['monto_pago_2'], Decimal('10.00'))


class PagosPorModuloFiltroTests(TestCase):
    def setUp(self):
        self.curso = Curso.objects.create(
            nombre='Curso Filtro Modulos',
            ofrece_presencial=True,
            valor_presencial=Decimal('90.00'),
            numero_modulos=3,
        )
        self.jornada = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='lun_mie_vie',
            fecha_inicio=date(2026, 7, 5),
        )

    def _crear_matricula(self, cedula, nombres, modulo_pagado=None,
                         tipo_matricula='reserva_abono', jornada=None,
                         fecha_pago_modulo=None):
        estudiante = Estudiante.objects.create(
            cedula=cedula,
            nombres=nombres,
        )
        matricula = Matricula.objects.create(
            estudiante=estudiante,
            curso=self.curso,
            jornada=jornada or self.jornada,
            modalidad='presencial',
            tipo_matricula=tipo_matricula,
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 5),
            valor_curso=Decimal('90.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
        )
        if modulo_pagado:
            Abono.objects.create(
                matricula=matricula,
                fecha=fecha_pago_modulo or date(2026, 7, 5),
                monto=Decimal('30.00'),
                tipo_pago='por_modulo',
                numero_modulo=modulo_pagado,
                cuenta_para_saldo=True,
            )
            matricula.refresh_from_db()
        return matricula

    def test_matriz_online_pago_unico_no_genera_columna_modulo_dos(self):
        curso = Curso.objects.create(
            nombre='Curso Matriz Pago Único',
            ofrece_online=True,
            valor_online=Decimal('35.00'),
            numero_modulos_online=2,
            es_ciclo_corto=True,
            pago_unico_online=True,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 8, 4),
        )
        estudiante = Estudiante.objects.create(
            cedula='0999999999',
            nombres='Estudiante Pago Único',
        )
        matricula = Matricula.objects.create(
            estudiante=estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 8, 1),
            valor_curso=Decimal('35.00'),
        )

        filas, modulos, _resumen, visibles = _construir_matriz_pagos(
            curso,
            modalidad='online',
        )

        self.assertEqual(modulos, [1])
        self.assertEqual(visibles, [1])
        self.assertEqual(len(filas), 1)
        self.assertTrue(filas[0]['pago_unico_online'])
        self.assertEqual(
            [m['numero'] for m in filas[0]['modulos_data']],
            [1],
        )
        self.assertEqual(filas[0]['matricula'].pk, matricula.pk)

    def test_matriz_identifica_en_modulo_el_pago_de_recuperacion(self):
        matricula = self._crear_matricula(
            '0988888888', 'Estudiante Recuperación'
        )
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 12),
            monto=Decimal('30.00'),
            tipo_pago='recuperacion',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 7, 10),
            fecha_programada=date(2026, 7, 12),
            saldo_pendiente_al_marcar=Decimal('90.00'),
            pagada=True,
            fecha_recuperacion=date(2026, 7, 12),
            abono=abono,
        )

        filas, _modulos, _resumen, _visibles = _construir_matriz_pagos(
            self.curso,
        )
        modulo = filas[0]['modulos_data'][0]

        self.assertTrue(modulo['es_recuperacion'])
        self.assertEqual(modulo['recuperacion_monto'], Decimal('30.00'))
        self.assertEqual(modulo['recuperacion_recibos'], abono.numero_recibo)
        self.assertIn('Ya pagada', modulo['recuperacion_detalle'])
        self.assertIn('Falta 10/07/2026', modulo['recuperacion_detalle'])
        self.assertIn('Recuperar 12/07/2026', modulo['recuperacion_detalle'])
        self.assertIn('Recuperó 12/07/2026', modulo['recuperacion_detalle'])
        self.assertIn('$30.00', modulo['recuperacion_detalle'])

    def test_filtro_estado_modulo_recuperacion_muestra_recuperaciones(self):
        matricula_recuperacion = self._crear_matricula(
            '0987777777', 'Estudiante Con Recuperación'
        )
        matricula_normal = self._crear_matricula(
            '0986666666', 'Estudiante Sin Recuperación'
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula_recuperacion,
            numero_modulo=2,
            fecha_marcada=date(2026, 7, 9),
            saldo_pendiente_al_marcar=Decimal('90.00'),
        )

        matriculas, modulos, _resumen, modulos_visibles = _construir_matriz_pagos(
            self.curso,
            filtro_modulo_estado='2_Recuperacion',
        )

        self.assertEqual(modulos, [1, 2, 3])
        self.assertEqual(modulos_visibles, [2])
        self.assertEqual(
            [x['matricula'].pk for x in matriculas],
            [matricula_recuperacion.pk],
        )
        self.assertNotIn(
            matricula_normal.pk,
            [x['matricula'].pk for x in matriculas],
        )
        modulo = matriculas[0]['modulos_visibles_data'][0]
        self.assertEqual(modulo['numero'], 2)
        self.assertEqual(modulo['estado'], 'Pendiente')
        self.assertTrue(modulo['es_recuperacion'])
        self.assertIn('Pendiente', modulo['recuperacion_detalle'])
        self.assertIn('Falta 09/07/2026', modulo['recuperacion_detalle'])
        self.assertIn('Saldo $90.00', modulo['recuperacion_detalle'])

    def test_filtro_fecha_modulo_incluye_recuperacion_programada_pendiente(self):
        matricula_dentro = self._crear_matricula(
            '0981234567', 'Estudiante Recuperación Programada'
        )
        matricula_fuera = self._crear_matricula(
            '0987654321', 'Estudiante Recuperación Fuera'
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula_dentro,
            numero_modulo=2,
            fecha_marcada=date(2026, 7, 9),
            fecha_programada=date(2026, 7, 14),
            saldo_pendiente_al_marcar=Decimal('90.00'),
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula_fuera,
            numero_modulo=2,
            fecha_marcada=date(2026, 7, 9),
            fecha_programada=date(2026, 7, 20),
            saldo_pendiente_al_marcar=Decimal('90.00'),
        )

        matriculas, _modulos, _resumen, modulos_visibles = _construir_matriz_pagos(
            self.curso,
            filtro_modulo_estado='2_Recuperacion',
            fecha_modulo_desde=date(2026, 7, 14),
            fecha_modulo_hasta=date(2026, 7, 14),
        )

        self.assertEqual(modulos_visibles, [2])
        self.assertEqual(
            [x['matricula'].pk for x in matriculas],
            [matricula_dentro.pk],
        )
        modulo = matriculas[0]['modulos_visibles_data'][0]
        self.assertTrue(modulo['coincide_fecha_modulo'])
        self.assertIn(date(2026, 7, 14), modulo['recuperacion_fechas'])

    def test_pagos_por_modulo_excel_exporta_detalle_de_recuperacion(self):
        admin = User.objects.create_superuser(
            username='admin_excel_modulo_recuperacion',
            password='clave12345',
        )
        matricula = self._crear_matricula(
            '0981212121', 'Estudiante Excel Recuperación'
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=2,
            fecha_marcada=date(2026, 7, 9),
            fecha_programada=date(2026, 7, 14),
            saldo_pendiente_al_marcar=Decimal('90.00'),
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:pagos_por_modulo_export_excel'),
            {
                'curso': str(self.curso.pk),
                'filtro_modulo_estado': '2_Recuperacion',
                'fecha_modulo_desde': '2026-07-14',
                'fecha_modulo_hasta': '2026-07-14',
            },
        )

        from openpyxl import load_workbook

        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content), read_only=True)
        ws = wb.active
        modulo_cell = ws.cell(row=3, column=12).value
        self.assertIn('Recuperación', modulo_cell)
        self.assertIn('Falta 09/07/2026', modulo_cell)
        self.assertIn('Recuperar 14/07/2026', modulo_cell)
        self.assertIn('Saldo $90.00', modulo_cell)

    def test_filtro_fecha_modulo_filtra_por_rango_de_fecha_pago(self):
        matricula_antes = self._crear_matricula(
            '0985555555', 'Estudiante Antes',
            modulo_pagado=1,
            fecha_pago_modulo=date(2026, 7, 8),
        )
        matricula_dentro = self._crear_matricula(
            '0984444444', 'Estudiante Dentro',
            modulo_pagado=1,
            fecha_pago_modulo=date(2026, 7, 12),
        )
        matricula_otro_modulo = self._crear_matricula(
            '0983333333', 'Estudiante Otro Modulo',
            modulo_pagado=2,
            fecha_pago_modulo=date(2026, 7, 14),
        )
        matricula_despues = self._crear_matricula(
            '0982222222', 'Estudiante Despues',
            modulo_pagado=1,
            fecha_pago_modulo=date(2026, 7, 22),
        )

        matriculas, _modulos, _resumen, modulos_visibles = _construir_matriz_pagos(
            self.curso,
            fecha_modulo_desde=date(2026, 7, 10),
            fecha_modulo_hasta=date(2026, 7, 15),
        )

        self.assertEqual(modulos_visibles, [1, 2])
        self.assertCountEqual(
            [x['matricula'].pk for x in matriculas],
            [matricula_dentro.pk, matricula_otro_modulo.pk],
        )
        self.assertNotIn(
            matricula_antes.pk,
            [x['matricula'].pk for x in matriculas],
        )
        self.assertNotIn(
            matricula_despues.pk,
            [x['matricula'].pk for x in matriculas],
        )
        modulos_por_matricula = {
            x['matricula'].pk: [
                mod['numero'] for mod in x['modulos_visibles_data']
                if mod.get('coincide_fecha_modulo')
            ]
            for x in matriculas
        }
        self.assertEqual(modulos_por_matricula[matricula_dentro.pk], [1])
        self.assertEqual(modulos_por_matricula[matricula_otro_modulo.pk], [2])

        matriculas, _modulos, _resumen, modulos_visibles = _construir_matriz_pagos(
            self.curso,
            filtro_modulo_estado='1_Pagado',
            fecha_modulo_desde=date(2026, 7, 10),
            fecha_modulo_hasta=date(2026, 7, 15),
        )

        self.assertEqual(modulos_visibles, [1])
        self.assertEqual(
            [x['matricula'].pk for x in matriculas],
            [matricula_dentro.pk],
        )

    def test_filtro_estado_modulo_muestra_solo_el_modulo_filtrado(self):
        matricula_modulo_1 = self._crear_matricula(
            '0911111111', 'Estudiante Modulo 1', modulo_pagado=1
        )
        matricula_modulo_2 = self._crear_matricula(
            '0922222222', 'Estudiante Modulo 2', modulo_pagado=2
        )

        matriculas, modulos, _resumen, modulos_visibles = _construir_matriz_pagos(
            self.curso,
            filtro_modulo_estado='1_Pagado',
        )

        self.assertEqual(modulos, [1, 2, 3])
        self.assertEqual(modulos_visibles, [1])
        self.assertEqual(
            [x['matricula'].pk for x in matriculas],
            [matricula_modulo_1.pk],
        )
        self.assertEqual(
            [mod['numero'] for mod in matriculas[0]['modulos_visibles_data']],
            [1],
        )

        matriculas, _modulos, _resumen, modulos_visibles = _construir_matriz_pagos(
            self.curso,
            filtro_modulo_estado='1_Pendiente',
        )

        self.assertEqual(modulos_visibles, [1])
        self.assertEqual(
            [x['matricula'].pk for x in matriculas],
            [matricula_modulo_2.pk],
        )
        self.assertEqual(
            [mod['numero'] for mod in matriculas[0]['modulos_visibles_data']],
            [1],
        )

    def test_lista_pagos_filtra_por_curso_modulo_y_estado(self):
        matricula_pagada = self._crear_matricula(
            '0955555555', 'Estudiante Modulo Pagado', modulo_pagado=1
        )
        matricula_pendiente = self._crear_matricula(
            '0966666666', 'Estudiante Modulo Pendiente', modulo_pagado=2
        )
        admin = User.objects.create_superuser(
            username='admin_lista_modulos',
            password='clave12345',
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:pagos_lista'),
            {
                'modulo_curso': str(self.curso.pk),
                'modulo_numero': '1',
                'modulo_estado': 'pagado',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context['modulo_filtro_activo'])
        self.assertEqual(
            [matricula.pk for matricula in response.context['matriculas']],
            [matricula_pagada.pk],
        )
        self.assertContains(response, 'Filtro separado por módulo de pago')
        self.assertContains(response, 'Módulo 1 · Pagado')
        self.assertNotContains(response, 'Estudiante Modulo Pendiente')

        response = self.client.get(
            reverse('academia:pagos_lista'),
            {
                'modulo_curso': str(self.curso.pk),
                'modulo_numero': '1',
                'modulo_estado': 'pendiente',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [matricula.pk for matricula in response.context['matriculas']],
            [matricula_pendiente.pk],
        )
        self.assertContains(response, 'Pendiente de pago')
        self.assertContains(response, 'Módulo 1 · Pendiente')
        self.assertNotContains(response, 'Estudiante Modulo Pagado')

    def test_filtro_lista_usa_solo_modalidades_y_modulos_del_curso(self):
        self.curso.ofrece_online = True
        self.curso.numero_modulos_online = 2
        self.curso.nombrar_modulos = True
        self.curso.nombres_modulos = {
            'presencial': ['Presencial 1', 'Presencial 2', 'Presencial 3'],
            'online': ['Virtual 1', 'Virtual 2'],
        }
        self.curso.save()
        self._crear_matricula(
            '0910101010',
            'Estudiante Presencial Modalidad',
        )
        jornada_online = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='online',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 7, 6),
            ciudad='Zoom',
        )
        estudiante_online = Estudiante.objects.create(
            cedula='0920202020',
            nombres='Estudiante Virtual Modalidad',
        )
        matricula_online = Matricula.objects.create(
            estudiante=estudiante_online,
            curso=self.curso,
            jornada=jornada_online,
            modalidad='online',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 6),
            valor_curso=Decimal('60.00'),
            tipo_registro='central_ia',
        )
        admin = User.objects.create_superuser(
            username='admin_modalidad_modulos',
            password='clave12345',
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:pagos_lista'),
            {
                'modulo_curso': str(self.curso.pk),
                'modulo_modalidad': 'online',
                'modulo_numero': '2',
                'modulo_estado': 'pendiente',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [matricula.pk for matricula in response.context['matriculas']],
            [matricula_online.pk],
        )
        payload = response.context['cursos_modulo_filtro'][str(self.curso.pk)]
        virtual = next(
            modalidad for modalidad in payload['modalidades']
            if modalidad['value'] == 'online'
        )
        self.assertEqual(
            [modulo['label'] for modulo in virtual['modulos']],
            ['Módulo 1 - Virtual 1', 'Módulo 2 - Virtual 2'],
        )
        self.assertEqual(virtual['campus'], [])
        self.assertContains(response, 'Virtual / Online')
        self.assertContains(response, 'La modalidad virtual no requiere campus.')

    def test_filtro_lista_limita_por_campus_y_limpia_solo_el_bloque_modulo(self):
        sede_guayaquil = Sede.objects.create(nombre='Guayaquil', pais='Ecuador')
        sede_quito = Sede.objects.create(nombre='Quito', pais='Ecuador')
        self.jornada.sede = sede_guayaquil
        self.jornada.save()
        matricula_guayaquil = self._crear_matricula(
            '0930303030',
            'Estudiante Campus Guayaquil',
        )
        jornada_quito = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 7, 7),
            sede=sede_quito,
        )
        self._crear_matricula(
            '0940404040',
            'Estudiante Campus Quito',
            jornada=jornada_quito,
        )
        admin = User.objects.create_superuser(
            username='admin_campus_modulos',
            password='clave12345',
        )
        self.client.force_login(admin)

        response = self.client.get(
            reverse('academia:pagos_lista'),
            {
                'q': 'Campus',
                'modulo_curso': str(self.curso.pk),
                'modulo_modalidad': 'presencial',
                'modulo_campus': f'sede:{sede_guayaquil.pk}',
                'modulo_numero': '1',
                'modulo_estado': 'pendiente',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [matricula.pk for matricula in response.context['matriculas']],
            [matricula_guayaquil.pk],
        )
        self.assertEqual(response.context['filtros_query_sin_modulo'], 'q=Campus')
        self.assertContains(response, 'Campus: Guayaquil')
        self.assertContains(response, 'Limpiar filtro')
        self.assertContains(response, '?q=Campus')
        self.assertNotContains(response, 'Estudiante Campus Quito')

    def test_vista_pagos_por_modulo_renderiza_solo_columna_filtrada(self):
        self._crear_matricula(
            '0933333333', 'Estudiante Vista Modulo 1', modulo_pagado=1
        )
        self._crear_matricula(
            '0944444444', 'Estudiante Vista Modulo 2', modulo_pagado=2
        )
        admin = User.objects.create_superuser(
            username='admin_filtro_modulos',
            password='clave12345',
        )

        self.client.force_login(admin)
        response = self.client.get(
            reverse('academia:pagos_por_modulo'),
            {
                'curso': str(self.curso.pk),
                'filtro_modulo_estado': '1_Pagado',
                'fecha_modulo_desde': '2026-07-05',
                'fecha_modulo_hasta': '2026-07-05',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['modulos'], [1, 2, 3])
        self.assertEqual(response.context['modulos_visibles'], [1])
        self.assertContains(response, 'Módulo 1 — Recuperación')
        self.assertContains(response, 'name="fecha_modulo_desde"')
        self.assertEqual(response.context['filtros']['fecha_modulo_label'], '05/07/2026')
        self.assertIn('fecha_modulo_desde=2026-07-05', response.context['export_querystring'])
        self.assertIn('fecha_modulo_hasta=2026-07-05', response.context['export_querystring'])
        self.assertEqual(len(response.context['matriculas_data']), 1)
        self.assertEqual(
            [
                mod['numero']
                for mod in response.context['matriculas_data'][0]['modulos_visibles_data']
            ],
            [1],
        )

    def test_filtro_tipo_matricula_no_muestra_reserva_modulo_1(self):
        self._crear_matricula(
            '0955555555', 'Estudiante Reserva Abono', tipo_matricula='reserva_abono'
        )
        self._crear_matricula(
            '0966666666', 'Estudiante Reserva Modulo Uno',
            tipo_matricula='reserva_modulo_1'
        )
        admin = User.objects.create_superuser(
            username='admin_filtro_tipo_matricula',
            password='clave12345',
        )

        self.client.force_login(admin)
        response = self.client.get(
            reverse('academia:pagos_por_modulo'),
            {
                'curso': str(self.curso.pk),
                'tipo_matricula': 'reserva_modulo_1',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotIn(
            ('reserva_modulo_1', 'Reserva + Módulo 1'),
            response.context['tipos_matricula'],
        )
        self.assertEqual(response.context['filtros']['tipo_matricula'], '')
        self.assertEqual(len(response.context['matriculas_data']), 2)

    def test_filtro_matricula_muestra_estudiante_con_jornada_y_filtra(self):
        jornada_dos = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='mar_jue',
            fecha_inicio=date(2026, 7, 12),
            ciudad='Guayaquil',
        )
        matricula_uno = self._crear_matricula(
            '0977777777', 'Ana Pagos Modulo'
        )
        matricula_dos = self._crear_matricula(
            '0988888888', 'Bruno Pagos Modulo', jornada=jornada_dos
        )
        admin = User.objects.create_superuser(
            username='admin_filtro_matricula',
            password='clave12345',
        )

        self.client.force_login(admin)
        response = self.client.get(
            reverse('academia:pagos_por_modulo'),
            {'curso': str(self.curso.pk)},
        )

        self.assertEqual(response.status_code, 200)
        opciones = response.context['estudiantes_jornada']
        self.assertEqual(len(opciones), 2)
        labels = [opcion['label'] for opcion in opciones]
        self.assertTrue(
            any('Ana Pagos Modulo' in label and 'Lun, Mié, Vie.' in label for label in labels)
        )
        self.assertTrue(
            any('Bruno Pagos Modulo' in label and 'Martes y Jueves' in label for label in labels)
        )

        response = self.client.get(
            reverse('academia:pagos_por_modulo'),
            {
                'curso': str(self.curso.pk),
                'matricula': str(matricula_dos.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['filtros']['matricula'], str(matricula_dos.pk))
        self.assertIn(f'matricula={matricula_dos.pk}', response.context['export_querystring'])
        self.assertEqual(
            [x['matricula'].pk for x in response.context['matriculas_data']],
            [matricula_dos.pk],
        )
        self.assertNotIn(
            matricula_uno.pk,
            [x['matricula'].pk for x in response.context['matriculas_data']],
        )


class PlanRecaudacionTests(TestCase):
    def test_sombreado_respeta_pagos_previos_y_vencimiento_siguiente(self):
        from .views_pagos import _construir_hoja_recaudacion
        for modalidad, es_corto, quincenal, proximo in [
            ('presencial', False, False, date(2026, 7, 15)),
            ('presencial', False, True, date(2026, 7, 22)),
            ('online', False, False, date(2026, 7, 15)),
            ('online', True, False, date(2026, 7, 15)),
            ('online', False, True, date(2026, 7, 21)),
        ]:
            m = self._matricula_con_adelanto(
                Decimal('110'), Decimal('35'), 4,
                modalidad=modalidad, es_ciclo_corto=es_corto,
            )
            m.curso.pagos_cada_dos_semanas = quincenal
            m.curso.save()
            for fecha, esperado in [(date(2026, 7, 9), False),
                                    (proximo - timedelta(days=1), False), (proximo, True),
                                    (proximo + timedelta(days=1), True)]:
                with self.subTest(modalidad=modalidad, corto=es_corto, quincenal=quincenal, fecha=fecha):
                    hoja = _construir_hoja_recaudacion(m.curso, [m], fecha)
                    self.assertEqual(hoja['items'][0]['pago_pendiente_fecha'], esperado)
                    alertas = _calcular_alertas_pago(fecha_actual=fecha, excluir_revisadas=False)
                    self.assertEqual(
                        hoja['items'][0]['pago_pendiente_fecha'],
                        any(a['matricula'].pk == m.pk for a in alertas),
                    )

    def test_sombreado_pago_parcial_y_curso_cancelado(self):
        from .views_pagos import _construir_hoja_recaudacion
        m = self._matricula_con_adelanto(Decimal('110'), Decimal('20'), 4)
        hoja = _construir_hoja_recaudacion(m.curso, [m], date(2026, 7, 8))
        self.assertTrue(hoja['items'][0]['pago_pendiente_fecha'])
        self.assertGreater(hoja['items'][0]['recaudado'], 0)
        CuotaManualRecaudacion.objects.create(matricula=m, fecha=date(2026, 7, 9), monto=0)
        hoja = _construir_hoja_recaudacion(m.curso, [m], date(2026, 7, 9))
        self.assertTrue(hoja['items'][0]['pago_pendiente_fecha'])
        pagado = self._matricula_con_adelanto(
            Decimal('35'), Decimal('35'), 2, modalidad='online',
            es_ciclo_corto=True, pago_unico_online=True,
        )
        hoja = _construir_hoja_recaudacion(pagado.curso, [pagado], date(2026, 8, 1))
        self.assertFalse(hoja['items'][0]['pago_pendiente_fecha'])

    def setUp(self):
        self._seq = 0

    def _matricula_con_adelanto(
        self, valor_curso, adelanto, semanas, modalidad='presencial',
        es_ciclo_corto=False, pago_unico_online=False,
    ):
        self._seq += 1
        curso = Curso.objects.create(
            nombre=f'Curso Recaudación Test {self._seq}',
            ofrece_presencial=True,
            ofrece_online=True,
            valor_presencial=valor_curso,
            valor_online=valor_curso,
            numero_modulos=semanas,
            numero_modulos_online=semanas,
            es_ciclo_corto=es_ciclo_corto,
            pago_unico_online=pago_unico_online,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad=modalidad,
            descripcion='domingos_intensivos' if modalidad == 'presencial' else 'mar_mie_jue',
            fecha_inicio=date(2026, 7, 8),
        )
        estudiante = Estudiante.objects.create(
            cedula=f'09{self._seq:08d}',
            nombres=f'Estudiante Recaudación {self._seq}',
        )
        matricula = Matricula.objects.create(
            estudiante=estudiante,
            curso=curso,
            jornada=jornada,
            modalidad=modalidad,
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 8),
            valor_curso=valor_curso,
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 8),
            monto=adelanto,
            tipo_pago='abono',
        )
        matricula.refresh_from_db()
        return matricula

    def _assert_plan(
        self, matricula, saldo_esperado, cuotas_esperadas,
        modulo_esperado=1, fecha_recaudacion=None,
    ):
        plan = _plan_recaudacion_matricula(matricula, fecha_recaudacion)
        self.assertEqual(plan['saldo_pendiente'], saldo_esperado)
        self.assertEqual(plan['cuotas'], cuotas_esperadas)
        self.assertEqual(sum(plan['cuotas'], Decimal('0.00')), saldo_esperado)
        self.assertEqual(plan['cuota_sugerida'], cuotas_esperadas[0])
        self.assertEqual(plan['saldo_modulo'], cuotas_esperadas[0])
        self.assertEqual(plan['modulo'], modulo_esperado)

    def test_presencial_90_reserva_10_en_4_semanas(self):
        matricula = self._matricula_con_adelanto(
            Decimal('90.00'), Decimal('10.00'), 4
        )

        self._assert_plan(matricula, Decimal('80.00'), [
            Decimal('20.00'), Decimal('20.00'),
            Decimal('20.00'), Decimal('20.00'),
        ])

    def test_virtual_60_reserva_10_en_4_semanas(self):
        matricula = self._matricula_con_adelanto(
            Decimal('60.00'), Decimal('10.00'), 4, modalidad='online'
        )

        self._assert_plan(matricula, Decimal('50.00'), [
            Decimal('13.00'), Decimal('13.00'),
            Decimal('12.00'), Decimal('12.00'),
        ])

    def test_online_60_reserva_10_en_2_modulos(self):
        matricula = self._matricula_con_adelanto(
            Decimal('60.00'), Decimal('10.00'), 2, modalidad='online'
        )

        self._assert_plan(matricula, Decimal('50.00'), [
            Decimal('25.00'), Decimal('25.00'),
        ])

    def test_curso_35_reserva_10_en_2_modulos_sin_centavos(self):
        matricula = self._matricula_con_adelanto(
            Decimal('35.00'), Decimal('10.00'), 2
        )

        self._assert_plan(matricula, Decimal('25.00'), [
            Decimal('13.00'), Decimal('12.00'),
        ])

    def test_ciclo_corto_online_35_reserva_10_en_un_solo_pago(self):
        matricula = self._matricula_con_adelanto(
            Decimal('35.00'), Decimal('10.00'), 2,
            modalidad='online', es_ciclo_corto=True,
            pago_unico_online=True,
        )

        self._assert_plan(
            matricula, Decimal('25.00'), [Decimal('25.00')]
        )
        self.assertEqual(
            matricula.curso.get_numero_modulos('online'),
            2,
        )

        request = RequestFactory().get('/pagos/hoja-recaudacion/', {
            'fecha': '2026-08-03',
            'curso': str(matricula.curso_id),
            'modalidad': 'online',
        })
        hojas, _filtros = _hojas_recaudacion_data(request)

        self.assertEqual(
            hojas[0]['items'][0]['cuota_sugerida'],
            Decimal('25.00'),
        )
        self.assertEqual(hojas[0]['items'][0]['modulo_label'], 'Módulo 1')
        self.assertEqual(hojas[0]['total_cuotas'], Decimal('25.00'))

    def test_hoja_recaudacion_identifica_pago_de_recuperacion(self):
        matricula = self._matricula_con_adelanto(
            Decimal('35.00'), Decimal('10.00'), 2,
            modalidad='online', es_ciclo_corto=True,
            pago_unico_online=True,
        )
        recuperacion = RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 8, 2),
            fecha_programada=date(2026, 8, 10),
            saldo_pendiente_al_marcar=Decimal('25.00'),
        )
        abono = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 2),
            monto=Decimal('25.00'),
            tipo_pago='recuperacion',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
        )
        recuperacion.pagada = True
        recuperacion.fecha_recuperacion = abono.fecha
        recuperacion.abono = abono
        recuperacion.save()

        request = RequestFactory().get('/pagos/hoja-recaudacion/', {
            'fecha': '2026-08-02',
            'curso': str(matricula.curso_id),
            'modalidad': 'online',
        })
        hojas, _filtros = _hojas_recaudacion_data(request)

        item = hojas[0]['items'][0]
        self.assertEqual(item['recaudado'], Decimal('25.00'))
        self.assertIn('Ya pagada', item['recuperacion'])
        self.assertIn('Módulo 1', item['recuperacion'])
        self.assertIn('Falta 02/08/2026', item['recuperacion'])
        self.assertIn('Recuperar 10/08/2026', item['recuperacion'])
        self.assertIn('Recuperó 02/08/2026', item['recuperacion'])
        self.assertIn('$25.00', item['recuperacion'])

    def test_hoja_recuperacion_pagada_muestra_cero_y_queda_al_final(self):
        from .views_pagos import _construir_hoja_recaudacion

        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('10.00'), 4,
            modalidad='presencial',
        )
        recuperacion = RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 8, 30),
            fecha_programada=date(2026, 9, 6),
            saldo_pendiente_al_marcar=Decimal('100.00'),
        )
        pago = Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 9, 6),
            monto=Decimal('25.00'),
            tipo_pago='recuperacion',
            numero_modulo=1,
            cuenta_para_saldo=True,
            metodo='efectivo',
        )
        recuperacion.pagada = True
        recuperacion.fecha_recuperacion = pago.fecha
        recuperacion.abono = pago
        recuperacion.save()

        otro_estudiante = Estudiante.objects.create(
            cedula='0999999991', nombres='Zeta Sin Recuperación',
        )
        otra_matricula = Matricula.objects.create(
            estudiante=otro_estudiante,
            curso=matricula.curso,
            jornada=matricula.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 8),
            valor_curso=Decimal('110.00'),
            valor_pagado=Decimal('0.00'),
            tipo_registro='central_ia',
        )
        Abono.objects.create(
            matricula=otra_matricula,
            fecha=date(2026, 7, 8),
            monto=Decimal('10.00'),
            tipo_pago='abono',
        )

        hoja = _construir_hoja_recaudacion(
            matricula.curso,
            [matricula, otra_matricula],
            date(2026, 9, 6),
            jornada=matricula.jornada,
        )

        self.assertEqual(
            [item['matricula_id'] for item in hoja['items']],
            [otra_matricula.pk, matricula.pk],
        )
        item = hoja['items'][-1]
        self.assertEqual(item['modulo'], 1)
        self.assertEqual(item['cuota_sugerida'], Decimal('0.00'))
        self.assertEqual(item['recaudado'], Decimal('25.00'))
        self.assertFalse(item['pago_pendiente_fecha'])
        self.assertIn('Ya pagada', item['recuperacion'])

    def test_hoja_recuperacion_pendiente_cobra_la_clase_y_muestra_debe(self):
        from .views_pagos import _construir_hoja_recaudacion

        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('10.00'), 4,
            modalidad='presencial',
        )
        RecuperacionPendiente.objects.create(
            matricula=matricula,
            numero_modulo=1,
            fecha_marcada=date(2026, 8, 30),
            fecha_programada=date(2026, 9, 6),
            saldo_pendiente_al_marcar=Decimal('100.00'),
        )

        hoja = _construir_hoja_recaudacion(
            matricula.curso,
            [matricula],
            date(2026, 9, 6),
            jornada=matricula.jornada,
        )
        item = hoja['items'][0]

        self.assertEqual(item['modulo'], 1)
        self.assertEqual(item['cuota_sugerida'], Decimal('25.00'))
        self.assertEqual(item['saldo_modulo'], Decimal('25.00'))
        self.assertEqual(item['recaudado'], Decimal('0.00'))
        self.assertTrue(item['pago_pendiente_fecha'])
        self.assertIn('Pendiente', item['recuperacion'])
        self.assertIn('Debe $25.00', item['recuperacion'])
        self.assertNotIn('Debe $100.00', item['recuperacion'])

    def test_hoja_recaudacion_usa_nombre_de_modulo_personalizado(self):
        matricula = self._matricula_con_adelanto(
            Decimal('35.00'), Decimal('10.00'), 2,
            modalidad='online', es_ciclo_corto=True,
            pago_unico_online=True,
        )
        matricula.curso.nombrar_modulos = True
        matricula.curso.nombres_modulos = {
            'online': ['Tributación inicial', 'Declaraciones'],
        }
        matricula.curso.save(update_fields=['nombrar_modulos', 'nombres_modulos'])

        request = RequestFactory().get('/pagos/hoja-recaudacion/', {
            'fecha': '2026-08-03',
            'curso': str(matricula.curso_id),
            'modalidad': 'online',
        })
        hojas, _filtros = _hojas_recaudacion_data(request)

        self.assertEqual(
            hojas[0]['items'][0]['modulo_label'],
            'Módulo 1 - Tributación inicial',
        )

    def test_pago_unico_online_no_cambia_las_dos_cuotas_presenciales(self):
        matricula = self._matricula_con_adelanto(
            Decimal('35.00'), Decimal('10.00'), 2,
            modalidad='presencial', es_ciclo_corto=True,
            pago_unico_online=True,
        )

        self._assert_plan(matricula, Decimal('25.00'), [
            Decimal('13.00'), Decimal('12.00'),
        ])

    def test_formulario_exige_ciclo_corto_para_pago_unico_online(self):
        datos = {
            'nombre': 'Curso Pago Único Formulario',
            'ofrece_online': 'on',
            'valor_presencial': '0.00',
            'valor_online': '35.00',
            'duracion': '2 semanas',
            'numero_modulos': '2',
            'numero_modulos_online': '2',
            'pago_unico_online': 'on',
            'activo': 'on',
        }

        form = CursoForm(data=datos)

        self.assertFalse(form.is_valid())
        self.assertIn('pago_unico_online', form.errors)

        datos['es_ciclo_corto'] = 'on'
        form = CursoForm(data=datos)

        self.assertTrue(form.is_valid(), form.errors)

    def test_pagos_cada_dos_semanas_es_independiente_del_ciclo_corto(self):
        datos = {
            'nombre': 'Curso Pagos Cada Dos Semanas',
            'ofrece_online': 'on',
            'valor_presencial': '0.00',
            'valor_online': '90.00',
            'duracion': '4 semanas',
            'numero_modulos': '4',
            'numero_modulos_online': '2',
            'pagos_cada_dos_semanas': 'on',
            'activo': 'on',
        }

        form = CursoForm(data=datos)

        self.assertTrue(form.is_valid(), form.errors)
        curso = form.save()
        self.assertTrue(curso.pagos_cada_dos_semanas)
        self.assertFalse(curso.es_ciclo_corto)

    def test_reserva_no_cubre_modulos_en_curso_economico(self):
        matricula = self._matricula_con_adelanto(
            Decimal('35.00'), Decimal('10.00'), 4
        )

        self._assert_plan(matricula, Decimal('25.00'), [
            Decimal('7.00'), Decimal('6.00'),
            Decimal('6.00'), Decimal('6.00'),
        ], modulo_esperado=1)

    def test_curso_de_dos_semanas_reparte_saldo_sin_medios_dolares(self):
        matricula = self._matricula_con_adelanto(
            Decimal('43.00'), Decimal('10.00'), 2
        )

        self._assert_plan(matricula, Decimal('33.00'), [
            Decimal('17.00'), Decimal('16.00'),
        ])

    def test_curso_corto_90_reserva_10_en_2_semanas(self):
        matricula = self._matricula_con_adelanto(
            Decimal('90.00'), Decimal('10.00'), 2
        )

        self._assert_plan(matricula, Decimal('80.00'), [
            Decimal('40.00'), Decimal('40.00'),
        ])

    def test_curso_110_reserva_10_en_4_semanas(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('10.00'), 4
        )

        self._assert_plan(matricula, Decimal('100.00'), [
            Decimal('25.00'), Decimal('25.00'),
            Decimal('25.00'), Decimal('25.00'),
        ])

    def test_hoja_atrasada_no_concentra_todo_el_saldo(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('10.00'), 4
        )

        self._assert_plan(
            matricula,
            Decimal('100.00'),
            [
                Decimal('25.00'), Decimal('25.00'),
                Decimal('25.00'), Decimal('25.00'),
            ],
            fecha_recaudacion=date(2026, 7, 31),
        )

    def test_pago_25_marca_modulo_completo_en_curso_110(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('10.00'), 4
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 15),
            monto=Decimal('25.00'),
            tipo_pago='por_modulo',
            numero_modulo=1,
        )

        modulo_1 = matricula.desglose_pagos_por_modulo()[0]

        self.assertEqual(modulo_1['esperado'], Decimal('25.00'))
        self.assertEqual(modulo_1['pagado'], Decimal('25.00'))
        self.assertEqual(modulo_1['estado'], 'Pagado')

    def test_curso_110_adelanto_35_en_4_semanas(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('35.00'), 4
        )

        # Pagó $35: reserva $10 + módulo 1 $25. $75 ÷ 3 = $25.00 exacto.
        self._assert_plan(matricula, Decimal('75.00'), [
            Decimal('25.00'), Decimal('25.00'), Decimal('25.00'),
        ], modulo_esperado=2)

    def test_curso_110_adelanto_40_cobra_solo_saldo_del_modulo_2(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('40.00'), 4
        )

        # Pagó $40: reserva $10 + módulo 1 $25 + $5 del módulo 2.
        # La hoja cobra solo los $20 que faltan del módulo 2; no redistribuye
        # los $70 pendientes entre módulos futuros.
        self._assert_plan(matricula, Decimal('70.00'), [
            Decimal('20.00'), Decimal('25.00'), Decimal('25.00'),
        ], modulo_esperado=2)

    def test_cuota_manual_no_supera_saldo_del_modulo_vigente(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('40.00'), 4
        )
        CuotaManualRecaudacion.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 15),
            monto=Decimal('70.00'),
        )

        plan = _plan_recaudacion_matricula(
            matricula, date(2026, 7, 15)
        )

        self.assertEqual(plan['saldo_pendiente'], Decimal('70.00'))
        self.assertEqual(plan['saldo_modulo'], Decimal('20.00'))
        self.assertEqual(plan['cuota_sugerida'], Decimal('20.00'))
        self.assertTrue(plan['cuota_manual'])

    def test_abono_parcial_deja_solo_el_saldo_del_modulo_vigente(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('40.00'), 4
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 15),
            monto=Decimal('20.00'),
            tipo_pago='abono',
        )
        matricula.refresh_from_db()

        self._assert_plan(matricula, Decimal('50.00'), [
            Decimal('25.00'), Decimal('25.00'),
        ], modulo_esperado=3)

        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 22),
            monto=Decimal('20.00'),
            tipo_pago='abono',
        )
        matricula.refresh_from_db()

        # Pagó $80 acumulados: reserva $10, dos módulos de $25 y $20 del
        # módulo 3. Solo se cobran los $5 faltantes del módulo 3.
        self._assert_plan(matricula, Decimal('30.00'), [
            Decimal('5.00'), Decimal('25.00'),
        ], modulo_esperado=3)

    def test_pago_mayor_avanza_sin_redistribuir_modulos_futuros(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('10.00'), 4
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 7, 15),
            monto=Decimal('40.00'),
            tipo_pago='abono',
        )
        matricula.refresh_from_db()

        # Pagó $50: reserva $10, módulo 1 $25 y $15 del módulo 2.
        # La hoja cobra los $10 que faltan del módulo 2 y no cambia los
        # valores de los módulos 3 y 4.
        self._assert_plan(matricula, Decimal('60.00'), [
            Decimal('10.00'), Decimal('25.00'), Decimal('25.00'),
        ], modulo_esperado=2)

        for idx in range(2):
            Abono.objects.create(
                matricula=matricula,
                fecha=date(2026, 7, 22 + idx * 7),
                monto=Decimal('25.00'),
                tipo_pago='abono',
            )
        matricula.refresh_from_db()

        self._assert_plan(matricula, Decimal('10.00'), [
            Decimal('10.00'),
        ], modulo_esperado=4)

    def test_cuota_nunca_supera_saldo_pendiente(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('100.00'), 4
        )

        self._assert_plan(matricula, Decimal('10.00'), [
            Decimal('10.00'),
        ], modulo_esperado=4)

    def test_pagos_exactos_de_cuotas_terminan_en_saldo_cero(self):
        matricula = self._matricula_con_adelanto(
            Decimal('110.00'), Decimal('10.00'), 4
        )
        for idx in range(4):
            Abono.objects.create(
                matricula=matricula,
                fecha=date(2026, 7, 15 + idx),
                monto=Decimal('25.00'),
                tipo_pago='abono',
            )
        matricula.refresh_from_db()

        plan = _plan_recaudacion_matricula(matricula)
        self.assertEqual(matricula.saldo, Decimal('0.00'))
        self.assertEqual(plan['saldo_pendiente'], Decimal('0.00'))
        self.assertEqual(plan['saldo_modulo'], Decimal('0.00'))
        self.assertEqual(sum(plan['cuotas'], Decimal('0.00')), Decimal('0.00'))
        self.assertEqual(plan['cuota_sugerida'], Decimal('0.00'))


class AlertasPagoPorJornadaTests(TestCase):
    def setUp(self):
        self.sede = Sede.objects.create(nombre='Guayaquil', orden=1)
        self.estudiante = Estudiante.objects.create(
            cedula='0999999999',
            nombres='Estudiante Alertas',
        )
        self.curso_presencial = Curso.objects.create(
            nombre='Curso Presencial Alertas',
            ofrece_presencial=True,
            valor_presencial=Decimal('80.00'),
            numero_modulos=4,
        )
        self.jornada_presencial = JornadaCurso.objects.create(
            curso=self.curso_presencial,
            modalidad='presencial',
            descripcion='otros',
            descripcion_otros='Miércoles',
            fecha_inicio=date(2026, 7, 1),
            sede=self.sede,
        )
        self.matricula_presencial = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso_presencial,
            jornada=self.jornada_presencial,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 6, 30),
            valor_curso=Decimal('80.00'),
            valor_pagado=Decimal('10.00'),
        )

    def _alerta_de(self, matricula, hoy):
        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = hoy
            alertas = _calcular_alertas_pago()
        return next(
            alerta for alerta in alertas
            if alerta['matricula'].pk == matricula.pk
        )

    def _crear_matricula_presencial(self, numero_modulos):
        curso = Curso.objects.create(
            nombre=f'Curso Alertas {numero_modulos} Módulos',
            ofrece_presencial=True,
            valor_presencial=Decimal(numero_modulos * 20),
            numero_modulos=numero_modulos,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='presencial',
            descripcion='otros',
            descripcion_otros='Miércoles',
            fecha_inicio=date(2026, 7, 1),
            sede=self.sede,
        )
        return Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 6, 30),
            valor_curso=Decimal(numero_modulos * 20),
            valor_pagado=Decimal('10.00'),
        )

    def test_calendario_presencial_se_ancla_al_inicio_de_jornada(self):
        calendario = _calendario_vencimientos(self.matricula_presencial)

        self.assertEqual(
            [calendario[k][0] for k in range(1, 5)],
            [
                date(2026, 7, 1),
                date(2026, 7, 8),
                date(2026, 7, 15),
                date(2026, 7, 22),
            ],
        )

    def test_alerta_presencial_mantiene_primer_modulo_hasta_pagarlo(self):
        alerta = self._alerta_de(
            self.matricula_presencial,
            date(2026, 7, 8),
        )

        self.assertEqual(alerta['numero_modulo'], 1)
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 7, 1))
        self.assertEqual(alerta['dias_atraso'], 7)
        self.assertEqual(alerta['saldo_total'], Decimal('70.00'))
        self.assertEqual(alerta['saldo_m1'], Decimal('18.00'))

    def test_pago_modulo_oculta_alerta_hasta_siguiente_fecha_semanal(self):
        # $10 de reserva + $18 del Módulo 1.
        self.matricula_presencial.valor_pagado = Decimal('28.00')
        self.matricula_presencial.save(update_fields=['valor_pagado'])

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 7, 7)
            ids_con_alerta = {
                alerta['matricula'].pk
                for alerta in _calcular_alertas_pago()
            }
        self.assertNotIn(self.matricula_presencial.pk, ids_con_alerta)

        alerta = self._alerta_de(
            self.matricula_presencial,
            date(2026, 7, 8),
        )
        self.assertEqual(alerta['numero_modulo'], 2)
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 7, 8))
        self.assertEqual(alerta['dias_atraso'], 0)
        self.assertEqual(alerta['saldo_m1'], Decimal('18.00'))

    def test_alerta_identifica_pago_de_recuperacion(self):
        abono = Abono.objects.create(
            matricula=self.matricula_presencial,
            fecha=date(2026, 7, 7),
            monto=Decimal('5.00'),
            tipo_pago='recuperacion',
            numero_modulo=1,
            cuenta_para_saldo=False,
            metodo='efectivo',
        )

        alerta = self._alerta_de(
            self.matricula_presencial,
            date(2026, 7, 8),
        )

        self.assertTrue(alerta['recuperacion_pagada'])
        self.assertEqual(alerta['recuperacion_modulos_label'], 'Módulo 1')
        self.assertEqual(alerta['recuperacion_fecha'], date(2026, 7, 7))
        self.assertEqual(alerta['recuperacion_recibo'], abono.numero_recibo)

    def test_dashboard_ofrece_filtro_detallado_por_curso(self):
        usuario = User.objects.create_superuser(
            username='admin_alertas_curso',
            password='clave-segura',
        )
        self.client.force_login(usuario)

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 7, 8)
            response = self.client.get(reverse('academia:bienvenida'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="gm-course-filter"')
        self.assertContains(response, 'data-curso-id="%s"' % self.curso_presencial.pk)
        self.assertContains(response, self.curso_presencial.nombre)
        self.assertEqual(response.context['alertas_cursos'], [{
            'id': self.curso_presencial.pk,
            'nombre': self.curso_presencial.nombre,
            'cantidad': 1,
        }])

    def test_dashboard_busqueda_de_alertas_ignora_tildes_y_mayusculas(self):
        usuario = User.objects.create_superuser(
            username='admin_alertas_busqueda',
            password='clave-segura',
        )
        self.estudiante.nombres = 'Osmár Alertas'
        self.estudiante.save(update_fields=['nombres'])
        self.client.force_login(usuario)

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 7, 8)
            response = self.client.get(reverse('academia:bienvenida'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'normalizarBusquedaGm')
        self.assertContains(response, ".normalize('NFD')")
        self.assertContains(response, r'[\u0300-\u036f]')
        self.assertContains(response, 'Osmár Alertas')

    def test_fecha_matricula_no_cambia_el_calendario_de_la_jornada(self):
        segunda_matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso_presencial,
            jornada=self.jornada_presencial,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 5, 10),
            valor_curso=Decimal('80.00'),
            valor_pagado=Decimal('10.00'),
        )

        alerta_original = self._alerta_de(
            self.matricula_presencial,
            date(2026, 7, 15),
        )
        alerta_anticipada = self._alerta_de(
            segunda_matricula,
            date(2026, 7, 15),
        )

        self.assertNotEqual(
            self.matricula_presencial.fecha_matricula,
            segunda_matricula.fecha_matricula,
        )
        self.assertEqual(alerta_original['numero_modulo'], 1)
        self.assertEqual(
            alerta_original['numero_modulo'],
            alerta_anticipada['numero_modulo'],
        )
        self.assertEqual(
            alerta_original['fecha_vencimiento'],
            alerta_anticipada['fecha_vencimiento'],
        )

    def test_matricula_anticipada_no_alerta_antes_de_iniciar_jornada(self):
        curso = Curso.objects.create(
            nombre='Curso con Jornada Futura',
            ofrece_presencial=True,
            valor_presencial=Decimal('80.00'),
            numero_modulos=4,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='presencial',
            descripcion='otros',
            descripcion_otros='Lunes',
            fecha_inicio=date(2026, 8, 10),
            sede=self.sede,
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 6, 1),
            valor_curso=Decimal('80.00'),
            valor_pagado=Decimal('10.00'),
        )

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 7, 29)
            matriculas_con_alerta = {
                alerta['matricula'].pk
                for alerta in _calcular_alertas_pago()
            }

        self.assertNotIn(matricula.pk, matriculas_con_alerta)

    def test_presencial_no_alerta_un_dia_antes_y_aparece_en_fecha_exacta(self):
        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 6, 30)
            ids_con_alerta = {
                alerta['matricula'].pk
                for alerta in _calcular_alertas_pago()
            }
        self.assertNotIn(self.matricula_presencial.pk, ids_con_alerta)

        alerta = self._alerta_de(
            self.matricula_presencial,
            date(2026, 7, 1),
        )
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 7, 1))
        self.assertEqual(alerta['dias_atraso'], 0)

    def test_alerta_presencial_se_detiene_en_el_ultimo_modulo(self):
        # Cubre reserva y los tres primeros módulos; queda pendiente el cuarto.
        self.matricula_presencial.valor_pagado = Decimal('63.00')
        self.matricula_presencial.save(update_fields=['valor_pagado'])
        alerta = self._alerta_de(
            self.matricula_presencial,
            date(2026, 7, 29),
        )

        self.assertEqual(alerta['numero_modulo'], 4)
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 7, 22))
        self.assertEqual(alerta['dias_atraso'], 7)

    def test_curso_de_dos_modulos_genera_exactamente_dos_semanas(self):
        matricula = self._crear_matricula_presencial(numero_modulos=2)
        matricula.valor_pagado = Decimal('25.00')
        matricula.save(update_fields=['valor_pagado'])

        calendario = _calendario_vencimientos(matricula)
        alerta = self._alerta_de(matricula, date(2026, 7, 29))

        self.assertEqual(len(calendario), 2)
        self.assertEqual(calendario[1][0], date(2026, 7, 1))
        self.assertEqual(calendario[2][0], date(2026, 7, 8))
        self.assertEqual(alerta['numero_modulo'], 2)
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 7, 8))

    def test_curso_de_cinco_modulos_genera_exactamente_cinco_semanas(self):
        matricula = self._crear_matricula_presencial(numero_modulos=5)
        matricula.valor_pagado = Decimal('82.00')
        matricula.save(update_fields=['valor_pagado'])

        calendario = _calendario_vencimientos(matricula)
        alerta = self._alerta_de(matricula, date(2026, 7, 29))

        self.assertEqual(len(calendario), 5)
        self.assertEqual(
            [calendario[k][0] for k in range(1, 6)],
            [
                date(2026, 7, 1),
                date(2026, 7, 8),
                date(2026, 7, 15),
                date(2026, 7, 22),
                date(2026, 7, 29),
            ],
        )
        self.assertEqual(alerta['numero_modulo'], 5)
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 7, 29))

    def test_alerta_online_aplica_segundo_modulo_a_los_siete_dias(self):
        curso = Curso.objects.create(
            nombre='Curso Online Alertas',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('90.00'),
            numero_modulos_online=2,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='otros',
            descripcion_otros='Online',
            fecha_inicio=date(2026, 7, 10),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 7, 9),
            valor_curso=Decimal('90.00'),
            valor_pagado=Decimal('45.00'),
        )

        # La reserva y el Módulo 1 ya están cubiertos: no debe reaparecer una
        # alerta anticipada por ese primer módulo.
        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 7, 9)
            ids_con_alerta = {
                alerta['matricula'].pk
                for alerta in _calcular_alertas_pago()
            }
        self.assertNotIn(matricula.pk, ids_con_alerta)

        alerta = self._alerta_de(matricula, date(2026, 7, 17))

        self.assertEqual(alerta['numero_modulo'], 2)
        self.assertEqual(alerta['hito'], 'modulo')
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 7, 17))

    def test_alerta_online_aparece_un_dia_antes_del_inicio(self):
        curso = Curso.objects.create(
            nombre='Curso Online Inicio Anticipado',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('60.00'),
            numero_modulos_online=2,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='otros',
            descripcion_otros='Online',
            fecha_inicio=date(2026, 8, 15),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 8, 10),
            valor_curso=Decimal('60.00'),
            valor_pagado=Decimal('10.00'),
        )

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 8, 13)
            ids_con_alerta = {
                alerta['matricula'].pk
                for alerta in _calcular_alertas_pago()
            }
        self.assertNotIn(matricula.pk, ids_con_alerta)

        alerta = self._alerta_de(matricula, date(2026, 8, 14))
        self.assertEqual(alerta['numero_modulo'], 1)
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 8, 14))
        self.assertEqual(alerta['dias_atraso'], 0)

    def test_online_pagado_no_alerta_un_dia_antes_del_inicio(self):
        curso = Curso.objects.create(
            nombre='Curso Online Pagado',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('60.00'),
            numero_modulos_online=2,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='otros',
            descripcion_otros='Online',
            fecha_inicio=date(2026, 8, 15),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 8, 10),
            valor_curso=Decimal('60.00'),
            valor_pagado=Decimal('60.00'),
        )

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 8, 14)
            ids_con_alerta = {
                alerta['matricula'].pk
                for alerta in _calcular_alertas_pago()
            }

        self.assertNotIn(matricula.pk, ids_con_alerta)

    def test_pago_cada_dos_semanas_alerta_el_17_para_pago_del_18(self):
        curso = Curso.objects.create(
            nombre='Asistente Contable Pagos Cada Dos Semanas',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('90.00'),
            numero_modulos_online=2,
            pagos_cada_dos_semanas=True,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 8, 4),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 8, 3),
            valor_curso=Decimal('90.00'),
            valor_pagado=Decimal('50.00'),
        )

        calendario = _calendario_vencimientos(matricula)
        self.assertEqual(calendario[1][0], date(2026, 8, 3))
        self.assertEqual(calendario[2][0], date(2026, 8, 18))

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 8, 16)
            ids_con_alerta = {
                alerta['matricula'].pk
                for alerta in _calcular_alertas_pago()
            }
        self.assertNotIn(matricula.pk, ids_con_alerta)

        alerta = self._alerta_de(matricula, date(2026, 8, 17))
        self.assertEqual(alerta['numero_modulo'], 2)
        self.assertEqual(alerta['fecha_alerta'], date(2026, 8, 17))
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 8, 17))
        self.assertEqual(alerta['fecha_pago'], date(2026, 8, 18))
        self.assertEqual(alerta['dias_atraso'], 0)
        self.assertTrue(alerta['pagos_cada_dos_semanas'])

    def test_pago_registrado_el_17_retira_alerta_hasta_la_siguiente_cuota(self):
        curso = Curso.objects.create(
            nombre='Curso Online Quincenal con Cuota Futura',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('120.00'),
            numero_modulos_online=3,
            pagos_cada_dos_semanas=True,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 8, 4),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 8, 3),
            valor_curso=Decimal('120.00'),
            valor_pagado=Decimal('0.00'),
        )
        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 3),
            monto=Decimal('40.00'),
            tipo_pago='solo_modulo',
            numero_modulo=1,
            cuenta_para_saldo=True,
        )

        alerta = self._alerta_de(matricula, date(2026, 8, 17))
        self.assertEqual(alerta['numero_modulo'], 2)
        self.assertEqual(alerta['fecha_pago'], date(2026, 8, 18))

        Abono.objects.create(
            matricula=matricula,
            fecha=date(2026, 8, 17),
            monto=Decimal('40.00'),
            tipo_pago='solo_modulo',
            numero_modulo=2,
            cuenta_para_saldo=True,
        )
        matricula.refresh_from_db()
        self.assertEqual(matricula.valor_pagado, Decimal('80.00'))
        self.assertEqual(matricula.saldo, Decimal('40.00'))

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 8, 17)
            ids_con_alerta = {
                item['matricula'].pk
                for item in _calcular_alertas_pago()
            }
        self.assertNotIn(matricula.pk, ids_con_alerta)

        siguiente_alerta = self._alerta_de(matricula, date(2026, 8, 31))
        self.assertEqual(siguiente_alerta['numero_modulo'], 3)
        self.assertEqual(siguiente_alerta['fecha_pago'], date(2026, 9, 1))

    def test_jornada_del_18_alerta_el_31_para_pago_del_01_de_septiembre(self):
        curso = Curso.objects.create(
            nombre='Curso Quincenal Segunda Jornada',
            ofrece_online=True,
            valor_online=Decimal('90.00'),
            numero_modulos_online=2,
            pagos_cada_dos_semanas=True,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 8, 18),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_modulo_1',
            forma_pago='abono_modulo',
            fecha_matricula=date(2026, 8, 17),
            valor_curso=Decimal('90.00'),
            valor_pagado=Decimal('50.00'),
        )

        calendario = _calendario_vencimientos(matricula)
        self.assertEqual(calendario[2][0], date(2026, 9, 1))

        alerta = self._alerta_de(matricula, date(2026, 8, 31))

        self.assertEqual(alerta['numero_modulo'], 2)
        self.assertEqual(alerta['fecha_alerta'], date(2026, 8, 31))
        self.assertEqual(alerta['fecha_vencimiento'], date(2026, 8, 31))
        self.assertEqual(alerta['fecha_pago'], date(2026, 9, 1))
        self.assertEqual(alerta['dias_atraso'], 0)

    def test_saldo_cero_elimina_alerta_definitivamente(self):
        self.matricula_presencial.valor_pagado = Decimal('80.00')
        self.matricula_presencial.save(update_fields=['valor_pagado'])

        with patch('academia.views_pagos.date') as fecha_mock:
            fecha_mock.today.return_value = date(2026, 8, 1)
            ids_con_alerta = {
                alerta['matricula'].pk
                for alerta in _calcular_alertas_pago()
            }

        self.assertNotIn(self.matricula_presencial.pk, ids_con_alerta)

    def test_ciclo_corto_online_genera_dos_cuotas_semanales(self):
        curso = Curso.objects.create(
            nombre='Curso Corto Online',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('35.00'),
            numero_modulos_online=2,
            es_ciclo_corto=True,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 8, 4),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 30),
            valor_curso=Decimal('35.00'),
            valor_pagado=Decimal('10.00'),
        )

        calendario = _calendario_vencimientos(matricula)
        plan = _plan_recaudacion_matricula(matricula, date(2026, 8, 3))
        alerta = self._alerta_de(matricula, date(2026, 8, 4))

        self.assertEqual(calendario[1][0], date(2026, 8, 3))
        self.assertEqual(calendario[2][0], date(2026, 8, 10))
        self.assertEqual(plan['cuotas'], [Decimal('13.00'), Decimal('12.00')])
        self.assertEqual(alerta['saldo_m1'], Decimal('13.00'))
        self.assertTrue(alerta['es_ciclo_corto'])
        self.assertEqual(
            alerta['ciclo_corto_label'],
            'Ciclo corto · 2 pagos por módulo',
        )

    def test_ciclo_corto_online_con_pago_unico_genera_un_solo_cobro(self):
        curso = Curso.objects.create(
            nombre='Curso Corto Online Pago Único',
            ofrece_presencial=False,
            ofrece_online=True,
            valor_online=Decimal('35.00'),
            numero_modulos_online=2,
            es_ciclo_corto=True,
            pago_unico_online=True,
        )
        jornada = JornadaCurso.objects.create(
            curso=curso,
            modalidad='online',
            descripcion='mar_mie_jue',
            fecha_inicio=date(2026, 8, 4),
        )
        matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=curso,
            jornada=jornada,
            modalidad='online',
            tipo_matricula='reserva_abono',
            forma_pago='abono',
            fecha_matricula=date(2026, 7, 30),
            valor_curso=Decimal('35.00'),
            valor_pagado=Decimal('10.00'),
        )

        calendario = _calendario_vencimientos(matricula)
        plan = _plan_recaudacion_matricula(matricula, date(2026, 8, 3))
        alerta = self._alerta_de(matricula, date(2026, 8, 4))

        self.assertEqual(
            calendario,
            {1: (date(2026, 8, 3), 'pago_unico')},
        )
        self.assertEqual(plan['cuotas'], [Decimal('25.00')])
        self.assertEqual(plan['cuota_sugerida'], Decimal('25.00'))
        self.assertEqual(alerta['hito'], 'pago_unico')
        self.assertEqual(alerta['saldo_m1'], Decimal('25.00'))
        self.assertEqual(
            alerta['ciclo_corto_label'],
            'Ciclo corto · Un solo pago',
        )


class PagosFiltroRecuperacionTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_superuser(
            username='admin_filtro_recuperacion',
            password='clave-segura',
        )
        self.client.force_login(self.usuario)
        self.curso = Curso.objects.create(
            nombre='Curso Filtro Recuperación',
            ofrece_presencial=True,
            valor_presencial=Decimal('80.00'),
            numero_modulos=4,
        )
        self.jornada = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='sabados_intensivos',
            fecha_inicio=date(2026, 8, 1),
        )
        estudiante_recuperacion = Estudiante.objects.create(
            cedula='0911111111',
            nombres='Estudiante Con Recuperación',
        )
        estudiante_normal = Estudiante.objects.create(
            cedula='0922222222',
            nombres='Estudiante Pago Normal',
        )
        self.matricula_recuperacion = Matricula.objects.create(
            estudiante=estudiante_recuperacion,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            fecha_matricula=date(2026, 8, 1),
            valor_curso=Decimal('80.00'),
            valor_pagado=Decimal('0.00'),
        )
        self.matricula_normal = Matricula.objects.create(
            estudiante=estudiante_normal,
            curso=self.curso,
            jornada=self.jornada,
            modalidad='presencial',
            tipo_matricula='reserva_abono',
            fecha_matricula=date(2026, 8, 1),
            valor_curso=Decimal('80.00'),
            valor_pagado=Decimal('0.00'),
        )
        Abono.objects.create(
            matricula=self.matricula_recuperacion,
            fecha=date(2026, 8, 1),
            monto=Decimal('12.00'),
            tipo_pago='recuperacion',
            numero_modulo=1,
            cuenta_para_saldo=False,
            metodo='efectivo',
        )
        Abono.objects.create(
            matricula=self.matricula_normal,
            fecha=date(2026, 8, 1),
            monto=Decimal('10.00'),
            tipo_pago='abono',
            cuenta_para_saldo=True,
            metodo='efectivo',
        )

    def test_estado_recuperacion_filtra_solo_matriculas_con_ese_pago(self):
        response = self.client.get(
            reverse('academia:pagos_lista'),
            {'estado': 'Recuperacion'},
        )

        self.assertEqual(response.status_code, 200)
        matriculas = response.context['matriculas']
        self.assertEqual(
            [matricula.pk for matricula in matriculas],
            [self.matricula_recuperacion.pk],
        )
        self.assertTrue(matriculas[0].tiene_pago_recuperacion)
        self.assertEqual(
            matriculas[0].total_pago_recuperacion,
            Decimal('12.00'),
        )
        self.assertContains(response, 'value="Recuperacion" selected')
        self.assertContains(response, '✱ Recuperación')
        self.assertNotContains(response, 'Estudiante Pago Normal')


class JornadaOrdenYFiltroTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_superuser(
            username='admin_orden_jornadas',
            email='orden@example.com',
            password='clave-segura',
        )
        self.client.force_login(self.usuario)
        self.curso = Curso.objects.create(
            nombre='Curso Orden Jornadas',
            ofrece_presencial=True,
            valor_presencial=Decimal('90.00'),
        )
        self.julio = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='mar_jue',
            fecha_inicio=date(2026, 7, 19),
            activo=True,
        )
        self.agosto = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='domingos_intensivos',
            fecha_inicio=date(2026, 8, 30),
            activo=False,
        )
        self.septiembre_primera = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='sabados_intensivos',
            fecha_inicio=date(2026, 9, 12),
            activo=True,
        )
        self.septiembre_nueva = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='otros',
            descripcion_otros='Jornada nueva',
            fecha_inicio=date(2026, 9, 12),
            activo=True,
        )
        self.url = reverse(
            'academia:curso_jornadas', args=[self.curso.pk],
        )

    def test_general_ordena_fecha_mas_nueva_primero_y_desempata_por_registro(self):
        response = self.client.get(self.url, {'modalidad': 'presencial'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [j.pk for j in response.context['jornadas_presencial']],
            [
                self.septiembre_nueva.pk,
                self.septiembre_primera.pk,
                self.agosto.pk,
                self.julio.pk,
            ],
        )
        self.assertEqual(response.context['estado_jornadas'], 'general')
        self.assertEqual(response.context['resumen_jornadas']['general'], 4)

    def test_filtros_separan_activas_e_inactivas_sin_perder_el_orden(self):
        activas = self.client.get(
            self.url, {'modalidad': 'presencial', 'estado': 'activa'},
        )
        inactivas = self.client.get(
            self.url, {'modalidad': 'presencial', 'estado': 'inactiva'},
        )

        self.assertEqual(
            [j.pk for j in activas.context['jornadas_presencial']],
            [self.septiembre_nueva.pk, self.septiembre_primera.pk, self.julio.pk],
        )
        self.assertEqual(
            [j.pk for j in inactivas.context['jornadas_presencial']],
            [self.agosto.pk],
        )
        self.assertEqual(activas.context['resumen_jornadas']['activas'], 3)
        self.assertEqual(inactivas.context['resumen_jornadas']['inactivas'], 1)
        self.assertContains(activas, 'aria-label="Filtrar jornadas por estado"')
        self.assertContains(activas, 'Activas <span>3</span>')
        self.assertContains(inactivas, 'Inactivas <span>1</span>')

    def test_filtro_invalido_vuelve_a_general(self):
        response = self.client.get(
            self.url, {'modalidad': 'presencial', 'estado': 'desconocido'},
        )

        self.assertEqual(response.context['estado_jornadas'], 'general')
        self.assertEqual(len(response.context['jornadas_presencial']), 4)


class JornadaFeriadoTests(TestCase):
    def setUp(self):
        self.usuario = User.objects.create_superuser(
            username='admin_feriados',
            email='admin@example.com',
            password='clave-segura',
        )
        self.client.force_login(self.usuario)
        self.curso = Curso.objects.create(
            nombre='Curso Intensivo Feriados',
            ofrece_presencial=True,
            valor_presencial=Decimal('110.00'),
        )
        self.sabado = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='sabados_intensivos',
            fecha_inicio=date(2026, 8, 1),
        )
        self.domingo = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='domingos_intensivos',
            fecha_inicio=date(2026, 8, 30),
        )
        self.regular = JornadaCurso.objects.create(
            curso=self.curso,
            modalidad='presencial',
            descripcion='mar_jue',
            fecha_inicio=date(2026, 8, 4),
        )
        self.estudiante = Estudiante.objects.create(
            cedula='0912345678',
            nombres='Estudiante Feriado',
        )
        self.matricula = Matricula.objects.create(
            estudiante=self.estudiante,
            curso=self.curso,
            jornada=self.sabado,
            modalidad='presencial',
            fecha_matricula=date(2026, 7, 25),
            valor_curso=Decimal('110.00'),
            valor_pagado=Decimal('10.00'),
        )

    def _url_feriado(self, jornada):
        return reverse(
            'academia:jornada_marcar_feriado',
            args=[self.curso.pk, jornada.pk],
        )

    def test_panel_muestra_check_solo_en_jornadas_intensivas_de_fin_de_semana(self):
        response = self.client.get(
            reverse('academia:curso_jornadas', args=[self.curso.pk]),
            {'modalidad': 'presencial'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Día no laboral')
        self.assertContains(response, 'Mover al 08/08/2026')
        self.assertContains(response, 'Mover al 06/09/2026')
        self.assertContains(response, self._url_feriado(self.sabado))
        self.assertContains(response, self._url_feriado(self.domingo))
        self.assertNotContains(response, self._url_feriado(self.regular))
        self.assertEqual(response.content.count(b'class="feriado-form"'), 2)

    def test_feriado_mueve_jornada_y_todos_sus_matriculados_siete_dias(self):
        momento = timezone.make_aware(datetime(2026, 8, 1, 9, 15))

        with patch('academia.views.timezone.now', return_value=momento):
            response = self.client.post(
                self._url_feriado(self.sabado),
                {'modalidad_activa': 'presencial'},
            )

        self.assertRedirects(
            response,
            f"{reverse('academia:curso_jornadas', args=[self.curso.pk])}?modalidad=presencial",
            fetch_redirect_response=False,
        )
        self.sabado.refresh_from_db()
        self.matricula.refresh_from_db()
        self.assertEqual(self.sabado.fecha_inicio, date(2026, 8, 8))
        self.assertEqual(self.sabado.feriado_aplicado_en, momento)
        self.assertEqual(self.matricula.jornada_id, self.sabado.pk)
        self.assertEqual(self.matricula.jornada.fecha_inicio, date(2026, 8, 8))

    def test_feriado_domingo_pasa_correctamente_al_mes_siguiente(self):
        response = self.client.post(self._url_feriado(self.domingo))

        self.assertEqual(response.status_code, 302)
        self.domingo.refresh_from_db()
        self.assertEqual(self.domingo.fecha_inicio, date(2026, 9, 6))
        self.assertEqual(self.domingo.fecha_inicio.weekday(), 6)

    def test_check_se_bloquea_y_se_reinicia_exactamente_a_las_24_horas(self):
        primer_click = timezone.make_aware(datetime(2026, 8, 1, 8, 0))

        with patch('academia.views.timezone.now', return_value=primer_click):
            self.client.post(self._url_feriado(self.sabado))

        with patch(
            'academia.views.timezone.now',
            return_value=primer_click + timedelta(hours=23, minutes=59),
        ):
            response_bloqueada = self.client.post(
                self._url_feriado(self.sabado),
                follow=True,
            )

        self.sabado.refresh_from_db()
        self.assertEqual(self.sabado.fecha_inicio, date(2026, 8, 8))
        self.assertContains(response_bloqueada, 'ya fue trasladada por feriado')
        self.assertContains(response_bloqueada, 'Feriado aplicado')
        self.assertContains(response_bloqueada, 'checked disabled')

        with patch(
            'academia.views.timezone.now',
            return_value=primer_click + timedelta(hours=24),
        ):
            self.client.post(self._url_feriado(self.sabado))

        self.sabado.refresh_from_db()
        self.assertEqual(self.sabado.fecha_inicio, date(2026, 8, 15))

    def test_endpoint_rechaza_jornadas_que_no_son_sabado_o_domingo_intensivo(self):
        response = self.client.post(self._url_feriado(self.regular), follow=True)

        self.regular.refresh_from_db()
        self.assertEqual(self.regular.fecha_inicio, date(2026, 8, 4))
        self.assertIsNone(self.regular.feriado_aplicado_en)
        self.assertContains(
            response,
            'solo se aplica a Sábados Intensivos o Domingos Intensivos',
        )
